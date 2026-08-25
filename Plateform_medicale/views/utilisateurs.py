"""
Gestion des utilisateurs : liste, exports, import Excel, CRUD, activation,
réinitialisation mot de passe.
"""

import csv
import datetime
import unicodedata

import openpyxl
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from ..forms import (
    UtilisateurCreationForm,
    UtilisateurModificationForm,
    generer_mot_de_passe,
    lier_fiche_medecin,
    lier_fiche_pharmacien,
)
from ..models import (
    JournalActivite,
    Medecin,
    Patient,
    Pharmacien,
    PlanCouverture,
    Prestataire,
    User,
    valider_telephone,
)
from ..services.onboarding import construire_bilan_onboarding, envoyer_activation_utilisateur
from .utils import _cellule_csv, _paginer, _trier, admin_required, journaliser


def _normaliser_texte_import(valeur):
    """Normalise une valeur de cellule pour une comparaison insensible aux accents/majuscules."""
    texte = "" if valeur is None else str(valeur).strip()
    texte = unicodedata.normalize("NFKD", texte).encode("ascii", "ignore").decode("ascii")
    return texte.upper()


def _filtrer_utilisateurs(request):
    """Filtres partages entre la liste et l'export Excel des utilisateurs."""
    utilisateurs = User.objects.all()

    role = request.GET.get("role", "")
    statut = request.GET.get("statut", "")
    recherche = request.GET.get("q", "").strip()

    if role:
        utilisateurs = utilisateurs.filter(role=role)
    if statut == "actif":
        utilisateurs = utilisateurs.filter(is_active=True)
    elif statut == "inactif":
        utilisateurs = utilisateurs.filter(is_active=False)
    if recherche:
        utilisateurs = utilisateurs.filter(
            Q(email__icontains=recherche)
            | Q(first_name__icontains=recherche)
            | Q(last_name__icontains=recherche)
        )

    utilisateurs = _trier(request, utilisateurs, ["last_name", "email", "role", "is_active"], ["last_name", "first_name"])

    return utilisateurs, {"role": role, "statut": statut, "recherche": recherche}


@admin_required
def liste_utilisateurs(request):
    utilisateurs, filtres = _filtrer_utilisateurs(request)
    tous = User.objects.all()
    contexte = {
        "utilisateurs": _paginer(request, utilisateurs),
        "roles": User.Role.choices,
        "role_selectionne": filtres["role"],
        "statut_selectionne": filtres["statut"],
        "recherche": filtres["recherche"],
        "total_utilisateurs": tous.count(),
        "total_medecins": tous.filter(role=User.Role.MEDECIN).count(),
        "total_assures": tous.filter(role=User.Role.ASSURE).count(),
        "total_pharmaciens_admins": tous.filter(role__in=[User.Role.PHARMACIEN, User.Role.ADMIN]).count(),
    }
    return render(request, "liste_utilisateurs.html", contexte)


@admin_required
def exporter_utilisateurs_excel(request):
    utilisateurs, _ = _filtrer_utilisateurs(request)

    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Utilisateurs"
    entetes = ["Email", "Prenom", "Nom", "Telephone", "Role", "Statut", "Date de creation"]
    feuille.append(entetes)

    for utilisateur in utilisateurs:
        feuille.append([
            utilisateur.email,
            utilisateur.first_name,
            utilisateur.last_name,
            utilisateur.phone_number,
            utilisateur.get_role_display(),
            "Actif" if utilisateur.is_active else "Inactif",
            utilisateur.date_joined.strftime("%d/%m/%Y %H:%M"),
        ])

    for index, nom_colonne in enumerate(entetes, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(len(nom_colonne), 18)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="utilisateurs_santesn.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def exporter_utilisateurs_csv(request):
    utilisateurs, _ = _filtrer_utilisateurs(request)

    reponse = HttpResponse(content_type="text/csv")
    reponse["Content-Disposition"] = 'attachment; filename="utilisateurs_santesn.csv"'
    reponse.write("\ufeff")  # BOM : Excel (FR) detecte l'UTF-8 sans le confondre avec l'encodage local.
    ecrivain = csv.writer(reponse, delimiter=";")
    ecrivain.writerow(["Email", "Prenom", "Nom", "Telephone", "Role", "Statut", "Date de creation"])
    for utilisateur in utilisateurs:
        ecrivain.writerow([_cellule_csv(v) for v in [
            utilisateur.email,
            utilisateur.first_name,
            utilisateur.last_name,
            utilisateur.phone_number,
            utilisateur.get_role_display(),
            "Actif" if utilisateur.is_active else "Inactif",
            utilisateur.date_joined.strftime("%d/%m/%Y %H:%M"),
        ]])
    return reponse


COLONNES_IMPORT_UTILISATEURS = [
    "Email", "Prenom", "Nom", "Telephone", "Role",
    "Date de naissance", "Specialite", "Prestataire", "Plan de couverture",
]

_ROLES_PAR_LIBELLE_IMPORT = {}
for _valeur_role, _label_role in User.Role.choices:
    _ROLES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_valeur_role)] = _valeur_role
    _ROLES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_label_role)] = _valeur_role


def _analyser_ligne_import_utilisateurs(numero_ligne, valeurs):
    """
    Valide une ligne du fichier d'import (voir COLONNES_IMPORT_UTILISATEURS).

    Retourne (donnees, None) si la ligne est valide, ou (None, message_erreur)
    sinon. Ne touche jamais la base : l'import est valide en integralite
    avant toute creation (regle "tout ou rien").
    """
    valeurs = (tuple(valeurs) + (None,) * len(COLONNES_IMPORT_UTILISATEURS))[:len(COLONNES_IMPORT_UTILISATEURS)]
    email, prenom, nom, telephone, role_brut, date_naissance_brute, specialite, prestataire_brut, plan_brut = valeurs

    email = (email or "").strip()
    prenom = (prenom or "").strip()
    nom = (nom or "").strip()
    telephone = (telephone or "").strip() if telephone else ""

    if not email or not prenom or not nom:
        return None, f"Ligne {numero_ligne} : email, prenom et nom sont obligatoires."

    role = _ROLES_PAR_LIBELLE_IMPORT.get(_normaliser_texte_import(role_brut))
    if not role:
        return None, (
            f"Ligne {numero_ligne} : role '{role_brut}' inconnu "
            "(attendu : Administrateur, Assure, Medecin ou Pharmacien)."
        )

    if telephone:
        try:
            valider_telephone(telephone)
        except ValidationError:
            return None, f"Ligne {numero_ligne} : numero de telephone invalide."

    prestataire = None
    prestataire_nom = (prestataire_brut or "").strip() if prestataire_brut else ""
    if prestataire_nom:
        prestataire = Prestataire.objects.filter(nom__iexact=prestataire_nom).first()
        if not prestataire:
            return None, f"Ligne {numero_ligne} : prestataire '{prestataire_nom}' introuvable."

    plan_couverture = None
    plan_nom = (plan_brut or "").strip() if plan_brut else ""
    if plan_nom:
        plan_couverture = PlanCouverture.objects.filter(nom__iexact=plan_nom).first()
        if not plan_couverture:
            return None, f"Ligne {numero_ligne} : plan de couverture '{plan_nom}' introuvable."

    donnees = {
        "email": email,
        "prenom": prenom,
        "nom": nom,
        "telephone": telephone,
        "role": role,
        "prestataire": prestataire,
        "plan_couverture": plan_couverture,
    }

    if role == User.Role.MEDECIN:
        if not telephone:
            return None, f"Ligne {numero_ligne} : le telephone est obligatoire pour un medecin."
        specialite = (specialite or "").strip()
        if not specialite:
            return None, f"Ligne {numero_ligne} : la specialite est obligatoire pour un medecin."
        donnees["specialite"] = specialite
    elif role == User.Role.ASSURE:
        if not date_naissance_brute:
            return None, f"Ligne {numero_ligne} : la date de naissance est obligatoire pour un assure."
        if isinstance(date_naissance_brute, datetime.datetime):
            donnees["date_naissance"] = date_naissance_brute.date()
        elif isinstance(date_naissance_brute, datetime.date):
            donnees["date_naissance"] = date_naissance_brute
        else:
            try:
                donnees["date_naissance"] = datetime.datetime.strptime(
                    str(date_naissance_brute).strip(), "%d/%m/%Y"
                ).date()
            except ValueError:
                return None, (
                    f"Ligne {numero_ligne} : date de naissance invalide (format attendu JJ/MM/AAAA)."
                )

    return donnees, None


def _creer_comptes_import_utilisateurs(lignes_validees, request=None):
    """Cree en une transaction tous les comptes (et fiches metier) valides par l'import et genere l'onboarding."""
    resultats = []
    with transaction.atomic():
        for donnees in lignes_validees:
            utilisateur = User.objects.create_user(
                email=donnees["email"],
                role=donnees["role"],
                first_name=donnees["prenom"],
                last_name=donnees["nom"],
                phone_number=donnees["telephone"],
            )
            if donnees["role"] == User.Role.MEDECIN:
                Medecin.objects.create(
                    user=utilisateur,
                    nom=donnees["nom"],
                    prenom=donnees["prenom"],
                    specialite=donnees["specialite"],
                    telephone=donnees["telephone"],
                    email=donnees["email"],
                    prestataire=donnees["prestataire"],
                )
            elif donnees["role"] == User.Role.PHARMACIEN:
                Pharmacien.objects.create(user=utilisateur, prestataire=donnees["prestataire"])
            elif donnees["role"] == User.Role.ASSURE:
                Patient.objects.create(
                    user=utilisateur,
                    nom=donnees["nom"],
                    prenom=donnees["prenom"],
                    date_naissance=donnees["date_naissance"],
                    telephone=donnees["telephone"],
                    type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL,
                    plan_couverture=donnees["plan_couverture"],
                )

            # Envoi du lien d'activation sécurisé (Email / WhatsApp)
            statut = envoyer_activation_utilisateur(utilisateur, request=request)
            bilan = statut.get("bilan") or construire_bilan_onboarding(statut, utilisateur, action="creation")
            resultats.append({
                "numero_ligne": donnees.get("numero_ligne", "-"),
                "email": utilisateur.email,
                "nom_complet": f"{utilisateur.first_name} {utilisateur.last_name}",
                "role": utilisateur.get_role_display(),
                "compte_cree": True,
                "email_envoye": statut["email_envoye"],
                "email_erreur": statut["email_erreur"],
                "whatsapp_envoye": statut["whatsapp_envoye"],
                "whatsapp_statut": statut["whatsapp_statut"],
                "whatsapp_message": statut["whatsapp_message"],
                "bilan": bilan,
            })
    return resultats


@admin_required
def importer_utilisateurs_excel(request):
    """
    Creation en masse de comptes (Assure principal / Medecin / Pharmacien /
    Administrateur) depuis un fichier Excel : voir COLONNES_IMPORT_UTILISATEURS.

    Regle "tout ou rien" : la moindre ligne invalide bloque tout l'import
    (aucun compte cree), pour eviter un import partiel difficile a auditer.
    """
    erreurs = []
    if request.method == "POST":
        fichier = request.FILES.get("fichier")
        if not fichier:
            erreurs.append("Choisissez un fichier Excel (.xlsx) a importer.")
        else:
            try:
                classeur = openpyxl.load_workbook(fichier, data_only=True)
            except Exception:
                erreurs.append(
                    "Fichier illisible : verifiez qu'il s'agit bien d'un fichier Excel (.xlsx) valide."
                )
            else:
                feuille = classeur.active
                entetes = next(feuille.iter_rows(min_row=1, max_row=1, values_only=True), ())
                entetes_normalisees = [_normaliser_texte_import(entete) for entete in entetes]
                entetes_attendues = [_normaliser_texte_import(colonne) for colonne in COLONNES_IMPORT_UTILISATEURS]

                if entetes_normalisees[:len(entetes_attendues)] != entetes_attendues:
                    erreurs.append(
                        "En-tetes de colonnes invalides : utilisez le modele telechargeable ci-dessous."
                    )
                else:
                    lignes_brutes = [
                        (numero, valeurs)
                        for numero, valeurs in enumerate(
                            feuille.iter_rows(min_row=2, values_only=True), start=2
                        )
                        if valeurs and not all(valeur in (None, "") for valeur in valeurs)
                    ]
                    if not lignes_brutes:
                        erreurs.append("Le fichier ne contient aucune ligne a importer.")
                    else:
                        donnees_valides = []
                        emails_vus = set()
                        for numero_ligne, valeurs in lignes_brutes:
                            donnees, erreur = _analyser_ligne_import_utilisateurs(numero_ligne, valeurs)
                            if erreur:
                                erreurs.append(erreur)
                                continue
                            email_normalise = donnees["email"].lower()
                            if email_normalise in emails_vus:
                                erreurs.append(
                                    f"Ligne {numero_ligne} : email '{donnees['email']}' en double dans le fichier."
                                )
                                continue
                            if User.objects.filter(email__iexact=donnees["email"]).exists():
                                erreurs.append(
                                    f"Ligne {numero_ligne} : email '{donnees['email']}' "
                                    "deja utilise par un compte existant."
                                )
                                continue
                            emails_vus.add(email_normalise)
                            donnees["numero_ligne"] = numero_ligne
                            donnees_valides.append(donnees)

                        if not erreurs:
                            resultats = _creer_comptes_import_utilisateurs(donnees_valides, request=request)
                            total_crees = len(resultats)
                            total_whatsapp = sum(1 for r in resultats if r["whatsapp_envoye"])
                            total_email = sum(1 for r in resultats if r["email_envoye"])

                            journaliser(
                                request, JournalActivite.Action.IMPORT,
                                "Import d'utilisateurs",
                                f"{total_crees} compte(s) créé(s) ({total_whatsapp} WhatsApp, {total_email} Email)",
                            )
                            return render(
                                request,
                                "resultat_import_utilisateurs.html",
                                {
                                    "resultats": resultats,
                                    "total_crees": total_crees,
                                    "total_whatsapp": total_whatsapp,
                                    "total_email": total_email,
                                },
                            )

    return render(request, "importer_utilisateurs.html", {"erreurs": erreurs})


@admin_required
def telecharger_modele_import_utilisateurs(request):
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Import utilisateurs"
    feuille.append(COLONNES_IMPORT_UTILISATEURS)
    feuille.append([
        "email@domaine.com", "Prenom", "Nom", "770000000",
        "Assure / Medecin / Pharmacien / Administrateur",
        "JJ/MM/AAAA (uniquement pour un Assure)",
        "Ex: Medecine generale (uniquement pour un Medecin)",
        "Nom exact d'un prestataire existant (optionnel)",
        "Nom exact d'un plan de couverture existant (optionnel)",
    ])

    for index, nom_colonne in enumerate(COLONNES_IMPORT_UTILISATEURS, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(len(nom_colonne), 20)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="modele_import_utilisateurs.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def ajouter_utilisateur(request):
    if request.method == "POST":
        form = UtilisateurCreationForm(request.POST)
        if form.is_valid():
            utilisateur = form.save()
            # Onboarding sécurisé : génération et expédition du lien d'activation (Email / WhatsApp)
            statut_onboarding = envoyer_activation_utilisateur(utilisateur, request=request)
            journaliser(request, JournalActivite.Action.CREATION, f"Utilisateur {utilisateur.email}",
                        f"role : {utilisateur.get_role_display()}")
            bilan = statut_onboarding.get("bilan") or construire_bilan_onboarding(statut_onboarding, utilisateur, action="creation")
            return render(
                request,
                "mot_de_passe_genere.html",
                {
                    "utilisateur": utilisateur,
                    "lien_activation": statut_onboarding["lien_activation"],
                    "email_envoye": statut_onboarding["email_envoye"],
                    "whatsapp_envoye": statut_onboarding["whatsapp_envoye"],
                    "whatsapp_statut": statut_onboarding["whatsapp_statut"],
                    "whatsapp_message": statut_onboarding["whatsapp_message"],
                    "whatsapp_erreur": statut_onboarding["whatsapp_erreur"],
                    "bilan": bilan,
                    "action": "creation",
                },
            )
    else:
        form = UtilisateurCreationForm()
    return render(request, "ajouter_utilisateur.html", {"form": form})


@admin_required
def modifier_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        form = UtilisateurModificationForm(request.POST, instance=utilisateur)
        if form.is_valid():
            nouveau_role = form.cleaned_data["role"]
            if utilisateur.pk == request.user.pk and nouveau_role != request.user.role:
                form.add_error("role", "Vous ne pouvez pas modifier votre propre rôle.")
            else:
                ancien_role = utilisateur.get_role_display()
                utilisateur_modifie = form.save()
                lier_fiche_medecin(utilisateur_modifie)
                lier_fiche_pharmacien(utilisateur_modifie)
                nouveau = utilisateur_modifie.get_role_display()
                journaliser(
                    request, JournalActivite.Action.MODIFICATION,
                    f"Utilisateur {utilisateur_modifie.email}",
                    "" if nouveau == ancien_role else f"role : {ancien_role} -> {nouveau}",
                )
                messages.success(request, "Utilisateur modifié.")
                return redirect("liste_utilisateurs")
    else:
        form = UtilisateurModificationForm(instance=utilisateur)
    return render(
        request,
        "modifier_utilisateur.html",
        {"form": form, "utilisateur": utilisateur},
    )


@admin_required
@require_POST
def activer_desactiver_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if utilisateur.pk == request.user.pk:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
        return redirect("liste_utilisateurs")

    if (
        utilisateur.role == User.Role.ADMIN
        and not utilisateur.is_active
        and User.objects.filter(role=User.Role.ADMIN, is_active=True).exclude(pk=utilisateur.pk).exists()
    ):
        messages.error(
            request,
            "Un seul administrateur actif autorisé.",
        )
        return redirect("liste_utilisateurs")

    utilisateur.is_active = not utilisateur.is_active
    utilisateur.save(update_fields=["is_active"])
    journaliser(
        request,
        JournalActivite.Action.ACTIVATION if utilisateur.is_active else JournalActivite.Action.DESACTIVATION,
        f"Utilisateur {utilisateur.email}",
    )
    if utilisateur.is_active:
        messages.success(request, f"Compte {utilisateur} activé.")
    else:
        messages.success(request, f"Compte {utilisateur} désactivé.")
    return redirect("liste_utilisateurs")


@admin_required
def reinitialiser_mot_de_passe(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        statut_onboarding = envoyer_activation_utilisateur(utilisateur, request=request)
        journaliser(request, JournalActivite.Action.MOT_DE_PASSE, f"Utilisateur {utilisateur.email}", "Génération du lien de réinitialisation sécurisé")
        bilan = statut_onboarding.get("bilan") or construire_bilan_onboarding(statut_onboarding, utilisateur, action="reinitialisation")
        return render(
            request,
            "mot_de_passe_genere.html",
            {
                "utilisateur": utilisateur,
                "lien_activation": statut_onboarding["lien_activation"],
                "email_envoye": statut_onboarding["email_envoye"],
                "whatsapp_envoye": statut_onboarding["whatsapp_envoye"],
                "whatsapp_statut": statut_onboarding["whatsapp_statut"],
                "whatsapp_message": statut_onboarding["whatsapp_message"],
                "whatsapp_erreur": statut_onboarding["whatsapp_erreur"],
                "bilan": bilan,
                "action": "reinitialisation",
            },
        )
    return render(request, "reinitialiser_mot_de_passe.html", {"utilisateur": utilisateur})


@admin_required
@require_POST
def renvoyer_activation(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    statut = envoyer_activation_utilisateur(utilisateur, request=request)
    journaliser(
        request,
        JournalActivite.Action.MOT_DE_PASSE,
        f"Utilisateur {utilisateur.email}",
        "Renvoi du lien d'activation sécurisé",
    )
    bilan = construire_bilan_onboarding(statut, utilisateur, action="renvoi")
    if bilan["niveau"] == "success":
        messages.success(request, bilan["texte_flash"])
    else:
        messages.warning(request, bilan["texte_flash"])

    return redirect("liste_utilisateurs")


@admin_required
def supprimer_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if utilisateur.pk == request.user.pk:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
        return redirect("liste_utilisateurs")

    if request.method == "POST":
        # Journalise AVANT delete() : apres, l'objet n'a plus d'email a citer.
        # L'entree, elle, survit -- elle ne porte que du texte fige.
        journaliser(request, JournalActivite.Action.SUPPRESSION, f"Utilisateur {utilisateur.email}",
                    f"role : {utilisateur.get_role_display()}")
        utilisateur.delete()
        messages.success(request, "Utilisateur supprimé.")
        return redirect("liste_utilisateurs")
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": utilisateur, "type": "Utilisateur"},
    )
