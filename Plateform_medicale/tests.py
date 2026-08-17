import datetime
import io
import json
import urllib.error
from decimal import Decimal
from unittest.mock import MagicMock, patch

import openpyxl
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.forms import modelform_factory
from django.test import Client, TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from .models import (
    Consultation,
    Delivrance,
    JournalActivite,
    LigneOrdonnance,
    Medecin,
    Notification,
    Ordonnance,
    Paiement,
    Patient,
    Pharmacien,
    PlanCouverture,
    Prestataire,
    PriseEnCharge,
    RendezVous,
    ServiceMedical,
    TentativeConnexion,
    User,
    distance_km,
)
from .views import SECTIONS_PARAMETRES as SECTIONS_PARAMETRES_REELLES
from .views import TAILLE_PAGE_LISTE

SECTIONS_TOUTES = ('general', 'apparence', 'securite')

PASSWORD = 'MotDePasseSolide2026!'


def creer_utilisateur(role, email):
    return User.objects.create_user(email=email, password=PASSWORD, role=role)


def creer_medecin(email, specialite='Medecine generale'):
    utilisateur = creer_utilisateur(User.Role.MEDECIN, email)
    return Medecin.objects.create(
        user=utilisateur,
        nom='Ndiaye',
        prenom='Awa',
        specialite=specialite,
        telephone='770000000',
        email=email,
    )


def creer_patient(nom='Diop', prenom='Moussa'):
    return Patient.objects.create(
        nom=nom,
        prenom=prenom,
        date_naissance=datetime.date(1990, 1, 1),
        telephone='770000001',
    )


def creer_pharmacien(email):
    utilisateur = creer_utilisateur(User.Role.PHARMACIEN, email)
    return Pharmacien.objects.create(user=utilisateur)


ENTETES_IMPORT_UTILISATEURS = [
    'Email', 'Prenom', 'Nom', 'Telephone', 'Role',
    'Date de naissance', 'Specialite', 'Prestataire', 'Plan de couverture',
]


def creer_fichier_import_utilisateurs(lignes, entetes=None):
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.append(entetes or ENTETES_IMPORT_UTILISATEURS)
    for ligne in lignes:
        feuille.append(ligne)
    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return SimpleUploadedFile(
        'import.xlsx',
        tampon.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def creer_ordonnance(patient, medecin, medicaments='Paracetamol 500mg - 3x/jour'):
    consultation = Consultation.objects.create(
        patient=patient,
        medecin=medecin,
        date_consultation=timezone.now(),
        diagnostic='Diagnostic de test',
    )
    return Ordonnance.objects.create(consultation=consultation, medicaments=medicaments)


class UserManagerTests(TestCase):
    def test_create_user_exige_un_email(self):
        with self.assertRaises(ValueError):
            User.objects.create_user(email='', password=PASSWORD)

    def test_create_superuser_a_le_role_admin(self):
        user = User.objects.create_superuser(email='admin@santesn.sn', password=PASSWORD)
        self.assertEqual(user.role, User.Role.ADMIN)
        self.assertTrue(user.is_staff)
        self.assertTrue(user.is_superuser)


class SetupWizardTests(TestCase):
    def test_wizard_accessible_sans_admin(self):
        response = self.client.get(reverse('setup_wizard'))
        self.assertEqual(response.status_code, 200)

    def test_login_redirige_vers_wizard_sans_admin(self):
        response = self.client.get(reverse('login'))
        self.assertRedirects(response, reverse('setup_wizard'))

    def test_wizard_cree_le_premier_super_admin(self):
        response = self.client.post(reverse('setup_wizard'), {
            'first_name': 'Awa',
            'last_name': 'Diop',
            'email': 'admin@santesn.sn',
            'phone_number': '770000000',
            'password1': PASSWORD,
            'password2': PASSWORD,
        })
        self.assertRedirects(
            response,
            reverse('post_login_redirect'),
            target_status_code=302,
        )
        admin = User.objects.get(email='admin@santesn.sn')
        self.assertEqual(admin.role, User.Role.ADMIN)
        self.assertTrue(admin.is_superuser)

    def test_wizard_desactive_apres_creation_admin(self):
        creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        response = self.client.get(reverse('setup_wizard'))
        self.assertRedirects(response, reverse('login'))


class LandingTests(TestCase):
    def test_page_publique_affiche_la_section_ecosysteme(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, "Pensé pour tout l'écosystème de la santé")
        self.assertContains(response, "Assurances santé")
        self.assertContains(response, "IPM")
        self.assertContains(response, "Hôpitaux")
        self.assertContains(response, "Cliniques")
        self.assertContains(response, "Pharmacies")
        self.assertContains(response, "Grandes entreprises")

    def test_cartes_services_ont_des_mini_visuels(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, 'class="motif-qr carte-service-visuel"')
        self.assertContains(response, 'class="mini-repartition"')
        self.assertContains(response, 'class="mini-famille"')

    def test_page_publique_affiche_la_section_verification_qr(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, "Vérification en un scan")
        self.assertContains(response, "Le médecin génère")
        self.assertContains(response, "QR Code unique")
        self.assertContains(response, "Le pharmacien scanne")

    def test_page_publique_affiche_le_showcase_des_dashboards(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, "Un espace dédié pour chaque rôle")
        self.assertContains(response, "Aperçu illustratif")
        self.assertContains(response, "Agenda — Dr Ndiaye")
        self.assertContains(response, "Scan d'ordonnance")
        self.assertContains(response, "Ordonnance disponible")
        self.assertContains(response, "Tableau de bord")

    def test_page_publique_affiche_devenir_partenaire(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, "Devenir partenaire SantéSN")
        self.assertContains(response, "Prenez contact")
        self.assertContains(response, "Nous configurons votre espace")
        self.assertContains(response, "Vous êtes opérationnel")
        self.assertContains(response, "aucune certification formelle")

    def test_page_publique_affiche_la_faq(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, "Tout ce qu'il faut savoir avant de commencer")
        self.assertContains(response, "Comment nos collaborateurs ou assurés obtiennent-ils un accès")
        self.assertContains(response, "Qui peut rejoindre le réseau de prestataires")
        self.assertContains(response, "Peut-on gérer les ayants droit")

    def test_footer_affiche_les_nouvelles_ancres_et_le_contact(self):
        response = self.client.get(reverse('landing'))
        self.assertContains(response, 'href="#partenaires"')
        self.assertContains(response, 'href="#faq"')
        self.assertContains(response, 'href="#dashboards"')
        self.assertContains(response, "La santé connectée, pensée pour l'écosystème sénégalais")


class LoginTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')

    def _login(self, email):
        return self.client.post(reverse('login'), {
            'email': email,
            'password': PASSWORD,
        })

    def test_connexion_admin_redirige_vers_dashboard_admin(self):
        response = self._login('admin@santesn.sn')
        self.assertRedirects(
            response,
            reverse('post_login_redirect'),
            target_status_code=302,
        )
        response = self.client.get(reverse('post_login_redirect'))
        self.assertRedirects(response, reverse('dashboard'))

    def test_connexion_assure_redirige_vers_espace_assure(self):
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        self._login('assure@santesn.sn')
        response = self.client.get(reverse('post_login_redirect'))
        self.assertRedirects(
            response,
            reverse('dashboard_assure'),
            target_status_code=302,
        )
        # Sans fiche Patient liee, le premier passage redirige vers la completion du profil.
        response = self.client.get(reverse('dashboard_assure'))
        self.assertRedirects(response, reverse('mon_profil_assure'))

    def test_connexion_medecin_redirige_vers_espace_medecin(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self._login('medecin@santesn.sn')
        response = self.client.get(reverse('post_login_redirect'))
        self.assertRedirects(response, reverse('dashboard_medecin'))

    def test_connexion_pharmacien_redirige_vers_espace_pharmacien(self):
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien@santesn.sn')
        self._login('pharmacien@santesn.sn')
        response = self.client.get(reverse('post_login_redirect'))
        self.assertRedirects(response, reverse('dashboard_pharmacien'))

    def test_lien_retour_accueil_present(self):
        """Lien de retour vers le landing page, au-dessus du formulaire de connexion."""
        response = self.client.get(reverse('login'))
        self.assertContains(response, 'class="lien-retour-accueil"')
        self.assertContains(response, f'href="{reverse("landing")}"')

    def test_mauvais_mot_de_passe_refuse(self):
        response = self.client.post(reverse('login'), {
            'email': 'admin@santesn.sn',
            'password': 'mauvais',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Email ou mot de passe incorrect.')
        # Plan de direction artistique, item 7 : erreur annoncee aux lecteurs d'ecran.
        self.assertContains(response, 'class="erreurs" role="alert"')


class LimitationTentativesConnexionTests(TestCase):
    def setUp(self):
        cache.clear()
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-brute@santesn.sn')

    def tearDown(self):
        cache.clear()

    def _mauvais_mot_de_passe(self):
        return self.client.post(reverse('login'), {
            'email': 'admin-brute@santesn.sn',
            'password': 'mauvais',
        })

    def test_blocage_apres_cinq_echecs(self):
        for _ in range(5):
            response = self._mauvais_mot_de_passe()
            self.assertContains(response, 'Email ou mot de passe incorrect.')

        response = self._mauvais_mot_de_passe()
        self.assertContains(response, 'Trop de tentatives de connexion')

    def test_connexion_reussie_reinitialise_le_compteur(self):
        for _ in range(3):
            self._mauvais_mot_de_passe()

        response = self.client.post(reverse('login'), {
            'email': 'admin-brute@santesn.sn',
            'password': PASSWORD,
        })
        self.assertRedirects(response, reverse('post_login_redirect'), target_status_code=302)

        self.client.logout()
        response = self._mauvais_mot_de_passe()
        self.assertContains(response, 'Email ou mot de passe incorrect.')
        self.assertNotContains(response, 'Trop de tentatives de connexion')


class ProtectionDesVuesTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.assure = creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')

    def test_dashboard_exige_connexion(self):
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_dashboard_interdit_aux_non_admins(self):
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_accessible_a_l_admin(self):
        self.client.login(username='admin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)

    def test_liste_patients_interdite_aux_non_admins(self):
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('liste_patients'))
        self.assertEqual(response.status_code, 403)

    def test_espace_assure_interdit_au_medecin(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard_assure'))
        self.assertEqual(response.status_code, 403)


class DashboardAdminTests(TestCase):
    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_sans_donnees_le_bandeau_financier_affiche_un_tiret(self):
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '—')
        self.assertContains(response, '0\xa0FCFA')

    def test_bandeau_financier_calcule_les_montants_regles_et_en_attente(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))

        consultation_reglee = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement_regle = Paiement.calculer_pour(consultation_reglee)
        paiement_regle.statut = Paiement.Statut.REGLE
        paiement_regle.save()

        consultation_non_reglee = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        Paiement.calculer_pour(consultation_non_reglee).save()

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['montant_regle'], Decimal('10000'))
        self.assertEqual(response.context['montant_non_regle'], Decimal('10000'))
        self.assertEqual(response.context['taux_reglement'], 50)

    def test_compte_les_pharmaciens(self):
        creer_pharmacien('pharmacien@santesn.sn')
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['total_pharmaciens'], 1)
        self.assertContains(response, 'Pharmaciens actifs')

    def test_ne_compte_que_les_prestataires_partenaires(self):
        Prestataire.objects.create(nom='Clinique partenaire', type_prestataire='CLINIQUE', partenaire=True)
        Prestataire.objects.create(nom='Cabinet non partenaire', type_prestataire='CABINET', partenaire=False)
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['total_prestataires'], 1)

    def test_gouvernance_compte_les_comptes_actifs_et_inactifs(self):
        inactif = creer_utilisateur(User.Role.MEDECIN, 'inactif@santesn.sn')
        inactif.is_active = False
        inactif.save()
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['total_comptes_inactifs'], 1)
        self.assertGreaterEqual(response.context['total_comptes_actifs'], 1)

    def test_gouvernance_stats_actifs_inactifs_sont_cliquables(self):
        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, f'href="{reverse("liste_utilisateurs")}?statut=actif"')
        self.assertContains(response, f'href="{reverse("liste_utilisateurs")}?statut=inactif"')

    def test_aujourd_hui_ne_compte_que_les_rendez_vous_et_consultations_du_jour(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('5000'))

        RendezVous.objects.create(
            patient=patient, medecin=medecin, date_heure=timezone.now(), statut='CONFIRME',
        )
        RendezVous.objects.create(
            patient=patient, medecin=medecin,
            date_heure=timezone.now() + datetime.timedelta(days=3), statut='CONFIRME',
        )
        Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['total_rendez_vous_aujourd_hui'], 1)
        self.assertEqual(response.context['total_consultations_aujourd_hui'], 1)

    def test_compte_les_prises_en_charge_en_attente(self):
        patient = creer_patient()
        PriseEnCharge.objects.create(patient=patient, motif='Test', statut='en_attente')
        PriseEnCharge.objects.create(patient=patient, motif='Test', statut='validee')

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['total_prises_en_charge_attente'], 1)

    def test_bandeau_urgent_affiche_l_anciennete_de_la_plus_ancienne_prise_en_charge(self):
        patient = creer_patient()
        ancienne = PriseEnCharge.objects.create(patient=patient, motif='Ancienne', statut='en_attente')
        PriseEnCharge.objects.filter(pk=ancienne.pk).update(
            date_demande=timezone.now() - datetime.timedelta(days=12)
        )
        PriseEnCharge.objects.create(patient=patient, motif='Recente', statut='en_attente')

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['jours_attente_max'], 12)
        self.assertContains(response, 'la plus ancienne : 12 j')

    def test_tuiles_portent_un_aria_label_libelle_avant_valeur(self):
        """Convention du projet : l'aria-label annonce le libelle puis la
        valeur. Les elements testes ont change avec la direction "Clinique
        claire" (le hero a 3 liens a laisse place au bandeau de files et aux
        cartes KPI), la convention, elle, reste la meme."""
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('5000'))

        consultation_non_reglee = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        Paiement.calculer_pour(consultation_non_reglee).save()

        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, 'aria-label="Consultations : 1"')
        self.assertContains(response, 'aria-label="Médecins : 1"')
        self.assertContains(
            response, 'aria-label="Règlements en attente : 1, soit 5 000 FCFA"'
        )
        self.assertContains(response, 'aria-label="Prises en charge en attente : 0"')

    def test_derniers_comptes_crees(self):
        creer_medecin('nouveau.medecin@santesn.sn')

        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, 'Derniers comptes cr')
        self.assertContains(response, 'nouveau.medecin@santesn.sn')

    def test_dashboard_n_affiche_plus_les_digests_assures_et_prestataires(self):
        """Regression : ces 2 digests sont retires du dashboard (deplaces
        vers liste_patients/liste_prestataires, cf. Taches 4 et 5)."""
        response = self.client.get(reverse('dashboard'))
        self.assertNotContains(response, 'Derniers assurés')
        self.assertNotContains(response, 'Derniers prestataires ajoutés')

    def test_derniers_comptes_exclut_les_assures(self):
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        response = self.client.get(reverse('dashboard'))
        emails = [u.email for u in response.context['derniers_comptes']]
        self.assertNotIn('assure@santesn.sn', emails)

    def test_dashboard_ouvre_sur_le_bandeau_des_files_d_attente(self):
        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, 'class="file-attente"')

    def test_bandeau_file_attente_vide_quand_rien_a_traiter(self):
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['file_totale'], 0)
        self.assertContains(response, 'class="file-vide"')

    def test_listes_du_dashboard_sont_en_grille_de_digests(self):
        patient = creer_patient()
        PriseEnCharge.objects.create(patient=patient, motif='Test validee', statut='validee')
        PriseEnCharge.objects.create(patient=patient, motif='Test refusee', statut='refusee')

        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, 'class="duo-listes"')
        self.assertContains(response, 'class="liste-lignes"')
        self.assertContains(response, 'badge validee')
        self.assertContains(response, 'badge refusee')

    def test_hero_paiements_affiche_le_delta_7_jours(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.statut = Paiement.Statut.REGLE
        paiement.date_reglement = timezone.now()
        paiement.save()

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['montant_regle_7j'], Decimal('10000'))
        self.assertContains(response, 'class="finances-delta"')

    def test_dashboard_inclut_la_tendance_des_paiements_regles(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('8000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.statut = Paiement.Statut.REGLE
        paiement.date_reglement = timezone.now()
        paiement.save()

        response = self.client.get(reverse('dashboard'))
        donnees = response.context['tendance_paiements']
        self.assertEqual(len(donnees['labels']), 30)
        self.assertEqual(len(donnees['totaux']), 30)
        self.assertEqual(donnees['totaux'][-1], Decimal('8000'))
        self.assertContains(response, 'graphe-tendance-paiements')

    def test_hero_paiements_ignore_les_reglements_hors_7_jours(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.statut = Paiement.Statut.REGLE
        paiement.date_reglement = timezone.now() - datetime.timedelta(days=8)
        paiement.save()

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['montant_regle_7j'], 0)
        self.assertNotContains(response, 'class="finances-delta"')

    def test_hero_paiements_inclut_les_reglements_a_6_jours(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('5000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.statut = Paiement.Statut.REGLE
        paiement.date_reglement = timezone.now() - datetime.timedelta(days=6)
        paiement.save()

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['montant_regle_7j'], Decimal('5000'))
        self.assertContains(response, 'class="finances-delta"')

    def test_kpi_consultations_et_ordonnances_affichent_le_delta_7_jours(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('5000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        Ordonnance.objects.create(consultation=consultation, medicaments='Paracetamol')

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['consultations_7j'], 1)
        self.assertEqual(response.context['ordonnances_7j'], 1)
        self.assertContains(response, 'class="kpi-delta"')

    def test_kpi_ignore_les_consultations_et_ordonnances_hors_7_jours(self):
        medecin = creer_medecin('medecin@santesn.sn')
        patient = creer_patient()
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('5000'))
        ancienne = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now() - datetime.timedelta(days=10), diagnostic='Test',
        )
        ancienne_ordonnance = Ordonnance.objects.create(consultation=ancienne, medicaments='Paracetamol')
        ancienne_ordonnance.date_creation = timezone.now() - datetime.timedelta(days=10)
        ancienne_ordonnance.save()

        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.context['consultations_7j'], 0)
        self.assertEqual(response.context['ordonnances_7j'], 0)

    def test_nav_admin_a_des_separateurs_de_section(self):
        response = self.client.get(reverse('dashboard'))
        self.assertContains(response, 'class="nav-section"')
        self.assertContains(response, '>Gestion<')
        self.assertContains(response, '>Opérations<')
        self.assertContains(response, '>Système<')


class GestionUtilisateursTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_liste_utilisateurs_interdite_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('liste_utilisateurs'))
        self.assertEqual(response.status_code, 403)

    def test_liste_utilisateurs_accessible_a_l_admin(self):
        response = self.client.get(reverse('liste_utilisateurs'))
        self.assertEqual(response.status_code, 200)

    def test_tableau_utilisateurs_accessible(self):
        """Plan de direction artistique, item 7 : legende + en-tetes scopes."""
        response = self.client.get(reverse('liste_utilisateurs'))
        self.assertContains(response, '<caption class="sr-only">')
        self.assertContains(response, '<th scope="col">')

    def test_desactivation_passe_par_la_modale_pas_par_confirm_natif(self):
        """Plan de direction artistique, item 2 : plus de confirm() natif."""
        actif = creer_utilisateur(User.Role.MEDECIN, 'actif@santesn.sn')
        inactif = creer_utilisateur(User.Role.MEDECIN, 'inactif@santesn.sn')
        inactif.is_active = False
        inactif.save(update_fields=['is_active'])

        response = self.client.get(reverse('liste_utilisateurs'))
        self.assertNotContains(response, 'onsubmit="return confirm(')
        self.assertContains(response, 'id="modale-confirmation"')
        self.assertContains(response, f'data-confirmation="Désactiver le compte de {actif} ?')
        self.assertNotContains(response, f'data-confirmation="Désactiver le compte de {inactif} ?')

    def test_creation_utilisateur_genere_un_mot_de_passe_fonctionnel(self):
        response = self.client.post(reverse('ajouter_utilisateur'), {
            'first_name': 'Fatou',
            'last_name': 'Ndiaye',
            'email': 'fatou.ndiaye@santesn.sn',
            'phone_number': '770001122',
            'role': User.Role.MEDECIN.value,
        })
        self.assertEqual(response.status_code, 200)
        mot_de_passe = response.context['mot_de_passe']
        utilisateur = User.objects.get(email='fatou.ndiaye@santesn.sn')
        self.assertEqual(utilisateur.role, User.Role.MEDECIN)

        self.client.logout()
        connecte = self.client.login(username='fatou.ndiaye@santesn.sn', password=mot_de_passe)
        self.assertTrue(connecte)

    def test_modification_role_utilisateur(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(reverse('modifier_utilisateur', args=[cible.pk]), {
            'first_name': cible.first_name,
            'last_name': cible.last_name,
            'email': cible.email,
            'phone_number': '',
            'role': User.Role.PHARMACIEN.value,
        })
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        cible.refresh_from_db()
        self.assertEqual(cible.role, User.Role.PHARMACIEN)

    def test_admin_ne_peut_pas_changer_son_propre_role(self):
        response = self.client.post(reverse('modifier_utilisateur', args=[self.admin.pk]), {
            'first_name': self.admin.first_name,
            'last_name': self.admin.last_name,
            'email': self.admin.email,
            'phone_number': '',
            'role': User.Role.MEDECIN.value,
        })
        self.assertEqual(response.status_code, 200)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.role, User.Role.ADMIN)

    def test_desactivation_utilisateur(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(reverse('activer_desactiver_utilisateur', args=[cible.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        cible.refresh_from_db()
        self.assertFalse(cible.is_active)

    def test_admin_ne_peut_pas_se_desactiver_lui_meme(self):
        response = self.client.post(reverse('activer_desactiver_utilisateur', args=[self.admin.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_active)

    def test_ajouter_utilisateur_ne_propose_pas_le_role_admin(self):
        response = self.client.get(reverse('ajouter_utilisateur'))
        self.assertNotContains(response, '<option value="ADMIN">')

    def test_creation_utilisateur_avec_role_admin_est_rejetee(self):
        response = self.client.post(reverse('ajouter_utilisateur'), {
            'first_name': 'Faux',
            'last_name': 'Admin',
            'email': 'faux.admin@santesn.sn',
            'phone_number': '',
            'role': User.Role.ADMIN.value,
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(User.objects.filter(email='faux.admin@santesn.sn').exists())

    def test_modifier_utilisateur_ne_permet_pas_de_promouvoir_admin(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(reverse('modifier_utilisateur', args=[cible.pk]), {
            'first_name': cible.first_name,
            'last_name': cible.last_name,
            'email': cible.email,
            'phone_number': '',
            'role': User.Role.ADMIN.value,
        })
        self.assertEqual(response.status_code, 200)
        cible.refresh_from_db()
        self.assertEqual(cible.role, User.Role.MEDECIN)

    def test_modifier_utilisateur_admin_garde_son_propre_role_dans_le_formulaire(self):
        response = self.client.get(reverse('modifier_utilisateur', args=[self.admin.pk]))
        choix = dict(response.context['form'].fields['role'].choices)
        self.assertIn(User.Role.ADMIN.value, choix)

    def test_reactivation_bloquee_si_un_autre_admin_est_deja_actif(self):
        ancien_admin = creer_utilisateur(User.Role.ADMIN, 'ancien-admin@santesn.sn')
        ancien_admin.is_active = False
        ancien_admin.save(update_fields=['is_active'])

        response = self.client.post(reverse('activer_desactiver_utilisateur', args=[ancien_admin.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        ancien_admin.refresh_from_db()
        self.assertFalse(ancien_admin.is_active)

    def test_admin_ne_peut_pas_se_supprimer_lui_meme(self):
        response = self.client.post(reverse('supprimer_utilisateur', args=[self.admin.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        self.assertTrue(User.objects.filter(pk=self.admin.pk).exists())

    def test_suppression_utilisateur(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(reverse('supprimer_utilisateur', args=[cible.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))
        self.assertFalse(User.objects.filter(pk=cible.pk).exists())

    def test_reinitialisation_mot_de_passe(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(reverse('reinitialiser_mot_de_passe', args=[cible.pk]))
        self.assertEqual(response.status_code, 200)
        nouveau_mot_de_passe = response.context['mot_de_passe']

        self.client.logout()
        connecte = self.client.login(username='cible@santesn.sn', password=nouveau_mot_de_passe)
        self.assertTrue(connecte)

    def test_filtre_par_role(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien@santesn.sn')
        response = self.client.get(reverse('liste_utilisateurs'), {'role': User.Role.MEDECIN.value})
        emails = [u.email for u in response.context['utilisateurs']]
        self.assertIn('medecin@santesn.sn', emails)
        self.assertNotIn('pharmacien@santesn.sn', emails)

    def test_filtre_par_statut(self):
        inactif = creer_utilisateur(User.Role.MEDECIN, 'inactif@santesn.sn')
        inactif.is_active = False
        inactif.save(update_fields=['is_active'])
        creer_utilisateur(User.Role.MEDECIN, 'actif@santesn.sn')

        response = self.client.get(reverse('liste_utilisateurs'), {'statut': 'inactif'})
        emails = [u.email for u in response.context['utilisateurs']]
        self.assertIn('inactif@santesn.sn', emails)
        self.assertNotIn('actif@santesn.sn', emails)

    def test_recherche_par_nom_ou_email(self):
        creer_utilisateur(User.Role.MEDECIN, 'ousmane.fall@santesn.sn')
        creer_utilisateur(User.Role.MEDECIN, 'autre@santesn.sn')
        response = self.client.get(reverse('liste_utilisateurs'), {'q': 'ousmane'})
        emails = [u.email for u in response.context['utilisateurs']]
        self.assertIn('ousmane.fall@santesn.sn', emails)
        self.assertNotIn('autre@santesn.sn', emails)


class ToastsMessagesTests(TestCase):
    """Plan de direction artistique, item 3 : messages Django en toasts."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_message_succes_rendu_en_toast(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'cible@santesn.sn')
        response = self.client.post(
            reverse('activer_desactiver_utilisateur', args=[cible.pk]), follow=True
        )
        self.assertContains(response, 'class="toasts"')
        self.assertContains(response, 'class="toast toast-success"')
        self.assertContains(response, 'toast-fermer')
        self.assertNotContains(response, '<ul style="list-style:none')

    def test_message_erreur_rendu_en_toast_error(self):
        response = self.client.post(
            reverse('activer_desactiver_utilisateur', args=[self.admin.pk]), follow=True
        )
        self.assertContains(response, 'class="toast toast-error"')


class EspaceMedecinTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.autre_medecin = creer_medecin('medecin2@santesn.sn')
        self.patient = creer_patient()
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

    def test_dashboard_interdit_aux_non_medecins(self):
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard_medecin'))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_accessible_au_medecin(self):
        response = self.client.get(reverse('dashboard_medecin'))
        self.assertEqual(response.status_code, 200)

    def test_medecin_sans_fiche_voit_message_dedie(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'orphelin@santesn.sn')
        self.client.login(username='orphelin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard_medecin'))
        self.assertContains(response, 'pas encore associe')

    def test_creation_rendez_vous_attribue_au_medecin_connecte(self):
        response = self.client.post(reverse('ajouter_rendez_vous'), {
            'patient': self.patient.pk,
            'prestataire': '',
            'date_heure': (timezone.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%dT%H:%M'),
            'motif': 'Controle',
        })
        self.assertRedirects(response, reverse('agenda_medecin'))
        rendez_vous = RendezVous.objects.get(patient=self.patient)
        self.assertEqual(rendez_vous.medecin, self.medecin)
        self.assertEqual(rendez_vous.statut, RendezVous.Statut.DEMANDE)

    def test_medecin_ne_peut_pas_modifier_le_rendez_vous_d_un_autre_medecin(self):
        rendez_vous = RendezVous.objects.create(
            patient=self.patient,
            medecin=self.autre_medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
        )
        response = self.client.post(
            reverse('changer_statut_rendez_vous', args=[rendez_vous.pk]),
            {'statut': 'CONFIRME'},
        )
        self.assertEqual(response.status_code, 404)
        rendez_vous.refresh_from_db()
        self.assertEqual(rendez_vous.statut, RendezVous.Statut.DEMANDE)

    def test_changement_statut_rendez_vous_propre(self):
        rendez_vous = RendezVous.objects.create(
            patient=self.patient,
            medecin=self.medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
        )
        response = self.client.post(
            reverse('changer_statut_rendez_vous', args=[rendez_vous.pk]),
            {'statut': 'CONFIRME'},
        )
        self.assertRedirects(response, reverse('agenda_medecin'))
        rendez_vous.refresh_from_db()
        self.assertEqual(rendez_vous.statut, RendezVous.Statut.CONFIRME)

    def test_creation_consultation_et_ordonnance_avec_qr(self):
        response = self.client.post(reverse('ajouter_consultation_medecin'), {
            'patient': self.patient.pk,
            'service': '',
            'prise_en_charge': '',
            'date_consultation': '2026-08-01T10:00',
            'diagnostic': 'Grippe saisonniere',
            'traitement': 'Repos et paracetamol',
        })
        consultation = Consultation.objects.get(patient=self.patient)
        self.assertEqual(consultation.medecin, self.medecin)
        self.assertRedirects(
            response,
            reverse('ajouter_ordonnance_medecin', args=[consultation.pk]),
        )

        response = self.client.post(
            reverse('ajouter_ordonnance_medecin', args=[consultation.pk]),
            {
                # Le formulaire attend desormais des LIGNES structurees.
                'lignes-TOTAL_FORMS': '1', 'lignes-INITIAL_FORMS': '0',
                'lignes-MIN_NUM_FORMS': '0', 'lignes-MAX_NUM_FORMS': '1000',
                'lignes-0-medicament': 'Paracetamol', 'lignes-0-dosage': '500 mg',
                'lignes-0-posologie': '3x/jour', 'lignes-0-duree': '5 jours',
                'lignes-0-quantite': '',
            },
        )
        ordonnance = Ordonnance.objects.get(consultation=consultation)
        self.assertTrue(ordonnance.code_qr.startswith('RX-'))
        self.assertRedirects(
            response,
            reverse('voir_ordonnance_medecin', args=[ordonnance.pk]),
        )

        reponse_qr = self.client.get(reverse('voir_ordonnance_medecin', args=[ordonnance.pk]))
        self.assertContains(reponse_qr, '<svg')

    def test_medecin_ne_peut_pas_creer_ordonnance_pour_consultation_d_un_autre(self):
        consultation_autre = Consultation.objects.create(
            patient=self.patient,
            medecin=self.autre_medecin,
            date_consultation=timezone.now(),
            diagnostic='Diagnostic confidentiel',
        )
        response = self.client.get(
            reverse('ajouter_ordonnance_medecin', args=[consultation_autre.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_mes_patients_scope_au_medecin_connecte(self):
        autre_patient = creer_patient(nom='Sow', prenom='Fatou')
        Consultation.objects.create(
            patient=self.patient,
            medecin=self.medecin,
            date_consultation=timezone.now(),
            diagnostic='RAS',
        )
        Consultation.objects.create(
            patient=autre_patient,
            medecin=self.autre_medecin,
            date_consultation=timezone.now(),
            diagnostic='RAS',
        )
        response = self.client.get(reverse('mes_patients'))
        patients = list(response.context['patients'])
        self.assertIn(self.patient, patients)
        self.assertNotIn(autre_patient, patients)

    def test_modification_profil_medecin_ne_touche_pas_email(self):
        response = self.client.post(reverse('modifier_profil_medecin'), {
            'specialite': 'Cardiologie',
            'telephone': '781234567',
        })
        self.assertRedirects(response, reverse('modifier_profil_medecin'))
        self.medecin.refresh_from_db()
        self.assertEqual(self.medecin.specialite, 'Cardiologie')
        self.assertEqual(self.medecin.telephone, '781234567')
        self.assertEqual(self.medecin.email, 'medecin1@santesn.sn')


class RecherchePatientsMedecinTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.autre_medecin = creer_medecin('medecin2@santesn.sn')
        self.patient = creer_patient(nom='Diop', prenom='Awa')
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

    def test_recherche_interdite_aux_non_medecins(self):
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        self.assertEqual(response.status_code, 403)

    def test_recherche_interdite_a_l_anonyme(self):
        self.client.logout()
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        self.assertEqual(response.status_code, 302)

    def test_recherche_moins_de_deux_caracteres_ne_renvoie_rien(self):
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'D'})
        self.assertEqual(response.json(), {'resultats': []})

    def test_recherche_par_nom_partiel_insensible_a_la_casse(self):
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'dio'})
        resultats = response.json()['resultats']
        self.assertEqual(len(resultats), 1)
        self.assertEqual(resultats[0]['id'], self.patient.pk)
        self.assertEqual(resultats[0]['numero_carte'], self.patient.numero_carte)
        # Le widget (mes_patients.html / dashboard_medecin.html) formate cette
        # valeur en JJ/MM/AAAA cote JS : le contrat JSON doit rester une date
        # ISO (AAAA-MM-JJ), sans quoi le formatage clic-a-clic serait fausse.
        self.assertEqual(resultats[0]['date_naissance'], self.patient.date_naissance.isoformat())

    def test_recherche_par_numero_de_carte_exact(self):
        response = self.client.get(
            reverse('rechercher_patients_medecin'), {'q': self.patient.numero_carte}
        )
        resultats = response.json()['resultats']
        self.assertEqual(len(resultats), 1)
        self.assertEqual(resultats[0]['id'], self.patient.pk)

    def test_recherche_plafonnee_a_huit_resultats(self):
        for i in range(10):
            creer_patient(nom='Diop%s' % i, prenom='Test')
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        self.assertEqual(len(response.json()['resultats']), 8)

    def test_recherche_priorise_toujours_la_correspondance_exacte_de_carte(self):
        carte_recherchee = 'SN-TESTPRIOR01'
        patient_carte = creer_patient(nom='Zzz', prenom='Zzz')
        patient_carte.numero_carte = carte_recherchee
        patient_carte.save()
        # 8 patients dont le nom contient litteralement le numero recherche,
        # tries alphabetiquement avant "Zzz" : sans priorisation explicite,
        # la correspondance exacte serait evincee du top 8 par le tri nom/prenom.
        for i in range(8):
            creer_patient(nom='Aaa%s%s' % (carte_recherchee, i), prenom='Test')
        response = self.client.get(
            reverse('rechercher_patients_medecin'), {'q': carte_recherchee}
        )
        resultats = response.json()['resultats']
        self.assertEqual(len(resultats), 8)
        self.assertEqual(resultats[0]['id'], patient_carte.pk)

    def test_recherche_ne_renvoie_aucune_donnee_medicale(self):
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        resultat = response.json()['resultats'][0]
        self.assertEqual(
            set(resultat.keys()),
            {'id', 'nom', 'prenom', 'numero_carte', 'type_beneficiaire', 'date_naissance', 'deja_vu'},
        )

    def test_recherche_indique_deja_vu(self):
        Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='RAS',
        )
        autre_patient = creer_patient(nom='Diopsy', prenom='Fatou')
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        resultats = {r['id']: r['deja_vu'] for r in response.json()['resultats']}
        self.assertTrue(resultats[self.patient.pk])
        self.assertFalse(resultats[autre_patient.pk])

    def test_recherche_trouve_un_patient_non_suivi_par_ce_medecin(self):
        """Portee actee dans la spec : tous les patients, pas seulement ceux du medecin connecte."""
        Consultation.objects.create(
            patient=self.patient, medecin=self.autre_medecin,
            date_consultation=timezone.now(), diagnostic='RAS',
        )
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        resultats = response.json()['resultats']
        self.assertEqual(len(resultats), 1)
        self.assertFalse(resultats[0]['deja_vu'])

    def test_medecin_sans_fiche_recoit_une_liste_vide(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'orphelin@santesn.sn')
        self.client.login(username='orphelin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': 'Diop'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'resultats': []})

    def test_recherche_par_pk_numerique(self):
        # pk force a 2 chiffres : la recherche ignore les requetes de moins
        # de 2 caracteres (test_recherche_moins_de_deux_caracteres_ne_renvoie_rien),
        # donc un pk a 1 chiffre ne suffirait pas a exercer la branche Q(pk=...).
        patient_cible = Patient.objects.create(
            pk=42, nom='Sow', prenom='Ndeye',
            date_naissance=datetime.date(1990, 1, 1), telephone='770000002',
        )
        response = self.client.get(
            reverse('rechercher_patients_medecin'), {'q': str(patient_cible.pk)}
        )
        resultats = response.json()['resultats']
        self.assertIn(patient_cible.pk, [r['id'] for r in resultats])

    def test_recherche_avec_chiffre_unicode_non_convertible_ne_plante_pas(self):
        """
        '²²' passe str.isdigit() mais fait planter int() (ValueError) : sans
        isdecimal(), Q(pk=requete) remontait un 500 sur Patient.objects.filter.
        """
        response = self.client.get(reverse('rechercher_patients_medecin'), {'q': '²²'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'resultats': []})


class PreRemplissagePatientConsultationTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.patient = creer_patient(nom='Diop', prenom='Awa')
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

    def test_patient_preselectionne_si_parametre_valide(self):
        response = self.client.get(
            reverse('ajouter_consultation_medecin'), {'patient': self.patient.pk}
        )
        self.assertEqual(response.context['form'].initial.get('patient'), str(self.patient.pk))

    def test_formulaire_vide_sans_parametre(self):
        response = self.client.get(reverse('ajouter_consultation_medecin'))
        self.assertNotIn('patient', response.context['form'].initial)

    def test_parametre_non_numerique_ignore_silencieusement(self):
        response = self.client.get(
            reverse('ajouter_consultation_medecin'), {'patient': 'abc'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('patient', response.context['form'].initial)

    def test_parametre_chiffre_unicode_non_convertible_ignore_silencieusement(self):
        """
        '²²' passe str.isdigit() mais fait planter int() (ValueError) : sans
        isdecimal(), Patient.objects.filter(pk='²²') remontait un 500.
        """
        response = self.client.get(
            reverse('ajouter_consultation_medecin'), {'patient': '²²'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('patient', response.context['form'].initial)

    def test_parametre_patient_inexistant_ignore_silencieusement(self):
        response = self.client.get(
            reverse('ajouter_consultation_medecin'), {'patient': '999999'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('patient', response.context['form'].initial)

    def test_soumission_post_inchangee_avec_parametre_dans_l_url(self):
        """Non-regression : le POST ignore le query-string, comme avant."""
        response = self.client.post(
            reverse('ajouter_consultation_medecin') + '?patient=%s' % self.patient.pk,
            {
                'patient': self.patient.pk,
                'service': '',
                'prise_en_charge': '',
                'date_consultation': '2026-08-01T10:00',
                'diagnostic': 'RAS',
                'traitement': '',
            },
        )
        consultation = Consultation.objects.get(patient=self.patient)
        self.assertRedirects(
            response, reverse('ajouter_ordonnance_medecin', args=[consultation.pk])
        )


class FichePatientMedecinTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.autre_medecin = creer_medecin('medecin2@santesn.sn')
        self.patient = creer_patient(nom='Diop', prenom='Awa')
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

    def test_fiche_interdite_aux_non_medecins(self):
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertEqual(response.status_code, 403)

    def test_fiche_interdite_a_l_anonyme(self):
        self.client.logout()
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertEqual(response.status_code, 302)

    def test_fiche_patient_inexistant_donne_404(self):
        response = self.client.get(reverse('fiche_patient_medecin', args=[999999]))
        self.assertEqual(response.status_code, 404)

    def test_fiche_accessible_pour_un_patient_jamais_vu(self):
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.patient.numero_carte)

    def test_historique_limite_aux_consultations_du_medecin_connecte(self):
        Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Vue par moi',
        )
        Consultation.objects.create(
            patient=self.patient, medecin=self.autre_medecin,
            date_consultation=timezone.now(), diagnostic='ConfidentielAutreMedecin',
        )
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        historique = list(response.context['historique'])
        self.assertEqual(len(historique), 1)
        self.assertEqual(historique[0].diagnostic, 'Vue par moi')
        self.assertNotContains(response, 'ConfidentielAutreMedecin')

    def test_bouton_nouvelle_consultation_pre_remplit_le_patient(self):
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        url_attendue = reverse('ajouter_consultation_medecin') + '?patient=%s' % self.patient.pk
        self.assertContains(response, url_attendue)

    def test_ayants_droit_affiches_pour_un_assure_principal(self):
        ayant_droit = creer_patient(nom='Diop', prenom='Petit')
        ayant_droit.type_beneficiaire = Patient.TypeBeneficiaire.AYANT_DROIT
        ayant_droit.assure_principal = self.patient
        ayant_droit.save()
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertContains(response, ayant_droit.numero_carte)

    def test_pas_d_ayants_droit_pour_un_ayant_droit(self):
        ayant_droit = creer_patient(nom='Diop', prenom='Petit')
        ayant_droit.type_beneficiaire = Patient.TypeBeneficiaire.AYANT_DROIT
        ayant_droit.assure_principal = self.patient
        ayant_droit.save()
        response = self.client.get(reverse('fiche_patient_medecin', args=[ayant_droit.pk]))
        self.assertEqual(len(response.context['ayants_droit']), 0)

    def test_badge_deja_suivi_si_relation_existante(self):
        Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='RAS',
        )
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertTrue(response.context['deja_vu'])
        self.assertContains(response, 'Déjà suivi')

    def test_pas_de_badge_deja_suivi_sans_relation(self):
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertFalse(response.context['deja_vu'])
        self.assertNotContains(response, 'Déjà suivi')

    def test_rendez_vous_termine_affiche_le_badge_ok(self):
        RendezVous.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
            statut=RendezVous.Statut.TERMINE,
        )
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertContains(response, 'dash-pill ok')

    def test_rendez_vous_demande_affiche_le_badge_attente(self):
        RendezVous.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
            statut=RendezVous.Statut.DEMANDE,
        )
        response = self.client.get(reverse('fiche_patient_medecin', args=[self.patient.pk]))
        self.assertContains(response, 'dash-pill attente')


class WidgetRecherchePatientsTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

    def test_widget_present_sur_mes_patients(self):
        response = self.client.get(reverse('mes_patients'))
        self.assertContains(response, 'id="recherche-patients-champ"')
        self.assertContains(response, reverse('rechercher_patients_medecin'))

    def test_widget_present_sur_dashboard_medecin(self):
        response = self.client.get(reverse('dashboard_medecin'))
        self.assertContains(response, 'id="recherche-patients-champ"')
        self.assertContains(response, reverse('rechercher_patients_medecin'))


class HistoriqueConsultationsTests(TestCase):
    """Plan de direction artistique, item 5 : filtres patient/date."""

    def setUp(self):
        self.medecin = creer_medecin('medecin1@santesn.sn')
        self.patient_a = creer_patient(nom='Diop', prenom='Aissatou')
        self.patient_b = creer_patient(nom='Fall', prenom='Ibrahima')
        self.client.login(username='medecin1@santesn.sn', password=PASSWORD)

        self.consultation_a = Consultation.objects.create(
            patient=self.patient_a, medecin=self.medecin,
            date_consultation=datetime.datetime(2026, 6, 1, 9, 0, tzinfo=datetime.timezone.utc),
            diagnostic='Rhume',
        )
        self.consultation_b = Consultation.objects.create(
            patient=self.patient_b, medecin=self.medecin,
            date_consultation=datetime.datetime(2026, 6, 15, 14, 0, tzinfo=datetime.timezone.utc),
            diagnostic='Entorse',
        )

    def test_formulaire_de_creation_dans_un_panel(self):
        # .panel-form porte le padding canonique des panneaux de contenu
        # (remplace les styles en ligne qui avaient derive en 13 valeurs).
        response = self.client.get(reverse('ajouter_consultation_medecin'))
        self.assertContains(response, 'class="panel panel-form"')

    def test_filtre_par_patient(self):
        response = self.client.get(reverse('historique_consultations'), {'patient': self.patient_a.pk})
        consultations = list(response.context['consultations'])
        self.assertEqual(consultations, [self.consultation_a])

    def test_filtre_par_date(self):
        response = self.client.get(reverse('historique_consultations'), {'date': '2026-06-15'})
        consultations = list(response.context['consultations'])
        self.assertEqual(consultations, [self.consultation_b])

    def test_date_invalide_ignoree_sans_erreur(self):
        response = self.client.get(reverse('historique_consultations'), {'date': 'pas-une-date'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['consultations']), 2)


class EspacePharmacienTests(TestCase):
    def setUp(self):
        self.pharmacien = creer_pharmacien('pharmacien1@santesn.sn')
        self.medecin = creer_medecin('medecin-rx@santesn.sn')
        self.patient = creer_patient()
        self.ordonnance = creer_ordonnance(self.patient, self.medecin)
        self.client.login(username='pharmacien1@santesn.sn', password=PASSWORD)

    def test_dashboard_interdit_aux_non_pharmaciens(self):
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard_pharmacien'))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_accessible_au_pharmacien(self):
        response = self.client.get(reverse('dashboard_pharmacien'))
        self.assertEqual(response.status_code, 200)

    def test_scan_code_valide_affiche_ordonnance(self):
        response = self.client.post(reverse('scanner_ordonnance'), {
            'code_qr': self.ordonnance.code_qr,
        })
        self.assertEqual(response.context['ordonnance'], self.ordonnance)
        self.assertContains(response, self.patient.nom)

    def test_scan_code_invalide_affiche_erreur(self):
        response = self.client.post(reverse('scanner_ordonnance'), {
            'code_qr': 'RX-INEXISTANT',
        })
        self.assertIsNone(response.context['ordonnance'])
        self.assertContains(response, 'Aucune ordonnance ne correspond à ce code.')

    def test_scan_camera_script_et_bouton_presents(self):
        """Plan de direction artistique, item 8 : scan QR par camera."""
        response = self.client.get(reverse('scanner_ordonnance'))
        self.assertContains(
            response,
            'src="https://cdn.jsdelivr.net/npm/jsqr@1.4.0/dist/jsQR.js"',
        )
        self.assertContains(
            response,
            'integrity="sha384-b5Ya4Bq3qCyz39m2ISh+4DxjAIljdeFwK/BsXLuj9gugaNwAcj/ia15fxNZL9Nlx"',
        )
        self.assertContains(
            response,
            '<button type="button" id="bouton-scan-camera" class="button btn" hidden>',
        )

    def test_scan_camera_panneau_et_gestion_erreurs(self):
        """Plan de direction artistique, item 8 : panneau video + repli si camera indisponible."""
        response = self.client.get(reverse('scanner_ordonnance'))
        self.assertContains(response, 'id="panneau-scan-camera"')
        self.assertContains(response, 'id="video-scan-camera"')
        self.assertContains(response, 'id="canvas-scan-camera"')
        self.assertContains(response, 'id="bouton-fermer-scan-camera"')
        self.assertContains(response, 'function demarrerScan')
        self.assertContains(response, 'function arreterScan')
        self.assertContains(response, "NotAllowedError")
        self.assertContains(response, "NotFoundError")
        self.assertContains(response, 'requestSubmit')

    def test_validation_delivrance(self):
        response = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]), {
            'code_qr': self.ordonnance.code_qr,
        })
        self.assertRedirects(response, reverse('historique_delivrances'))
        delivrance = Delivrance.objects.get(ordonnance=self.ordonnance)
        self.assertEqual(delivrance.pharmacien, self.pharmacien)

    def test_double_delivrance_refusee(self):
        Delivrance.objects.create(ordonnance=self.ordonnance, pharmacien=self.pharmacien)
        response = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]), {
            'code_qr': self.ordonnance.code_qr,
        })
        self.assertRedirects(response, reverse('historique_delivrances'))
        self.assertEqual(Delivrance.objects.filter(ordonnance=self.ordonnance).count(), 1)

    def test_validation_delivrance_sans_scan_prealable_refusee(self):
        """Contourner le scan en POSTant directement sur l'ordonnance (sans code_qr) doit echouer."""
        response = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]))
        self.assertEqual(response.status_code, 404)
        self.assertFalse(Delivrance.objects.filter(ordonnance=self.ordonnance).exists())

    def test_validation_delivrance_code_qr_incorrect_refusee(self):
        response = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]), {
            'code_qr': 'RX-AUTRECODE1',
        })
        self.assertEqual(response.status_code, 404)
        self.assertFalse(Delivrance.objects.filter(ordonnance=self.ordonnance).exists())

    def test_validation_delivrance_pharmacien_sans_fiche(self):
        self.client.logout()
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien-sans-fiche@santesn.sn')
        self.client.login(username='pharmacien-sans-fiche@santesn.sn', password=PASSWORD)
        response = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]), {
            'code_qr': self.ordonnance.code_qr,
        })
        self.assertTemplateUsed(response, 'pharmacien_fiche_manquante.html')
        self.assertFalse(Delivrance.objects.filter(ordonnance=self.ordonnance).exists())

    def test_historique_scope_au_pharmacien_connecte(self):
        autre_pharmacien = creer_pharmacien('pharmacien2@santesn.sn')
        autre_ordonnance = creer_ordonnance(creer_patient(nom='Sow', prenom='Awa'), self.medecin)
        Delivrance.objects.create(ordonnance=self.ordonnance, pharmacien=self.pharmacien)
        Delivrance.objects.create(ordonnance=autre_ordonnance, pharmacien=autre_pharmacien)

        response = self.client.get(reverse('historique_delivrances'))
        delivrances = list(response.context['delivrances'])
        self.assertEqual(len(delivrances), 1)
        self.assertEqual(delivrances[0].ordonnance, self.ordonnance)


class EspaceAssureTests(TestCase):
    def setUp(self):
        self.utilisateur = creer_utilisateur(User.Role.ASSURE, 'assure1@santesn.sn')
        self.client.login(username='assure1@santesn.sn', password=PASSWORD)

    def _completer_profil(self):
        self.client.post(reverse('mon_profil_assure'), {
            'nom': 'Diop',
            'prenom': 'Moussa',
            'date_naissance': '1988-04-12',
            'telephone': '770001122',
            'adresse': 'Dakar',
        })
        return Patient.objects.get(user=self.utilisateur)

    def test_dashboard_interdit_aux_non_assures(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('dashboard_assure'))
        self.assertEqual(response.status_code, 403)

    def test_dashboard_redirige_vers_profil_sans_fiche(self):
        response = self.client.get(reverse('dashboard_assure'))
        self.assertRedirects(response, reverse('mon_profil_assure'))

    def test_completion_profil_cree_patient_principal(self):
        patient = self._completer_profil()
        self.assertEqual(patient.type_beneficiaire, Patient.TypeBeneficiaire.PRINCIPAL)
        self.assertTrue(patient.numero_carte)
        response = self.client.get(reverse('dashboard_assure'))
        self.assertEqual(response.status_code, 200)

    def test_ajout_ayant_droit(self):
        patient = self._completer_profil()
        response = self.client.post(reverse('ajouter_ayant_droit'), {
            'nom': 'Diop',
            'prenom': 'Fatou',
            'date_naissance': '2015-06-01',
            'telephone': '',
            'lien_parente': 'ENFANT',
        })
        self.assertRedirects(response, reverse('liste_ayants_droit'))
        ayant_droit = Patient.objects.get(nom='Diop', prenom='Fatou')
        self.assertEqual(ayant_droit.assure_principal, patient)
        self.assertEqual(ayant_droit.type_beneficiaire, Patient.TypeBeneficiaire.AYANT_DROIT)
        self.assertNotEqual(ayant_droit.numero_carte, patient.numero_carte)

    def test_ayant_droit_herite_du_plan_de_couverture(self):
        plan = PlanCouverture.objects.create(nom='Standard', taux_couverture=Decimal('80.00'))
        patient = self._completer_profil()
        patient.plan_couverture = plan
        patient.save(update_fields=['plan_couverture'])

        self.client.post(reverse('ajouter_ayant_droit'), {
            'nom': 'Diop', 'prenom': 'Fatou', 'date_naissance': '2015-06-01',
            'telephone': '', 'lien_parente': 'ENFANT',
        })
        ayant_droit = Patient.objects.get(nom='Diop', prenom='Fatou')
        self.assertEqual(ayant_droit.taux_couverture, plan.taux_couverture)

    def test_assure_ne_peut_pas_modifier_ayant_droit_dun_autre_compte(self):
        self._completer_profil()
        autre_assure = creer_utilisateur(User.Role.ASSURE, 'assure2@santesn.sn')
        autre_patient = Patient.objects.create(
            user=autre_assure, nom='Sow', prenom='Awa',
            date_naissance=datetime.date(1980, 1, 1), telephone='770000002',
        )
        autre_ayant_droit = Patient.objects.create(
            nom='Sow', prenom='Ibra', date_naissance=datetime.date(2010, 1, 1),
            telephone='', type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            assure_principal=autre_patient,
        )
        response = self.client.get(reverse('modifier_ayant_droit', args=[autre_ayant_droit.pk]))
        self.assertEqual(response.status_code, 404)

    def test_suppression_ayant_droit(self):
        self._completer_profil()
        self.client.post(reverse('ajouter_ayant_droit'), {
            'nom': 'Diop', 'prenom': 'Fatou', 'date_naissance': '2015-06-01',
            'telephone': '', 'lien_parente': 'ENFANT',
        })
        ayant_droit = Patient.objects.get(nom='Diop', prenom='Fatou')
        response = self.client.post(reverse('supprimer_ayant_droit', args=[ayant_droit.pk]))
        self.assertRedirects(response, reverse('liste_ayants_droit'))
        self.assertFalse(Patient.objects.filter(pk=ayant_droit.pk).exists())

    def test_assure_ne_peut_pas_supprimer_ayant_droit_dun_autre_compte(self):
        self._completer_profil()
        autre_assure = creer_utilisateur(User.Role.ASSURE, 'assure2@santesn.sn')
        autre_patient = Patient.objects.create(
            user=autre_assure, nom='Sow', prenom='Awa',
            date_naissance=datetime.date(1980, 1, 1), telephone='770000002',
        )
        autre_ayant_droit = Patient.objects.create(
            nom='Sow', prenom='Ibra', date_naissance=datetime.date(2010, 1, 1),
            telephone='', type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            assure_principal=autre_patient,
        )
        response = self.client.post(reverse('supprimer_ayant_droit', args=[autre_ayant_droit.pk]))
        self.assertEqual(response.status_code, 404)
        self.assertTrue(Patient.objects.filter(pk=autre_ayant_droit.pk).exists())

    def test_creation_rendez_vous_pour_beneficiaire(self):
        patient = self._completer_profil()
        medecin = creer_medecin('medecin-rdv@santesn.sn')
        response = self.client.post(reverse('ajouter_rendez_vous_assure'), {
            'patient': patient.pk,
            'medecin': medecin.pk,
            'prestataire': '',
            'date_heure': (timezone.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%dT%H:%M'),
            'motif': 'Controle',
        })
        self.assertRedirects(response, reverse('mes_rendez_vous_assure'))
        rendez_vous = RendezVous.objects.get(patient=patient)
        self.assertEqual(rendez_vous.medecin, medecin)
        self.assertEqual(rendez_vous.statut, RendezVous.Statut.DEMANDE)

    def test_prestataire_preselectionne_depuis_le_lien(self):
        patient = self._completer_profil()
        prestataire = Prestataire.objects.create(
            nom='Clinique Test', type_prestataire='CLINIQUE', partenaire=True,
        )
        response = self.client.get(reverse('ajouter_rendez_vous_assure'), {'prestataire': prestataire.pk})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['form'].initial.get('prestataire'), str(prestataire.pk))

    def test_prestataire_invalide_dans_lurl_est_ignore(self):
        self._completer_profil()
        response = self.client.get(reverse('ajouter_rendez_vous_assure'), {'prestataire': 'abc'})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('prestataire', response.context['form'].initial)

    def test_ne_peut_pas_prendre_rendez_vous_pour_un_patient_hors_famille(self):
        self._completer_profil()
        medecin = creer_medecin('medecin-rdv2@santesn.sn')
        autre_patient = creer_patient(nom='Sow', prenom='Awa')
        response = self.client.post(reverse('ajouter_rendez_vous_assure'), {
            'patient': autre_patient.pk,
            'medecin': medecin.pk,
            'prestataire': '',
            'date_heure': (timezone.now() + datetime.timedelta(days=1)).strftime('%Y-%m-%dT%H:%M'),
            'motif': 'Controle',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(RendezVous.objects.filter(patient=autre_patient).exists())

    def test_annulation_rendez_vous(self):
        patient = self._completer_profil()
        medecin = creer_medecin('medecin-rdv3@santesn.sn')
        rendez_vous = RendezVous.objects.create(
            patient=patient, medecin=medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
        )
        response = self.client.post(reverse('annuler_rendez_vous_assure', args=[rendez_vous.pk]))
        self.assertRedirects(response, reverse('mes_rendez_vous_assure'))
        rendez_vous.refresh_from_db()
        self.assertEqual(rendez_vous.statut, RendezVous.Statut.ANNULE)

    def test_ordonnances_et_historique_scopes_a_la_famille(self):
        patient = self._completer_profil()
        medecin = creer_medecin('medecin-rdv4@santesn.sn')
        ordonnance = creer_ordonnance(patient, medecin)

        autre_patient = creer_patient(nom='Kane', prenom='Modou')
        creer_ordonnance(autre_patient, medecin)

        response = self.client.get(reverse('mes_ordonnances_assure'))
        ordonnances = list(response.context['ordonnances'])
        self.assertEqual(ordonnances, [ordonnance])

        response = self.client.get(reverse('mon_historique_assure'))
        consultations = list(response.context['consultations'])
        self.assertEqual([c.patient for c in consultations], [patient])

    def test_ne_peut_pas_voir_ordonnance_dun_autre_foyer(self):
        self._completer_profil()
        medecin = creer_medecin('medecin-rdv5@santesn.sn')
        autre_patient = creer_patient(nom='Kane', prenom='Modou')
        autre_ordonnance = creer_ordonnance(autre_patient, medecin)

        response = self.client.get(reverse('voir_ordonnance_assure', args=[autre_ordonnance.pk]))
        self.assertEqual(response.status_code, 404)


class AdminPrestatairesTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_liste_prestataires_interdite_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('liste_prestataires'))
        self.assertEqual(response.status_code, 403)

    def test_formulaire_prestataire_carte_accessible(self):
        """Plan de direction artistique, item 7 : passe accessibilite."""
        response = self.client.get(reverse('ajouter_prestataire'))
        self.assertContains(response, 'role="group"')
        self.assertContains(response, 'aria-labelledby="titre-carte-prestataire"')
        self.assertContains(response, 'aria-live="polite"')

        prestataire = Prestataire.objects.create(nom='Hopital Accessible', type_prestataire='HOPITAL')
        response = self.client.get(reverse('modifier_prestataire', args=[prestataire.pk]))
        self.assertContains(response, 'role="group"')
        self.assertContains(response, 'aria-live="polite"')

    def test_creation_prestataire(self):
        response = self.client.post(reverse('ajouter_prestataire'), {
            'nom': 'Clinique Pasteur',
            'type_prestataire': 'CLINIQUE',
            'adresse': 'Dakar',
            'ville': 'Dakar',
            'telephone': '338000000',
            'partenaire': 'on',
            'date_conventionnement': '',
        })
        self.assertRedirects(response, reverse('liste_prestataires'))
        self.assertTrue(Prestataire.objects.filter(nom='Clinique Pasteur').exists())

    def test_modification_et_suppression_prestataire(self):
        prestataire = Prestataire.objects.create(nom='Hopital Test', type_prestataire='HOPITAL')
        response = self.client.post(reverse('modifier_prestataire', args=[prestataire.pk]), {
            'nom': 'Hopital Renomme',
            'type_prestataire': 'HOPITAL',
            'adresse': '', 'ville': '', 'telephone': '',
            'partenaire': 'on', 'date_conventionnement': '',
        })
        self.assertRedirects(response, reverse('liste_prestataires'))
        prestataire.refresh_from_db()
        self.assertEqual(prestataire.nom, 'Hopital Renomme')

        response = self.client.post(reverse('supprimer_prestataire', args=[prestataire.pk]))
        self.assertRedirects(response, reverse('liste_prestataires'))
        self.assertFalse(Prestataire.objects.filter(pk=prestataire.pk).exists())

    def test_liste_prestataires_filtre_sans_localisation(self):
        Prestataire.objects.create(nom='Avec position', type_prestataire='HOPITAL',
                                    latitude=Decimal('14.6928'), longitude=Decimal('-17.4467'))
        Prestataire.objects.create(nom='Sans position', type_prestataire='HOPITAL')

        response = self.client.get(reverse('liste_prestataires'), {'localisation': 'sans'})
        noms = [prestataire.nom for prestataire in response.context['prestataires']]
        self.assertEqual(noms, ['Sans position'])

        response = self.client.get(reverse('liste_prestataires'), {'localisation': 'avec'})
        noms = [prestataire.nom for prestataire in response.context['prestataires']]
        self.assertEqual(noms, ['Avec position'])

    def test_carte_reseau_ignore_les_prestataires_sans_coordonnees(self):
        Prestataire.objects.create(nom='Sans coordonnees', type_prestataire='HOPITAL', partenaire=True)
        Prestataire.objects.create(
            nom='Avec coordonnees', type_prestataire='HOPITAL', partenaire=True,
            latitude=Decimal('14.6928'), longitude=Decimal('-17.4467'),
        )
        response = self.client.get(reverse('liste_prestataires'))
        self.assertEqual(len(response.context['prestataires_carte']), 1)
        self.assertContains(response, 'carte-reseau-admin')

    def test_pas_de_carte_sans_prestataire_geolocalise(self):
        response = self.client.get(reverse('liste_prestataires'))
        self.assertNotContains(response, 'carte-reseau-admin')

    def test_liste_prestataires_triable_par_numero(self):
        ancien = Prestataire.objects.create(nom='Ancien', type_prestataire='HOPITAL')
        recent = Prestataire.objects.create(nom='Recent', type_prestataire='HOPITAL')

        response = self.client.get(reverse('liste_prestataires'), {'tri': '-id'})
        prestataires = list(response.context['prestataires'])
        self.assertEqual(prestataires[0].pk, recent.pk)
        self.assertEqual(prestataires[1].pk, ancien.pk)
        self.assertContains(response, '?tri=id')


def _reponse_nominatim(payload):
    """Simule le context manager renvoye par urllib.request.urlopen."""
    mock_reponse = MagicMock()
    mock_reponse.__enter__.return_value.read.return_value = json.dumps(payload).encode('utf-8')
    return mock_reponse


class RechercheLieuPrestataireTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_interdite_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('recherche_lieu_prestataire'), {'q': 'Dakar'})
        self.assertEqual(response.status_code, 403)

    def test_sans_query_renvoie_non_trouve(self):
        response = self.client.get(reverse('recherche_lieu_prestataire'))
        self.assertEqual(response.json(), {'trouve': False})

    @patch('Plateform_medicale.views.urllib.request.urlopen')
    def test_lieu_trouve(self, mock_urlopen):
        mock_urlopen.return_value = _reponse_nominatim([
            {'lat': '14.7645', 'lon': '-16.9557', 'display_name': 'Thies, Senegal'},
        ])
        response = self.client.get(reverse('recherche_lieu_prestataire'), {'q': 'Thies'})
        self.assertEqual(response.json(), {
            'trouve': True, 'lat': '14.7645', 'lon': '-16.9557', 'nom': 'Thies, Senegal',
        })

    @patch('Plateform_medicale.views.urllib.request.urlopen')
    def test_lieu_introuvable(self, mock_urlopen):
        mock_urlopen.return_value = _reponse_nominatim([])
        response = self.client.get(reverse('recherche_lieu_prestataire'), {'q': 'Zzznotarealplace'})
        self.assertEqual(response.json(), {'trouve': False})

    @patch('Plateform_medicale.views.urllib.request.urlopen',
           side_effect=urllib.error.URLError('offline'))
    def test_service_indisponible_ne_plante_pas(self, mock_urlopen):
        response = self.client.get(reverse('recherche_lieu_prestataire'), {'q': 'Dakar'})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['trouve'])


class DistanceKmTests(TestCase):
    def test_meme_point_distance_nulle(self):
        self.assertEqual(distance_km(14.6928, -17.4467, 14.6928, -17.4467), 0)

    def test_un_degre_de_latitude(self):
        resultat = distance_km(0, 0, 1, 0)
        self.assertAlmostEqual(resultat, 111.19, delta=0.5)

    def test_un_degre_de_longitude_a_l_equateur(self):
        resultat = distance_km(0, 0, 0, 1)
        self.assertAlmostEqual(resultat, 111.19, delta=0.5)


class PrestataireCoordonneesTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_creation_prestataire_avec_coordonnees(self):
        response = self.client.post(reverse('ajouter_prestataire'), {
            'nom': 'Hopital Principal', 'type_prestataire': 'HOPITAL',
            'adresse': 'Dakar', 'ville': 'Dakar', 'telephone': '338000001',
            'partenaire': 'on', 'date_conventionnement': '',
            'latitude': '14.6928', 'longitude': '-17.4467',
        })
        self.assertRedirects(response, reverse('liste_prestataires'))
        prestataire = Prestataire.objects.get(nom='Hopital Principal')
        self.assertAlmostEqual(float(prestataire.latitude), 14.6928, places=4)
        self.assertAlmostEqual(float(prestataire.longitude), -17.4467, places=4)

    def test_creation_prestataire_sans_coordonnees_reste_valide(self):
        response = self.client.post(reverse('ajouter_prestataire'), {
            'nom': 'Cabinet Sans Pin', 'type_prestataire': 'CABINET',
            'adresse': '', 'ville': '', 'telephone': '',
            'partenaire': 'on', 'date_conventionnement': '',
            'latitude': '', 'longitude': '',
        })
        self.assertRedirects(response, reverse('liste_prestataires'))
        prestataire = Prestataire.objects.get(nom='Cabinet Sans Pin')
        self.assertIsNone(prestataire.latitude)
        self.assertIsNone(prestataire.longitude)


class AdminPharmaciensTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_affectation_pharmacien_a_une_pharmacie(self):
        pharmacien = creer_pharmacien('pharmacien@santesn.sn')
        pharmacie = Prestataire.objects.create(nom='Pharmacie Centrale', type_prestataire='PHARMACIE')
        response = self.client.post(reverse('modifier_pharmacien', args=[pharmacien.pk]), {
            'prestataire': pharmacie.pk,
        })
        self.assertRedirects(response, reverse('liste_pharmaciens'))
        pharmacien.refresh_from_db()
        self.assertEqual(pharmacien.prestataire, pharmacie)


class AdminPatientFormTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_liste_patients_triable_par_numero(self):
        ancien = creer_patient(nom='Ancien', prenom='A')
        recent = creer_patient(nom='Recent', prenom='B')

        response = self.client.get(reverse('liste_patients'), {'tri': '-id'})
        patients = list(response.context['patients'])
        self.assertEqual(patients[0].pk, recent.pk)
        self.assertEqual(patients[1].pk, ancien.pk)
        self.assertContains(response, '?tri=id')

    def test_admin_peut_attribuer_un_plan_de_couverture(self):
        patient = creer_patient()
        plan = PlanCouverture.objects.create(nom='Premium', taux_couverture=Decimal('90.00'))
        response = self.client.post(reverse('modifier_patient', args=[patient.pk]), {
            'nom': patient.nom, 'prenom': patient.prenom,
            'date_naissance': '1990-01-01', 'telephone': patient.telephone, 'adresse': '',
            'type_beneficiaire': 'PRINCIPAL', 'assure_principal': '', 'lien_parente': '',
            'plan_couverture': plan.pk,
        })
        self.assertRedirects(response, reverse('liste_patients'))
        patient.refresh_from_db()
        self.assertEqual(patient.plan_couverture, plan)

    def test_filtre_liste_patients_par_type(self):
        principal = creer_patient(nom='Diop', prenom='Moussa')
        Patient.objects.create(
            nom='Diop', prenom='Petit', date_naissance=datetime.date(2015, 1, 1),
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT, assure_principal=principal,
        )
        response = self.client.get(reverse('liste_patients'), {'type': 'AYANT_DROIT'})
        patients = list(response.context['patients'])
        self.assertEqual(len(patients), 1)
        self.assertEqual(patients[0].prenom, 'Petit')

    def test_creation_assure_principal_cree_son_compte(self):
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Ndiaye', 'prenom': 'Fatou', 'date_naissance': '1985-05-05',
            'telephone': '', 'adresse': '', 'type_beneficiaire': 'PRINCIPAL',
            'assure_principal': '', 'lien_parente': '', 'plan_couverture': '',
            'email': 'fatou.ndiaye@santesn.sn',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'fatou.ndiaye@santesn.sn')
        patient = Patient.objects.get(nom='Ndiaye')
        self.assertIsNotNone(patient.user)
        self.assertEqual(patient.user.role, User.Role.ASSURE)

    def test_creation_assure_principal_sans_email_refuse(self):
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Ndiaye', 'prenom': 'Fatou', 'date_naissance': '1985-05-05',
            'telephone': '', 'adresse': '', 'type_beneficiaire': 'PRINCIPAL',
            'assure_principal': '', 'lien_parente': '', 'plan_couverture': '', 'email': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.filter(nom='Ndiaye').exists())

    def test_creation_ayant_droit_ne_cree_pas_de_compte(self):
        principal = creer_patient(nom='Diop', prenom='Moussa')
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Diop', 'prenom': 'Petit', 'date_naissance': '2015-01-01',
            'telephone': '', 'adresse': '', 'type_beneficiaire': 'AYANT_DROIT',
            'assure_principal': principal.pk, 'lien_parente': 'ENFANT', 'plan_couverture': '',
            'email': '',
        })
        self.assertRedirects(response, reverse('liste_patients'))
        ayant_droit = Patient.objects.get(nom='Diop', prenom='Petit')
        self.assertIsNone(ayant_droit.user)

    def test_suppression_assure_principal_desactive_le_compte_lie(self):
        utilisateur = creer_utilisateur(User.Role.ASSURE, 'assure.a.supprimer@santesn.sn')
        patient = Patient.objects.create(
            user=utilisateur, nom='Sarr', prenom='Khady',
            date_naissance=datetime.date(1990, 1, 1), telephone='770000009',
        )
        self.assertTrue(utilisateur.is_active)
        response = self.client.post(reverse('supprimer_patient', args=[patient.pk]))
        self.assertRedirects(response, reverse('liste_patients'))
        utilisateur.refresh_from_db()
        self.assertFalse(utilisateur.is_active)

    def test_suppression_ayant_droit_par_admin_ne_touche_aucun_compte(self):
        principal = creer_patient(nom='Diop', prenom='Moussa')
        ayant_droit = Patient.objects.create(
            nom='Diop', prenom='Petit', date_naissance=datetime.date(2015, 1, 1),
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT, assure_principal=principal,
        )
        response = self.client.post(reverse('supprimer_patient', args=[ayant_droit.pk]))
        self.assertRedirects(response, reverse('liste_patients'))
        self.assertFalse(Patient.objects.filter(pk=ayant_droit.pk).exists())


class NotificationsTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_envoi_a_un_utilisateur_precis(self):
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        response = self.client.post(reverse('envoyer_notification'), {
            'destinataire': medecin_user.pk,
            'role': '',
            'message': 'Reunion demain a 9h',
        })
        self.assertRedirects(response, reverse('liste_notifications_envoyees'))
        self.assertEqual(Notification.objects.filter(destinataire=medecin_user).count(), 1)

    def test_envoi_a_tout_un_role(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin1@santesn.sn')
        creer_utilisateur(User.Role.MEDECIN, 'medecin2@santesn.sn')
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien1@santesn.sn')

        response = self.client.post(reverse('envoyer_notification'), {
            'destinataire': '',
            'role': User.Role.MEDECIN.value,
            'message': 'Mise a jour du protocole',
        })
        self.assertRedirects(response, reverse('liste_notifications_envoyees'))
        self.assertEqual(Notification.objects.count(), 2)

    def test_ni_destinataire_ni_role_refuse(self):
        response = self.client.post(reverse('envoyer_notification'), {
            'destinataire': '', 'role': '', 'message': 'Test',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Notification.objects.count(), 0)

    def test_utilisateur_voit_et_marque_lue_sa_propre_notification(self):
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        notification = Notification.objects.create(destinataire=medecin_user, message='Bienvenue')

        self.client.logout()
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('mes_notifications'))
        self.assertContains(response, 'Bienvenue')

        response = self.client.post(reverse('marquer_notification_lue', args=[notification.pk]))
        self.assertRedirects(response, reverse('mes_notifications'))
        notification.refresh_from_db()
        self.assertTrue(notification.lue)

    def test_ne_peut_pas_marquer_la_notification_dun_autre(self):
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        autre_user = creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien@santesn.sn')
        notification = Notification.objects.create(destinataire=medecin_user, message='Confidentiel')

        self.client.logout()
        self.client.login(username='pharmacien@santesn.sn', password=PASSWORD)
        response = self.client.post(reverse('marquer_notification_lue', args=[notification.pk]))
        self.assertEqual(response.status_code, 404)

    def test_filtre_lue_sur_mes_notifications(self):
        """Plan de direction artistique, item 6 : filtres lu/non-lu + recherche."""
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        lue = Notification.objects.create(destinataire=medecin_user, message='Deja lue', lue=True)
        Notification.objects.create(destinataire=medecin_user, message='Pas encore lue', lue=False)

        self.client.logout()
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)

        response = self.client.get(reverse('mes_notifications'), {'lue': 'non'})
        self.assertContains(response, 'Pas encore lue')
        self.assertNotContains(response, 'Deja lue')

        response = self.client.get(reverse('mes_notifications'), {'lue': 'oui'})
        self.assertContains(response, 'Deja lue')
        self.assertNotContains(response, 'Pas encore lue')

    def test_filtre_lue_sur_liste_notifications_envoyees(self):
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        Notification.objects.create(destinataire=medecin_user, message='Deja lue', lue=True)
        Notification.objects.create(destinataire=medecin_user, message='Pas encore lue', lue=False)

        response = self.client.get(reverse('liste_notifications_envoyees'), {'lue': 'non'})
        self.assertContains(response, 'Pas encore lue')
        self.assertNotContains(response, 'Deja lue')

    def test_recherche_sur_liste_notifications_envoyees(self):
        medecin_user = creer_utilisateur(User.Role.MEDECIN, 'ousmane.fall@santesn.sn')
        autre_user = creer_utilisateur(User.Role.PHARMACIEN, 'autre@santesn.sn')
        Notification.objects.create(destinataire=medecin_user, message='Reunion demain')
        Notification.objects.create(destinataire=autre_user, message='Livraison de stock')

        response = self.client.get(reverse('liste_notifications_envoyees'), {'q': 'ousmane'})
        self.assertContains(response, 'Reunion demain')
        self.assertNotContains(response, 'Livraison de stock')

        response = self.client.get(reverse('liste_notifications_envoyees'), {'q': 'livraison'})
        self.assertContains(response, 'Livraison de stock')
        self.assertNotContains(response, 'Reunion demain')


class RapportsTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_rapports_interdit_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('rapports'))
        self.assertEqual(response.status_code, 403)

    def test_rapports_accessible_a_l_admin(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        response = self.client.get(reverse('rapports'))
        self.assertEqual(response.status_code, 200)
        roles = {ligne['label']: ligne['total'] for ligne in response.context['utilisateurs_par_role']}
        self.assertEqual(roles['Médecin'], 1)

    def test_rapports_inclut_consultations_par_mois(self):
        patient = creer_patient()
        medecin = creer_medecin('medecin@santesn.sn')
        Consultation.objects.create(
            patient=patient,
            medecin=medecin,
            date_consultation=timezone.now(),
            diagnostic='Controle',
        )
        response = self.client.get(reverse('rapports'))
        donnees = response.context['consultations_par_mois']
        self.assertEqual(len(donnees['labels']), 6)
        self.assertEqual(len(donnees['totaux']), 6)
        self.assertEqual(donnees['totaux'][-1], 1)

    def test_rapports_inclut_consultations_par_jour_et_par_annee(self):
        """Bascule de periode (jour/mois/annee) sur le graphique des consultations."""
        patient = creer_patient()
        medecin = creer_medecin('medecin@santesn.sn')
        Consultation.objects.create(
            patient=patient,
            medecin=medecin,
            date_consultation=timezone.now(),
            diagnostic='Controle',
        )
        response = self.client.get(reverse('rapports'))

        donnees_jour = response.context['consultations_par_jour']
        self.assertEqual(len(donnees_jour['labels']), 30)
        self.assertEqual(len(donnees_jour['totaux']), 30)
        self.assertEqual(donnees_jour['totaux'][-1], 1)

        donnees_annee = response.context['consultations_par_annee']
        self.assertEqual(len(donnees_annee['labels']), 5)
        self.assertEqual(len(donnees_annee['totaux']), 5)
        self.assertEqual(donnees_annee['totaux'][-1], 1)
        self.assertEqual(donnees_annee['labels'][-1], str(timezone.now().year))

    def test_rapports_boutons_periode_consultations_presents(self):
        response = self.client.get(reverse('rapports'))
        self.assertContains(response, 'id="boutons-periode-consultations"')
        self.assertContains(response, 'data-periode="jour"')
        self.assertContains(response, 'data-periode="mois"')
        self.assertContains(response, 'data-periode="annee"')
        self.assertContains(response, 'id="graphe-consultations"')
        self.assertContains(response, 'donnees-consultations-jour')
        self.assertContains(response, 'donnees-consultations-annee')

    def test_rapports_ne_contient_pas_de_commentaire_django_non_analyse(self):
        """Regression : {# ... #} ne supporte pas les commentaires multi-lignes
        (contrairement a {% comment %}{% endcomment %}) -- s'il en reste un, le
        texte brut (accolades comprises) fuit dans le HTML rendu, ce qui casse
        ensuite le parsing HTML en aval (le <template> du repli graphique et le
        <script> qui dessine les graphiques Chart.js finissent avales dans du
        contenu inerte, aucun des 5 graphiques ne s'affiche)."""
        response = self.client.get(reverse('rapports'))
        self.assertNotContains(response, '{#')
        self.assertNotContains(response, '#}')

    def test_export_rapports_excel_interdit_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('exporter_rapports_excel'))
        self.assertEqual(response.status_code, 403)

    def test_export_rapports_excel_contient_un_onglet_par_tableau(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        response = self.client.get(reverse('exporter_rapports_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', response['Content-Disposition'])

        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            classeur.sheetnames,
            [
                'Chiffres cles',
                'Utilisateurs par role',
                'Assures par type',
                'Rendez-vous par statut',
                'Prises en charge par statut',
                'Consultations par mois',
            ],
        )
        feuille_roles = classeur['Utilisateurs par role']
        lignes = {ligne[0].value: ligne[1].value for ligne in feuille_roles.iter_rows(min_row=2)}
        self.assertEqual(lignes['Médecin'], 1)

    def test_export_rapports_pdf_interdit_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('exporter_rapports_pdf'))
        self.assertEqual(response.status_code, 403)

    def test_export_rapports_pdf_genere_un_fichier_pdf(self):
        response = self.client.get(reverse('exporter_rapports_pdf'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))


class AdminMedecinsFormTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_creation_medecin_via_formulaire_cree_aussi_son_compte(self):
        response = self.client.post(reverse('ajouter_medecin'), {
            'nom': 'Sarr', 'prenom': 'Ibrahima', 'specialite': 'Pediatrie',
            'telephone': '770001122', 'email': 'ibrahima.sarr@santesn.sn', 'prestataire': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ibrahima.sarr@santesn.sn')
        medecin = Medecin.objects.get(email='ibrahima.sarr@santesn.sn')
        self.assertIsNotNone(medecin.user)
        self.assertEqual(medecin.user.email, 'ibrahima.sarr@santesn.sn')
        self.assertEqual(medecin.user.role, User.Role.MEDECIN)
        self.assertTrue(User.objects.filter(email='ibrahima.sarr@santesn.sn').exists())

    def test_email_duplique_refuse_proprement(self):
        Medecin.objects.create(
            nom='Ba', prenom='Ousmane', specialite='Cardiologie',
            telephone='770002233', email='dup@santesn.sn',
        )
        response = self.client.post(reverse('ajouter_medecin'), {
            'nom': 'Diop', 'prenom': 'Awa', 'specialite': 'Dermatologie',
            'telephone': '770003344', 'email': 'dup@santesn.sn', 'prestataire': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Medecin.objects.filter(nom='Diop').exists())

    def test_email_deja_utilise_par_un_compte_refuse(self):
        creer_utilisateur(User.Role.ASSURE, 'compte.existant@santesn.sn')
        response = self.client.post(reverse('ajouter_medecin'), {
            'nom': 'Diop', 'prenom': 'Awa', 'specialite': 'Dermatologie',
            'telephone': '770003344', 'email': 'compte.existant@santesn.sn', 'prestataire': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Medecin.objects.filter(nom='Diop').exists())

    def test_modification_et_suppression_medecin(self):
        medecin = Medecin.objects.create(
            nom='Fall', prenom='Modou', specialite='Generaliste',
            telephone='770004455', email='fall@santesn.sn',
        )
        response = self.client.post(reverse('modifier_medecin', args=[medecin.pk]), {
            'nom': 'Fall', 'prenom': 'Modou', 'specialite': 'Chirurgie',
            'telephone': '770004455', 'email': 'fall@santesn.sn', 'prestataire': '',
        })
        self.assertRedirects(response, reverse('liste_medecins'))
        medecin.refresh_from_db()
        self.assertEqual(medecin.specialite, 'Chirurgie')

        response = self.client.post(reverse('supprimer_medecin', args=[medecin.pk]))
        self.assertRedirects(response, reverse('liste_medecins'))
        self.assertFalse(Medecin.objects.filter(pk=medecin.pk).exists())

    def test_modifier_medecin_inexistant_donne_404(self):
        response = self.client.get(reverse('modifier_medecin', args=[9999]))
        self.assertEqual(response.status_code, 404)

    def test_suppression_medecin_desactive_le_compte_lie(self):
        medecin = creer_medecin('medecin.a.supprimer@santesn.sn')
        utilisateur = medecin.user
        self.assertTrue(utilisateur.is_active)
        response = self.client.post(reverse('supprimer_medecin', args=[medecin.pk]))
        self.assertRedirects(response, reverse('liste_medecins'))
        utilisateur.refresh_from_db()
        self.assertFalse(utilisateur.is_active)


class AdminServicesFormTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_creation_et_suppression_service(self):
        response = self.client.post(reverse('ajouter_service'), {
            'nom': 'Radiographie', 'description': 'Radio standard', 'prix': '15000', 'prestataire': '',
        })
        self.assertRedirects(response, reverse('liste_services'))
        service = ServiceMedical.objects.get(nom='Radiographie')

        response = self.client.post(reverse('supprimer_service', args=[service.pk]))
        self.assertRedirects(response, reverse('liste_services'))
        self.assertFalse(ServiceMedical.objects.filter(pk=service.pk).exists())

    def test_prix_invalide_refuse_proprement(self):
        response = self.client.post(reverse('ajouter_service'), {
            'nom': 'Analyse', 'description': '', 'prix': 'gratuit', 'prestataire': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(ServiceMedical.objects.filter(nom='Analyse').exists())

    def test_supprimer_service_inexistant_donne_404(self):
        response = self.client.post(reverse('supprimer_service', args=[9999]))
        self.assertEqual(response.status_code, 404)


class AdminPriseEnChargeFormTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)
        self.patient = creer_patient()

    def test_creation_prise_en_charge_statut_en_attente_par_defaut(self):
        response = self.client.post(reverse('ajouter_prise_en_charge'), {
            'patient': self.patient.pk, 'motif': 'Consultation generale',
        })
        self.assertRedirects(response, reverse('liste_prises_en_charge'))
        prise = PriseEnCharge.objects.get(patient=self.patient)
        self.assertEqual(prise.statut, 'en_attente')

    def test_modification_statut_prise_en_charge(self):
        prise = PriseEnCharge.objects.create(patient=self.patient, motif='Test', statut='en_attente')
        response = self.client.post(reverse('modifier_prise_en_charge', args=[prise.pk]), {
            'patient': self.patient.pk, 'motif': 'Test modifie', 'statut': 'validee',
        })
        self.assertRedirects(response, reverse('liste_prises_en_charge'))
        prise.refresh_from_db()
        self.assertEqual(prise.statut, 'validee')
        self.assertEqual(prise.motif, 'Test modifie')

    def test_suppression_prise_en_charge(self):
        prise = PriseEnCharge.objects.create(patient=self.patient, motif='Test', statut='en_attente')
        response = self.client.post(reverse('supprimer_prise_en_charge', args=[prise.pk]))
        self.assertRedirects(response, reverse('liste_prises_en_charge'))
        self.assertFalse(PriseEnCharge.objects.filter(pk=prise.pk).exists())

    def test_suppression_prise_en_charge_inexistante_donne_404(self):
        response = self.client.post(reverse('supprimer_prise_en_charge', args=[9999]))
        self.assertEqual(response.status_code, 404)

    def test_liste_prises_en_charge_filtre_par_statut(self):
        PriseEnCharge.objects.create(patient=self.patient, motif='Motif attente', statut='en_attente')
        PriseEnCharge.objects.create(patient=self.patient, motif='Motif valide', statut='validee')

        response = self.client.get(reverse('liste_prises_en_charge'), {'statut': 'en_attente'})
        self.assertContains(response, 'Motif attente')
        self.assertNotContains(response, 'Motif valide')


class PaiementTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin@santesn.sn')
        self.patient = creer_patient()
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        self.service = ServiceMedical.objects.create(nom='Consultation generale', prix=Decimal('10000'))

    def test_consultation_avec_prise_en_charge_validee_calcule_les_parts(self):
        plan = PlanCouverture.objects.create(nom='Standard', taux_couverture=Decimal('80.00'))
        self.patient.plan_couverture = plan
        self.patient.save()
        prise = PriseEnCharge.objects.create(patient=self.patient, motif='Test', statut='validee')

        self.client.post(reverse('ajouter_consultation_medecin'), {
            'patient': self.patient.pk,
            'service': self.service.pk,
            'prise_en_charge': prise.pk,
            'date_consultation': '2026-08-01T10:00',
            'diagnostic': 'Controle',
            'traitement': '',
        })
        paiement = Consultation.objects.get(patient=self.patient).paiement
        self.assertEqual(paiement.montant_total, Decimal('10000'))
        self.assertEqual(paiement.taux_applique, Decimal('80.00'))
        self.assertEqual(paiement.montant_part_assurance, Decimal('8000.00'))
        self.assertEqual(paiement.montant_part_patient, Decimal('2000.00'))
        self.assertEqual(paiement.statut, Paiement.Statut.NON_REGLE)

    def test_consultation_sans_prise_en_charge_validee_patient_paie_tout(self):
        prise_en_attente = PriseEnCharge.objects.create(patient=self.patient, motif='Test')
        self.client.post(reverse('ajouter_consultation_medecin'), {
            'patient': self.patient.pk,
            'service': self.service.pk,
            'prise_en_charge': prise_en_attente.pk,
            'date_consultation': '2026-08-01T10:00',
            'diagnostic': 'Controle',
            'traitement': '',
        })
        paiement = Consultation.objects.get(patient=self.patient).paiement
        self.assertEqual(paiement.montant_part_assurance, Decimal('0'))
        self.assertEqual(paiement.montant_part_patient, Decimal('10000'))

    def test_liste_paiements_interdite_aux_non_admins(self):
        response = self.client.get(reverse('liste_paiements'))
        self.assertEqual(response.status_code, 403)

    def test_liste_paiements_affiche_le_paiement_a_l_admin(self):
        self.client.logout()
        creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

        consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=self.service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        Paiement.calculer_pour(consultation).save()

        response = self.client.get(reverse('liste_paiements'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Marquer réglé')

    def test_marquer_paiement_regle(self):
        self.client.logout()
        creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

        consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=self.service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.save()

        response = self.client.post(reverse('marquer_paiement_regle', args=[paiement.pk]), {
            'mode_reglement': 'ESPECES',
        })
        self.assertRedirects(response, reverse('liste_paiements'))
        paiement.refresh_from_db()
        self.assertEqual(paiement.statut, Paiement.Statut.REGLE)
        self.assertEqual(paiement.mode_reglement, 'ESPECES')
        self.assertIsNotNone(paiement.date_reglement)

    def test_marquer_paiement_regle_exige_mode_reglement(self):
        self.client.logout()
        creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

        consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=self.service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        paiement = Paiement.calculer_pour(consultation)
        paiement.save()

        response = self.client.post(reverse('marquer_paiement_regle', args=[paiement.pk]), {
            'mode_reglement': '',
        })
        self.assertEqual(response.status_code, 200)
        paiement.refresh_from_db()
        self.assertEqual(paiement.statut, Paiement.Statut.NON_REGLE)


class PlanCouvertureAdminTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_liste_interdite_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('liste_plans_couverture'))
        self.assertEqual(response.status_code, 403)

    def test_creation_modification_suppression_plan(self):
        response = self.client.post(reverse('ajouter_plan_couverture'), {
            'nom': 'Essentiel', 'taux_couverture': '70.00', 'plafond_annuel': '',
        })
        self.assertRedirects(response, reverse('liste_plans_couverture'))
        plan = PlanCouverture.objects.get(nom='Essentiel')

        response = self.client.post(reverse('modifier_plan_couverture', args=[plan.pk]), {
            'nom': 'Essentiel+', 'taux_couverture': '75.00', 'plafond_annuel': '',
        })
        self.assertRedirects(response, reverse('liste_plans_couverture'))
        plan.refresh_from_db()
        self.assertEqual(plan.nom, 'Essentiel+')

        response = self.client.post(reverse('supprimer_plan_couverture', args=[plan.pk]))
        self.assertRedirects(response, reverse('liste_plans_couverture'))
        self.assertFalse(PlanCouverture.objects.filter(pk=plan.pk).exists())


class SuppressionCascadeTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_avertissement_suppression_patient_avec_ayant_droit(self):
        patient = creer_patient()
        Patient.objects.create(
            nom='Diop', prenom='Petit', date_naissance=datetime.date(2015, 1, 1),
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT, assure_principal=patient,
        )
        response = self.client.get(reverse('supprimer_patient', args=[patient.pk]))
        self.assertContains(response, 'ayant(s) droit')

    def test_pas_davertissement_si_aucune_donnee_liee(self):
        patient = creer_patient()
        response = self.client.get(reverse('supprimer_patient', args=[patient.pk]))
        self.assertNotContains(response, 'Seront aussi supprimés')


class PharmacienSuppressionCompteTests(TestCase):
    def test_suppression_compte_pharmacien_preserve_la_fiche(self):
        admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        pharmacien = creer_pharmacien('pharmacien@santesn.sn')
        utilisateur_pharmacien = pharmacien.user

        self.client.login(username='admin@santesn.sn', password=PASSWORD)
        response = self.client.post(reverse('supprimer_utilisateur', args=[utilisateur_pharmacien.pk]))
        self.assertRedirects(response, reverse('liste_utilisateurs'))

        pharmacien.refresh_from_db()
        self.assertIsNone(pharmacien.user)


class ValidationFormulairesTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_patient_ayant_droit_sans_principal_refuse(self):
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Diop', 'prenom': 'Fatou', 'date_naissance': '2015-01-01',
            'telephone': '', 'adresse': '',
            'type_beneficiaire': 'AYANT_DROIT', 'assure_principal': '', 'lien_parente': 'ENFANT',
            'plan_couverture': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.filter(nom='Diop', prenom='Fatou').exists())

    def test_patient_principal_avec_principal_refuse(self):
        autre_principal = creer_patient(nom='Sow', prenom='Awa')
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Diop', 'prenom': 'Fatou', 'date_naissance': '1990-01-01',
            'telephone': '', 'adresse': '',
            'type_beneficiaire': 'PRINCIPAL', 'assure_principal': autre_principal.pk,
            'lien_parente': '', 'plan_couverture': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.filter(nom='Diop', prenom='Fatou').exists())

    def test_telephone_invalide_refuse(self):
        response = self.client.post(reverse('ajouter_patient'), {
            'nom': 'Diop', 'prenom': 'Moussa', 'date_naissance': '1990-01-01',
            'telephone': 'pas-un-numero!!', 'adresse': '',
            'type_beneficiaire': 'PRINCIPAL', 'assure_principal': '', 'lien_parente': '',
            'plan_couverture': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.filter(nom='Diop', prenom='Moussa').exists())


class RendezVousDateValidationTests(TestCase):
    def setUp(self):
        self.medecin_utilisateur = creer_medecin('medecin@santesn.sn')
        self.patient = creer_patient()
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)

    def test_rendez_vous_dans_le_passe_refuse(self):
        date_passee = (timezone.now() - datetime.timedelta(days=1)).strftime('%Y-%m-%dT%H:%M')
        response = self.client.post(reverse('ajouter_rendez_vous'), {
            'patient': self.patient.pk, 'prestataire': '', 'date_heure': date_passee, 'motif': 'Test',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(RendezVous.objects.filter(patient=self.patient).exists())


class ConsultationPriseEnChargeValidationTests(TestCase):
    def setUp(self):
        self.medecin = creer_medecin('medecin@santesn.sn')
        self.patient = creer_patient()
        self.autre_patient = creer_patient(nom='Sow', prenom='Awa')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)

    def test_prise_en_charge_dun_autre_patient_refusee(self):
        prise_en_charge = PriseEnCharge.objects.create(patient=self.autre_patient, motif='Autre')
        response = self.client.post(reverse('ajouter_consultation_medecin'), {
            'patient': self.patient.pk, 'service': '', 'prise_en_charge': prise_en_charge.pk,
            'date_consultation': '2026-08-01T10:00', 'diagnostic': 'Test', 'traitement': '',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Consultation.objects.filter(patient=self.patient).exists())


class ChangerMotDePasseTests(TestCase):
    def setUp(self):
        self.utilisateur = creer_utilisateur(User.Role.ASSURE, 'assure@santesn.sn')

    def test_exige_connexion(self):
        response = self.client.get(reverse('changer_mot_de_passe'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('login'), response.url)

    def test_changement_reussi_reste_connecte(self):
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        nouveau_mot_de_passe = 'UnAutreMotDePasseSolide2026!'
        response = self.client.post(reverse('changer_mot_de_passe'), {
            'old_password': PASSWORD,
            'new_password1': nouveau_mot_de_passe,
            'new_password2': nouveau_mot_de_passe,
        })
        # Sans fiche Patient liee, post_login_redirect renchaine vers dashboard_assure
        # puis mon_profil_assure (302 en cascade) : sans rapport avec le mot de passe.
        self.assertRedirects(response, reverse('post_login_redirect'), target_status_code=302)

        # Toujours connecte (la session n'a pas ete invalidee par le changement).
        response = self.client.get(reverse('mon_profil_assure'))
        self.assertNotEqual(response.status_code, 302)

        self.utilisateur.refresh_from_db()
        self.assertTrue(self.utilisateur.check_password(nouveau_mot_de_passe))

    def test_ancien_mot_de_passe_incorrect_refuse(self):
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.post(reverse('changer_mot_de_passe'), {
            'old_password': 'mauvais',
            'new_password1': 'UnAutreMotDePasseSolide2026!',
            'new_password2': 'UnAutreMotDePasseSolide2026!',
        })
        self.assertEqual(response.status_code, 200)
        self.utilisateur.refresh_from_db()
        self.assertTrue(self.utilisateur.check_password(PASSWORD))
        # Plan de direction artistique, item 7 : erreur annoncee aux lecteurs d'ecran.
        self.assertContains(response, 'class="erreurs" role="alert"')

    def test_confirmation_differente_refusee(self):
        self.client.login(username='assure@santesn.sn', password=PASSWORD)
        response = self.client.post(reverse('changer_mot_de_passe'), {
            'old_password': PASSWORD,
            'new_password1': 'UnAutreMotDePasseSolide2026!',
            'new_password2': 'UnMotDePasseDifferent2026!',
        })
        self.assertEqual(response.status_code, 200)
        self.utilisateur.refresh_from_db()
        self.assertTrue(self.utilisateur.check_password(PASSWORD))


class ExportUtilisateursExcelTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_interdit_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('exporter_utilisateurs_excel'))
        self.assertEqual(response.status_code, 403)

    def test_export_contient_le_bon_type_de_contenu_et_les_utilisateurs(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin1@santesn.sn')
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien1@santesn.sn')

        response = self.client.get(reverse('exporter_utilisateurs_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', response['Content-Disposition'])

        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        feuille = classeur.active
        emails = [ligne[0].value for ligne in feuille.iter_rows(min_row=2)]
        self.assertIn('medecin1@santesn.sn', emails)
        self.assertIn('pharmacien1@santesn.sn', emails)
        self.assertIn('admin@santesn.sn', emails)

    def test_export_respecte_le_filtre_par_role(self):
        creer_utilisateur(User.Role.MEDECIN, 'medecin1@santesn.sn')
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien1@santesn.sn')

        response = self.client.get(reverse('exporter_utilisateurs_excel'), {'role': 'MEDECIN'})
        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        feuille = classeur.active
        emails = [ligne[0].value for ligne in feuille.iter_rows(min_row=2)]
        self.assertIn('medecin1@santesn.sn', emails)
        self.assertNotIn('pharmacien1@santesn.sn', emails)


class ImportUtilisateursExcelTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_page_interdite_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('importer_utilisateurs_excel'))
        self.assertEqual(response.status_code, 403)

    def test_page_accessible_a_l_admin(self):
        response = self.client.get(reverse('importer_utilisateurs_excel'))
        self.assertEqual(response.status_code, 200)

    def test_modele_interdit_aux_non_admins(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('telecharger_modele_import_utilisateurs'))
        self.assertEqual(response.status_code, 403)

    def test_modele_contient_les_bonnes_entetes(self):
        response = self.client.get(reverse('telecharger_modele_import_utilisateurs'))
        self.assertEqual(response.status_code, 200)
        classeur = openpyxl.load_workbook(io.BytesIO(response.content))
        entetes = [cellule.value for cellule in next(classeur.active.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(entetes, ENTETES_IMPORT_UTILISATEURS)

    def test_import_multi_role_cree_les_comptes_et_fiches(self):
        prestataire = Prestataire.objects.create(nom='Hopital Test', type_prestataire='HOPITAL')
        plan = PlanCouverture.objects.create(nom='Standard', taux_couverture=Decimal('80.00'))
        fichier = creer_fichier_import_utilisateurs([
            ['fatou@ex.sn', 'Fatou', 'Ndiaye', '770000001', 'Assure', '15/03/1990', '', '', 'Standard'],
            ['jean@ex.sn', 'Jean', 'Diallo', '770000002', 'Medecin', '', 'Cardiologie', 'Hopital Test', ''],
            ['awa@ex.sn', 'Awa', 'Sow', '770000003', 'Pharmacien', '', '', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'fatou@ex.sn')

        patient = Patient.objects.get(user__email='fatou@ex.sn')
        self.assertEqual(patient.type_beneficiaire, Patient.TypeBeneficiaire.PRINCIPAL)
        self.assertEqual(patient.plan_couverture, plan)

        medecin = Medecin.objects.get(email='jean@ex.sn')
        self.assertEqual(medecin.specialite, 'Cardiologie')
        self.assertEqual(medecin.prestataire, prestataire)

        pharmacien = Pharmacien.objects.get(user__email='awa@ex.sn')
        self.assertIsNotNone(pharmacien.user)

    def test_import_bloque_tout_si_une_ligne_est_invalide(self):
        fichier = creer_fichier_import_utilisateurs([
            ['valide@ex.sn', 'Valide', 'Test', '770000001', 'Pharmacien', '', '', '', ''],
            ['invalide@ex.sn', 'Invalide', 'Test', '770000002', 'RoleInconnu', '', '', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(User.objects.filter(email='valide@ex.sn').exists())
        self.assertFalse(User.objects.filter(email='invalide@ex.sn').exists())

    def test_import_email_deja_existant_bloque_tout(self):
        creer_utilisateur(User.Role.MEDECIN, 'existant@ex.sn')
        fichier = creer_fichier_import_utilisateurs([
            ['nouveau@ex.sn', 'Nouveau', 'Test', '770000001', 'Pharmacien', '', '', '', ''],
            ['existant@ex.sn', 'Existant', 'Test', '770000002', 'Pharmacien', '', '', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertFalse(User.objects.filter(email='nouveau@ex.sn').exists())

    def test_import_email_duplique_dans_le_fichier_bloque_tout(self):
        fichier = creer_fichier_import_utilisateurs([
            ['double@ex.sn', 'Un', 'Test', '770000001', 'Pharmacien', '', '', '', ''],
            ['double@ex.sn', 'Deux', 'Test', '770000002', 'Pharmacien', '', '', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertEqual(User.objects.filter(email='double@ex.sn').count(), 0)

    def test_import_assure_sans_date_naissance_bloque_tout(self):
        fichier = creer_fichier_import_utilisateurs([
            ['sans.date@ex.sn', 'Sans', 'Date', '770000001', 'Assure', '', '', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertFalse(User.objects.filter(email='sans.date@ex.sn').exists())

    def test_import_medecin_sans_telephone_bloque_tout(self):
        fichier = creer_fichier_import_utilisateurs([
            ['sans.tel@ex.sn', 'Sans', 'Tel', '', 'Medecin', '', 'Cardiologie', '', ''],
        ])
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertFalse(User.objects.filter(email='sans.tel@ex.sn').exists())

    def test_import_entetes_invalides_est_refuse(self):
        fichier = creer_fichier_import_utilisateurs(
            [['x@ex.sn', 'X', 'Y', '770000001', 'Pharmacien', '', '', '', '']],
            entetes=['Colonne1', 'Colonne2'],
        )
        response = self.client.post(reverse('importer_utilisateurs_excel'), {'fichier': fichier})
        self.assertFalse(User.objects.filter(email='x@ex.sn').exists())


class PrestatairesProchesTests(TestCase):
    def setUp(self):
        self.utilisateur = creer_utilisateur(User.Role.ASSURE, 'assure1@santesn.sn')
        self.client.login(username='assure1@santesn.sn', password=PASSWORD)
        self.proche = Prestataire.objects.create(
            nom='Clinique Proche', type_prestataire='CLINIQUE', partenaire=True,
            ville='Dakar', latitude=Decimal('14.6928'), longitude=Decimal('-17.4467'),
        )
        self.loin = Prestataire.objects.create(
            nom='Hopital Lointain', type_prestataire='HOPITAL', partenaire=True,
            ville='Saint-Louis', latitude=Decimal('16.0179'), longitude=Decimal('-16.4896'),
        )
        self.sans_coordonnees = Prestataire.objects.create(
            nom='Cabinet Sans Pin', type_prestataire='CABINET', partenaire=True, ville='Dakar',
        )
        self.non_partenaire = Prestataire.objects.create(
            nom='Ancien Partenaire', type_prestataire='CLINIQUE', partenaire=False,
            ville='Dakar', latitude=Decimal('14.70'), longitude=Decimal('-17.44'),
        )

    def test_interdit_aux_non_assures(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'medecin@santesn.sn')
        self.client.login(username='medecin@santesn.sn', password=PASSWORD)
        response = self.client.get(reverse('prestataires_proches'))
        self.assertEqual(response.status_code, 403)

    def test_sans_localisation_liste_non_triee(self):
        response = self.client.get(reverse('prestataires_proches'))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['localisation_active'])
        noms = [p.nom for p, distance in response.context['prestataires_tries']]
        self.assertIn('Clinique Proche', noms)
        self.assertIn('Hopital Lointain', noms)

    def test_avec_localisation_tri_par_distance(self):
        response = self.client.get(reverse('prestataires_proches'), {'lat': '14.6928', 'lng': '-17.4467'})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['localisation_active'])
        resultats = response.context['prestataires_tries']
        self.assertEqual(resultats[0][0], self.proche)
        self.assertEqual(resultats[1][0], self.loin)
        self.assertLess(resultats[0][1], resultats[1][1])

    def test_prestataire_sans_coordonnees_affiche_a_part(self):
        response = self.client.get(reverse('prestataires_proches'))
        noms_tries = [p.nom for p, distance in response.context['prestataires_tries']]
        self.assertNotIn('Cabinet Sans Pin', noms_tries)
        noms_sans_coordonnees = [p.nom for p in response.context['prestataires_sans_coordonnees']]
        self.assertIn('Cabinet Sans Pin', noms_sans_coordonnees)

    def test_prestataire_non_partenaire_absent(self):
        response = self.client.get(reverse('prestataires_proches'))
        noms_tries = [p.nom for p, distance in response.context['prestataires_tries']]
        noms_sans_coordonnees = [p.nom for p in response.context['prestataires_sans_coordonnees']]
        self.assertNotIn('Ancien Partenaire', noms_tries)
        self.assertNotIn('Ancien Partenaire', noms_sans_coordonnees)

    def test_lat_lng_invalides_ignores(self):
        response = self.client.get(reverse('prestataires_proches'), {'lat': 'abc', 'lng': 'def'})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['localisation_active'])

    def test_geojson_contient_les_champs_enrichis(self):
        Medecin.objects.create(
            nom='Diallo', prenom='Awa', specialite='Generaliste',
            telephone='338001122', email='awa.diallo@example.sn',
            prestataire=self.proche,
        )
        response = self.client.get(reverse('prestataires_proches'))
        geojson = response.context['prestataires_geojson']
        item = next(p for p in geojson if p['nom'] == 'Clinique Proche')
        self.assertEqual(item['pk'], self.proche.pk)
        self.assertEqual(item['type_code'], 'CLINIQUE')
        self.assertEqual(item['telephone'], '')
        self.assertEqual(item['medecin_count'], 1)


class PaginationListesAdminTests(TestCase):
    """Plan de direction artistique, item 1 : pagination des listes admin."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin@santesn.sn')
        self.client.login(username='admin@santesn.sn', password=PASSWORD)

    def test_liste_services_repartie_sur_deux_pages(self):
        for i in range(25):
            ServiceMedical.objects.create(nom=f'Service {i:02d}', prix=1000)

        premiere_page = self.client.get(reverse('liste_services'))
        page = premiere_page.context['services']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertTrue(page.has_next())
        self.assertEqual(page.paginator.num_pages, 2)
        self.assertContains(premiere_page, 'page=2')

        deuxieme_page = self.client.get(reverse('liste_services'), {'page': 2})
        page = deuxieme_page.context['services']
        self.assertEqual(len(page), 5)
        self.assertFalse(page.has_next())

    def test_pas_de_navigation_si_une_seule_page(self):
        ServiceMedical.objects.create(nom='Seul service', prix=1000)
        response = self.client.get(reverse('liste_services'))
        self.assertNotContains(response, 'class="pagination"')

    def test_page_hors_limites_repli_sur_la_derniere_page(self):
        for i in range(25):
            ServiceMedical.objects.create(nom=f'Service {i:02d}', prix=1000)
        response = self.client.get(reverse('liste_services'), {'page': 999})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['services'].number, 2)

    def test_filtre_preserve_en_changeant_de_page(self):
        for i in range(25):
            creer_utilisateur(User.Role.MEDECIN, f'medecin{i:02d}@santesn.sn')
        creer_utilisateur(User.Role.PHARMACIEN, 'pharmacien@santesn.sn')

        response = self.client.get(
            reverse('liste_utilisateurs'), {'role': User.Role.MEDECIN.value, 'page': 2}
        )
        emails = [u.email for u in response.context['utilisateurs']]
        self.assertNotIn('pharmacien@santesn.sn', emails)
        self.assertEqual(response.context['utilisateurs'].number, 2)
        self.assertContains(response, f'role={User.Role.MEDECIN.value}')


class ListeRendezVousAdminTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-rdv@santesn.sn')
        self.medecin = creer_medecin('medecin-rdv@santesn.sn')
        self.patient = creer_patient(nom='Sarr', prenom='Mariama')
        self.autre_patient = creer_patient(nom='Fall', prenom='Ousmane')
        maintenant = timezone.now()
        self.demande = RendezVous.objects.create(
            patient=self.patient,
            medecin=self.medecin,
            date_heure=maintenant + datetime.timedelta(days=1),
            statut=RendezVous.Statut.DEMANDE,
        )
        self.confirme = RendezVous.objects.create(
            patient=self.autre_patient,
            medecin=self.medecin,
            date_heure=maintenant + datetime.timedelta(days=2),
            statut=RendezVous.Statut.CONFIRME,
        )
        self.client.login(username='admin-rdv@santesn.sn', password=PASSWORD)

    def test_liste_accessible_a_l_admin(self):
        reponse = self.client.get(reverse('liste_rendez_vous'))
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, 'Mariama')
        self.assertContains(reponse, 'Ousmane')

    def test_filtre_par_statut(self):
        reponse = self.client.get(reverse('liste_rendez_vous'), {'statut': 'DEMANDE'})
        self.assertEqual(list(reponse.context['rendez_vous']), [self.demande])

    def test_recherche_par_nom_de_patient(self):
        reponse = self.client.get(reverse('liste_rendez_vous'), {'q': 'Mariama'})
        self.assertEqual(list(reponse.context['rendez_vous']), [self.demande])

    def test_role_non_admin_refuse(self):
        self.client.logout()
        creer_utilisateur(User.Role.MEDECIN, 'autre-medecin@santesn.sn')
        self.client.login(username='autre-medecin@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('liste_rendez_vous'))
        self.assertEqual(reponse.status_code, 403)

    def test_anonyme_redirige_vers_connexion(self):
        self.client.logout()
        reponse = self.client.get(reverse('liste_rendez_vous'))
        self.assertEqual(reponse.status_code, 302)


class ListeOrdonnancesAdminTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-ord@santesn.sn')
        self.medecin = creer_medecin('medecin-ord@santesn.sn')
        self.patient = creer_patient(nom='Ba', prenom='Aminata')
        self.pharmacien = creer_pharmacien('pharmacien-ord@santesn.sn')
        self.non_delivree = creer_ordonnance(self.patient, self.medecin)
        self.delivree = creer_ordonnance(self.patient, self.medecin, medicaments='Ibuprofene')
        Delivrance.objects.create(ordonnance=self.delivree, pharmacien=self.pharmacien)
        self.client.login(username='admin-ord@santesn.sn', password=PASSWORD)

    def test_liste_accessible_a_l_admin(self):
        reponse = self.client.get(reverse('liste_ordonnances'))
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(len(reponse.context['ordonnances']), 2)

    def test_filtre_non_delivrees(self):
        reponse = self.client.get(reverse('liste_ordonnances'), {'delivrance': 'non'})
        self.assertEqual(list(reponse.context['ordonnances']), [self.non_delivree])

    def test_filtre_delivrees(self):
        reponse = self.client.get(reverse('liste_ordonnances'), {'delivrance': 'oui'})
        self.assertEqual(list(reponse.context['ordonnances']), [self.delivree])

    def test_recherche_par_code_qr(self):
        reponse = self.client.get(
            reverse('liste_ordonnances'), {'q': self.non_delivree.code_qr}
        )
        self.assertEqual(list(reponse.context['ordonnances']), [self.non_delivree])

    def test_recherche_par_nom_de_patient(self):
        reponse = self.client.get(reverse('liste_ordonnances'), {'q': 'Aminata'})
        self.assertEqual(len(reponse.context['ordonnances']), 2)

    def test_role_non_admin_refuse(self):
        self.client.logout()
        self.client.login(username='pharmacien-ord@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('liste_ordonnances'))
        self.assertEqual(reponse.status_code, 403)


class ContexteShellTests(TestCase):
    def test_admin_recoit_les_compteurs_de_file(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-shell@santesn.sn')
        patient = creer_patient(nom='Ndour', prenom='Khady')
        PriseEnCharge.objects.create(patient=patient, motif='Test', statut='en_attente')
        self.client.login(username='admin-shell@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.context['nb_prises_en_charge_attente'], 1)
        self.assertEqual(reponse.context['nb_paiements_non_regles'], 0)

    def test_non_admin_n_a_pas_les_compteurs_admin(self):
        creer_medecin('medecin-shell@santesn.sn')
        self.client.login(username='medecin-shell@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('dashboard_medecin'))
        self.assertIsNone(reponse.context.get('nb_prises_en_charge_attente'))

    def test_anonyme_ne_declenche_aucune_requete_de_compteur(self):
        from .views import user_role
        requete = MagicMock()
        requete.user.is_authenticated = False
        contexte = user_role(requete)
        self.assertIsNone(contexte.get('nb_prises_en_charge_attente'))
        self.assertEqual(contexte['notifications_non_lues'], 0)

    def test_recherche_topbar_reservee_a_l_admin(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-rech@santesn.sn')
        self.client.login(username='admin-rech@santesn.sn', password=PASSWORD)
        self.assertContains(self.client.get(reverse('dashboard')), 'recherche-globale')

        self.client.logout()
        creer_medecin('medecin-rech@santesn.sn')
        self.client.login(username='medecin-rech@santesn.sn', password=PASSWORD)
        self.assertNotContains(
            self.client.get(reverse('dashboard_medecin')), 'recherche-globale'
        )

    def test_fil_ariane_route_connue_et_inconnue(self):
        from .templatetags.formats import libelle_page
        self.assertEqual(libelle_page('liste_ordonnances'), 'Ordonnances')
        self.assertEqual(libelle_page('route_qui_n_existe_pas'), '')


class DashboardAdminContexteTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-dash@santesn.sn')
        self.medecin = creer_medecin('medecin-dash@santesn.sn')
        self.principal = creer_patient(nom='Diallo', prenom='Abdoulaye')
        Patient.objects.create(
            nom='Diallo',
            prenom='Fatou',
            date_naissance=datetime.date(2015, 5, 5),
            telephone='770000002',
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            lien_parente=Patient.LienParente.ENFANT,
            assure_principal=self.principal,
        )
        creer_ordonnance(self.principal, self.medecin)
        RendezVous.objects.create(
            patient=self.principal,
            medecin=self.medecin,
            date_heure=timezone.now() + datetime.timedelta(days=1),
            statut=RendezVous.Statut.DEMANDE,
        )
        Prestataire.objects.create(
            nom='Hopital Test', type_prestataire=Prestataire.Type.HOPITAL, ville='Dakar'
        )
        self.client.login(username='admin-dash@santesn.sn', password=PASSWORD)

    def test_nouvelles_cles_de_contexte(self):
        contexte = self.client.get(reverse('dashboard')).context
        self.assertEqual(contexte['rdv_a_confirmer'], 1)
        self.assertEqual(contexte['ordonnances_non_delivrees'], 1)
        self.assertEqual(contexte['patients_principaux'], 1)
        self.assertEqual(contexte['ayants_droit'], 1)
        self.assertEqual(contexte['assures_sans_plan'], 1)
        self.assertEqual(contexte['medecins_sans_prestataire'], 1)
        self.assertEqual(contexte['prestataires_sans_coordonnees'], 1)
        self.assertEqual(contexte['paiements_non_regles_nb'], 0)

    def test_file_totale_agrege_les_quatre_files(self):
        contexte = self.client.get(reverse('dashboard')).context
        self.assertEqual(
            contexte['file_totale'],
            contexte['total_prises_en_charge_attente']
            + contexte['rdv_a_confirmer']
            + contexte['ordonnances_non_delivrees']
            + contexte['paiements_non_regles_nb'],
        )

    def test_prises_en_charge_en_attente_remontent_en_premier(self):
        ancienne_validee = PriseEnCharge.objects.create(
            patient=self.principal, motif='Validee recente', statut='validee'
        )
        attente = PriseEnCharge.objects.create(
            patient=self.principal, motif='En attente', statut='en_attente'
        )
        PriseEnCharge.objects.filter(pk=attente.pk).update(
            date_demande=timezone.now() - datetime.timedelta(days=30)
        )
        contexte = self.client.get(reverse('dashboard')).context
        premieres = list(contexte['dernieres_prises_en_charge'])
        self.assertEqual(premieres[0].pk, attente.pk)
        self.assertIn(ancienne_validee, premieres)

    def test_base_vide_se_rend_sans_erreur(self):
        Ordonnance.objects.all().delete()
        Consultation.objects.all().delete()
        RendezVous.objects.all().delete()
        Patient.objects.all().delete()
        Medecin.objects.all().delete()
        Prestataire.objects.all().delete()
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.context['file_totale'], 0)

    def test_non_admin_refuse(self):
        self.client.logout()
        self.client.login(username='medecin-dash@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.status_code, 403)


class CompteursFilesAttenteTests(TestCase):
    """Les pastilles du menu (context processor user_role) et le bandeau du
    dashboard affichent les memes compteurs : ils ne doivent etre calcules
    qu'une seule fois par rendu, et rester coherents entre les deux."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-files@santesn.sn')
        self.patient = creer_patient(nom='Sow', prenom='Awa')
        PriseEnCharge.objects.create(patient=self.patient, motif='A', statut='en_attente')
        PriseEnCharge.objects.create(patient=self.patient, motif='B', statut='en_attente')
        PriseEnCharge.objects.create(patient=self.patient, motif='C', statut='validee')
        self.client.login(username='admin-files@santesn.sn', password=PASSWORD)

    def test_le_compteur_n_est_pas_calcule_deux_fois(self):
        with CaptureQueriesContext(connection) as contexte:
            self.client.get(reverse('dashboard'))
        requetes_attente = [
            q['sql'] for q in contexte.captured_queries
            if 'COUNT' in q['sql'].upper()
            and 'priseencharge' in q['sql'].lower()
            and 'en_attente' in q['sql'].lower()
        ]
        self.assertEqual(
            len(requetes_attente), 1,
            f"le comptage des prises en charge en attente est execute "
            f"{len(requetes_attente)} fois : {requetes_attente}",
        )

    def test_menu_et_bandeau_affichent_le_meme_nombre(self):
        contexte = self.client.get(reverse('dashboard')).context
        self.assertEqual(
            contexte['nb_prises_en_charge_attente'],
            contexte['total_prises_en_charge_attente'],
        )
        self.assertEqual(contexte['nb_paiements_non_regles'], contexte['paiements_non_regles_nb'])
        self.assertEqual(contexte['nb_prises_en_charge_attente'], 2)

    def test_compteurs_toujours_disponibles_hors_dashboard(self):
        """Les pastilles du menu vivent sur toutes les pages admin, pas
        seulement sur le dashboard qui, lui, alimente le cache de requete."""
        contexte = self.client.get(reverse('liste_utilisateurs')).context
        self.assertEqual(contexte['nb_prises_en_charge_attente'], 2)
        self.assertEqual(contexte['nb_paiements_non_regles'], 0)


class ParametresEtMonCompteTests(TestCase):
    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-param@santesn.sn')
        self.admin.first_name = 'Awa'
        self.admin.last_name = 'Ndiaye'
        self.admin.save()
        self.client.login(username='admin-param@santesn.sn', password=PASSWORD)

    def test_page_parametres_accessible(self):
        reponse = self.client.get(reverse('parametres'))
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.context['section'], 'general')
        self.assertContains(reponse, 'Configuration de la plateforme')

    def test_chaque_categorie_ouvre_sa_propre_page(self):
        for slug in ('general', 'apparence', 'securite'):
            reponse = self.client.get(reverse('parametres_section', args=[slug]))
            self.assertEqual(reponse.status_code, 200, slug)
            self.assertEqual(reponse.context['section'], slug)

    def test_section_inconnue_renvoie_404(self):
        self.assertEqual(
            self.client.get(reverse('parametres_section', args=['inexistante'])).status_code,
            404)

    def test_aucune_section_nest_reservee_aujourdhui(self):
        # Constat, pas un souhait : depuis le retrait de "Donnees", les quatre
        # sections restantes sont ouvertes a tous les roles. Si une section
        # reservee reapparait, ce test tombe et rappelle de couvrir son acces.
        self.assertEqual(
            [s[0] for s in SECTIONS_PARAMETRES_REELLES if s[3] is not None], [])

    def test_une_section_reservee_renvoie_404_et_nest_pas_seulement_masquee(self):
        # Le mecanisme de restriction par role est teste sur un registre
        # SUBSTITUE : inventer une section reservee dans l'interface reelle
        # pour le seul besoin du test reviendrait a afficher une chose qui
        # n'existe pas.
        registre = SECTIONS_PARAMETRES_REELLES + [
            ("reservee", "Reservee", "lock", "ADMIN")]
        with patch('Plateform_medicale.views.SECTIONS_PARAMETRES', registre):
            self.assertEqual(
                self.client.get(reverse('parametres_section', args=['reservee'])).status_code,
                200)
            self.client.logout()
            creer_medecin('medecin-param@santesn.sn')
            self.client.login(username='medecin-param@santesn.sn', password=PASSWORD)
            self.assertEqual(
                self.client.get(reverse('parametres_section', args=['reservee'])).status_code,
                404)
            menu = self.client.get(reverse('parametres'))
            self.assertNotContains(menu, reverse('parametres_section', args=['reservee']))

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(self.client.get(reverse('parametres')).status_code, 302)
        self.assertEqual(self.client.get(reverse('mon_compte')).status_code, 302)
        self.assertEqual(
            self.client.get(reverse('parametres_section', args=['securite'])).status_code, 302)

    def test_modifier_son_nom_sans_mot_de_passe(self):
        reponse = self.client.post(reverse('mon_compte'), {
            'first_name': 'Awa', 'last_name': 'Sarr',
            'phone_number': '770000009', 'email': self.admin.email,
            'mot_de_passe_actuel': '',
        })
        self.assertEqual(reponse.status_code, 302)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.last_name, 'Sarr')

    def test_changer_email_exige_le_mot_de_passe(self):
        reponse = self.client.post(reverse('mon_compte'), {
            'first_name': 'Awa', 'last_name': 'Ndiaye',
            'phone_number': '', 'email': 'nouvelle@santesn.sn',
            'mot_de_passe_actuel': '',
        })
        self.assertEqual(reponse.status_code, 200)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.email, 'admin-param@santesn.sn')

    def test_changer_email_avec_le_bon_mot_de_passe(self):
        reponse = self.client.post(reverse('mon_compte'), {
            'first_name': 'Awa', 'last_name': 'Ndiaye',
            'phone_number': '', 'email': 'nouvelle@santesn.sn',
            'mot_de_passe_actuel': PASSWORD,
        })
        self.assertEqual(reponse.status_code, 302)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.email, 'nouvelle@santesn.sn')

    def test_email_deja_pris_refuse(self):
        creer_utilisateur(User.Role.MEDECIN, 'occupe@santesn.sn')
        reponse = self.client.post(reverse('mon_compte'), {
            'first_name': 'Awa', 'last_name': 'Ndiaye',
            'phone_number': '', 'email': 'occupe@santesn.sn',
            'mot_de_passe_actuel': PASSWORD,
        })
        self.assertEqual(reponse.status_code, 200)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.email, 'admin-param@santesn.sn')

    def test_on_ne_peut_pas_changer_son_role(self):
        self.client.post(reverse('mon_compte'), {
            'first_name': 'Awa', 'last_name': 'Ndiaye', 'phone_number': '',
            'email': self.admin.email, 'role': User.Role.MEDECIN,
        })
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.role, User.Role.ADMIN)

    def test_selecteur_de_theme_present(self):
        reponse = self.client.get(reverse('parametres_section', args=['apparence']))
        for choix in ('clair', 'sombre', 'systeme'):
            self.assertContains(reponse, f'data-theme-choix="{choix}"')


class DeconnecterPartoutTests(TestCase):
    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-sessions@santesn.sn')
        self.client.login(username='admin-sessions@santesn.sn', password=PASSWORD)

    def test_ferme_la_session_courante(self):
        from django.contrib.sessions.models import Session
        self.assertTrue(Session.objects.exists())
        reponse = self.client.post(reverse('deconnecter_partout'))
        self.assertRedirects(reponse, reverse('login'))
        # La page suivante doit renvoyer un anonyme vers la connexion.
        self.assertEqual(self.client.get(reverse('parametres')).status_code, 302)

    def test_ferme_les_autres_sessions_du_meme_compte(self):
        from django.contrib.sessions.models import Session
        autre = Client()
        autre.login(username='admin-sessions@santesn.sn', password=PASSWORD)
        self.assertEqual(autre.get(reverse('parametres')).status_code, 200)

        self.client.post(reverse('deconnecter_partout'))
        self.assertEqual(autre.get(reverse('parametres')).status_code, 302)

    def test_ne_touche_pas_aux_sessions_des_autres_comptes(self):
        creer_medecin('medecin-sessions@santesn.sn')
        voisin = Client()
        voisin.login(username='medecin-sessions@santesn.sn', password=PASSWORD)
        self.assertEqual(voisin.get(reverse('parametres')).status_code, 200)

        self.client.post(reverse('deconnecter_partout'))
        self.assertEqual(voisin.get(reverse('parametres')).status_code, 200)

    def test_get_refuse(self):
        self.assertEqual(self.client.get(reverse('deconnecter_partout')).status_code, 405)

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(self.client.post(reverse('deconnecter_partout')).status_code, 302)


class ParametresContenuTests(TestCase):
    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-contenu@santesn.sn')
        self.client.login(username='admin-contenu@santesn.sn', password=PASSWORD)

    def test_duree_de_session_vient_de_la_configuration(self):
        from django.conf import settings as reglages
        contexte = self.client.get(reverse('parametres_section', args=['securite'])).context
        self.assertEqual(contexte['duree_session_heures'], reglages.SESSION_COOKIE_AGE // 3600)

    def test_configuration_plateforme_vient_de_settings(self):
        from django.conf import settings as reglages
        contexte = self.client.get(reverse('parametres')).context
        self.assertEqual(contexte['fuseau_horaire'], reglages.TIME_ZONE)


    def test_sections_admin_absentes_pour_un_medecin(self):
        self.client.logout()
        creer_medecin('medecin-contenu@santesn.sn')
        self.client.login(username='medecin-contenu@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('parametres'))
        self.assertEqual(reponse.status_code, 200)
        slugs = [s['slug'] for s in reponse.context['sections']]
        self.assertNotIn('notifications', slugs)
        self.assertNotIn('donnees', slugs)
        self.assertEqual(slugs, ['general', 'apparence', 'securite'])

    def test_chaque_section_pointe_une_icone_qui_existe(self):
        """Panne silencieuse : _ICONES.get(nom, "") renvoie une chaine VIDE
        pour un nom errone. Le SVG se rend, sans dessin et sans erreur -- rien
        ne le signale a l'ecran. Meme famille que la table de couleurs des
        rapports indexee par libelle."""
        from .templatetags.icones import _ICONES
        for slug, _libelle, icone, _role in SECTIONS_PARAMETRES_REELLES:
            self.assertIn(icone, _ICONES, slug)

    def test_general_signale_l_absence_de_backend_email(self):
        """La section 'Avancé' a ete fusionnee dans 'Général' : elle ne
        portait qu'un panneau, et 'Général' n'en portait qu'un depuis que la
        carte du compte l'a quittee. Une section de plus pour un seul panneau
        n'ajoutait qu'un clic."""
        reponse = self.client.get(reverse('parametres_section', args=['general']))
        self.assertContains(reponse, 'Envoi d')
        self.assertContains(reponse, 'inactif')
        self.assertContains(reponse, 'Configuration de la plateforme')

    def test_section_avancee_nexiste_plus(self):
        self.assertEqual(
            self.client.get(reverse('parametres_section', args=['avance'])).status_code,
            404)


class RechercheOrdonnancePharmacienTests(TestCase):
    """Repli quand le code QR est inutilisable. Regle de securite : la
    recherche manuelle ne selectionne JAMAIS une ordonnance a la place du
    pharmacien, meme s'il n'y a qu'un seul resultat."""

    def setUp(self):
        self.pharmacien = creer_pharmacien('pharma-recherche@santesn.sn')
        self.medecin = creer_medecin('medecin-recherche@santesn.sn')
        self.diop = creer_patient(nom='Diop', prenom='Awa')
        self.diallo = creer_patient(nom='Diallo', prenom='Moussa')
        self.ord_diop = creer_ordonnance(self.diop, self.medecin, medicaments='Paracetamol')
        self.ord_diallo = creer_ordonnance(self.diallo, self.medecin, medicaments='Ibuprofene')
        self.client.login(username='pharma-recherche@santesn.sn', password=PASSWORD)

    # --- chemin normal : le code exact ouvre directement ---

    def test_code_exact_ouvre_l_ordonnance(self):
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': self.ord_diop.code_qr})
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.context['ordonnance'], self.ord_diop)
        self.assertIsNone(reponse.context['resultats'])

    def test_code_inexistant_n_ouvre_rien(self):
        reponse = self.client.post(reverse('scanner_ordonnance'), {'code_qr': 'RX-INEXISTANT'})
        self.assertEqual(reponse.status_code, 200)
        self.assertIsNone(reponse.context['ordonnance'])
        self.assertContains(reponse, 'Aucune ordonnance ne correspond')

    # --- repli : recherche manuelle ---

    def test_recherche_par_nom_liste_sans_selectionner(self):
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Diop'})
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(list(reponse.context['resultats']), [self.ord_diop])
        # LE POINT CRITIQUE : un seul resultat, mais rien n'est ouvert.
        self.assertIsNone(reponse.context['ordonnance'])

    def test_recherche_ambigue_ne_selectionne_jamais(self):
        autre_diop = creer_patient(nom='Diop', prenom='Fatou')
        creer_ordonnance(autre_diop, self.medecin, medicaments='Amoxicilline')
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Diop'})
        self.assertEqual(len(reponse.context['resultats']), 2)
        self.assertIsNone(reponse.context['ordonnance'])

    def test_recherche_par_fragment_de_code(self):
        fragment = self.ord_diallo.code_qr[3:9]
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': fragment})
        self.assertIn(self.ord_diallo, reponse.context['resultats'])
        self.assertIsNone(reponse.context['ordonnance'])

    def test_recherche_trop_courte_refusee(self):
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Di'})
        self.assertIsNone(reponse.context['resultats'])
        self.assertContains(reponse, 'au moins 3 caract')

    def test_recherche_sans_resultat(self):
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Zzzzz'})
        self.assertIsNone(reponse.context['ordonnance'])
        self.assertContains(reponse, 'Aucune ordonnance ne correspond')

    def test_resultats_plafonnes_et_signales(self):
        from .views import RECHERCHE_ORDONNANCE_MAX
        patient = creer_patient(nom='Ndiayeprolifique', prenom='Test')
        for _ in range(RECHERCHE_ORDONNANCE_MAX + 3):
            creer_ordonnance(patient, self.medecin)
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'recherche': 'Ndiayeprolifique'})
        self.assertEqual(len(reponse.context['resultats']), RECHERCHE_ORDONNANCE_MAX)
        self.assertTrue(reponse.context['trop_de_resultats'])

    # --- la selection d'un resultat repasse par le chemin exact ---

    def test_selectionner_un_resultat_ouvre_la_bonne_ordonnance(self):
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': self.ord_diallo.code_qr})
        self.assertEqual(reponse.context['ordonnance'], self.ord_diallo)

    # --- permissions ---

    def test_role_non_pharmacien_refuse(self):
        self.client.logout()
        creer_utilisateur(User.Role.ADMIN, 'admin-recherche@santesn.sn')
        self.client.login(username='admin-recherche@santesn.sn', password=PASSWORD)
        self.assertEqual(
            self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Diop'}).status_code, 403)

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(
            self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Diop'}).status_code, 302)

    def test_workflow_complet_scan_puis_delivrance(self):
        """Le parcours normal du pharmacien reste intact."""
        self.client.post(reverse('scanner_ordonnance'), {'code_qr': self.ord_diop.code_qr})
        # valider_delivrance exige le pk ET le code correspondant : on ne peut
        # pas valider une delivrance en devinant un identifiant.
        reponse = self.client.post(reverse('valider_delivrance', args=[self.ord_diop.pk]),
                                   {'code_qr': self.ord_diop.code_qr})
        self.assertEqual(reponse.status_code, 302)
        self.assertTrue(Delivrance.objects.filter(ordonnance=self.ord_diop).exists())

    def test_delivrance_refusee_si_le_code_ne_correspond_pas_au_pk(self):
        """Garde-fou existant, couvert ici : le pk seul ne suffit pas."""
        reponse = self.client.post(
            reverse('valider_delivrance', args=[self.ord_diop.pk]),
            {'code_qr': self.ord_diallo.code_qr},
        )
        self.assertEqual(reponse.status_code, 404)
        self.assertFalse(Delivrance.objects.filter(ordonnance=self.ord_diop).exists())


class LibellesAccentuesTests(TestCase):
    """Les libelles affiches sont accentues, SANS toucher aux valeurs stockees
    en base (qui servent de filtres et de constantes metier)."""

    def test_valeurs_en_base_inchangees(self):
        self.assertEqual(Prestataire.Type.HOPITAL.value, 'HOPITAL')
        self.assertEqual(Patient.TypeBeneficiaire.PRINCIPAL.value, 'PRINCIPAL')
        self.assertEqual(RendezVous.Statut.CONFIRME.value, 'CONFIRME')
        self.assertEqual(Paiement.Statut.NON_REGLE.value, 'non_regle')
        self.assertEqual(PriseEnCharge.STATUT_CHOICES[1][0], 'validee')

    def test_libelles_affiches_accentues(self):
        self.assertEqual(Prestataire.Type.HOPITAL.label, 'Hôpital')
        self.assertEqual(Prestataire.Type.CABINET.label, 'Cabinet médical')
        self.assertEqual(Patient.TypeBeneficiaire.PRINCIPAL.label, 'Assuré principal')
        self.assertEqual(Paiement.ModeReglement.ESPECES.label, 'Espèces')
        self.assertEqual(RendezVous.Statut.CONFIRME.label, 'Confirmé')
        self.assertEqual(RendezVous.Statut.ANNULE.label, 'Annulé')
        self.assertEqual(RendezVous.Statut.TERMINE.label, 'Terminé')
        self.assertEqual(Paiement.Statut.NON_REGLE.label, 'Non réglé')
        self.assertEqual(Paiement.Statut.REGLE.label, 'Réglé')
        self.assertEqual(dict(PriseEnCharge.STATUT_CHOICES)['validee'], 'Validée')
        self.assertEqual(dict(PriseEnCharge.STATUT_CHOICES)['refusee'], 'Refusée')
        self.assertEqual(dict(PriseEnCharge.STATUT_CHOICES)['terminee'], 'Terminée')

    def test_filtre_par_statut_fonctionne_toujours(self):
        """Garde-fou : les filtres GET utilisent la VALEUR, pas le libelle."""
        creer_utilisateur(User.Role.ADMIN, 'admin-libelles@santesn.sn')
        patient = creer_patient(nom='Fall', prenom='Ndeye')
        PriseEnCharge.objects.create(patient=patient, motif='A', statut='validee')
        PriseEnCharge.objects.create(patient=patient, motif='B', statut='refusee')
        self.client.login(username='admin-libelles@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('liste_prises_en_charge'), {'statut': 'validee'})
        self.assertEqual(len(reponse.context['prises_en_charge']), 1)
        self.assertContains(reponse, 'Validée')

    def test_rapports_ne_porte_plus_quune_seule_visualisation(self):
        """Remplace test_table_de_couleurs_des_rapports_suit_les_libelles.

        Quatre repartitions etaient tracees en plus de la courbe : 2 a 4
        valeurs chacune, posees AU-DESSUS d'un tableau replie qui portait les
        memes chiffres. Elles ont ete remplacees par ce tableau, desormais
        visible sans rien deplier.

        La table COULEURS_STATUT, indexee par LIBELLE, ne servait qu'a elles :
        elle disparait, et avec elle le piege qu'elle imposait de maintenir
        (changer un libelle dans models.py repassait les graphiques en gris,
        sans aucune erreur visible). Ce test verrouille les deux absences.
        """
        creer_utilisateur(User.Role.ADMIN, 'admin-couleurs@santesn.sn')
        self.client.login(username='admin-couleurs@santesn.sn', password=PASSWORD)
        contenu = self.client.get(reverse('rapports')).content.decode()
        self.assertEqual(contenu.count('<canvas'), 1)
        self.assertIn('graphe-consultations', contenu)
        self.assertNotIn('COULEURS_STATUT', contenu)
        # Un seul "Voir le tableau" subsiste : celui de la courbe, ou le
        # tableau accompagne un graphique reel au lieu de le remplacer.
        self.assertEqual(contenu.count('Voir le tableau'), 1)
        # Les repartitions restent lisibles, en tableau plein.
        for titre in ('Utilisateurs par rôle', 'Assurés par type',
                      'Rendez-vous par statut', 'Prises en charge par statut'):
            self.assertIn(titre, contenu)


class PaginationHorsAdminTests(TestCase):
    """Phase 2 : les 7 listes des espaces Assure / Medecin / Pharmacien
    paginaient pas, contrairement aux 13 listes admin. Une carriere de
    consultations sur une seule page finit par rendre l'ecran inutilisable."""

    def setUp(self):
        self.medecin = creer_medecin('medecin-pagination@santesn.sn')
        self.pharmacien = creer_pharmacien('pharma-pagination@santesn.sn')

        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-pagination@santesn.sn')
        self.principal = Patient.objects.create(
            user=self.assure_user, nom='Sarr', prenom='Ousmane',
            date_naissance=datetime.date(1985, 3, 3), telephone='770000010',
        )

    def _remplir(self, combien):
        """Cree assez d'elements pour depasser une page."""
        for i in range(combien):
            consultation = Consultation.objects.create(
                patient=self.principal, medecin=self.medecin,
                date_consultation=timezone.now() - datetime.timedelta(hours=i),
                diagnostic=f'Diagnostic {i}',
            )
            ordonnance = Ordonnance.objects.create(
                consultation=consultation, medicaments=f'Medicament {i}')
            RendezVous.objects.create(
                patient=self.principal, medecin=self.medecin,
                date_heure=timezone.now() + datetime.timedelta(days=i + 1),
                statut=RendezVous.Statut.CONFIRME,
            )
            Delivrance.objects.create(ordonnance=ordonnance, pharmacien=self.pharmacien)

    # --- Espace Assure ---

    def test_listes_assure_paginees(self):
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='assure-pagination@santesn.sn', password=PASSWORD)
        for nom_url, cle in [('mes_rendez_vous_assure', 'rendez_vous'),
                             ('mes_ordonnances_assure', 'ordonnances'),
                             ('mon_historique_assure', 'consultations')]:
            reponse = self.client.get(reverse(nom_url))
            page = reponse.context[cle]
            self.assertEqual(len(page), TAILLE_PAGE_LISTE, nom_url)
            self.assertEqual(page.paginator.count, TAILLE_PAGE_LISTE + 4, nom_url)
            self.assertTrue(page.has_next(), nom_url)

    def test_seconde_page_assure_accessible(self):
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='assure-pagination@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('mon_historique_assure'), {'page': 2})
        self.assertEqual(reponse.context['consultations'].number, 2)
        self.assertEqual(len(reponse.context['consultations']), 4)

    # --- Espace Medecin ---

    def test_listes_medecin_paginees(self):
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='medecin-pagination@santesn.sn', password=PASSWORD)
        for nom_url, cle in [('agenda_medecin', 'rendez_vous'),
                             ('historique_consultations', 'consultations')]:
            page = self.client.get(reverse(nom_url)).context[cle]
            self.assertEqual(len(page), TAILLE_PAGE_LISTE, nom_url)
            self.assertTrue(page.has_next(), nom_url)

    def test_mes_patients_pagine(self):
        for i in range(TAILLE_PAGE_LISTE + 3):
            patient = creer_patient(nom=f'Patient{i:02d}', prenom='Test')
            Consultation.objects.create(
                patient=patient, medecin=self.medecin,
                date_consultation=timezone.now(), diagnostic='X')
        self.client.login(username='medecin-pagination@santesn.sn', password=PASSWORD)
        page = self.client.get(reverse('mes_patients')).context['patients']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertEqual(page.paginator.count, TAILLE_PAGE_LISTE + 3)

    def test_filtres_du_medecin_survivent_au_changement_de_page(self):
        """Le filtre par patient doit rester actif page 2 (prefixe_pagination)."""
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='medecin-pagination@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('historique_consultations'),
                                  {'patient': self.principal.pk, 'page': 2})
        self.assertEqual(reponse.context['consultations'].number, 2)
        self.assertEqual(reponse.context['patient_selectionne'], str(self.principal.pk))
        self.assertContains(reponse, f'patient={self.principal.pk}')

    # --- Espace Pharmacien ---

    def test_historique_delivrances_pagine(self):
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='pharma-pagination@santesn.sn', password=PASSWORD)
        page = self.client.get(reverse('historique_delivrances')).context['delivrances']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertTrue(page.has_next())

    # --- Ordre deterministe ---

    def test_ordre_stable_entre_les_pages(self):
        """Sans order_by, la repartition entre pages serait instable : un meme
        element pourrait apparaitre deux fois ou disparaitre."""
        self._remplir(TAILLE_PAGE_LISTE + 4)
        self.client.login(username='assure-pagination@santesn.sn', password=PASSWORD)
        page1 = self.client.get(reverse('mes_rendez_vous_assure')).context['rendez_vous']
        page2 = self.client.get(reverse('mes_rendez_vous_assure'),
                                {'page': 2}).context['rendez_vous']
        ids1 = [r.pk for r in page1]
        ids2 = [r.pk for r in page2]
        self.assertEqual(len(set(ids1) & set(ids2)), 0, "chevauchement entre les pages")
        self.assertEqual(len(set(ids1) | set(ids2)), TAILLE_PAGE_LISTE + 4)

    # --- Pas de regression sur les petites listes ---

    def test_petite_liste_sans_barre_de_pagination(self):
        self._remplir(3)
        self.client.login(username='assure-pagination@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('mes_rendez_vous_assure'))
        self.assertEqual(reponse.context['rendez_vous'].paginator.num_pages, 1)
        self.assertNotContains(reponse, 'aria-label="Pagination"')


ENTETES_IMPORT_REGLEMENTS = [
    'Reference', 'Patient', 'Date de consultation', 'Part patient',
    'Mode de reglement', 'Date de reglement',
]


def creer_fichier_import_reglements(lignes, entetes=None):
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.append(entetes or ENTETES_IMPORT_REGLEMENTS)
    for ligne in lignes:
        feuille.append(ligne)
    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return SimpleUploadedFile(
        'reglements.xlsx', tampon.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class ImportReglementsTests(TestCase):
    """L'import n'ajoute pas de paiements (impossible : Paiement est en 1-1
    avec Consultation et ses montants sont derives). Il enregistre le
    reglement de paiements existants, en tout ou rien."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-reglements@santesn.sn')
        self.medecin = creer_medecin('medecin-reglements@santesn.sn')
        self.patient = creer_patient(nom='Sow', prenom='Awa')
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))
        consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test',
        )
        self.paiement = Paiement.calculer_pour(consultation)
        self.paiement.save()
        self.client.login(username='admin-reglements@santesn.sn', password=PASSWORD)

    def _importer(self, lignes, entetes=None):
        return self.client.post(reverse('importer_reglements_excel'),
                                {'fichier': creer_fichier_import_reglements(lignes, entetes)})

    # --- cas nominal ---

    def test_import_valide(self):
        reponse = self._importer([[self.paiement.pk, 'Awa Sow', '', '', 'Especes', '10/08/2026']])
        self.assertRedirects(reponse, reverse('liste_paiements'))
        self.paiement.refresh_from_db()
        self.assertEqual(self.paiement.statut, Paiement.Statut.REGLE)
        self.assertEqual(self.paiement.mode_reglement, Paiement.ModeReglement.ESPECES)
        self.assertEqual(self.paiement.date_reglement.date(), datetime.date(2026, 8, 10))

    def test_libelle_accentue_du_mode_accepte(self):
        self._importer([[self.paiement.pk, '', '', '', 'Espèces', '10/08/2026']])
        self.paiement.refresh_from_db()
        self.assertEqual(self.paiement.mode_reglement, Paiement.ModeReglement.ESPECES)

    def test_montant_de_controle_correct_accepte(self):
        self._importer([[self.paiement.pk, '', '', self.paiement.montant_part_patient,
                         'Virement', '10/08/2026']])
        self.paiement.refresh_from_db()
        self.assertEqual(self.paiement.statut, Paiement.Statut.REGLE)

    # --- cas d'erreur ---

    def _erreurs(self, lignes, entetes=None):
        reponse = self._importer(lignes, entetes)
        self.assertEqual(reponse.status_code, 200)
        self.paiement.refresh_from_db()
        self.assertEqual(self.paiement.statut, Paiement.Statut.NON_REGLE,
                         "aucun reglement ne doit passer si une ligne est invalide")
        return reponse

    def test_fichier_vide(self):
        self.assertContains(self._erreurs([]), 'aucune ligne')

    def test_colonne_manquante(self):
        reponse = self._erreurs([[self.paiement.pk, '', '', '']],
                                entetes=['Reference', 'Patient', 'Part patient', 'Mode'])
        self.assertContains(reponse, 'En-t')

    def test_reference_inexistante(self):
        self.assertContains(self._erreurs([[999999, '', '', '', 'Especes', '10/08/2026']]),
                            'aucun paiement ne porte la r')

    def test_reference_invalide(self):
        self.assertContains(self._erreurs([['abc', '', '', '', 'Especes', '10/08/2026']]),
                            'invalide')

    def test_mode_inconnu(self):
        self.assertContains(self._erreurs([[self.paiement.pk, '', '', '', 'Bitcoin', '10/08/2026']]),
                            'inconnu')

    def test_mode_absent(self):
        self.assertContains(self._erreurs([[self.paiement.pk, '', '', '', '', '10/08/2026']]),
                            'mode de r')

    def test_date_invalide(self):
        self.assertContains(self._erreurs([[self.paiement.pk, '', '', '', 'Especes', '32/13/2026']]),
                            'invalide')

    def test_date_absente(self):
        self.assertContains(self._erreurs([[self.paiement.pk, '', '', '', 'Especes', '']]),
                            'obligatoire')

    def test_date_future_refusee(self):
        futur = (timezone.localdate() + datetime.timedelta(days=5)).strftime('%d/%m/%Y')
        self.assertContains(self._erreurs([[self.paiement.pk, '', '', '', 'Especes', futur]]),
                            'futur')

    def test_montant_de_controle_incoherent(self):
        self.assertContains(
            self._erreurs([[self.paiement.pk, '', '', '999999', 'Especes', '10/08/2026']]),
            'ne correspond pas')

    def test_montant_invalide(self):
        self.assertContains(
            self._erreurs([[self.paiement.pk, '', '', 'beaucoup', 'Especes', '10/08/2026']]),
            'invalide')

    def test_doublon_dans_le_fichier(self):
        self.assertContains(
            self._erreurs([[self.paiement.pk, '', '', '', 'Especes', '10/08/2026'],
                           [self.paiement.pk, '', '', '', 'Virement', '11/08/2026']]),
            'appara')

    def test_paiement_deja_regle(self):
        self.paiement.statut = Paiement.Statut.REGLE
        self.paiement.save()
        reponse = self._importer([[self.paiement.pk, '', '', '', 'Especes', '10/08/2026']])
        self.assertContains(reponse, 'est d')

    def test_plusieurs_erreurs_signalees_ensemble(self):
        reponse = self._erreurs([
            [999999, '', '', '', 'Especes', '10/08/2026'],
            [self.paiement.pk, '', '', '', 'Bitcoin', '10/08/2026'],
        ])
        self.assertContains(reponse, 'aucun paiement ne porte la r')
        self.assertContains(reponse, 'inconnu')

    def test_tout_ou_rien(self):
        """Une ligne valide suivie d'une invalide : rien ne doit passer."""
        autre = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Autre')
        paiement2 = Paiement.calculer_pour(autre)
        paiement2.save()
        self._importer([[self.paiement.pk, '', '', '', 'Especes', '10/08/2026'],
                        [paiement2.pk, '', '', '', 'Bitcoin', '10/08/2026']])
        self.paiement.refresh_from_db()
        paiement2.refresh_from_db()
        self.assertEqual(self.paiement.statut, Paiement.Statut.NON_REGLE)
        self.assertEqual(paiement2.statut, Paiement.Statut.NON_REGLE)

    # --- permissions et aller-retour ---

    def test_role_non_admin_refuse(self):
        self.client.logout()
        creer_medecin('autre-medecin-regl@santesn.sn')
        self.client.login(username='autre-medecin-regl@santesn.sn', password=PASSWORD)
        for nom in ('importer_reglements_excel', 'telecharger_modele_import_reglements'):
            self.assertEqual(self.client.get(reverse(nom)).status_code, 403, nom)

    def test_export_csv_contient_la_reference(self):
        """Sans identifiant dans l'export, l'aller-retour serait impossible."""
        reponse = self.client.get(reverse('exporter_paiements_csv'))
        contenu = reponse.content.decode('utf-8-sig')
        self.assertIn('Reference', contenu.splitlines()[0])
        self.assertIn(str(self.paiement.pk), contenu)

    def test_modele_telechargeable(self):
        reponse = self.client.get(reverse('telecharger_modele_import_reglements'))
        self.assertEqual(reponse.status_code, 200)
        self.assertIn('spreadsheetml', reponse['Content-Type'])


class NavigationCoherenteTests(TestCase):
    """Chaque fonctionnalite a UN emplacement logique : on verifie qu'on n'a
    pas reintroduit de doublon."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-nav@santesn.sn')
        self.client.login(username='admin-nav@santesn.sn', password=PASSWORD)

    def test_notifications_ne_sont_plus_une_section_de_parametres(self):
        reponse = self.client.get(reverse('parametres'))
        slugs = [s['slug'] for s in reponse.context['sections']]
        self.assertNotIn('notifications', slugs)
        self.assertEqual(
            self.client.get(reverse('parametres_section', args=['notifications'])).status_code,
            404)

    def test_notifications_restent_dans_le_menu(self):
        self.assertContains(self.client.get(reverse('dashboard')),
                            reverse('envoyer_notification'))

    def test_import_utilisateurs_absent_des_actions_rapides(self):
        reponse = self.client.get(reverse('dashboard'))
        self.assertNotContains(reponse, reverse('importer_utilisateurs_excel'))

    def test_import_utilisateurs_uniquement_sur_sa_page_metier(self):
        # UNE fonctionnalite = UN emplacement principal. L'import vivait aussi
        # dans Parametres -> Donnees : le meme bouton, a deux endroits.
        self.assertContains(self.client.get(reverse('liste_utilisateurs')),
                            reverse('importer_utilisateurs_excel'))
        for section in SECTIONS_TOUTES:
            self.assertNotContains(
                self.client.get(reverse('parametres_section', args=[section])),
                reverse('importer_utilisateurs_excel'))

    def test_import_reglements_uniquement_sur_la_page_paiements(self):
        self.assertContains(self.client.get(reverse('liste_paiements')),
                            reverse('importer_reglements_excel'))
        for section in SECTIONS_TOUTES:
            self.assertNotContains(
                self.client.get(reverse('parametres_section', args=[section])),
                reverse('importer_reglements_excel'))

    def test_exports_uniquement_sur_leur_page_metier(self):
        # Les exports proposes dans Parametres ignoraient les filtres alors
        # que le sous-titre affirmait le contraire : deux chemins, dont un
        # qui mentait.
        for nom, page in (
            ('exporter_utilisateurs_excel', 'liste_utilisateurs'),
            ('exporter_rapports_excel', 'rapports'),
            ('exporter_paiements_csv', 'liste_paiements'),
        ):
            self.assertContains(self.client.get(reverse(page)), reverse(nom))
            for section in SECTIONS_TOUTES:
                self.assertNotContains(
                    self.client.get(reverse('parametres_section', args=[section])),
                    reverse(nom))

    def test_menu_du_compte_porte_la_deconnexion(self):
        """La deconnexion appartient au compte, pas a la navigation metier."""
        reponse = self.client.get(reverse('dashboard'))
        self.assertContains(reponse, 'menu-compte-panneau')
        self.assertContains(reponse, reverse('logout'))
        self.assertNotContains(reponse, 'class="logout-button"')

    def test_le_bloc_compte_ouvre_le_compte_pas_le_mot_de_passe(self):
        reponse = self.client.get(reverse('dashboard'))
        self.assertNotContains(reponse, f'class="topbar-compte" href="{reverse("changer_mot_de_passe")}"')


class ConfirmationsActionsSensiblesTests(TestCase):
    """Audit des confirmations : seules les actions destructives en ont une.
    Le mecanisme existe deja (data-confirmation intercepte par la modale) ;
    ces tests verifient qu'il est pose la ou il faut, et NULLE PART ailleurs."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-confirm@santesn.sn')
        self.client.login(username='admin-confirm@santesn.sn', password=PASSWORD)

    def test_deconnecter_partout_demande_confirmation(self):
        reponse = self.client.get(reverse('parametres_section', args=['securite']))
        self.assertContains(reponse, 'data-confirmation=')
        self.assertContains(reponse, 'Oui, déconnecter partout')

    def test_desactivation_confirmee_mais_pas_l_activation(self):
        actif = creer_utilisateur(User.Role.MEDECIN, 'actif-confirm@santesn.sn')
        inactif = creer_utilisateur(User.Role.MEDECIN, 'inactif-confirm@santesn.sn')
        inactif.is_active = False
        inactif.save()

        html = self.client.get(reverse('liste_utilisateurs')).content.decode()
        # La ligne du compte actif porte la confirmation (on va le desactiver),
        # celle du compte inactif non (le reactiver n'a rien de destructif).
        self.assertIn('Oui, désactiver le compte', html)
        bloc_inactif = html[html.find('inactif-confirm@santesn.sn'):]
        bloc_inactif = bloc_inactif[:bloc_inactif.find('</tr>')]
        self.assertNotIn('data-confirmation', bloc_inactif)

    def test_annulation_de_rendez_vous_confirmee(self):
        self.client.logout()
        assure = creer_utilisateur(User.Role.ASSURE, 'assure-confirm@santesn.sn')
        patient = Patient.objects.create(
            user=assure, nom='Ba', prenom='Oumar',
            date_naissance=datetime.date(1990, 2, 2), telephone='770000020')
        medecin = creer_medecin('medecin-confirm@santesn.sn')
        RendezVous.objects.create(
            patient=patient, medecin=medecin,
            date_heure=timezone.now() + datetime.timedelta(days=3),
            statut=RendezVous.Statut.CONFIRME)
        self.client.login(username='assure-confirm@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('mes_rendez_vous_assure'))
        self.assertContains(reponse, 'Oui, annuler le rendez-vous')

    def test_actions_de_lecture_sans_confirmation(self):
        """Filtrer, exporter, consulter : aucune confirmation ne doit gener."""
        for nom in ('liste_utilisateurs', 'liste_paiements', 'rapports', 'dashboard'):
            html = self.client.get(reverse(nom)).content.decode()
            debut = html.find('<form method="get"')
            if debut != -1:
                bloc = html[debut:html.find('</form>', debut)]
                self.assertNotIn('data-confirmation', bloc, nom)

    def test_icone_engrenage_pour_les_parametres(self):
        """Une roue dentee, pas un soleil.

        Le menu a d'abord utilise une cle (symbole de securite), puis un
        cercle entoure de 8 rayons droits et diagonaux -- dessine comme "un
        engrenage en segments", mais qui se lisait a l'ecran comme un SOLEIL.
        L'ancien test se contentait de chercher un "circle", que le soleil
        possedait aussi : il ne pouvait pas voir l'erreur.

        On teste desormais ce qui SEPARE les deux formes : une roue dentee a
        un contour ferme (z), un soleil n'a que des rayons ouverts.
        """
        from .templatetags.icones import _ICONES
        svg = _ICONES['settings']
        self.assertIn('<circle', svg)          # le moyeu
        self.assertIn('z"', svg)               # le contour dente se referme
        self.assertNotIn('M12 2.6v3.1', svg)   # plus aucun rayon droit
        html = self.client.get(reverse('dashboard')).content.decode()
        debut = html.find('aria-label="Paramètres"')
        self.assertNotEqual(debut, -1)
        self.assertIn('<circle', html[debut:debut + 1400])


class BlocageMecanismeTests(TestCase):
    """Le mecanisme lui-meme : compteur, seuil, expiration, remise a zero.
    Il a ete DEPLACE du cache vers la base, sa regle est inchangee."""

    def setUp(self):
        # Indispensable : login_view redirige vers l'assistant d'installation
        # tant qu'aucun administrateur n'existe.
        creer_utilisateur(User.Role.ADMIN, 'admin-meca@santesn.sn')
        self.utilisateur = creer_utilisateur(User.Role.ASSURE, 'bloc-meca@santesn.sn')

    def _echouer(self, fois, email='bloc-meca@santesn.sn'):
        for _ in range(fois):
            self.client.post(reverse('login'), {'email': email, 'password': 'faux'})

    def test_compte_non_bloque_avant_le_seuil(self):
        self._echouer(TentativeConnexion.MAX_TENTATIVES - 1)
        self.assertFalse(TentativeConnexion.bloque('bloc-meca@santesn.sn'))
        self.assertEqual(TentativeConnexion.comptes_bloques(), [])

    def test_compte_bloque_au_seuil(self):
        self._echouer(TentativeConnexion.MAX_TENTATIVES)
        self.assertTrue(TentativeConnexion.bloque('bloc-meca@santesn.sn'))
        self.assertEqual(len(TentativeConnexion.comptes_bloques()), 1)

    def test_bon_mot_de_passe_refuse_pendant_le_blocage(self):
        self._echouer(TentativeConnexion.MAX_TENTATIVES)
        reponse = self.client.post(reverse('login'),
                                   {'email': 'bloc-meca@santesn.sn', 'password': PASSWORD})
        self.assertContains(reponse, 'Trop de tentatives')
        self.assertFalse(reponse.wsgi_request.user.is_authenticated)

    def test_expiration_automatique(self):
        self._echouer(TentativeConnexion.MAX_TENTATIVES)
        ligne = TentativeConnexion.objects.get(email='bloc-meca@santesn.sn')
        # On recule le dernier echec au-dela de la duree de blocage.
        TentativeConnexion.objects.filter(pk=ligne.pk).update(
            dernier_echec=timezone.now() - TentativeConnexion.DUREE_BLOCAGE
            - datetime.timedelta(seconds=1))
        self.assertFalse(TentativeConnexion.bloque('bloc-meca@santesn.sn'))
        self.assertEqual(TentativeConnexion.comptes_bloques(), [])

    def test_reconnexion_possible_apres_expiration(self):
        self._echouer(TentativeConnexion.MAX_TENTATIVES)
        TentativeConnexion.objects.filter(email='bloc-meca@santesn.sn').update(
            dernier_echec=timezone.now() - TentativeConnexion.DUREE_BLOCAGE
            - datetime.timedelta(seconds=1))
        reponse = self.client.post(reverse('login'),
                                   {'email': 'bloc-meca@santesn.sn', 'password': PASSWORD})
        self.assertEqual(reponse.status_code, 302)

    def test_connexion_reussie_efface_le_compteur(self):
        self._echouer(2)
        self.client.post(reverse('login'),
                         {'email': 'bloc-meca@santesn.sn', 'password': PASSWORD})
        self.assertFalse(
            TentativeConnexion.objects.filter(email='bloc-meca@santesn.sn').exists())

    def test_compteur_repart_de_un_apres_expiration(self):
        """Sans cela, 5 echecs espaces dans le temps bloqueraient a vie."""
        self._echouer(TentativeConnexion.MAX_TENTATIVES - 1)
        TentativeConnexion.objects.filter(email='bloc-meca@santesn.sn').update(
            dernier_echec=timezone.now() - TentativeConnexion.DUREE_BLOCAGE
            - datetime.timedelta(seconds=1))
        self._echouer(1)
        ligne = TentativeConnexion.objects.get(email='bloc-meca@santesn.sn')
        self.assertEqual(ligne.tentatives, 1)

    def test_adresse_inexistante_freinee_mais_absente_de_la_liste(self):
        """Ne pas freiner les adresses inconnues offrirait un oracle
        permettant de deviner quels comptes existent. Mais elles n'ont rien
        a faire dans une liste de COMPTES bloques."""
        self._echouer(TentativeConnexion.MAX_TENTATIVES, email='inconnu@santesn.sn')
        self.assertTrue(TentativeConnexion.bloque('inconnu@santesn.sn'))
        self.assertEqual(TentativeConnexion.comptes_bloques(), [])

    def test_compte_desactive_indiscernable_d_identifiants_faux(self):
        """Le backend par defaut rejette les comptes inactifs AVANT que le
        formulaire ne puisse les distinguer : l'utilisateur voit le message
        generique et la tentative est comptee. C'est le comportement voulu --
        annoncer "ce compte est desactive" confirmerait qu'une adresse existe."""
        self.utilisateur.is_active = False
        self.utilisateur.save()
        reponse = self.client.post(reverse('login'),
                                   {'email': 'bloc-meca@santesn.sn', 'password': PASSWORD})
        self.assertContains(reponse, 'incorrect')
        self.assertNotContains(reponse, 'désactivé')
        self.assertTrue(
            TentativeConnexion.objects.filter(email='bloc-meca@santesn.sn').exists())


class ComptesBloquesAdminTests(TestCase):
    """Interface d'administration : liste, compteur, filtre, recherche,
    deblocage, permissions."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-bloc@santesn.sn')
        self.assure = creer_utilisateur(User.Role.ASSURE, 'awa.diallo@santesn.sn')
        self.assure.first_name, self.assure.last_name = 'Awa', 'Diallo'
        self.assure.save()
        self.medecin_user = creer_utilisateur(User.Role.MEDECIN, 'moussa.sow@santesn.sn')
        self.medecin_user.first_name, self.medecin_user.last_name = 'Moussa', 'Sow'
        self.medecin_user.save()
        self.client.login(username='admin-bloc@santesn.sn', password=PASSWORD)

    def _bloquer(self, email):
        for _ in range(TentativeConnexion.MAX_TENTATIVES):
            TentativeConnexion.enregistrer_echec(email)

    def _page(self, **params):
        return self.client.get(reverse('parametres_section', args=['securite']), params)

    # --- liste et compteur ---

    def test_aucun_compte_bloque(self):
        reponse = self._page()
        self.assertEqual(reponse.context['comptes_bloques'], [])
        self.assertContains(reponse, 'Aucun compte temporairement bloqué')

    def test_compte_normal_absent_de_la_liste(self):
        self._bloquer('awa.diallo@santesn.sn')
        emails = [c['utilisateur'].email for c in self._page().context['comptes_bloques']]
        self.assertNotIn('moussa.sow@santesn.sn', emails)

    def test_compte_bloque_present_avec_ses_informations(self):
        self._bloquer('awa.diallo@santesn.sn')
        reponse = self._page()
        comptes = reponse.context['comptes_bloques']
        self.assertEqual(len(comptes), 1)
        self.assertEqual(comptes[0]['utilisateur'], self.assure)
        self.assertEqual(comptes[0]['tentatives'], TentativeConnexion.MAX_TENTATIVES)
        self.assertGreater(comptes[0]['minutes_restantes'], 0)
        self.assertContains(reponse, 'Awa Diallo')
        self.assertContains(reponse, 'Bloqué')

    def test_compteur_reflete_le_nombre_reel(self):
        self._bloquer('awa.diallo@santesn.sn')
        self._bloquer('moussa.sow@santesn.sn')
        page = self._page()
        self.assertEqual(len(page.context['comptes_bloques']), 2)
        self.assertContains(page, '2 comptes')

    # --- filtre par role ---

    def test_filtre_par_role(self):
        self._bloquer('awa.diallo@santesn.sn')
        self._bloquer('moussa.sow@santesn.sn')
        for role, attendu in ((User.Role.ASSURE, 'awa.diallo@santesn.sn'),
                              (User.Role.MEDECIN, 'moussa.sow@santesn.sn')):
            comptes = self._page(role=role).context['comptes_bloques']
            self.assertEqual([c['utilisateur'].email for c in comptes], [attendu], role)

    def test_filtre_sur_un_role_sans_compte_bloque(self):
        self._bloquer('awa.diallo@santesn.sn')
        self.assertEqual(self._page(role=User.Role.PHARMACIEN).context['comptes_bloques'], [])

    # --- recherche ---

    def test_recherche_par_nom(self):
        self._bloquer('awa.diallo@santesn.sn')
        self._bloquer('moussa.sow@santesn.sn')
        comptes = self._page(q='Diallo').context['comptes_bloques']
        self.assertEqual([c['utilisateur'].email for c in comptes], ['awa.diallo@santesn.sn'])

    def test_recherche_par_email(self):
        self._bloquer('awa.diallo@santesn.sn')
        comptes = self._page(q='awa.diallo').context['comptes_bloques']
        self.assertEqual(len(comptes), 1)

    def test_recherche_et_role_combines(self):
        self._bloquer('awa.diallo@santesn.sn')
        self._bloquer('moussa.sow@santesn.sn')
        self.assertEqual(self._page(q='Diallo', role=User.Role.MEDECIN)
                         .context['comptes_bloques'], [])
        self.assertEqual(len(self._page(q='Sow', role=User.Role.MEDECIN)
                             .context['comptes_bloques']), 1)

    # --- deblocage ---

    def test_deblocage_retire_le_compte_et_autorise_la_connexion(self):
        self._bloquer('awa.diallo@santesn.sn')
        reponse = self.client.post(reverse('debloquer_compte', args=[self.assure.pk]))
        self.assertEqual(reponse.status_code, 302)
        self.assertFalse(TentativeConnexion.bloque('awa.diallo@santesn.sn'))
        self.assertEqual(self._page().context['comptes_bloques'], [])

        autre = Client()
        self.assertEqual(
            autre.post(reverse('login'),
                       {'email': 'awa.diallo@santesn.sn', 'password': PASSWORD}).status_code,
            302)

    def test_deblocage_demande_confirmation(self):
        self._bloquer('awa.diallo@santesn.sn')
        reponse = self._page()
        self.assertContains(reponse, 'data-confirmation=')
        self.assertContains(reponse, 'Oui, débloquer le compte')

    def test_deblocage_en_get_refuse(self):
        self._bloquer('awa.diallo@santesn.sn')
        self.assertEqual(
            self.client.get(reverse('debloquer_compte', args=[self.assure.pk])).status_code, 405)
        self.assertTrue(TentativeConnexion.bloque('awa.diallo@santesn.sn'))

    def test_deblocage_d_un_compte_deja_libre_ne_casse_rien(self):
        reponse = self.client.post(reverse('debloquer_compte', args=[self.assure.pk]))
        self.assertEqual(reponse.status_code, 302)

    def test_pas_de_deblocage_groupe(self):
        """Choix assume : un blocage massif signale souvent une attaque."""
        self._bloquer('awa.diallo@santesn.sn')
        self.assertNotContains(self._page(), 'Débloquer tous')

    # --- permissions et donnees sensibles ---

    def test_roles_non_admin_refuses(self):
        self._bloquer('awa.diallo@santesn.sn')
        self.client.logout()
        creer_medecin('medecin-bloc@santesn.sn')
        self.client.login(username='medecin-bloc@santesn.sn', password=PASSWORD)
        self.assertEqual(
            self.client.get(reverse('parametres_section', args=['securite'])).status_code, 200)
        self.assertIsNone(self._page().context.get('comptes_bloques'))
        self.assertEqual(
            self.client.post(reverse('debloquer_compte', args=[self.assure.pk])).status_code, 403)
        self.assertTrue(TentativeConnexion.bloque('awa.diallo@santesn.sn'))

    def test_anonyme_refuse(self):
        self.client.logout()
        self.assertEqual(
            self.client.post(reverse('debloquer_compte', args=[self.assure.pk])).status_code, 302)

    def test_aucune_donnee_sensible_exposee(self):
        self._bloquer('awa.diallo@santesn.sn')
        html = self._page().content.decode()
        self.assertNotIn(self.assure.password, html)
        self.assertNotIn('pbkdf2', html)
        self.assertNotIn('csrftoken', html.replace('csrfmiddlewaretoken', ''))

    # --- tableau de bord ---

    def test_alerte_dashboard_seulement_si_compte_bloque(self):
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.context['nb_comptes_bloques'], 0)
        self.assertNotContains(reponse, 'class="alerte-securite"')

        self._bloquer('awa.diallo@santesn.sn')
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.context['nb_comptes_bloques'], 1)
        self.assertContains(reponse, 'class="alerte-securite"')
        self.assertContains(reponse, 'Voir les comptes')

    def test_alerte_dashboard_disparait_apres_deblocage(self):
        self._bloquer('awa.diallo@santesn.sn')
        self.client.post(reverse('debloquer_compte', args=[self.assure.pk]))
        reponse = self.client.get(reverse('dashboard'))
        self.assertEqual(reponse.context['nb_comptes_bloques'], 0)
        self.assertNotContains(reponse, 'class="alerte-securite"')


class MesPrisesEnChargeAssureTests(TestCase):
    """L'assure voyait la CONSEQUENCE (sa part a payer) sans jamais voir la
    CAUSE : l'etat de sa prise en charge. Ces tests verrouillent le perimetre
    (ses beneficiaires seulement) et le fait que les montants proviennent des
    consultations rattachees -- PriseEnCharge n'en porte aucun."""

    def setUp(self):
        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-pec@santesn.sn')
        self.principal = Patient.objects.create(
            user=self.assure_user,
            nom='Sow',
            prenom='Ousmane',
            date_naissance=datetime.date(1980, 1, 1),
            telephone='770000030',
        )
        self.enfant = Patient.objects.create(
            nom='Sow',
            prenom='Awa',
            date_naissance=datetime.date(2015, 1, 1),
            telephone='770000031',
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            lien_parente=Patient.LienParente.ENFANT,
            assure_principal=self.principal,
        )
        self.etranger = creer_patient(nom='Ndiaye', prenom='Fatou')
        self.medecin = creer_medecin('medecin-pec@santesn.sn')

        self.pec = PriseEnCharge.objects.create(
            patient=self.principal, motif='Suivi cardiaque', statut='validee')
        self.pec_enfant = PriseEnCharge.objects.create(
            patient=self.enfant, motif='Vaccination', statut='en_attente')
        self.pec_etrangere = PriseEnCharge.objects.create(
            patient=self.etranger, motif='Ne me regarde pas', statut='validee')

        self.client.login(username='assure-pec@santesn.sn', password=PASSWORD)

    def _page(self, **params):
        return self.client.get(reverse('mes_prises_en_charge_assure'), params)

    def _ligne(self, prise, reponse=None):
        reponse = reponse or self._page()
        return next(l for l in reponse.context['lignes'] if l['prise'] == prise)

    def test_voit_les_siennes_et_celles_de_ses_ayants_droit(self):
        motifs = [l['prise'].motif for l in self._page().context['lignes']]
        self.assertIn('Suivi cardiaque', motifs)
        self.assertIn('Vaccination', motifs)

    def test_ne_voit_pas_celles_des_autres_assures(self):
        reponse = self._page()
        motifs = [l['prise'].motif for l in reponse.context['lignes']]
        self.assertNotIn('Ne me regarde pas', motifs)
        self.assertNotContains(reponse, 'Ne me regarde pas')

    def test_filtre_par_statut(self):
        lignes = self._page(statut='en_attente').context['lignes']
        self.assertEqual([l['prise'].motif for l in lignes], ['Vaccination'])

    def test_montants_issus_des_consultations_rattachees(self):
        service = ServiceMedical.objects.create(nom='Cardiologie', prix=Decimal('20000'))
        # Le taux n'est pas un champ de Patient : c'est une propriete lue sur le
        # plan de couverture du titulaire.
        self.principal.plan_couverture = PlanCouverture.objects.create(
            nom='Standard', taux_couverture=Decimal('80'), plafond_annuel=Decimal('1000000'))
        self.principal.save()
        consultation = Consultation.objects.create(
            patient=self.principal,
            medecin=self.medecin,
            service=service,
            prise_en_charge=self.pec,
            date_consultation=timezone.now(),
            diagnostic='Controle',
        )
        Paiement.calculer_pour(consultation).save()

        ligne = self._ligne(self.pec)
        self.assertEqual(ligne['montant_couvert'], Decimal('16000'))
        self.assertEqual(ligne['montant_a_charge'], Decimal('4000'))

    def test_demande_sans_consultation_n_affiche_aucun_montant(self):
        reponse = self._page()
        ligne = self._ligne(self.pec_enfant, reponse)
        self.assertEqual(ligne['consultations'], [])
        self.assertEqual(ligne['montant_a_charge'], Decimal('0'))
        self.assertContains(reponse, 'Aucune consultation')

    def test_message_specifique_selon_le_statut(self):
        self.assertContains(self._page(statut='en_attente'), "en cours d'examen")
        PriseEnCharge.objects.filter(pk=self.pec.pk).update(statut='refusee')
        self.assertContains(self._page(statut='refusee'), 'restent à votre')

    def test_pagination(self):
        for i in range(TAILLE_PAGE_LISTE + 3):
            PriseEnCharge.objects.create(patient=self.principal, motif=f'Demande {i}')
        page = self._page().context['page']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertTrue(page.has_next())

    def test_role_non_assure_refuse(self):
        self.client.logout()
        creer_utilisateur(User.Role.ADMIN, 'admin-pec@santesn.sn')
        self.client.login(username='admin-pec@santesn.sn', password=PASSWORD)
        self.assertEqual(self._page().status_code, 403)

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(self._page().status_code, 302)

    def test_assure_sans_fiche_redirige_vers_son_profil(self):
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'assure-sans-fiche@santesn.sn')
        self.client.login(username='assure-sans-fiche@santesn.sn', password=PASSWORD)
        self.assertRedirects(self._page(), reverse('mon_profil_assure'))

    def test_entree_de_menu_presente(self):
        self.assertContains(self._page(), reverse('mes_prises_en_charge_assure'))

    def test_etat_vide_distingue_aucune_demande_et_filtre_trop_strict(self):
        PriseEnCharge.objects.filter(patient__in=[self.principal, self.enfant]).delete()
        self.assertContains(self._page(), 'Aucune prise en charge enregistrée')
        self.assertContains(self._page(statut='validee'), 'Aucune prise en charge avec ce statut')


class CoherencePatientPriseEnChargeTests(TestCase):
    """Invariant : consultation.prise_en_charge.patient == consultation.patient.

    Il etait verifie a la CREATION par ConsultationForm, mais restait cassable
    apres coup par deux chemins : reattribuer la prise en charge a un autre
    patient (le champ est editable en modification), et /admin/, qui enregistre
    Consultation sans ModelAdmin dedie et n'utilise donc aucun formulaire de
    l'app. La regle vit desormais dans les modeles, ou tout ModelForm la
    rencontre."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-coherence@santesn.sn')
        self.client.login(username='admin-coherence@santesn.sn', password=PASSWORD)
        self.patient = creer_patient(nom='Ba', prenom='Cheikh')
        self.autre = creer_patient(nom='Fall', prenom='Astou')
        self.medecin = creer_medecin('medecin-coherence@santesn.sn')
        self.prise = PriseEnCharge.objects.create(patient=self.patient, motif='Suivi')

    def _consultation(self, prise=None):
        return Consultation.objects.create(
            patient=self.patient,
            medecin=self.medecin,
            prise_en_charge=self.prise if prise is None else prise,
            date_consultation=timezone.now(),
            diagnostic='Controle',
        )

    def _modifier(self, patient):
        return self.client.post(
            reverse('modifier_prise_en_charge', args=[self.prise.pk]),
            {'patient': patient.pk, 'motif': 'Suivi', 'statut': 'en_attente'},
        )

    def test_reattribution_refusee_quand_des_consultations_existent(self):
        self._consultation()
        reponse = self._modifier(self.autre)
        self.assertEqual(reponse.status_code, 200)
        self.prise.refresh_from_db()
        self.assertEqual(self.prise.patient_id, self.patient.pk)
        self.assertContains(reponse, 'Des consultations sont déjà rattachées')

    def test_reattribution_permise_tant_quaucune_consultation_nexiste(self):
        # Corriger une erreur de saisie doit rester possible : c'est le seul
        # cas legitime de changement de patient.
        reponse = self._modifier(self.autre)
        self.assertRedirects(reponse, reverse('liste_prises_en_charge'))
        self.prise.refresh_from_db()
        self.assertEqual(self.prise.patient_id, self.autre.pk)

    def test_modification_sans_changer_de_patient_reste_possible(self):
        self._consultation()
        reponse = self.client.post(
            reverse('modifier_prise_en_charge', args=[self.prise.pk]),
            {'patient': self.patient.pk, 'motif': 'Suivi', 'statut': 'validee'},
        )
        self.assertRedirects(reponse, reverse('liste_prises_en_charge'))
        self.prise.refresh_from_db()
        self.assertEqual(self.prise.statut, 'validee')

    def test_regle_consultation_atteint_tout_modelform_donc_admin(self):
        # /admin/ construit un ModelForm nu (admin.site.register(Consultation)
        # sans ModelAdmin) : c'est ce formulaire-la qu'on reproduit ici.
        prise_etrangere = PriseEnCharge.objects.create(patient=self.autre, motif='Autre')
        Formulaire = modelform_factory(Consultation, fields='__all__')
        form = Formulaire(data={
            'patient': self.patient.pk,
            'medecin': self.medecin.pk,
            'prise_en_charge': prise_etrangere.pk,
            'date_consultation': '2026-08-01 10:00:00',
            'diagnostic': 'Test',
            'traitement': '',
        })
        self.assertFalse(form.is_valid())
        self.assertIn('prise_en_charge', form.errors)

    def test_regle_prise_en_charge_atteint_tout_modelform_donc_admin(self):
        self._consultation()
        Formulaire = modelform_factory(PriseEnCharge, fields='__all__')
        form = Formulaire(
            data={'patient': self.autre.pk, 'motif': 'Suivi', 'statut': 'en_attente'},
            instance=self.prise,
        )
        self.assertFalse(form.is_valid())
        self.assertIn('patient', form.errors)


class ListeConsultationsAdminTests(TestCase):
    """Dernier angle mort du suivi admin : apres les rendez-vous et les
    ordonnances, l'acte de soin lui-meme n'apparaissait sur aucun ecran.

    Ecran en LECTURE SEULE et SANS donnees medicales : diagnostic et
    traitement restent au medecin qui les a saisis (meme regle que
    fiche_patient_medecin)."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-consult@santesn.sn')
        self.client.login(username='admin-consult@santesn.sn', password=PASSWORD)

        self.patient = creer_patient(nom='Diallo', prenom='Mariama')
        self.medecin = creer_medecin('medecin-consult@santesn.sn')
        self.service = ServiceMedical.objects.create(nom='Radiologie', prix=Decimal('30000'))

        self.couverte = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=self.service,
            prise_en_charge=PriseEnCharge.objects.create(
                patient=self.patient, motif='Couverte', statut='validee'),
            date_consultation=timezone.now(),
            diagnostic='Fracture du poignet', traitement='Immobilisation 6 semaines',
        )
        self.en_attente = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            prise_en_charge=PriseEnCharge.objects.create(
                patient=self.patient, motif='Demandee', statut='en_attente'),
            date_consultation=timezone.now(), diagnostic='Suivi',
        )
        self.sans_prise_en_charge = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Hors couverture',
        )

    def _page(self, **params):
        return self.client.get(reverse('liste_consultations'), params)

    def _lignes(self, **params):
        return list(self._page(**params).context['consultations'])

    def test_liste_toutes_les_consultations(self):
        self.assertEqual(len(self._lignes()), 3)

    def test_aucune_donnee_medicale_affichee(self):
        # La regle centrale de cet ecran : l'admin voit l'acte et sa
        # facturation, jamais son contenu.
        reponse = self._page()
        self.assertNotContains(reponse, 'Fracture du poignet')
        self.assertNotContains(reponse, 'Immobilisation')
        self.assertNotContains(reponse, 'Diagnostic')
        self.assertNotContains(reponse, 'Traitement')

    def test_filtre_couverture_oui_ne_garde_que_les_prises_en_charge_validees(self):
        self.assertEqual(self._lignes(couverture='oui'), [self.couverte])

    def test_filtre_couverture_non_inclut_les_consultations_sans_prise_en_charge(self):
        # Point verifie explicitement : exclude() sur une FK nullable doit
        # RETENIR les lignes sans prise en charge (LEFT OUTER JOIN + le
        # IS NOT NULL que Django ajoute dans le NOT). Ce sont justement les
        # soins restes a 100% a la charge du patient.
        lignes = self._lignes(couverture='non')
        self.assertIn(self.sans_prise_en_charge, lignes)
        self.assertIn(self.en_attente, lignes)
        self.assertNotIn(self.couverte, lignes)

    def test_recherche_par_patient_et_par_medecin(self):
        self.assertEqual(len(self._lignes(q='Diallo')), 3)
        self.assertEqual(len(self._lignes(q='Ndiaye')), 3)
        self.assertEqual(len(self._lignes(q='Introuvable')), 0)

    def test_filtre_par_date(self):
        hier = (timezone.now() - datetime.timedelta(days=1)).date()
        self.assertEqual(len(self._lignes(date=hier.isoformat())), 0)
        aujourdhui = timezone.localtime(self.couverte.date_consultation).date()
        self.assertEqual(len(self._lignes(date=aujourdhui.isoformat())), 3)

    def test_date_invalide_ignoree_sans_erreur(self):
        reponse = self._page(date='pas-une-date')
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(len(reponse.context['consultations']), 3)

    def test_tri_restreint_aux_champs_autorises(self):
        self.assertEqual(self._page(tri='diagnostic').status_code, 200)
        self.assertEqual(len(self._lignes(tri='diagnostic')), 3)

    def test_ecran_en_lecture_seule(self):
        # Aucune action d'ecriture propre a l'ecran : ni .action-ligne (le
        # formulaire POST en ligne de tableau du projet), ni lien de
        # suppression. On ne teste pas method="post" en brut : base.html en
        # contient pour la deconnexion et les notifications.
        # On assertionne sur le BALISAGE (class="...") : le nom de classe nu
        # apparait aussi dans la feuille de style inline de base.html.
        contenu = self._page().content.decode()
        self.assertNotIn('class="action-ligne"', contenu)
        self.assertNotIn('confirmer_suppression', contenu)
        self.assertNotIn('>Modifier<', contenu)

    def test_reserve_a_l_administrateur(self):
        for role, email in [
            (User.Role.MEDECIN, 'med-refuse@santesn.sn'),
            (User.Role.ASSURE, 'assure-refuse@santesn.sn'),
            (User.Role.PHARMACIEN, 'pharma-refuse@santesn.sn'),
        ]:
            self.client.logout()
            creer_utilisateur(role, email)
            self.client.login(username=email, password=PASSWORD)
            self.assertEqual(self._page().status_code, 403, role)

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(self._page().status_code, 302)

    def test_pagination(self):
        for _ in range(TAILLE_PAGE_LISTE):
            Consultation.objects.create(
                patient=self.patient, medecin=self.medecin,
                date_consultation=timezone.now(), diagnostic='En masse',
            )
        page = self._page().context['consultations']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertTrue(page.has_next())

    def test_entree_de_menu_presente(self):
        self.assertContains(self._page(), reverse('liste_consultations'))


class FiltresListesNonAdminTests(TestCase):
    """Le tri par colonne (_trier) et les filtres GET servaient les 11 listes
    admin et aucune liste non-admin. Les 4 ecrans ou un filtre repond a une
    vraie question en ont un desormais.

    mes_patients en est volontairement exclu : la page porte deja une
    recherche (combobox JS rechercher_patients_medecin), un second champ
    serait un doublon."""

    def setUp(self):
        self.medecin = creer_medecin('medecin-filtres@santesn.sn')
        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-filtres@santesn.sn')
        self.patient = Patient.objects.create(
            user=self.assure_user, nom='Kane', prenom='Ibrahima',
            date_naissance=datetime.date(1985, 5, 5), telephone='770000040',
        )
        maintenant = timezone.now()
        self.futur = RendezVous.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_heure=maintenant + datetime.timedelta(days=7),
            motif='Controle', statut=RendezVous.Statut.DEMANDE,
        )
        self.tres_futur = RendezVous.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_heure=maintenant + datetime.timedelta(days=30),
            motif='Bilan', statut=RendezVous.Statut.CONFIRME,
        )
        self.passe = RendezVous.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_heure=maintenant - datetime.timedelta(days=7),
            motif='Ancien', statut=RendezVous.Statut.TERMINE,
        )

    # --- Agenda du medecin ------------------------------------------------

    def _agenda(self, **params):
        self.client.login(username='medecin-filtres@santesn.sn', password=PASSWORD)
        return self.client.get(reverse('agenda_medecin'), params)

    def test_agenda_filtre_a_venir(self):
        lignes = list(self._agenda(periode='a_venir').context['rendez_vous'])
        self.assertEqual(lignes, [self.futur, self.tres_futur])

    def test_agenda_a_venir_affiche_le_plus_proche_en_premier(self):
        # Le coeur du changement : en tri unique "-date_heure", demander
        # "a venir" mettait le rendez-vous le plus LOINTAIN en tete.
        lignes = list(self._agenda(periode='a_venir').context['rendez_vous'])
        self.assertEqual(lignes[0], self.futur)

    def test_agenda_filtre_passes(self):
        lignes = list(self._agenda(periode='passes').context['rendez_vous'])
        self.assertEqual(lignes, [self.passe])

    def test_agenda_filtre_statut(self):
        lignes = list(self._agenda(statut=RendezVous.Statut.CONFIRME).context['rendez_vous'])
        self.assertEqual(lignes, [self.tres_futur])

    def test_agenda_sans_filtre_reste_du_plus_recent_au_plus_ancien(self):
        lignes = list(self._agenda().context['rendez_vous'])
        self.assertEqual(lignes, [self.tres_futur, self.futur, self.passe])

    def test_agenda_statut_inconnu_ignore_et_non_reaffiche(self):
        reponse = self._agenda(statut='PAS_UN_STATUT')
        self.assertEqual(len(reponse.context['rendez_vous']), 3)
        self.assertEqual(reponse.context['statut_choisi'], '')

    def test_agenda_libelle_jamais_accepte_comme_valeur(self):
        # Les TextChoices separent valeur stockee et libelle affiche : filtrer
        # sur le libelle ne doit rien retenir, et surtout pas planter.
        reponse = self._agenda(statut='Confirmé')
        self.assertEqual(len(reponse.context['rendez_vous']), 3)

    def test_agenda_etat_vide_distingue_le_filtre(self):
        self.assertContains(
            self._agenda(statut=RendezVous.Statut.ANNULE),
            'Aucun rendez-vous ne correspond',
        )

    # --- Rendez-vous de l'assure ------------------------------------------

    def _mes_rdv(self, **params):
        self.client.login(username='assure-filtres@santesn.sn', password=PASSWORD)
        return self.client.get(reverse('mes_rendez_vous_assure'), params)

    def test_assure_filtre_a_venir_le_plus_proche_en_premier(self):
        lignes = list(self._mes_rdv(periode='a_venir').context['rendez_vous'])
        self.assertEqual(lignes, [self.futur, self.tres_futur])

    def test_assure_filtre_statut(self):
        lignes = list(self._mes_rdv(statut=RendezVous.Statut.TERMINE).context['rendez_vous'])
        self.assertEqual(lignes, [self.passe])

    # --- Ordonnances de l'assure ------------------------------------------

    def _mes_ordonnances(self, **params):
        self.client.login(username='assure-filtres@santesn.sn', password=PASSWORD)
        return self.client.get(reverse('mes_ordonnances_assure'), params)

    def test_assure_filtre_ordonnances_a_retirer(self):
        a_retirer = creer_ordonnance(self.patient, self.medecin)
        retiree = creer_ordonnance(self.patient, self.medecin)
        pharmacien = creer_pharmacien('pharma-filtres@santesn.sn')
        Delivrance.objects.create(ordonnance=retiree, pharmacien=pharmacien)

        self.assertEqual(list(self._mes_ordonnances(delivrance='non').context['ordonnances']),
                         [a_retirer])
        self.assertEqual(list(self._mes_ordonnances(delivrance='oui').context['ordonnances']),
                         [retiree])
        self.assertEqual(len(self._mes_ordonnances().context['ordonnances']), 2)

    # --- Historique du pharmacien -----------------------------------------

    def _historique(self, **params):
        self.client.login(username='pharma-histo@santesn.sn', password=PASSWORD)
        return self.client.get(reverse('historique_delivrances'), params)

    def test_pharmacien_recherche_et_date(self):
        pharmacien = creer_pharmacien('pharma-histo@santesn.sn')
        ordonnance = creer_ordonnance(self.patient, self.medecin)
        delivrance = Delivrance.objects.create(ordonnance=ordonnance, pharmacien=pharmacien)

        self.assertEqual(len(self._historique(q='Kane').context['delivrances']), 1)
        self.assertEqual(len(self._historique(q=ordonnance.code_qr[:6]).context['delivrances']), 1)
        self.assertEqual(len(self._historique(q='Introuvable').context['delivrances']), 0)

        jour = timezone.localtime(delivrance.date_delivrance).date()
        self.assertEqual(len(self._historique(date=jour.isoformat()).context['delivrances']), 1)
        veille = (jour - datetime.timedelta(days=1)).isoformat()
        self.assertEqual(len(self._historique(date=veille).context['delivrances']), 0)

    def test_pharmacien_date_invalide_ignoree(self):
        creer_pharmacien('pharma-histo@santesn.sn')
        reponse = self._historique(date='pas-une-date')
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.context['date_choisie'], '')

    # --- Non-regression ---------------------------------------------------

    def test_mes_patients_na_pas_de_seconde_barre_de_recherche(self):
        # La recherche de cette page est le combobox JS deja en place : un
        # formulaire GET .filtres en plus ferait deux champs pour une seule
        # question.
        self.client.login(username='medecin-filtres@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('mes_patients'))
        self.assertNotContains(reponse, 'class="filtres"')
        self.assertContains(reponse, 'recherche-patients-champ')


class JournalActiviteTests(TestCase):
    """Journal des DECISIONS administratives et des DESTRUCTIONS.

    Un journal incomplet est pire qu'aucun journal : il donne l'illusion
    d'une trace exhaustive. Ces tests couvrent donc les deux sens -- ce qui
    DOIT etre journalise, et ce qui ne doit PAS l'etre."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-journal@santesn.sn')
        self.admin.first_name = 'Awa'
        self.admin.last_name = 'Ndiaye'
        self.admin.save()
        self.client.login(username='admin-journal@santesn.sn', password=PASSWORD)

    def _entrees(self, **filtres):
        return list(JournalActivite.objects.filter(**filtres))

    # --- Ce qui doit etre journalise -------------------------------------

    def test_suppression_dutilisateur(self):
        cible = creer_utilisateur(User.Role.ASSURE, 'a-supprimer@santesn.sn')
        self.client.post(reverse('supprimer_utilisateur', args=[cible.pk]))
        entree = JournalActivite.objects.get(action=JournalActivite.Action.SUPPRESSION)
        self.assertIn('a-supprimer@santesn.sn', entree.objet)
        self.assertEqual(entree.auteur_libelle, 'Awa Ndiaye')

    def test_lentree_survit_a_la_suppression_de_son_objet(self):
        # Le coeur du modele : du texte fige, pas de cle etrangere. Une cle
        # en CASCADE aurait efface l'entree en meme temps que l'objet -- or
        # c'est precisement la suppression qu'on veut garder.
        cible = creer_utilisateur(User.Role.ASSURE, 'efface@santesn.sn')
        self.client.post(reverse('supprimer_utilisateur', args=[cible.pk]))
        self.assertFalse(User.objects.filter(pk=cible.pk).exists())
        self.assertTrue(JournalActivite.objects.filter(objet__contains='efface@santesn.sn').exists())

    def test_lentree_survit_a_la_suppression_de_son_auteur(self):
        # Supprimer le compte d'un administrateur ne doit pas effacer la
        # trace de ce qu'il a fait : la cle passe a NULL, le libelle reste.
        cible = creer_utilisateur(User.Role.ASSURE, 'cible@santesn.sn')
        self.client.post(reverse('supprimer_utilisateur', args=[cible.pk]))
        self.admin.delete()
        entree = JournalActivite.objects.get(objet__contains='cible@santesn.sn')
        self.assertIsNone(entree.auteur)
        self.assertEqual(entree.auteur_libelle, 'Awa Ndiaye')

    def test_desactivation_puis_activation(self):
        cible = creer_utilisateur(User.Role.MEDECIN, 'bascule@santesn.sn')
        self.client.post(reverse('activer_desactiver_utilisateur', args=[cible.pk]))
        self.client.post(reverse('activer_desactiver_utilisateur', args=[cible.pk]))
        actions = list(JournalActivite.objects.order_by('pk').values_list('action', flat=True))
        self.assertEqual(actions, [JournalActivite.Action.DESACTIVATION,
                                   JournalActivite.Action.ACTIVATION])

    def test_reinitialisation_de_mot_de_passe(self):
        cible = creer_utilisateur(User.Role.ASSURE, 'oubli@santesn.sn')
        self.client.post(reverse('reinitialiser_mot_de_passe', args=[cible.pk]))
        self.assertEqual(len(self._entrees(action=JournalActivite.Action.MOT_DE_PASSE)), 1)

    def test_deblocage_manuel(self):
        cible = creer_utilisateur(User.Role.ASSURE, 'bloque@santesn.sn')
        TentativeConnexion.objects.create(
            email='bloque@santesn.sn', tentatives=TentativeConnexion.MAX_TENTATIVES)
        self.client.post(reverse('debloquer_compte', args=[cible.pk]))
        entree = JournalActivite.objects.get(action=JournalActivite.Action.DEBLOCAGE)
        self.assertIn('bloque@santesn.sn', entree.objet)

    def test_deblocage_sans_effet_nest_pas_journalise(self):
        # Le compte n'etait deja plus bloque : rien ne s'est passe, donc rien
        # a tracer. Une entree ici ferait croire a une intervention.
        cible = creer_utilisateur(User.Role.ASSURE, 'jamais-bloque@santesn.sn')
        self.client.post(reverse('debloquer_compte', args=[cible.pk]))
        self.assertEqual(self._entrees(action=JournalActivite.Action.DEBLOCAGE), [])

    def test_decision_sur_une_prise_en_charge(self):
        # C'est ce statut qui decide si le patient paie 10% ou 100% :
        # l'auteur de la decision doit rester connu.
        patient = creer_patient(nom='Ba', prenom='Fatou')
        prise = PriseEnCharge.objects.create(patient=patient, motif='Suivi')
        self.client.post(reverse('modifier_prise_en_charge', args=[prise.pk]),
                         {'patient': patient.pk, 'motif': 'Suivi', 'statut': 'validee'})
        entree = JournalActivite.objects.get(action=JournalActivite.Action.DECISION)
        self.assertIn('En attente', entree.details)
        self.assertIn('Validée', entree.details)

    def test_modification_sans_changement_de_statut_nest_pas_une_decision(self):
        patient = creer_patient(nom='Sy', prenom='Modou')
        prise = PriseEnCharge.objects.create(patient=patient, motif='Avant')
        self.client.post(reverse('modifier_prise_en_charge', args=[prise.pk]),
                         {'patient': patient.pk, 'motif': 'Après', 'statut': 'en_attente'})
        self.assertEqual(self._entrees(action=JournalActivite.Action.DECISION), [])
        self.assertEqual(len(self._entrees(action=JournalActivite.Action.MODIFICATION)), 1)

    def test_reglement_dun_paiement(self):
        # Paiement porte date_reglement et mode_reglement, mais PAS qui l'a
        # marque regle : le journal est la seule trace de l'auteur.
        patient = creer_patient(nom='Fall', prenom='Ndeye')
        medecin = creer_medecin('medecin-journal@santesn.sn')
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test')
        paiement = Paiement.calculer_pour(consultation)
        paiement.save()
        self.client.post(reverse('marquer_paiement_regle', args=[paiement.pk]),
                         {'mode_reglement': Paiement.ModeReglement.ESPECES})
        entree = JournalActivite.objects.get(action=JournalActivite.Action.REGLEMENT)
        self.assertIn('Espèces', entree.details)

    def test_toutes_les_suppressions_metier_sont_journalisees(self):
        """Couverture : un journal partiel donnerait l'illusion d'une trace
        exhaustive. Chaque ecran de suppression doit laisser une entree."""
        patient = creer_patient(nom='Kane', prenom='Awa')
        medecin = creer_medecin('medecin-suppr@santesn.sn')
        cas = [
            ('supprimer_prise_en_charge',
             PriseEnCharge.objects.create(patient=patient, motif='X').pk),
            ('supprimer_prestataire',
             Prestataire.objects.create(nom='Hopital X',
                                        type_prestataire=Prestataire.Type.HOPITAL).pk),
            ('supprimer_plan_couverture',
             PlanCouverture.objects.create(nom='Plan X',
                                           taux_couverture=Decimal('50')).pk),
            ('supprimer_service',
             ServiceMedical.objects.create(nom='Service X', prix=Decimal('1')).pk),
            ('supprimer_medecin', medecin.pk),
            ('supprimer_patient', patient.pk),
        ]
        for nom_vue, pk in cas:
            JournalActivite.objects.all().delete()
            self.client.post(reverse(nom_vue, args=[pk]))
            self.assertEqual(
                len(self._entrees(action=JournalActivite.Action.SUPPRESSION)), 1, nom_vue)

    # --- Ce qui ne doit PAS etre journalise -------------------------------

    def test_la_consultation_decrans_ne_laisse_aucune_trace(self):
        for nom in ('dashboard', 'liste_utilisateurs', 'rapports', 'journal_activite'):
            self.client.get(reverse(nom))
        self.assertEqual(JournalActivite.objects.count(), 0)

    def test_les_actes_de_soin_ne_sont_pas_journalises(self):
        """Une Consultation porte deja son medecin et sa date, une Delivrance
        son pharmacien : les reecrire ici les dupliquerait sans rien
        apprendre, et noierait les entrees qui comptent."""
        patient = creer_patient(nom='Diop', prenom='Alioune')
        medecin = creer_medecin('medecin-actes@santesn.sn')
        ordonnance = creer_ordonnance(patient, medecin)
        pharmacien = creer_pharmacien('pharma-actes@santesn.sn')
        self.client.logout()
        self.client.login(username='pharma-actes@santesn.sn', password=PASSWORD)
        self.client.post(reverse('valider_delivrance', args=[ordonnance.pk]),
                         {'code_qr': ordonnance.code_qr})
        self.assertEqual(JournalActivite.objects.count(), 0)

    # --- L'ecran ----------------------------------------------------------

    def test_ecran_reserve_a_l_administrateur(self):
        for role, email in (
            (User.Role.MEDECIN, 'med-journal@santesn.sn'),
            (User.Role.ASSURE, 'assure-journal@santesn.sn'),
            (User.Role.PHARMACIEN, 'pharma-journal@santesn.sn'),
        ):
            self.client.logout()
            creer_utilisateur(role, email)
            self.client.login(username=email, password=PASSWORD)
            self.assertEqual(
                self.client.get(reverse('journal_activite')).status_code, 403, role)

    def test_ecran_en_lecture_seule(self):
        cible = creer_utilisateur(User.Role.ASSURE, 'trace@santesn.sn')
        self.client.post(reverse('activer_desactiver_utilisateur', args=[cible.pk]))
        contenu = self.client.get(reverse('journal_activite')).content.decode()
        self.assertNotIn('class="action-ligne"', contenu)
        self.assertNotIn('confirmer_suppression', contenu)

    def test_filtres_action_recherche_et_date(self):
        cible = creer_utilisateur(User.Role.ASSURE, 'filtrable@santesn.sn')
        self.client.post(reverse('activer_desactiver_utilisateur', args=[cible.pk]))
        self.client.post(reverse('reinitialiser_mot_de_passe', args=[cible.pk]))

        def lignes(**params):
            return list(self.client.get(reverse('journal_activite'), params).context['entrees'])

        self.assertEqual(len(lignes()), 2)
        self.assertEqual(len(lignes(action=JournalActivite.Action.MOT_DE_PASSE)), 1)
        self.assertEqual(len(lignes(action='PAS_UNE_ACTION')), 2)
        self.assertEqual(len(lignes(q='filtrable')), 2)
        self.assertEqual(len(lignes(q='introuvable')), 0)
        aujourdhui = timezone.localtime(timezone.now()).date()
        self.assertEqual(len(lignes(date=aujourdhui.isoformat())), 2)
        veille = (aujourdhui - datetime.timedelta(days=1)).isoformat()
        self.assertEqual(len(lignes(date=veille)), 0)
        self.assertEqual(len(lignes(date='pas-une-date')), 2)

    def test_admin_django_refuse_toute_ecriture_sur_le_journal(self):
        """Un journal d'audit qu'un administrateur peut retoucher ne vaut
        rien : /admin/ enregistre le modele en lecture seule stricte."""
        from django.contrib import admin as django_admin
        from .models import JournalActivite as Modele
        options = django_admin.site._registry[Modele]
        requete = type('R', (), {'user': self.admin})()
        self.assertFalse(options.has_add_permission(requete))
        self.assertFalse(options.has_change_permission(requete))
        self.assertFalse(options.has_delete_permission(requete))

    def test_raccourci_depuis_les_comptes_bloques(self):
        # Un lien vers le journal DEJA FILTRE : il ne debloque rien et ne
        # duplique aucune fonctionnalite.
        reponse = self.client.get(reverse('parametres_section', args=['securite']))
        self.assertContains(reponse, f"{reverse('journal_activite')}?action=DEBLOCAGE")

    def test_entree_depuis_parametres_securite_et_pas_dans_le_menu(self):
        """Le journal est une fonction d'ADMINISTRATION, pas un module metier :
        il n'a pas d'entree propre dans le menu lateral, on y entre par
        Parametres > Securite."""
        self.assertNotContains(self.client.get(reverse('dashboard')),
                               'data-tooltip="Journal d\'activité"')
        securite = self.client.get(reverse('parametres_section', args=['securite']))
        self.assertContains(securite, reverse('journal_activite'))
        self.assertContains(securite, "Journal d'activité")

    def test_journal_invisible_dans_parametres_pour_un_non_admin(self):
        self.client.logout()
        creer_medecin('medecin-journal-param@santesn.sn')
        self.client.login(username='medecin-journal-param@santesn.sn', password=PASSWORD)
        self.assertNotContains(
            self.client.get(reverse('parametres_section', args=['securite'])),
            reverse('journal_activite'))

    def test_pagination(self):
        for i in range(TAILLE_PAGE_LISTE + 2):
            JournalActivite.objects.create(
                auteur=self.admin, auteur_libelle='Awa Ndiaye',
                action=JournalActivite.Action.MODIFICATION, objet=f'Objet {i}')
        page = self.client.get(reverse('journal_activite')).context['entrees']
        self.assertEqual(len(page), TAILLE_PAGE_LISTE)
        self.assertTrue(page.has_next())


class CartePriseEnChargeTests(TestCase):
    """Carte de prise en charge : edition, impression, QR et scan.

    La carte appartient a un BENEFICIAIRE (numero_carte est porte par
    Patient), jamais a un compte : un ayant droit a une carte sans avoir de
    compte, un medecin a un compte sans avoir de carte."""

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-carte@santesn.sn')
        self.admin.first_name = 'Awa'
        self.admin.last_name = 'Ndiaye'
        self.admin.save()

        self.plan = PlanCouverture.objects.create(
            nom='Famille 80%', taux_couverture=Decimal('80'), plafond_annuel=Decimal('1000000'))
        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-carte@santesn.sn')
        self.principal = Patient.objects.create(
            user=self.assure_user, nom='Diop', prenom='Moussa',
            date_naissance=datetime.date(1980, 4, 12), telephone='770000050',
            plan_couverture=self.plan)
        self.enfant = Patient.objects.create(
            nom='Diop', prenom='Awa', date_naissance=datetime.date(2014, 2, 2),
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            lien_parente=Patient.LienParente.ENFANT, assure_principal=self.principal)

        self.medecin = creer_medecin('medecin-carte@santesn.sn')
        self.autre_medecin = creer_medecin('autre-medecin-carte@santesn.sn')
        self.pharmacien = creer_pharmacien('pharma-carte@santesn.sn')

    def _carte(self, patient=None):
        return self.client.get(reverse('carte_patient', args=[(patient or self.principal).pk]))

    def _scan(self, patient=None):
        return self.client.get(reverse('carte_scan', args=[(patient or self.principal).numero_carte]))

    # --- 1 a 5 : generation, informations, recto, verso -------------------

    def test_carte_generee_pour_un_assure_principal(self):
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        reponse = self._carte()
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, self.principal.numero_carte)

    def test_carte_generee_pour_un_ayant_droit_qui_na_pas_de_compte(self):
        # Le cas qui interdisait de placer la carte dans "Utilisateurs".
        self.assertIsNone(self.enfant.user)
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        reponse = self._carte(self.enfant)
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, self.enfant.numero_carte)
        self.assertContains(reponse, 'Enfant')

    def test_recto_porte_les_informations_reelles(self):
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        reponse = self._carte()
        self.assertContains(reponse, 'Carte de prise en charge')
        self.assertContains(reponse, 'Moussa')
        self.assertContains(reponse, 'Diop')
        self.assertContains(reponse, 'Assuré principal')
        self.assertContains(reponse, 'Famille 80%')

    def test_recto_ninvente_ni_photo_ni_expiration(self):
        """Patient ne porte ni photo ni date d'expiration : les afficher
        reviendrait a imprimer du faux sur une piece presentee au guichet."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        contenu = self._carte().content.decode()
        self.assertNotIn("Date d'expiration", contenu)
        self.assertNotIn('Expire le', contenu)
        self.assertNotIn('<img', contenu)

    def test_numero_de_carte_correct_et_unique(self):
        self.assertTrue(self.principal.numero_carte.startswith('SN-'))
        self.assertNotEqual(self.principal.numero_carte, self.enfant.numero_carte)

    # --- 6 a 7 : un VRAI QR ----------------------------------------------

    def test_verso_porte_un_vrai_qr_svg(self):
        """Pas une image decorative : un SVG genere par la meme fabrique que
        celui des ordonnances, comportant de vrais modules."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        contenu = self._carte().content.decode()
        self.assertIn('<svg', contenu)
        self.assertGreater(contenu.count('<path'), 10)

    def test_le_qr_encode_ladresse_de_scan_et_aucune_donnee_medicale(self):
        service = ServiceMedical.objects.create(nom='Cardiologie', prix=Decimal('20000'))
        consultation = Consultation.objects.create(
            patient=self.principal, medecin=self.medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Hypertension',
            traitement='Amlodipine 5 mg')
        Ordonnance.objects.create(consultation=consultation,
                                  medicaments='Amlodipine 5 mg - 1x/jour')

        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        reponse = self._carte()
        # L'adresse n'est plus ECRITE sur la page : elle n'existe que dans le
        # code lui-meme. La carte ne s'explique pas, elle se presente.
        self.assertEqual(
            reponse.context['url_scan'],
            'http://testserver' + reverse('carte_scan', args=[self.principal.numero_carte]))
        contenu = reponse.content.decode()
        for secret in ('Hypertension', 'Amlodipine'):
            self.assertNotIn(secret, contenu)

    # --- 8 a 10 : scan, autorisations ------------------------------------

    def test_scan_anonyme_renvoie_vers_la_connexion(self):
        reponse = self._scan()
        self.assertEqual(reponse.status_code, 302)
        self.assertIn(reverse('login'), reponse.url)

    def test_scan_par_un_pharmacien_autorise(self):
        self.client.login(username='pharma-carte@santesn.sn', password=PASSWORD)
        reponse = self._scan()
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, self.principal.numero_carte)

    def test_scan_refuse_a_lassure_et_a_ladministrateur(self):
        # Un QR ne doit jamais contourner les permissions : meme connecte,
        # qui n'a pas a soigner ni a delivrer n'entre pas.
        for email in ('assure-carte@santesn.sn', 'admin-carte@santesn.sn'):
            self.client.logout()
            self.client.login(username=email, password=PASSWORD)
            self.assertEqual(self._scan().status_code, 403, email)

    def test_scan_dun_numero_inconnu_donne_404(self):
        self.client.login(username='pharma-carte@santesn.sn', password=PASSWORD)
        self.assertEqual(
            self.client.get(reverse('carte_scan', args=['SN-INEXISTANT'])).status_code, 404)

    # --- 11 a 12 : portee des ordonnances --------------------------------

    def _ordonnance(self, medecin, delivree=False, medicament='Paracétamol 500 mg'):
        consultation = Consultation.objects.create(
            patient=self.principal, medecin=medecin,
            date_consultation=timezone.now(), diagnostic='Bilan')
        ordonnance = Ordonnance.objects.create(consultation=consultation,
                                               medicaments=medicament)
        if delivree:
            Delivrance.objects.create(ordonnance=ordonnance, pharmacien=self.pharmacien)
        return ordonnance

    def test_pharmacien_ne_voit_que_les_ordonnances_non_delivrees(self):
        a_servir = self._ordonnance(self.medecin, medicament='Ibuprofène 400 mg')
        self._ordonnance(self.medecin, delivree=True, medicament='Aspirine 100 mg')

        self.client.login(username='pharma-carte@santesn.sn', password=PASSWORD)
        reponse = self._scan()
        self.assertEqual(list(reponse.context['ordonnances']), [a_servir])
        self.assertNotContains(reponse, 'Aspirine')

    def test_medecin_ne_voit_que_ses_propres_prescriptions(self):
        sienne = self._ordonnance(self.medecin, medicament='Ibuprofène 400 mg')
        self._ordonnance(self.autre_medecin, medicament='Tramadol 50 mg')

        self.client.login(username='medecin-carte@santesn.sn', password=PASSWORD)
        reponse = self._scan()
        self.assertEqual(list(reponse.context['ordonnances']), [sienne])
        self.assertNotContains(reponse, 'Tramadol')

    def test_le_scan_nouvre_pas_le_diagnostic(self):
        consultation = Consultation.objects.create(
            patient=self.principal, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Diabète de type 2',
            traitement='Metformine')
        Ordonnance.objects.create(consultation=consultation, medicaments='Metformine 850 mg')
        self.client.login(username='pharma-carte@santesn.sn', password=PASSWORD)
        self.assertNotContains(self._scan(), 'Diabète de type 2')

    # --- 13 a 14 : impression, apercu ------------------------------------

    def test_la_page_prevoit_une_impression_dediee(self):
        """On imprime une carte, pas la page : le decor applicatif est masque
        et la carte sort au format ISO 7810 ID-1 (85,6 x 54 mm)."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        contenu = self._carte().content.decode()
        self.assertIn('@media print', contenu)
        self.assertIn('85.6mm', contenu)
        self.assertIn('54mm', contenu)
        self.assertIn('carte-sans-impression', contenu)

    def test_apercu_recto_et_verso(self):
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        reponse = self._carte()
        self.assertContains(reponse, 'Recto')
        self.assertContains(reponse, 'Verso')

    # --- 15 : permissions sur l'edition ----------------------------------

    def test_edition_de_carte_reservee_a_ladministrateur(self):
        for role, email in ((User.Role.MEDECIN, 'medecin-carte@santesn.sn'),
                            (User.Role.PHARMACIEN, 'pharma-carte@santesn.sn'),
                            (User.Role.ASSURE, 'assure-carte@santesn.sn')):
            self.client.logout()
            self.client.login(username=email, password=PASSWORD)
            self.assertEqual(self._carte().status_code, 403, role)

    def test_action_carte_presente_sur_la_liste_des_assures(self):
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        self.assertContains(self.client.get(reverse('liste_patients')),
                            reverse('carte_patient', args=[self.principal.pk]))

    # --- 16 : journalisation ---------------------------------------------

    def test_edition_de_carte_journalisee(self):
        """Editer une carte, c'est delivrer une piece : on trace qui et pour
        qui. Le serveur ne voit pas la boite d'impression du navigateur --
        c'est donc l'edition qui est enregistree, pas l'impression."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        self._carte()
        entree = JournalActivite.objects.get(action=JournalActivite.Action.CARTE)
        self.assertIn('Moussa Diop', entree.objet)
        self.assertIn(self.principal.numero_carte, entree.details)
        self.assertEqual(entree.auteur_libelle, 'Awa Ndiaye')

    def test_le_scan_nest_pas_journalise(self):
        """Consulter une carte au comptoir est un acte de soin courant, pas
        une decision administrative : le journaliser noierait les entrees qui
        comptent (meme regle que les delivrances)."""
        self.client.login(username='pharma-carte@santesn.sn', password=PASSWORD)
        self._scan()
        self.assertEqual(JournalActivite.objects.count(), 0)

    # --- Le QR domine le verso, sans notice technique --------------------

    def test_le_qr_occupe_lessentiel_du_verso(self):
        """Un QR minuscule perdu dans du texte ne se scanne pas au comptoir.
        44 mm sur une carte de 54 mm de haut : le code domine, en gardant la
        marge de silence necessaire a une lecture fiable."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        import re as _re
        contenu = self._carte().content.decode()
        debut = contenu.find('carte-verso-qr')
        largeur = _re.search(r'<svg width="([0-9.]+)mm"', contenu[debut:])
        self.assertIsNotNone(largeur, 'aucun QR trouve au verso')
        self.assertGreaterEqual(float(largeur.group(1)), 40)
        self.assertLessEqual(float(largeur.group(1)), 50)

    def test_aucune_notice_technique_sur_lecran_admin(self):
        """L'interface montre les actions utiles, pas le fonctionnement du
        systeme : le token, l'encodage et la chaine de permissions relevent
        du cahier des charges, pas de l'ecran."""
        self.client.login(username='admin-carte@santesn.sn', password=PASSWORD)
        contenu = self._carte().content.decode()
        # Pas de mot generique ici : "token" matcherait csrfmiddlewaretoken,
        # "permissions" un libelle legitime. On cible les phrases de notice.
        for jargon in ('Ce que fait ce QR', 'aucun droit par elle-même',
                       'Pharmacien — voit les ordonnances',
                       'identifiant sécurisé'):
            self.assertNotIn(jargon, contenu)
        # Les deux actions utiles restent.
        self.assertIn('Imprimer la carte', contenu)


class MonQrCodeAssureTests(TestCase):
    """L'assure consulte son propre QR depuis son profil -- le MEME que celui
    de sa carte imprimee, pas un second identifiant a garder synchronise."""

    def setUp(self):
        self.user_a = creer_utilisateur(User.Role.ASSURE, 'qr-a@santesn.sn')
        self.a = Patient.objects.create(
            user=self.user_a, nom='Ba', prenom='Awa',
            date_naissance=datetime.date(1990, 5, 5), telephone='770000070')
        self.user_b = creer_utilisateur(User.Role.ASSURE, 'qr-b@santesn.sn')
        self.b = Patient.objects.create(
            user=self.user_b, nom='Sy', prenom='Modou',
            date_naissance=datetime.date(1988, 2, 2), telephone='770000071')
        self.client.login(username='qr-a@santesn.sn', password=PASSWORD)

    def test_le_profil_ouvre_le_qr_par_un_bouton(self):
        """Un bouton, pas le code en permanence : on ne presente son QR
        qu'au comptoir. Le code reste dans la page (dialogue natif) pour
        s'ouvrir sans aller-retour serveur."""
        reponse = self.client.get(reverse('mon_profil_assure'))
        self.assertContains(reponse, 'Mon QR code')
        self.assertContains(reponse, 'bouton-voir-qr')
        contenu = reponse.content.decode()
        self.assertIn('<svg', contenu)
        self.assertGreater(contenu.count('<rect'), 20)

    def test_le_numero_de_carte_nest_pas_repete_par_la_section_qr(self):
        """Il est deja sur la carte juste au-dessus, avec son bouton de
        copie. Le repeter en dessous serait un doublon."""
        contenu = self.client.get(reverse('mon_profil_assure')).content.decode()
        # Les deux occurrences legitimes sont dans le bloc carte : le numero
        # affiche et l'attribut data-copier de son bouton de copie. La
        # section QR, qui vient apres, ne doit pas en ajouter une troisieme.
        section_qr = contenu[contenu.find('panel panel-bloc mon-qr'):]
        self.assertNotIn(self.a.numero_carte, section_qr)

    def test_le_qr_du_profil_est_celui_de_la_carte(self):
        """Meme adresse encodee : un second identifiant "pour le profil"
        serait un second systeme, donc une desynchronisation un jour."""
        contenu = self.client.get(reverse('mon_profil_assure')).content.decode()
        attendu = self.a.qr_svg('http://testserver'
                                + reverse('carte_scan', args=[self.a.numero_carte]),
                                taille_mm=52)
        self.assertIn(attendu, contenu)

    def test_un_assure_ne_voit_pas_le_numero_ni_le_qr_dun_autre(self):
        contenu = self.client.get(reverse('mon_profil_assure')).content.decode()
        self.assertNotIn(self.b.numero_carte, contenu)

    def test_un_assure_natteint_pas_la_carte_imprimable_dun_autre(self):
        self.assertEqual(
            self.client.get(reverse('carte_patient', args=[self.b.pk])).status_code, 403)

    def test_le_profil_ne_montre_aucune_notice_technique(self):
        contenu = self.client.get(reverse('mon_profil_assure')).content.decode()
        for jargon in ('Ce que fait ce QR', 'identifiant sécurisé',
                       'aucun droit par elle-même'):
            self.assertNotIn(jargon, contenu)

    def test_profil_sans_fiche_ne_plante_pas(self):
        """Un assure qui n'a pas encore complete son profil n'a pas de numero
        de carte : la section ne doit pas s'afficher a moitie."""
        self.client.logout()
        creer_utilisateur(User.Role.ASSURE, 'qr-sans-fiche@santesn.sn')
        self.client.login(username='qr-sans-fiche@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('mon_profil_assure'))
        self.assertEqual(reponse.status_code, 200)
        # L'invariant qui compte : aucun QR n'est fabrique sans numero de
        # carte. Verifie sur le contexte ET sur le balisage de la section.
        self.assertIsNone(reponse.context['patient'])
        self.assertIsNone(reponse.context['qr_svg'])
        self.assertNotContains(reponse, 'class="panel panel-bloc mon-qr"')



class ColonnesOptionnellesMobileTests(TestCase):
    """Sous 900 px, les listes admin masquent leurs colonnes secondaires.

    Mesure AVANT correctif : un tableau de 9 colonnes faisait 1063 px dans une
    carte de 356 px et la colonne Actions se retrouvait HORS ECRAN -- on ne
    pouvait plus ni modifier ni desactiver un compte depuis un telephone.
    Mesure APRES : aucun debordement sur 12 pages x 3 largeurs."""

    LISTES = [
        'liste_utilisateurs', 'liste_patients', 'liste_medecins',
        'liste_pharmaciens', 'liste_prestataires', 'liste_plans_couverture',
        'liste_prises_en_charge', 'liste_rendez_vous', 'liste_consultations',
        'liste_ordonnances', 'liste_paiements', 'journal_activite',
        'liste_notifications_envoyees',
    ]

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-colonnes@santesn.sn')
        self.client.login(username='admin-colonnes@santesn.sn', password=PASSWORD)
        # De quoi faire rendre au moins une ligne dans chaque liste.
        patient = creer_patient(nom='Sow', prenom='Awa')
        medecin = creer_medecin('medecin-colonnes@santesn.sn')
        creer_pharmacien('pharma-colonnes@santesn.sn')
        Prestataire.objects.create(nom='Hôpital X',
                                   type_prestataire=Prestataire.Type.HOPITAL, ville='Dakar')
        PlanCouverture.objects.create(nom='Plan X', taux_couverture=Decimal('70'))
        PriseEnCharge.objects.create(patient=patient, motif='Suivi')
        RendezVous.objects.create(patient=patient, medecin=medecin,
                                  date_heure=timezone.now(), motif='Contrôle')
        service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('10000'))
        consultation = Consultation.objects.create(
            patient=patient, medecin=medecin, service=service,
            date_consultation=timezone.now(), diagnostic='Test')
        Paiement.calculer_pour(consultation).save()
        Ordonnance.objects.create(consultation=consultation, medicaments='Paracétamol')
        JournalActivite.objects.create(auteur_libelle='Awa', objet='Test',
                                       action=JournalActivite.Action.MODIFICATION)

    def test_chaque_liste_marque_des_colonnes_secondaires(self):
        for nom in self.LISTES:
            self.assertContains(self.client.get(reverse(nom)), 'col-optionnelle',
                                msg_prefix=nom)

    def test_la_colonne_actions_nest_jamais_optionnelle(self):
        """L'invariant central : c'est precisement elle qui disparaissait de
        l'ecran. Elle doit rester visible a toute largeur."""
        for nom in self.LISTES:
            contenu = self.client.get(reverse(nom)).content.decode()
            position = contenu.find('>Actions<')
            if position == -1:
                continue  # liste sans colonne d'actions (lecture seule)
            entete = contenu[contenu.rfind('<th', 0, position):position]
            self.assertNotIn('col-optionnelle', entete, nom)

    def test_les_colonnes_masquees_le_sont_en_tete_ET_en_cellule(self):
        """Masquer l'en-tete sans masquer les cellules decalerait toutes les
        colonnes d'une liste."""
        contenu = self.client.get(reverse('liste_patients')).content.decode()
        entetes = contenu.count('<th scope="col" class="col-optionnelle"') \
            + contenu.count('<th class="col-optionnelle"')
        self.assertGreater(entetes, 0)
        self.assertGreater(contenu.count('<td class="col-optionnelle"')
                           + contenu.count('col-optionnelle">'), entetes)

    def test_la_regle_css_existe_sous_900px(self):
        contenu = self.client.get(reverse('liste_utilisateurs')).content.decode()
        self.assertIn('.col-optionnelle', contenu)
        self.assertIn('@media (max-width: 900px)', contenu)


class RechercheMedecinsAdminTests(TestCase):
    """Seul referentiel admin qui grandit sans filtre : au-dela d'une page,
    retrouver un medecin obligeait a feuilleter. Pharmaciens et plans de
    couverture n'en ont volontairement pas -- quelques lignes chacun."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-med@santesn.sn')
        self.client.login(username='admin-med@santesn.sn', password=PASSWORD)
        self.cardiologue = creer_medecin('cardio@santesn.sn', specialite='Cardiologie')
        self.pediatre = creer_medecin('pediatre@santesn.sn', specialite='Pédiatrie')
        Medecin.objects.filter(pk=self.pediatre.pk).update(nom='Sarr', prenom='Ousmane')
        self.pediatre.refresh_from_db()

    def _lignes(self, **params):
        return list(self.client.get(reverse('liste_medecins'), params).context['medecins'])

    def test_recherche_par_nom(self):
        self.assertEqual(self._lignes(q='Sarr'), [self.pediatre])

    def test_recherche_par_specialite(self):
        self.assertEqual(self._lignes(q='Cardio'), [self.cardiologue])

    def test_recherche_par_email(self):
        self.assertEqual(self._lignes(q='pediatre@'), [self.pediatre])

    def test_sans_recherche_tout_est_liste(self):
        self.assertEqual(len(self._lignes()), 2)

    def test_etat_vide_distingue_la_recherche(self):
        self.assertContains(self.client.get(reverse('liste_medecins'), {'q': 'introuvable'}),
                            'Aucun médecin ne correspond')

    def test_pharmaciens_et_plans_nont_pas_de_barre_de_filtres(self):
        """Quelques lignes chacun : un filtre y serait de la quantite."""
        for nom in ('liste_pharmaciens', 'liste_plans_couverture'):
            self.assertNotContains(self.client.get(reverse(nom)), 'class="filtres"', msg_prefix=nom)


class CloisonnementParUrlTests(TestCase):
    """Audit de securite : changer un identifiant dans l'URL ne doit jamais
    donner acces aux donnees d'un autre assure.

    Toutes ces vues filtrent le proprietaire DANS la requete
    (get_object_or_404(..., patient__in=beneficiaires)) plutot qu'apres
    coup : la reponse est un 404, qui ne revele meme pas que l'objet
    existe. Ces tests verrouillent ce choix -- c'est le genre de garde-fou
    qu'un refactor casse sans s'en apercevoir."""

    def setUp(self):
        self.user_a = creer_utilisateur(User.Role.ASSURE, 'assure-a@santesn.sn')
        self.a = Patient.objects.create(
            user=self.user_a, nom='Diop', prenom='Awa',
            date_naissance=datetime.date(1980, 1, 1), telephone='770000060')
        self.enfant_a = Patient.objects.create(
            nom='Diop', prenom='Fatou', date_naissance=datetime.date(2015, 1, 1),
            type_beneficiaire=Patient.TypeBeneficiaire.AYANT_DROIT,
            lien_parente=Patient.LienParente.ENFANT, assure_principal=self.a)

        self.user_b = creer_utilisateur(User.Role.ASSURE, 'assure-b@santesn.sn')
        self.b = Patient.objects.create(
            user=self.user_b, nom='Sarr', prenom='Modou',
            date_naissance=datetime.date(1985, 1, 1), telephone='770000061')

        self.medecin = creer_medecin('medecin-idor@santesn.sn')

        # Donnees appartenant a B, que A ne doit jamais atteindre.
        self.consultation_b = Consultation.objects.create(
            patient=self.b, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Confidentiel B')
        self.ordonnance_b = Ordonnance.objects.create(
            consultation=self.consultation_b, medicaments='Traitement de B')
        self.rdv_b = RendezVous.objects.create(
            patient=self.b, medecin=self.medecin,
            date_heure=timezone.now() + datetime.timedelta(days=3), motif='RDV de B')
        self.pec_b = PriseEnCharge.objects.create(patient=self.b, motif='Demande de B')
        self.notification_b = Notification.objects.create(
            destinataire=self.user_b, message='Message pour B')

        self.client.login(username='assure-a@santesn.sn', password=PASSWORD)

    # --- Lecture ---------------------------------------------------------

    def test_ordonnance_dun_autre_assure_inaccessible(self):
        reponse = self.client.get(reverse('voir_ordonnance_assure', args=[self.ordonnance_b.pk]))
        self.assertEqual(reponse.status_code, 404)

    def test_ses_propres_ordonnances_restent_accessibles(self):
        """Contre-epreuve : le cloisonnement ne doit pas tout fermer."""
        consultation = Consultation.objects.create(
            patient=self.enfant_a, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Suivi')
        ordonnance = Ordonnance.objects.create(consultation=consultation,
                                               medicaments='Traitement de A')
        self.assertEqual(
            self.client.get(reverse('voir_ordonnance_assure', args=[ordonnance.pk])).status_code,
            200)

    def test_listes_de_lassure_ne_montrent_que_ses_beneficiaires(self):
        for nom, cle in (('mes_ordonnances_assure', 'ordonnances'),
                         ('mes_rendez_vous_assure', 'rendez_vous'),
                         ('mes_prises_en_charge_assure', 'lignes')):
            contenu = self.client.get(reverse(nom)).content.decode()
            self.assertNotIn('Traitement de B', contenu, nom)
            self.assertNotIn('RDV de B', contenu, nom)
            self.assertNotIn('Demande de B', contenu, nom)

    # --- Ecriture --------------------------------------------------------

    def test_annuler_le_rendez_vous_dun_autre_assure_refuse(self):
        reponse = self.client.post(reverse('annuler_rendez_vous_assure', args=[self.rdv_b.pk]))
        self.assertEqual(reponse.status_code, 404)
        self.rdv_b.refresh_from_db()
        self.assertNotEqual(self.rdv_b.statut, RendezVous.Statut.ANNULE)

    def test_modifier_layant_droit_dun_autre_assure_refuse(self):
        # B n'a pas d'ayant droit : on tente sur le patient B lui-meme, qui
        # n'est rattache a personne.
        for nom in ('modifier_ayant_droit', 'supprimer_ayant_droit'):
            self.assertEqual(
                self.client.get(reverse(nom, args=[self.b.pk])).status_code, 404, nom)

    def test_marquer_lue_la_notification_dun_autre_refuse(self):
        reponse = self.client.post(
            reverse('marquer_notification_lue', args=[self.notification_b.pk]))
        self.assertEqual(reponse.status_code, 404)
        self.notification_b.refresh_from_db()
        self.assertFalse(self.notification_b.lue)

    # --- Escalade de role ------------------------------------------------

    def test_un_assure_natteint_aucun_ecran_admin(self):
        for nom in ('dashboard', 'liste_utilisateurs', 'liste_patients',
                    'liste_paiements', 'rapports', 'journal_activite'):
            self.assertEqual(self.client.get(reverse(nom)).status_code, 403, nom)

    def test_un_assure_natteint_ni_lespace_medecin_ni_pharmacien(self):
        for nom in ('agenda_medecin', 'mes_patients', 'historique_consultations',
                    'scanner_ordonnance', 'historique_delivrances'):
            self.assertEqual(self.client.get(reverse(nom)).status_code, 403, nom)

    def test_un_medecin_natteint_pas_la_carte_dun_patient(self):
        """Editer une carte est une delivrance de piece : admin seulement."""
        self.client.logout()
        self.client.login(username='medecin-idor@santesn.sn', password=PASSWORD)
        self.assertEqual(
            self.client.get(reverse('carte_patient', args=[self.b.pk])).status_code, 403)

    def test_une_url_admin_forgee_ne_modifie_rien(self):
        """POST direct sur une action d'ecriture admin, sans passer par l'ecran."""
        reponse = self.client.post(reverse('activer_desactiver_utilisateur', args=[self.user_b.pk]))
        self.assertEqual(reponse.status_code, 403)
        self.user_b.refresh_from_db()
        self.assertTrue(self.user_b.is_active)


class OrdonnanceDocumentTests(TestCase):
    """L'ordonnance est un DOCUMENT, pas un ecran : meme mise en page A4 a
    l'ecran et sur papier, et rien d'autre que le document a l'impression.

    Contrainte du modele : Ordonnance.medicaments est UN SEUL champ de texte
    libre. Ni dosage, ni posologie, ni duree, ni quantite n'existent en base
    -- les lignes saisies par le medecin sont donc rendues telles quelles."""

    def setUp(self):
        self.prestataire = Prestataire.objects.create(
            nom='Hôpital Principal de Dakar',
            type_prestataire=Prestataire.Type.HOPITAL,
            ville='Dakar', telephone='338391010')
        self.medecin = creer_medecin('medecin-doc@santesn.sn', specialite='Cardiologie')
        Medecin.objects.filter(pk=self.medecin.pk).update(prestataire=self.prestataire)
        self.medecin.refresh_from_db()

        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-doc@santesn.sn')
        self.patient = Patient.objects.create(
            user=self.assure_user, nom='Sow', prenom='Moussa',
            date_naissance=datetime.date(1975, 1, 1), telephone='770000080')

        self.consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Hypertension')
        self.ordonnance = Ordonnance.objects.create(
            consultation=self.consultation,
            medicaments='Paracétamol 500 mg — 3×/jour\nAmoxicilline 1 g — 2×/jour')

        self.client.login(username='medecin-doc@santesn.sn', password=PASSWORD)

    def _document(self):
        return self.client.get(reverse('voir_ordonnance_medecin', args=[self.ordonnance.pk]))

    def test_le_document_porte_les_donnees_reelles(self):
        reponse = self._document()
        self.assertContains(reponse, 'SantéSN')
        self.assertContains(reponse, 'Cardiologie')
        self.assertContains(reponse, 'Hôpital Principal de Dakar')
        self.assertContains(reponse, 'Moussa')
        self.assertContains(reponse, self.patient.numero_carte)
        self.assertContains(reponse, '01/01/1975')
        self.assertContains(reponse, self.ordonnance.code_qr)

    def test_une_ligne_saisie_donne_une_ligne_prescrite(self):
        """Pas de colonnes inventees : le modele n'a qu'un champ de texte."""
        lignes = self._document().context['lignes_prescription']
        self.assertEqual(lignes, ['Paracétamol 500 mg — 3×/jour',
                                  'Amoxicilline 1 g — 2×/jour'])

    def test_les_lignes_vides_sont_ignorees(self):
        Ordonnance.objects.filter(pk=self.ordonnance.pk).update(
            medicaments='Ibuprofène 400 mg\n\n\n  \nOméprazole 20 mg\n')
        self.assertEqual(self._document().context['lignes_prescription'],
                         ['Ibuprofène 400 mg', 'Oméprazole 20 mg'])

    def test_ordonnance_sans_medicament_affiche_un_etat_vide(self):
        Ordonnance.objects.filter(pk=self.ordonnance.pk).update(medicaments='')
        reponse = self._document()
        self.assertEqual(reponse.context['lignes_prescription'], [])
        self.assertContains(reponse, "Aucun médicament n'a été saisi")

    def test_beaucoup_de_medicaments_ne_cassent_pas_le_document(self):
        Ordonnance.objects.filter(pk=self.ordonnance.pk).update(
            medicaments='\n'.join(f'Médicament {i} — 1×/jour' for i in range(25)))
        reponse = self._document()
        self.assertEqual(len(reponse.context['lignes_prescription']), 25)
        # Une prescription ne doit jamais etre tranchee entre deux pages.
        self.assertContains(reponse, 'page-break-inside: avoid')

    def test_un_vrai_qr_pas_une_image_decorative(self):
        contenu = self._document().content.decode()
        # On ancre sur le BALISAGE : le nom de classe nu matche d'abord la
        # regle CSS inline de base.html.
        debut = contenu.find('class="feuille-qr"')
        self.assertNotEqual(debut, -1)
        self.assertIn('<svg', contenu[debut:debut + 400])
        self.assertGreater(contenu[debut:].count('<rect'), 20)

    def test_le_qr_nexpose_pas_les_medicaments(self):
        """Il encode le code de verification, pas le contenu medical."""
        contenu = self._document().content.decode()
        debut = contenu.find('class="feuille-qr"')
        fin = contenu.find('</svg>', debut)
        self.assertNotIn('Paracétamol', contenu[debut:fin])
        self.assertNotIn('Amoxicilline', contenu[debut:fin])

    def test_limpression_ne_sort_que_le_document(self):
        contenu = self._document().content.decode()
        self.assertIn('@media print', contenu)
        self.assertIn('size: A4', contenu)
        # Le decor applicatif est masque, y compris les actions de la page.
        self.assertIn('.sans-impression', contenu)
        self.assertIn('class="page-title sans-impression"', contenu)

    def test_espace_de_signature_sans_signature_inventee(self):
        """Aucune signature n'est stockee : on reserve l'espace, on n'en
        fabrique pas une."""
        reponse = self._document()
        self.assertContains(reponse, 'Signature et cachet du médecin')
        self.assertContains(reponse, 'feuille-cadre-signature')

    def test_lassure_voit_le_meme_document(self):
        self.client.logout()
        self.client.login(username='assure-doc@santesn.sn', password=PASSWORD)
        reponse = self.client.get(
            reverse('voir_ordonnance_assure', args=[self.ordonnance.pk]))
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, 'feuille-prescription')
        self.assertEqual(len(reponse.context['lignes_prescription']), 2)

    def test_un_medecin_natteint_pas_lordonnance_dun_confrere(self):
        autre = creer_medecin('autre-doc@santesn.sn')
        self.client.logout()
        self.client.login(username='autre-doc@santesn.sn', password=PASSWORD)
        self.assertEqual(self._document().status_code, 404)
        self.assertIsNotNone(autre)


class LisibiliteDesChampsTests(TestCase):
    """Le texte saisi doit rester lisible DANS LES DEUX THEMES.

    Defaut constate : la regle input/select/textarea ne fixait aucune couleur
    de texte et codait son fond en dur (#fbfdfc). En theme sombre le champ
    gardait donc un fond quasi blanc pendant que le texte heritait de la
    couleur claire de body -- contraste mesure ~1,07:1, on tapait sans rien
    voir. Apres correction : 16,44:1 en clair, 14,42:1 en sombre."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-champs@santesn.sn')
        self.client.login(username='admin-champs@santesn.sn', password=PASSWORD)

    def _feuille(self):
        return self.client.get(reverse('liste_utilisateurs')).content.decode()

    def test_les_champs_fixent_leur_couleur_de_texte(self):
        """Un champ de saisie ne doit jamais dependre de ce dont il herite."""
        feuille = self._feuille()
        debut = feuille.find('        input,\n        select,\n        textarea {')
        self.assertNotEqual(debut, -1, "regle des champs introuvable")
        regle = feuille[debut:feuille.find('}', debut)]
        self.assertIn('color: var(--titre)', regle)
        self.assertIn('background: var(--champ-bg)', regle)
        self.assertNotIn('#fbfdfc', regle)

    def test_le_fond_des_champs_est_defini_dans_les_deux_themes(self):
        """Un jeton defini dans un seul theme est exactement ce qui a cause
        le defaut : le champ gardait sa valeur claire en sombre."""
        feuille = self._feuille()
        self.assertEqual(feuille.count('--champ-bg:'), 2)
        sombre = feuille[feuille.find(':root[data-theme="sombre"]'):]
        self.assertIn('--champ-bg:', sombre[:sombre.find('}')])

    def test_les_listes_deroulantes_aussi(self):
        """Sur certains navigateurs les <option> ne suivent pas le <select>."""
        self.assertIn('select option', self._feuille())

    def test_le_texte_indicatif_reste_distinct_du_texte_saisi(self):
        feuille = self._feuille()
        debut = feuille.find('input::placeholder')
        self.assertNotEqual(debut, -1)
        self.assertIn('color: var(--muted)', feuille[debut:feuille.find('}', debut)])


class ContrasteThemeSombreTests(TestCase):
    """Defauts trouves par MESURE du contraste reel, theme par theme, sur
    12 pages. Deux etaient reels ; les autres signalements etaient des faux
    positifs de la sonde (fond en degrade, texte en background-clip)."""

    def setUp(self):
        creer_utilisateur(User.Role.ADMIN, 'admin-contraste@santesn.sn')
        self.client.login(username='admin-contraste@santesn.sn', password=PASSWORD)

    def _feuille(self):
        return self.client.get(reverse('liste_utilisateurs')).content.decode()

    def test_le_bouton_destructif_a_son_propre_jeton_de_fond(self):
        """--danger sert AUSSI de couleur de texte : en sombre il s'eclaircit
        pour rester lisible sur fond sombre, et un aplat clair sous un libelle
        blanc tombait alors a 3,26:1. Un jeton, un role."""
        feuille = self._feuille()
        debut = feuille.find('.button.btn-danger {')
        self.assertNotEqual(debut, -1)
        regle = feuille[debut:feuille.find('}', debut)]
        self.assertIn('background: var(--danger-surface)', regle)
        self.assertNotIn('background: var(--danger)', regle)

    def test_le_jeton_de_fond_destructif_existe_dans_les_deux_themes(self):
        feuille = self._feuille()
        self.assertEqual(feuille.count('--danger-surface:'), 2)
        sombre = feuille[feuille.find(':root[data-theme="sombre"]'):]
        self.assertIn('--danger-surface:', sombre[:sombre.find('}')])

    def test_la_mention_du_verso_fixe_sa_couleur(self):
        """Le verso de la carte est CLAIR alors que .carte-assure impose du
        blanc pour le recto : sans couleur explicite la mention sortait blanc
        sur blanc (1,08:1). Une seule regle pour cet element -- il y en avait
        deux, ce qui rendait la cascade illisible."""
        feuille = self._feuille()
        self.assertEqual(feuille.count('.carte-verso-mention {'), 1)
        debut = feuille.find('.carte-verso-mention {')
        self.assertIn('color:', feuille[debut:feuille.find('}', debut)])


class RequetesConstantesTests(TestCase):
    """Le nombre de requetes d'une liste ne doit pas suivre son nombre de
    lignes. Audit mene sur 22 ecrans : 21 etaient deja constants, un seul
    presentait un vrai N+1."""

    def setUp(self):
        self.medecin = creer_medecin('medecin-n1@santesn.sn')
        self.client.login(username='medecin-n1@santesn.sn', password=PASSWORD)

    def _consultations(self, combien):
        for i in range(combien):
            patient = creer_patient(nom=f'N{i}', prenom=f'P{i}')
            consultation = Consultation.objects.create(
                patient=patient, medecin=self.medecin,
                date_consultation=timezone.now(), diagnostic='D')
            Ordonnance.objects.create(consultation=consultation, medicaments='X')

    def test_historique_consultations_ne_requete_pas_par_ligne(self):
        """.first() sur une relation prechargee emet une NOUVELLE requete
        LIMIT 1 et ignore le cache du prefetch : une par ligne affichee.
        Le filtre |first travaille sur la liste deja chargee.
        Mesure avant correctif : 11 requetes a 3 lignes, 28 a 30."""
        self._consultations(3)
        with CaptureQueriesContext(connection) as petit:
            self.client.get(reverse('historique_consultations'))
        self._consultations(15)
        with CaptureQueriesContext(connection) as grand:
            self.client.get(reverse('historique_consultations'))
        self.assertEqual(len(petit.captured_queries), len(grand.captured_queries))

    def test_le_gabarit_nappelle_pas_first_sur_la_relation_prechargee(self):
        self._consultations(1)
        contenu = self.client.get(reverse('historique_consultations')).content.decode()
        self.assertNotIn('ordonnance_set.first', contenu)


class ParcoursCompletsTests(TestCase):
    """Parcours de BOUT EN BOUT, role par role.

    Les ecrans etaient testes un par un. Ces tests verifient les TRANSITIONS :
    chaque etape doit etre atteignable depuis la precedente par un lien
    reellement present dans la page, et non seulement par son URL. C'est le
    defaut que des tests par ecran ne voient pas -- A fonctionne, B fonctionne,
    mais A ne mene pas a B.
    """

    def setUp(self):
        self.admin = creer_utilisateur(User.Role.ADMIN, 'admin-parcours@santesn.sn')
        self.admin.first_name, self.admin.last_name = 'Awa', 'Ndiaye'
        self.admin.save()

        self.plan = PlanCouverture.objects.create(
            nom='Famille 80%', taux_couverture=Decimal('80'))
        self.assure_user = creer_utilisateur(User.Role.ASSURE, 'assure-parcours@santesn.sn')
        self.patient = Patient.objects.create(
            user=self.assure_user, nom='Diop', prenom='Awa',
            date_naissance=datetime.date(1985, 6, 6), telephone='770000090',
            plan_couverture=self.plan)

        self.prestataire = Prestataire.objects.create(
            nom='Hôpital Principal', type_prestataire=Prestataire.Type.HOPITAL,
            ville='Dakar')
        self.medecin = creer_medecin('medecin-parcours@santesn.sn')
        Medecin.objects.filter(pk=self.medecin.pk).update(prestataire=self.prestataire)
        self.pharmacien = creer_pharmacien('pharma-parcours@santesn.sn')

        self.service = ServiceMedical.objects.create(nom='Consultation', prix=Decimal('15000'))
        self.pec = PriseEnCharge.objects.create(
            patient=self.patient, motif='Suivi', statut='validee')
        self.consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin, service=self.service,
            prise_en_charge=self.pec, date_consultation=timezone.now(),
            diagnostic='Bilan', traitement='Repos')
        Paiement.calculer_pour(self.consultation).save()
        self.ordonnance = Ordonnance.objects.create(
            consultation=self.consultation, medicaments='Paracétamol 500 mg')

    # ------------------------------------------------------------------
    def _connexion(self, email):
        """Passe par le VRAI formulaire de connexion, pas par force_login :
        c'est la premiere transition du parcours."""
        reponse = self.client.post(reverse('login'),
                                   {'email': email, 'password': PASSWORD}, follow=True)
        self.assertEqual(reponse.status_code, 200, f"connexion refusee : {email}")
        return reponse

    def _etape(self, precedente, url, nom_etape):
        """Verifie que `url` est ATTEIGNABLE depuis la page precedente (le lien
        y figure) puis l'ouvre."""
        self.assertContains(precedente, url,
                            msg_prefix=f"aucun lien vers {nom_etape}")
        reponse = self.client.get(url, follow=True)
        self.assertEqual(reponse.status_code, 200, nom_etape)
        return reponse

    # --- ADMINISTRATEUR -----------------------------------------------
    def test_parcours_administrateur(self):
        page = self._connexion('admin-parcours@santesn.sn')
        # La connexion doit deposer sur le tableau de bord du role.
        self.assertContains(page, 'Tableau de bord')

        page = self._etape(page, reverse('liste_utilisateurs'), 'Utilisateurs')
        page = self._etape(page, reverse('liste_patients'), 'Assurés')
        page = self._etape(page, reverse('carte_patient', args=[self.patient.pk]),
                           'carte du bénéficiaire')
        self.assertContains(page, self.patient.numero_carte)
        self.assertContains(page, '<svg')                      # le QR y est

        page = self._etape(page, reverse('liste_patients'), 'retour Assurés')

        parametres = self.client.get(reverse('parametres'), follow=True)
        page = self._etape(parametres, reverse('parametres_section', args=['securite']),
                           'Paramètres → Sécurité')
        self.assertContains(page, 'Comptes temporairement bloqués')

        page = self._etape(page, reverse('journal_activite'), 'Journal d\'activité')
        # L'edition de carte faite plus haut doit y figurer : le parcours
        # laisse une trace, ce n'est pas qu'un enchainement d'ecrans.
        self.assertContains(page, 'Carte éditée')

        page = self._etape(page, reverse('parametres_section', args=['securite']),
                           'retour Sécurité')
        retour = self.client.get(reverse('dashboard'), follow=True)
        self.assertEqual(retour.status_code, 200)

    def test_parcours_administrateur_deblocage(self):
        """Blocage reel -> alerte du tableau de bord -> Securite -> deblocage."""
        TentativeConnexion.objects.create(
            email='assure-parcours@santesn.sn',
            tentatives=TentativeConnexion.MAX_TENTATIVES)
        self._connexion('admin-parcours@santesn.sn')

        tableau = self.client.get(reverse('dashboard'))
        self.assertContains(tableau, 'temporairement bloqué')
        page = self._etape(tableau, reverse('parametres_section', args=['securite']),
                           'alerte → Sécurité')
        self.assertContains(page, 'assure-parcours@santesn.sn')

        reponse = self.client.post(
            reverse('debloquer_compte', args=[self.assure_user.pk]), follow=True)
        self.assertEqual(reponse.status_code, 200)
        self.assertFalse(TentativeConnexion.bloque('assure-parcours@santesn.sn'))
        # L'action doit etre tracee, et visible depuis le journal.
        journal = self.client.get(reverse('journal_activite'))
        self.assertContains(journal, 'Déblocage')

    # --- ASSURE --------------------------------------------------------
    def test_parcours_assure(self):
        page = self._connexion('assure-parcours@santesn.sn')
        page = self._etape(page, reverse('mon_profil_assure'), 'Mon profil')
        self.assertContains(page, 'Mon QR code')
        self.assertContains(page, 'bouton-voir-qr')

        page = self._etape(page, reverse('mon_historique_assure'), 'Mon historique')
        page = self._etape(page, reverse('mes_prises_en_charge_assure'),
                           'Mes prises en charge')
        self.assertContains(page, 'Suivi')

        page = self._etape(page, reverse('mes_ordonnances_assure'), 'Mes ordonnances')
        page = self._etape(page, reverse('voir_ordonnance_assure', args=[self.ordonnance.pk]),
                           'ordonnance')
        self.assertContains(page, 'Paracétamol 500 mg')

    # --- MEDECIN -------------------------------------------------------
    def test_parcours_medecin(self):
        page = self._connexion('medecin-parcours@santesn.sn')
        page = self._etape(page, reverse('mes_patients'), 'Mes patients')
        page = self._etape(page, reverse('fiche_patient_medecin', args=[self.patient.pk]),
                           'fiche patient')
        self.assertContains(page, self.patient.numero_carte)

        page = self._etape(page, reverse('historique_consultations'), 'Consultations')
        page = self._etape(page, reverse('voir_ordonnance_medecin', args=[self.ordonnance.pk]),
                           'ordonnance')
        # Le document doit etre pret a imprimer, pas seulement affiche.
        self.assertContains(page, 'Imprimer l\'ordonnance')
        self.assertContains(page, 'size: A4')

    def test_parcours_medecin_creation_dune_ordonnance(self):
        """Consultation -> ordonnance -> apercu, en passant par les vrais
        formulaires."""
        self._connexion('medecin-parcours@santesn.sn')
        creation = self.client.post(reverse('ajouter_consultation_medecin'), {
            'patient': self.patient.pk, 'service': self.service.pk,
            'prise_en_charge': self.pec.pk,
            'date_consultation': '2026-08-17T09:00',
            'diagnostic': 'Angine', 'traitement': 'Repos',
        }, follow=True)
        self.assertEqual(creation.status_code, 200)
        consultation = Consultation.objects.filter(diagnostic='Angine').first()
        self.assertIsNotNone(consultation, "la consultation n'a pas ete enregistree")

        ordonnance = self.client.post(
            reverse('ajouter_ordonnance_medecin', args=[consultation.pk]),
            {
                # Creation par lignes structurees : le champ de texte libre
                # n'existe plus dans le formulaire du medecin.
                'lignes-TOTAL_FORMS': '1', 'lignes-INITIAL_FORMS': '0',
                'lignes-MIN_NUM_FORMS': '0', 'lignes-MAX_NUM_FORMS': '1000',
                'lignes-0-medicament': 'Amoxicilline', 'lignes-0-dosage': '1 g',
                'lignes-0-posologie': '', 'lignes-0-duree': '', 'lignes-0-quantite': '',
            }, follow=True)
        self.assertEqual(ordonnance.status_code, 200)
        creee = Ordonnance.objects.filter(consultation=consultation).first()
        self.assertIsNotNone(creee, "l'ordonnance n'a pas ete enregistree")
        document = self.client.get(reverse('voir_ordonnance_medecin', args=[creee.pk]))
        self.assertContains(document, 'Amoxicilline')
        self.assertContains(document, '1 g')

    # --- PHARMACIEN ----------------------------------------------------
    def test_parcours_pharmacien(self):
        page = self._connexion('pharma-parcours@santesn.sn')
        page = self._etape(page, reverse('scanner_ordonnance'), 'Scanner')

        # Le code exact : chemin normal du comptoir.
        trouvee = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': self.ordonnance.code_qr})
        self.assertEqual(trouvee.status_code, 200)
        self.assertContains(trouvee, 'Paracétamol 500 mg')

        validee = self.client.post(reverse('valider_delivrance', args=[self.ordonnance.pk]),
                                   {'code_qr': self.ordonnance.code_qr}, follow=True)
        self.assertEqual(validee.status_code, 200)
        self.ordonnance.refresh_from_db()
        self.assertTrue(Delivrance.objects.filter(ordonnance=self.ordonnance).exists())
        # La validation doit deposer sur l'historique, ou la delivrance figure.
        self.assertContains(validee, 'Diop')

    def test_parcours_pharmacien_repli_code_illisible(self):
        """QR froisse : la recherche par nom doit mener a la meme ordonnance."""
        self._connexion('pharma-parcours@santesn.sn')
        reponse = self.client.post(reverse('scanner_ordonnance'), {'recherche': 'Diop'})
        self.assertEqual(reponse.status_code, 200)
        self.assertContains(reponse, self.ordonnance.code_qr)

    # --- Transitions refusees ------------------------------------------
    def test_la_deconnexion_ferme_reellement_le_parcours(self):
        self._connexion('admin-parcours@santesn.sn')
        self.assertEqual(self.client.get(reverse('dashboard')).status_code, 200)
        self.client.post(reverse('logout'))
        suite = self.client.get(reverse('dashboard'))
        self.assertEqual(suite.status_code, 302)
        self.assertIn(reverse('login'), suite.url)


class LigneOrdonnanceTests(TestCase):
    """Phase 1 : structure et affichage. Le formulaire medecin multi-lignes
    est une phase 2 dediee.

    Regle centrale : une ordonnance a SOIT des lignes structurees, SOIT son
    texte historique -- jamais les deux melanges, et aucune conversion de
    l'un vers l'autre."""

    def setUp(self):
        self.patient = creer_patient(nom='Ba', prenom='Awa')
        self.medecin = creer_medecin('medecin-ligne@santesn.sn')
        self.consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='D')

    def _ordonnance(self, texte=''):
        return Ordonnance.objects.create(consultation=self.consultation,
                                         medicaments=texte)

    # --- Modele ---------------------------------------------------------

    def test_creation_dune_ligne(self):
        ordonnance = self._ordonnance()
        ligne = LigneOrdonnance.objects.create(
            ordonnance=ordonnance, medicament='Paracétamol', dosage='500 mg',
            posologie='3×/jour', duree='5 jours', quantite='1 boîte')
        self.assertEqual(list(ordonnance.lignes.all()), [ligne])
        self.assertIn('Paracétamol', str(ligne))

    def test_plusieurs_lignes_et_ordre_conserve(self):
        ordonnance = self._ordonnance()
        troisieme = LigneOrdonnance.objects.create(
            ordonnance=ordonnance, medicament='C', ordre=3)
        premiere = LigneOrdonnance.objects.create(
            ordonnance=ordonnance, medicament='A', ordre=1)
        deuxieme = LigneOrdonnance.objects.create(
            ordonnance=ordonnance, medicament='B', ordre=2)
        self.assertEqual(list(ordonnance.lignes.all()),
                         [premiere, deuxieme, troisieme])

    def test_ordre_egal_conserve_lordre_de_creation(self):
        """Meta.ordering finit par pk : deux lignes de meme ordre ne doivent
        jamais s'echanger d'un affichage a l'autre."""
        ordonnance = self._ordonnance()
        a = LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='A')
        b = LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='B')
        self.assertEqual([l.pk for l in ordonnance.lignes.all()], [a.pk, b.pk])

    def test_modification_et_suppression_dune_ligne(self):
        ordonnance = self._ordonnance()
        ligne = LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='A')
        ligne.medicament = 'B'
        ligne.save()
        self.assertEqual(ordonnance.lignes.first().medicament, 'B')
        ligne.delete()
        self.assertEqual(ordonnance.lignes.count(), 0)

    def test_ligne_sans_medicament_refusee(self):
        """Le seul champ obligatoire : une ligne sans medicament n'a pas de
        sens. Les autres restent facultatifs -- imposer une posologie
        pousserait a remplir du vide pour passer la validation."""
        ligne = LigneOrdonnance(ordonnance=self._ordonnance(), medicament='')
        with self.assertRaises(ValidationError):
            ligne.full_clean()

    def test_les_autres_champs_sont_facultatifs(self):
        ligne = LigneOrdonnance(ordonnance=self._ordonnance(), medicament='Paracétamol')
        ligne.full_clean()          # ne doit pas lever
        ligne.save()
        self.assertEqual(ligne.dosage, '')

    def test_supprimer_lordonnance_supprime_ses_lignes(self):
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='A')
        ordonnance.delete()
        self.assertEqual(LigneOrdonnance.objects.count(), 0)

    # --- Affichage : les deux formats ne se melangent pas ---------------

    def test_ordonnance_structuree_rend_le_tableau(self):
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(
            ordonnance=ordonnance, medicament='Amoxicilline', dosage='1 g',
            posologie='2×/jour', duree='7 jours', quantite='14 comprimés')
        self.client.login(username='medecin-ligne@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('voir_ordonnance_medecin', args=[ordonnance.pk]))
        self.assertEqual(len(reponse.context['lignes_structurees']), 1)
        self.assertEqual(reponse.context['lignes_prescription'], [])
        for valeur in ('Amoxicilline', '1 g', '2×/jour', '7 jours', '14 comprimés'):
            self.assertContains(reponse, valeur)
        self.assertContains(reponse, 'class=\"feuille-table\"')

    def test_ordonnance_historique_rend_son_texte_inchange(self):
        texte = "| Médicament | Dosage |\n| Paracétamol | 500 mg |"
        ordonnance = self._ordonnance(texte)
        self.client.login(username='medecin-ligne@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('voir_ordonnance_medecin', args=[ordonnance.pk]))
        self.assertEqual(reponse.context['lignes_structurees'], [])
        self.assertEqual(len(reponse.context['lignes_prescription']), 2)
        # Le nom de classe figure aussi dans la CSS inline : on cible le balisage.
        self.assertNotContains(reponse, 'class=\"feuille-table\"')

    def test_les_deux_formats_ne_se_melangent_jamais(self):
        """Une ordonnance qui aurait a la fois du texte ET des lignes affiche
        les lignes, sans dupliquer le texte."""
        ordonnance = self._ordonnance('Ancien texte libre')
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='Nouveau')
        self.client.login(username='medecin-ligne@santesn.sn', password=PASSWORD)
        reponse = self.client.get(reverse('voir_ordonnance_medecin', args=[ordonnance.pk]))
        self.assertContains(reponse, 'Nouveau')
        self.assertNotContains(reponse, 'Ancien texte libre')

    def test_champ_facultatif_vide_affiche_un_tiret_et_rien_dinvente(self):
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='Ibuprofène')
        self.client.login(username='medecin-ligne@santesn.sn', password=PASSWORD)
        self.assertContains(
            self.client.get(reverse('voir_ordonnance_medecin', args=[ordonnance.pk])), '—')

    # --- Par role -------------------------------------------------------

    def test_assure_voit_le_tableau_structure(self):
        assure = creer_utilisateur(User.Role.ASSURE, 'assure-ligne@santesn.sn')
        Patient.objects.filter(pk=self.patient.pk).update(user=assure)
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='Doliprane')
        self.client.login(username='assure-ligne@santesn.sn', password=PASSWORD)
        self.assertContains(
            self.client.get(reverse('voir_ordonnance_assure', args=[ordonnance.pk])),
            'Doliprane')

    def test_pharmacien_voit_les_lignes_apres_scan(self):
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='Amoxicilline',
                                       dosage='1 g', posologie='2×/jour')
        creer_pharmacien('pharma-ligne@santesn.sn')
        self.client.login(username='pharma-ligne@santesn.sn', password=PASSWORD)
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': ordonnance.code_qr})
        self.assertContains(reponse, 'Amoxicilline')
        self.assertContains(reponse, '2×/jour')
        self.assertContains(reponse, 'class=\"table-prescription\"')

    def test_pharmacien_voit_le_texte_dune_ordonnance_historique(self):
        ordonnance = self._ordonnance('Traitement historique')
        creer_pharmacien('pharma-histo-ligne@santesn.sn')
        self.client.login(username='pharma-histo-ligne@santesn.sn', password=PASSWORD)
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': ordonnance.code_qr})
        self.assertContains(reponse, 'Traitement historique')
        self.assertNotContains(reponse, 'class=\"table-prescription\"')

    # --- Non-regression sur les donnees existantes ----------------------

    def test_le_texte_historique_nest_ni_transforme_ni_efface(self):
        """La garantie de la migration : ajouter des lignes a une ordonnance
        ne touche jamais son champ medicaments."""
        texte = 'Medicament A 500mg - 2x/jour pendant 5 jours'
        ordonnance = self._ordonnance(texte)
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='B')
        ordonnance.refresh_from_db()
        self.assertEqual(ordonnance.medicaments, texte)

    def test_une_ordonnance_peut_navoir_aucune_ligne(self):
        """C'est le cas de TOUTES les ordonnances anterieures."""
        ordonnance = self._ordonnance('Texte')
        self.assertEqual(ordonnance.lignes.count(), 0)
        self.assertFalse(ordonnance.lignes.exists())

    def test_le_qr_reste_inchange(self):
        """Phase 1 ne touche pas au QR : il encode le code de verification,
        pas le contenu medical."""
        ordonnance = self._ordonnance()
        LigneOrdonnance.objects.create(ordonnance=ordonnance, medicament='Secret')
        self.assertNotIn('Secret', ordonnance.qr_svg)
        self.assertTrue(ordonnance.code_qr.startswith('RX-'))


class FormulaireOrdonnanceMedecinTests(TestCase):
    """Phase 2 : le medecin saisit une ordonnance STRUCTUREE.

    Il n'existe pas de modification d'ordonnance dans SantéSN -- seulement
    la creation. Ces tests ne verifient donc pas une modification : en
    inventer une serait ajouter une regle metier au passage."""

    PREFIXE = "lignes"

    def setUp(self):
        self.patient = creer_patient(nom='Sy', prenom='Awa')
        self.medecin = creer_medecin('medecin-formset@santesn.sn')
        self.autre_medecin = creer_medecin('autre-formset@santesn.sn')
        self.consultation = Consultation.objects.create(
            patient=self.patient, medecin=self.medecin,
            date_consultation=timezone.now(), diagnostic='Angine')
        self.client.login(username='medecin-formset@santesn.sn', password=PASSWORD)

    def _url(self, consultation=None):
        return reverse('ajouter_ordonnance_medecin',
                       args=[(consultation or self.consultation).pk])

    def _donnees(self, lignes):
        data = {f'{self.PREFIXE}-TOTAL_FORMS': str(len(lignes)),
                f'{self.PREFIXE}-INITIAL_FORMS': '0',
                f'{self.PREFIXE}-MIN_NUM_FORMS': '0',
                f'{self.PREFIXE}-MAX_NUM_FORMS': '1000'}
        for i, ligne in enumerate(lignes):
            for champ in ('medicament', 'dosage', 'posologie', 'duree', 'quantite'):
                data[f'{self.PREFIXE}-{i}-{champ}'] = ligne.get(champ, '')
        return data

    # --- Affichage ------------------------------------------------------

    def test_le_formulaire_souvre_sur_une_ligne_vide(self):
        reponse = self.client.get(self._url())
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.context['formset'].total_form_count(), 1)
        for libelle in ('Médicament', 'Dosage', 'Posologie', 'Durée', 'Quantité'):
            self.assertContains(reponse, libelle)
        self.assertContains(reponse, 'Ajouter un médicament')

    def test_le_gabarit_de_ligne_vierge_est_fourni_au_navigateur(self):
        """Le bouton clone ce gabarit : sans lui, l'ajout dynamique est mort."""
        reponse = self.client.get(self._url())
        self.assertContains(reponse, 'id="gabarit-ligne"')
        self.assertContains(reponse, '__prefix__')
        self.assertContains(reponse, f'id_{self.PREFIXE}-TOTAL_FORMS')

    def test_chaque_champ_a_un_vrai_label(self):
        contenu = self.client.get(self._url()).content.decode()
        for champ in ('medicament', 'dosage', 'posologie', 'duree', 'quantite'):
            self.assertIn(f'for="id_{self.PREFIXE}-0-{champ}"', contenu)

    # --- Enregistrement -------------------------------------------------

    def test_creation_avec_une_ligne(self):
        reponse = self.client.post(self._url(), self._donnees([
            {'medicament': 'Paracétamol', 'dosage': '500 mg',
             'posologie': '3×/jour', 'duree': '5 jours', 'quantite': '1 boîte'},
        ]), follow=True)
        self.assertEqual(reponse.status_code, 200)
        ordonnance = Ordonnance.objects.get(consultation=self.consultation)
        ligne = ordonnance.lignes.get()
        self.assertEqual(ligne.medicament, 'Paracétamol')
        self.assertEqual(ligne.duree, '5 jours')
        self.assertEqual(ligne.ordre, 1)

    def test_creation_avec_plusieurs_lignes_et_ordre_de_saisie(self):
        self.client.post(self._url(), self._donnees([
            {'medicament': 'Amoxicilline'},
            {'medicament': 'Ibuprofène'},
            {'medicament': 'Vitamine C'},
        ]), follow=True)
        ordonnance = Ordonnance.objects.get(consultation=self.consultation)
        self.assertEqual([l.medicament for l in ordonnance.lignes.all()],
                         ['Amoxicilline', 'Ibuprofène', 'Vitamine C'])
        self.assertEqual([l.ordre for l in ordonnance.lignes.all()], [1, 2, 3])

    def test_les_champs_facultatifs_peuvent_rester_vides(self):
        self.client.post(self._url(), self._donnees([{'medicament': 'Doliprane'}]),
                         follow=True)
        ligne = Ordonnance.objects.get(consultation=self.consultation).lignes.get()
        self.assertEqual(ligne.dosage, '')
        self.assertEqual(ligne.posologie, '')

    def test_la_nouvelle_ordonnance_na_pas_de_texte_libre(self):
        """medicaments ne porte plus que le contenu des ordonnances
        anterieures a LigneOrdonnance."""
        self.client.post(self._url(), self._donnees([{'medicament': 'X'}]), follow=True)
        self.assertEqual(Ordonnance.objects.get(consultation=self.consultation).medicaments, '')

    def test_rien_nest_complete_automatiquement(self):
        """'500 mg' ecrit dans le nom NE DOIT PAS glisser dans dosage."""
        self.client.post(self._url(),
                         self._donnees([{'medicament': 'Paracétamol 500 mg'}]), follow=True)
        ligne = Ordonnance.objects.get(consultation=self.consultation).lignes.get()
        self.assertEqual(ligne.medicament, 'Paracétamol 500 mg')
        self.assertEqual(ligne.dosage, '')

    # --- Validation serveur ---------------------------------------------

    def test_formulaire_entierement_vide_refuse(self):
        reponse = self.client.post(self._url(), self._donnees([{}]))
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(Ordonnance.objects.count(), 0)
        self.assertContains(reponse, 'au moins un médicament')

    def test_ligne_sans_medicament_mais_avec_dosage_refusee(self):
        """Un dosage sans medicament est une prescription incomprehensible."""
        reponse = self.client.post(self._url(),
                                   self._donnees([{'dosage': '500 mg'}]))
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(Ordonnance.objects.count(), 0)

    def test_ligne_vide_ignoree_si_une_autre_est_remplie(self):
        self.client.post(self._url(), self._donnees([
            {'medicament': 'Amoxicilline'}, {},
        ]), follow=True)
        ordonnance = Ordonnance.objects.get(consultation=self.consultation)
        self.assertEqual(ordonnance.lignes.count(), 1)

    def test_aucune_ordonnance_creee_si_le_formset_est_invalide(self):
        """La transaction protege : pas d'ordonnance orpheline sans ligne."""
        self.client.post(self._url(), self._donnees([{}]))
        self.assertEqual(Ordonnance.objects.count(), 0)
        self.assertEqual(LigneOrdonnance.objects.count(), 0)

    def test_total_forms_incoherent_ne_casse_pas_la_page(self):
        donnees = self._donnees([{'medicament': 'A'}])
        donnees[f'{self.PREFIXE}-TOTAL_FORMS'] = '99'
        reponse = self.client.post(self._url(), donnees)
        self.assertIn(reponse.status_code, (200, 302))

    # --- Permissions ----------------------------------------------------

    def test_un_medecin_ne_prescrit_pas_sur_la_consultation_dun_confrere(self):
        autre = Consultation.objects.create(
            patient=self.patient, medecin=self.autre_medecin,
            date_consultation=timezone.now(), diagnostic='Autre')
        reponse = self.client.post(self._url(autre),
                                   self._donnees([{'medicament': 'X'}]))
        self.assertEqual(reponse.status_code, 404)
        self.assertEqual(Ordonnance.objects.count(), 0)

    def test_consultation_inexistante_donne_404(self):
        self.assertEqual(
            self.client.get(reverse('ajouter_ordonnance_medecin', args=[999999])).status_code,
            404)

    def test_les_autres_roles_sont_refuses(self):
        for role, email in ((User.Role.ASSURE, 'assure-formset@santesn.sn'),
                            (User.Role.PHARMACIEN, 'pharma-formset@santesn.sn'),
                            (User.Role.ADMIN, 'admin-formset@santesn.sn')):
            self.client.logout()
            creer_utilisateur(role, email)
            self.client.login(username=email, password=PASSWORD)
            self.assertEqual(self.client.get(self._url()).status_code, 403, role)

    def test_anonyme_redirige(self):
        self.client.logout()
        self.assertEqual(self.client.get(self._url()).status_code, 302)

    # --- Parcours de bout en bout ---------------------------------------

    def test_parcours_creation_puis_apercu_a4(self):
        reponse = self.client.post(self._url(), self._donnees([
            {'medicament': 'Amoxicilline', 'dosage': '1 g', 'posologie': '2×/jour',
             'duree': '7 jours', 'quantite': '14 comprimés'},
            {'medicament': 'Paracétamol', 'dosage': '500 mg'},
        ]), follow=True)
        # La creation depose directement sur le document.
        self.assertContains(reponse, 'Imprimer l\'ordonnance')
        self.assertContains(reponse, 'class="feuille-table"')
        for valeur in ('Amoxicilline', '1 g', '14 comprimés', 'Paracétamol'):
            self.assertContains(reponse, valeur)

    def test_le_pharmacien_lit_ensuite_les_lignes_saisies(self):
        self.client.post(self._url(), self._donnees([
            {'medicament': 'Amoxicilline', 'posologie': '2×/jour'},
        ]), follow=True)
        ordonnance = Ordonnance.objects.get(consultation=self.consultation)
        self.client.logout()
        creer_pharmacien('pharma-lecture@santesn.sn')
        self.client.login(username='pharma-lecture@santesn.sn', password=PASSWORD)
        reponse = self.client.post(reverse('scanner_ordonnance'),
                                   {'code_qr': ordonnance.code_qr})
        self.assertContains(reponse, 'Amoxicilline')
        self.assertContains(reponse, '2×/jour')

    def test_le_qr_reste_intact_et_nexpose_pas_les_medicaments(self):
        self.client.post(self._url(),
                         self._donnees([{'medicament': 'Morphine'}]), follow=True)
        ordonnance = Ordonnance.objects.get(consultation=self.consultation)
        self.assertTrue(ordonnance.code_qr.startswith('RX-'))
        self.assertNotIn('Morphine', ordonnance.qr_svg)
