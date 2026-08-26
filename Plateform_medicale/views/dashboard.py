"""
Dashboard administrateur, rapports d'activité et exports (Excel, PDF).
"""

import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Line, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import HRFlowable, KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.db.models import Case, Count, IntegerField, Q, Sum, Value, When
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from ..models import (
    Consultation,
    Delivrance,
    JournalActivite,
    Medecin,
    Ordonnance,
    Paiement,
    Patient,
    Pharmacien,
    PriseEnCharge,
    Prestataire,
    RendezVous,
    TentativeConnexion,
    User,
)
from .utils import (
    MOIS_ABREGES,
    admin_required,
    compteurs_files_attente,
    _paginer,
    _trier,
)


@admin_required
def dashboard(request):
    # Les demandes en attente remontent en tete : triee par date seule, la
    # liste noyait l'urgence sous des dossiers deja clos.
    dernieres_prises_en_charge = (
        PriseEnCharge.objects.select_related("patient")
        .annotate(
            priorite=Case(
                When(statut="en_attente", then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
        .order_by("priorite", "-date_demande")[:5]
    )

    # Bandeau financier : meme agregat que liste_paiements (Sum filtre par
    # statut), pour un signal de sante financiere absent jusqu'ici du
    # dashboard alors que c'est la donnee la plus parlante pour une
    # compagnie d'assurance/IPM qui evaluerait la plateforme.
    totaux_paiements = Paiement.objects.aggregate(
        total_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.REGLE)),
        total_non_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.NON_REGLE)),
    )
    montant_regle = totaux_paiements["total_regle"] or 0
    montant_non_regle = totaux_paiements["total_non_regle"] or 0
    montant_total_paiements = montant_regle + montant_non_regle
    taux_reglement = round((montant_regle / montant_total_paiements) * 100) if montant_total_paiements else None

    # Activite du jour, plateforme entiere (pas un seul medecin)
    maintenant = timezone.now()
    debut_jour = maintenant.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_jour = debut_jour + datetime.timedelta(days=1)
    total_rendez_vous_aujourd_hui = (
        RendezVous.objects.filter(date_heure__gte=debut_jour, date_heure__lt=fin_jour)
        .exclude(statut=RendezVous.Statut.ANNULE)
        .count()
    )
    total_consultations_aujourd_hui = Consultation.objects.filter(
        date_consultation__gte=debut_jour, date_consultation__lt=fin_jour
    ).count()

    # Tendance des paiements regles sur les 7 derniers jours
    il_y_a_7_jours = maintenant - datetime.timedelta(days=7)
    montant_regle_7j = Paiement.objects.filter(
        statut=Paiement.Statut.REGLE, date_reglement__gte=il_y_a_7_jours
    ).aggregate(total=Sum("montant_part_patient"))["total"] or 0

    # Borne haute explicite pour eviter les consultations futures mal datees
    consultations_7j = Consultation.objects.filter(
        date_consultation__gte=il_y_a_7_jours, date_consultation__lte=maintenant
    ).count()
    ordonnances_7j = Ordonnance.objects.filter(date_creation__gte=il_y_a_7_jours).count()

    # Anciennete de la file d'attente
    plus_ancienne_attente = (
        PriseEnCharge.objects.filter(statut="en_attente").order_by("date_demande").first()
    )
    jours_attente_max = (
        (maintenant - plus_ancienne_attente.date_demande).days if plus_ancienne_attente else None
    )

    # Derniers comptes crees, hors assures
    derniers_comptes = (
        User.objects.exclude(role=User.Role.ASSURE).order_by("-date_joined")[:5]
    )

    tendance_paiements = _montants_regles_par_jour()

    # Un seul aller-retour pour les trois comptages de Patient
    repartition_patients = Patient.objects.aggregate(
        total=Count("id"),
        principaux=Count("id", filter=Q(type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL)),
        sans_plan=Count(
            "id",
            filter=Q(
                type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL,
                plan_couverture__isnull=True,
            ),
        ),
    )
    patients_principaux = repartition_patients["principaux"]
    ayants_droit = repartition_patients["total"] - patients_principaux

    # Files d'attente du bandeau "A traiter"
    compteurs = compteurs_files_attente(request)
    total_prises_en_charge_attente = compteurs["prises_en_charge_attente"]
    paiements_non_regles_nb = compteurs["paiements_non_regles"]
    rdv_a_confirmer = RendezVous.objects.filter(statut=RendezVous.Statut.DEMANDE).count()
    ordonnances_non_delivrees = Ordonnance.objects.filter(delivrance__isnull=True).count()

    # Libelles au pluriel
    pluriels_prestataire = {
        Prestataire.Type.HOPITAL: "hôpitaux",
        Prestataire.Type.CLINIQUE: "cliniques",
        Prestataire.Type.PHARMACIE: "pharmacies",
        Prestataire.Type.CABINET: "cabinets",
    }
    nb_comptes_bloques = len(TentativeConnexion.comptes_bloques())

    prestataires_par_type = [
        {
            "libelle": pluriels_prestataire.get(ligne["type_prestataire"], "autres"),
            "total": ligne["total"],
        }
        for ligne in Prestataire.objects.values("type_prestataire")
        .annotate(total=Count("id"))
        .order_by("-total")
    ]

    nb_pec_urgentes_48h = compteurs.get("nb_pec_urgentes_48h", 0)

    contexte = {
        "total_patients": Patient.objects.count(),
        "total_medecins": Medecin.objects.count(),
        "total_pharmaciens": Pharmacien.objects.count(),
        "total_prestataires": Prestataire.objects.filter(partenaire=True).count(),
        "total_prises_en_charge_attente": total_prises_en_charge_attente,
        "jours_attente_max": jours_attente_max,
        "nb_pec_urgentes_48h": nb_pec_urgentes_48h,
        "total_consultations": Consultation.objects.count(),
        "total_ordonnances": Ordonnance.objects.count(),
        "montant_regle": montant_regle,
        "montant_regle_7j": montant_regle_7j,
        "montant_non_regle": montant_non_regle,
        "taux_reglement": taux_reglement,
        "taux_recouvrement": round((montant_regle / (montant_regle + montant_non_regle)) * 100, 1) if (montant_regle + montant_non_regle) > 0 else 100.0,
        "tendance_paiements": tendance_paiements,
        "consultations_7j": consultations_7j,
        "ordonnances_7j": ordonnances_7j,
        "total_comptes_actifs": User.objects.filter(is_active=True).count(),
        "total_comptes_inactifs": User.objects.filter(is_active=False).count(),
        "total_rendez_vous_aujourd_hui": total_rendez_vous_aujourd_hui,
        "total_consultations_aujourd_hui": total_consultations_aujourd_hui,
        "dernieres_prises_en_charge": dernieres_prises_en_charge,
        "derniers_comptes": derniers_comptes,
        "patients_principaux": patients_principaux,
        "ayants_droit": ayants_droit,
        "assures_sans_plan": repartition_patients["sans_plan"],
        "rdv_a_confirmer": rdv_a_confirmer,
        "ordonnances_non_delivrees": ordonnances_non_delivrees,
        "total_delivrances": Delivrance.objects.count(),
        "paiements_non_regles_nb": paiements_non_regles_nb,
        "montant_total_facture": montant_total_paiements,
        "medecins_sans_prestataire": Medecin.objects.filter(prestataire__isnull=True).count(),
        "pharmaciens_sans_prestataire": Pharmacien.objects.filter(prestataire__isnull=True).count(),
        "prestataires_sans_coordonnees": Prestataire.objects.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        ).count(),
        "prestataires_par_type": prestataires_par_type,
        "nb_comptes_bloques": nb_comptes_bloques,
        "file_totale": (
            total_prises_en_charge_attente
            + rdv_a_confirmer
            + ordonnances_non_delivrees
            + paiements_non_regles_nb
        ),
    }
    return render(request, "dashboard.html", contexte)


def _consultations_par_mois(nombre_mois=6):
    """Nombre de consultations par mois, sur les `nombre_mois` derniers mois (mois courant inclus)."""
    annee, mois = timezone.now().year, timezone.now().month
    mois_reference = []
    for _ in range(nombre_mois):
        mois_reference.append((annee, mois))
        mois -= 1
        if mois == 0:
            mois, annee = 12, annee - 1
    mois_reference.reverse()

    comptages = (
        Consultation.objects.annotate(mois=TruncMonth("date_consultation"))
        .values("mois")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {(c["mois"].year, c["mois"].month): c["total"] for c in comptages if c["mois"]}

    return {
        "labels": [f"{MOIS_ABREGES[m - 1]} {a}" for a, m in mois_reference],
        "totaux": [totaux_par_cle.get(cle, 0) for cle in mois_reference],
    }


def _consultations_par_jour(nombre_jours=30, queryset=None):
    """Nombre de consultations par jour, sur les `nombre_jours` derniers jours (jour courant inclus).

    `queryset` permet de restreindre le comptage (ex. aux consultations d'un
    seul medecin) ; par defaut, porte sur toutes les consultations.
    """
    aujourd_hui = timezone.now().date()
    jours_reference = [aujourd_hui - datetime.timedelta(days=delta) for delta in range(nombre_jours - 1, -1, -1)]

    base = Consultation.objects.all() if queryset is None else queryset
    comptages = (
        base.annotate(jour=TruncDate("date_consultation"))
        .values("jour")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {c["jour"]: c["total"] for c in comptages if c["jour"]}

    return {
        "labels": [jour.strftime("%d/%m") for jour in jours_reference],
        "totaux": [totaux_par_cle.get(jour, 0) for jour in jours_reference],
    }


def _montants_regles_par_jour(nombre_jours=30):
    """Montant total regle (Paiement.montant_part_patient) par jour, sur les
    `nombre_jours` derniers jours (jour courant inclus) -- meme forme que
    _consultations_par_jour, sur Paiement.date_reglement plutot que
    Consultation.date_consultation."""
    aujourd_hui = timezone.now().date()
    jours_reference = [aujourd_hui - datetime.timedelta(days=delta) for delta in range(nombre_jours - 1, -1, -1)]

    montants = (
        Paiement.objects.filter(statut=Paiement.Statut.REGLE, date_reglement__date__gte=jours_reference[0])
        .annotate(jour=TruncDate("date_reglement"))
        .values("jour")
        .annotate(total=Sum("montant_part_patient"))
    )
    totaux_par_cle = {m["jour"]: m["total"] for m in montants if m["jour"]}

    return {
        "labels": [jour.strftime("%d/%m") for jour in jours_reference],
        "totaux": [float(totaux_par_cle.get(jour, 0)) for jour in jours_reference],
    }


def _consultations_par_annee(nombre_annees=5):
    """Nombre de consultations par annee, sur les `nombre_annees` dernieres annees (annee courante incluse)."""
    annee_courante = timezone.now().year
    annees_reference = list(range(annee_courante - nombre_annees + 1, annee_courante + 1))

    comptages = (
        Consultation.objects.annotate(annee=TruncYear("date_consultation"))
        .values("annee")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {c["annee"].year: c["total"] for c in comptages if c["annee"]}

    return {
        "labels": [str(annee) for annee in annees_reference],
        "totaux": [totaux_par_cle.get(annee, 0) for annee in annees_reference],
    }


def _donnees_rapports():
    """Comptages et agregats de synthese de l'activite de la plateforme (Phase 13)."""
    return {
        "utilisateurs_par_role": [
            {"label": label, "total": User.objects.filter(role=value).count()}
            for value, label in User.Role.choices
        ],
        "patients_par_type": [
            {"label": label, "total": Patient.objects.filter(type_beneficiaire=value).count()}
            for value, label in Patient.TypeBeneficiaire.choices
        ],
        "rendez_vous_par_statut": [
            {"label": label, "total": RendezVous.objects.filter(statut=value).count()}
            for value, label in RendezVous.Statut.choices
        ],
        "prises_en_charge_par_statut": [
            {"label": label, "total": PriseEnCharge.objects.filter(statut=value).count()}
            for value, label in PriseEnCharge.STATUT_CHOICES
        ],
        "prises_en_charge_chart": {
            "labels": [label for value, label in PriseEnCharge.STATUT_CHOICES],
            "totaux": [PriseEnCharge.objects.filter(statut=value).count() for value, label in PriseEnCharge.STATUT_CHOICES],
        },
        "total_consultations": Consultation.objects.count(),
        "total_ordonnances": Ordonnance.objects.count(),
        "total_delivrances": Delivrance.objects.count(),
        "total_prestataires_partenaires": Prestataire.objects.filter(partenaire=True).count(),
        "consultations_par_jour": _consultations_par_jour(),
        "consultations_par_mois": _consultations_par_mois(),
        "consultations_par_annee": _consultations_par_annee(),
    }


@admin_required
def rapports(request):
    """Synthese de l'activite de la plateforme : comptages et graphiques (Phase 13)."""
    return render(request, "rapports.html", _donnees_rapports())


@admin_required
def exporter_rapports_excel(request):
    """
    Exporte un classeur Excel multi-feuilles professionnel avec styles SantéSN,
    formats numériques et cohérence 1:1 avec le rapport PDF.
    """
    donnees = _donnees_rapports()
    classeur = openpyxl.Workbook()
    classeur.remove(classeur.active)

    style_entete = PatternFill(start_color="0E7C86", end_color="0E7C86", fill_type="solid")
    font_entete = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_titre = Font(name="Calibri", size=13, bold=True, color="0B2027")
    font_sous_titre = Font(name="Calibri", size=9.5, italic=True, color="5F6F7D")
    font_section = Font(name="Calibri", size=11, bold=True, color="0B2027")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_fine = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    maintenant = timezone.now().strftime("%d/%m/%Y à %H:%M")
    auteur = request.user.get_full_name() or request.user.email

    # 1. FEUILLE SYNTHESE
    ws_synthese = classeur.create_sheet(title="SYNTHESE")
    ws_synthese.views.sheetView[0].showGridLines = True
    ws_synthese["A1"] = "SantéSN — Synthèse Administrative & Métriques Clés"
    ws_synthese["A1"].font = font_titre
    ws_synthese["A2"] = f"Document généré le {maintenant} par {auteur} · Données consolidées"
    ws_synthese["A2"].font = font_sous_titre

    ws_synthese["A4"] = "Indicateur Clé"
    ws_synthese["B4"] = "Valeur Consolidée"
    for col in ["A4", "B4"]:
        ws_synthese[col].fill = style_entete
        ws_synthese[col].font = font_entete
        ws_synthese[col].alignment = align_center

    kpi_lignes = [
        ("Consultations médicales réalisées", donnees["total_consultations"]),
        ("Ordonnances prescrites", donnees["total_ordonnances"]),
        ("Délivrances pharmaceutiques effectuées", donnees["total_delivrances"]),
        ("Prestataires partenaires conventionnés", donnees["total_prestataires_partenaires"]),
        ("Nombre total d'utilisateurs", sum(r["total"] for r in donnees["utilisateurs_par_role"])),
        ("Nombre total d'assurés enregistrés", sum(p["total"] for p in donnees["patients_par_type"])),
        ("Rendez-vous programmés / traités", sum(rv["total"] for rv in donnees["rendez_vous_par_statut"])),
        ("Demandes de prise en charge", sum(pec["total"] for pec in donnees["prises_en_charge_par_statut"])),
    ]

    for row_idx, (libelle, val) in enumerate(kpi_lignes, start=5):
        c_a = ws_synthese.cell(row=row_idx, column=1, value=libelle)
        c_b = ws_synthese.cell(row=row_idx, column=2, value=val)
        c_a.border = border_fine
        c_b.border = border_fine
        c_b.alignment = align_right
        c_b.number_format = '#,##0'

    # 2. FEUILLE CONSULTATIONS
    ws_consultations = classeur.create_sheet(title="CONSULTATIONS")
    ws_consultations.views.sheetView[0].showGridLines = True
    ws_consultations.freeze_panes = "A2"
    ws_consultations.append(["Mois / Période", "Nombre de consultations"])
    for cell in ws_consultations[1]:
        cell.fill = style_entete
        cell.font = font_entete
        cell.alignment = align_center

    total_cons = 0
    for label, total in zip(donnees["consultations_par_mois"]["labels"], donnees["consultations_par_mois"]["totaux"]):
        ws_consultations.append([label, total])
        total_cons += total

    for row in ws_consultations.iter_rows(min_row=2, max_row=ws_consultations.max_row, min_col=1, max_col=2):
        row[0].border = border_fine
        row[1].border = border_fine
        row[1].alignment = align_right
        row[1].number_format = '#,##0'

    row_tot = ws_consultations.max_row + 1
    c_tot_lbl = ws_consultations.cell(row=row_tot, column=1, value="Total (12 derniers mois)")
    c_tot_val = ws_consultations.cell(row=row_tot, column=2, value=total_cons)
    c_tot_lbl.font = Font(name="Calibri", bold=True)
    c_tot_val.font = Font(name="Calibri", bold=True)
    c_tot_lbl.border = border_fine
    c_tot_val.border = border_fine
    c_tot_val.alignment = align_right
    c_tot_val.number_format = '#,##0'

    # 3. FEUILLE PRISES_EN_CHARGE
    ws_pec = classeur.create_sheet(title="PRISES_EN_CHARGE")
    ws_pec.views.sheetView[0].showGridLines = True
    ws_pec.freeze_panes = "A2"
    ws_pec.append(["Statut de prise en charge", "Nombre de dossiers", "Pourcentage"])
    for cell in ws_pec[1]:
        cell.fill = style_entete
        cell.font = font_entete
        cell.alignment = align_center

    total_dossiers_pec = sum(ligne["total"] for ligne in donnees["prises_en_charge_par_statut"])
    for ligne in donnees["prises_en_charge_par_statut"]:
        pct = (ligne["total"] / total_dossiers_pec) if total_dossiers_pec > 0 else 0.0
        ws_pec.append([ligne["label"], ligne["total"], pct])

    for row in ws_pec.iter_rows(min_row=2, max_row=ws_pec.max_row, min_col=1, max_col=3):
        row[0].border = border_fine
        row[1].border = border_fine
        row[1].alignment = align_right
        row[1].number_format = '#,##0'
        row[2].border = border_fine
        row[2].alignment = align_right
        row[2].number_format = '0.0%'

    row_tot_pec = ws_pec.max_row + 1
    c_lbl_pec = ws_pec.cell(row=row_tot_pec, column=1, value="Total")
    c_lbl_pec.font = Font(name="Calibri", bold=True)
    c_lbl_pec.border = border_fine
    c_tot_pec = ws_pec.cell(row=row_tot_pec, column=2, value=total_dossiers_pec)
    c_tot_pec.font = Font(name="Calibri", bold=True)
    c_tot_pec.border = border_fine
    c_tot_pec.alignment = align_right
    c_tot_pec.number_format = '#,##0'
    c_pct_tot = ws_pec.cell(row=row_tot_pec, column=3, value=1.0 if total_dossiers_pec > 0 else 0.0)
    c_pct_tot.font = Font(name="Calibri", bold=True)
    c_pct_tot.border = border_fine
    c_pct_tot.alignment = align_right
    c_pct_tot.number_format = '0.0%'

    # 4. FEUILLE ORDONNANCES_ET_DELIVRANCES
    ws_ord = classeur.create_sheet(title="ORDONNANCES_DELIVRANCES")
    ws_ord.views.sheetView[0].showGridLines = True
    ws_ord.freeze_panes = "A2"
    ws_ord.append(["Flux Pharmaceutique", "Volume", "Taux de Délivrance"])
    for cell in ws_ord[1]:
        cell.fill = style_entete
        cell.font = font_entete
        cell.alignment = align_center

    tot_ord = donnees["total_ordonnances"]
    tot_del = donnees["total_delivrances"]
    taux_del = (tot_del / tot_ord) if tot_ord > 0 else 0.0
    ws_ord.append(["Ordonnances prescrites", tot_ord, "-"])
    ws_ord.append(["Délivrances réalisées en officine", tot_del, taux_del])
    ws_ord.cell(row=2, column=2).number_format = '#,##0'
    ws_ord.cell(row=3, column=2).number_format = '#,##0'
    ws_ord.cell(row=3, column=3).number_format = '0.0%'
    for r in [2, 3]:
        for c in [1, 2, 3]:
            ws_ord.cell(row=r, column=c).border = border_fine

    # 5. FEUILLE UTILISATEURS_ET_ASSURES
    ws_users = classeur.create_sheet(title="UTILISATEURS_ASSURES")
    ws_users.views.sheetView[0].showGridLines = True
    ws_users.freeze_panes = "A2"
    ws_users.append(["Catégorie / Segment", "Total Enregistré"])
    for cell in ws_users[1]:
        cell.fill = style_entete
        cell.font = font_entete
        cell.alignment = align_center

    ws_users.append(["--- UTILISATEURS PAR RÔLE ---", ""])
    ws_users.cell(row=2, column=1).font = font_section
    for ligne in donnees["utilisateurs_par_role"]:
        ws_users.append([f"Rôle : {ligne['label']}", ligne["total"]])

    ws_users.append(["--- ASSURÉS PAR BÉNÉFICIAIRE ---", ""])
    ws_users.cell(row=ws_users.max_row, column=1).font = font_section
    for ligne in donnees["patients_par_type"]:
        ws_users.append([f"Assuré : {ligne['label']}", ligne["total"]])

    for row in ws_users.iter_rows(min_row=2, max_row=ws_users.max_row, min_col=1, max_col=2):
        row[0].border = border_fine
        row[1].border = border_fine
        if isinstance(row[1].value, (int, float)):
            row[1].alignment = align_right
            row[1].number_format = '#,##0'

    # Auto-ajustement des largeurs de colonne sur toutes les feuilles
    for ws in classeur.worksheets:
        for col in ws.columns:
            longueur_max = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(longueur_max + 4, 16)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="rapports_santesn.xlsx"'
    classeur.save(reponse)
    return reponse


def _creer_graphique_consultations_pdf(labels, totaux):
    """Génère un graphique à barres vectoriel ReportLab pour les consultations."""
    d = Drawing(480, 135)
    if not totaux or sum(totaux) == 0:
        d.add(String(240, 60, "Aucune consultation sur la période", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748b")))
        return d

    chart = VerticalBarChart()
    chart.x = 35
    chart.y = 20
    chart.height = 95
    chart.width = 430
    chart.data = [totaux]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 7
    chart.categoryAxis.labels.dy = -10
    chart.valueAxis.valueMin = 0
    max_val = max(totaux)
    chart.valueAxis.valueMax = max_val + (1 if max_val < 5 else int(max_val * 0.2))
    chart.valueAxis.labels.fontSize = 8
    chart.bars[0].fillColor = colors.HexColor("#0e7c86")
    chart.bars[0].strokeColor = colors.HexColor("#095059")
    chart.bars[0].strokeWidth = 0.5
    d.add(chart)
    return d


def _creer_graphique_pec_pdf(statuts):
    """Génère un graphique camembert vectoriel ReportLab pour la répartition des prises en charge."""
    totaux = [s["total"] for s in statuts]
    labels = [s["label"] for s in statuts]
    d = Drawing(480, 125)
    if not totaux or sum(totaux) == 0:
        d.add(String(240, 60, "Aucune demande de prise en charge enregistrée", textAnchor="middle", fontSize=9, fillColor=colors.HexColor("#64748b")))
        return d

    pie = Pie()
    pie.x = 60
    pie.y = 10
    pie.width = 105
    pie.height = 105
    pie.data = totaux
    pie.labels = [str(t) for t in totaux]
    pie.simpleLabels = 0
    palette = [colors.HexColor("#d97706"), colors.HexColor("#0e7c86"), colors.HexColor("#e11d48"), colors.HexColor("#64748b")]
    for idx, color in enumerate(palette[:len(totaux)]):
        pie.slices[idx].fillColor = color
        pie.slices[idx].strokeColor = colors.white
        pie.slices[idx].strokeWidth = 1

    legend = Legend()
    legend.x = 220
    legend.y = 85
    legend.dx = 10
    legend.dy = 10
    legend.fontName = "Helvetica"
    legend.fontSize = 8
    legend.boxAnchor = "nw"
    legend.columnMaximum = 5
    legend.colorNamePairs = [(palette[i % len(palette)], f"{labels[i]} : {totaux[i]}") for i in range(len(labels))]

    d.add(pie)
    d.add(legend)
    return d


class NumberedCanvas:
    """Générateur de Canvas ReportLab avec numérotation 'Page X sur Y' et en-tête officiel."""

    def __init__(self, *args, **kwargs):
        from reportlab.pdfgen import canvas
        self._canvas_class = canvas.Canvas
        self._underlying = canvas.Canvas(*args, **kwargs)
        self._saved_page_states = []

    def __getattr__(self, name):
        return getattr(self._underlying, name)

    def showPage(self):
        self._saved_page_states.append(dict(self._underlying.__dict__))
        self._underlying._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self._underlying.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            self._underlying.showPage()
        self._underlying.save()

    def draw_page_decorations(self, page_count):
        self._underlying.saveState()

        # En-tête institutionnel
        self._underlying.setFont("Helvetica-Bold", 9)
        self._underlying.setFillColor(colors.HexColor("#0e7c86"))
        self._underlying.drawString(54, 804, "SANTÉSN")
        self._underlying.setFont("Helvetica", 8)
        self._underlying.setFillColor(colors.HexColor("#5f6f7d"))
        self._underlying.drawString(108, 804, "— Plateforme Médicale & Système National de Tiers-Payant")

        self._underlying.setStrokeColor(colors.HexColor("#cbd5e1"))
        self._underlying.setLineWidth(0.5)
        self._underlying.line(54, 796, 541, 796)

        # Pied de page institutionnel
        self._underlying.line(54, 45, 541, 45)
        self._underlying.drawString(54, 32, "SantéSN — Document officiel de synthèse administrative")
        texte_page = f"Page {self._underlying._pageNumber} sur {page_count}"
        self._underlying.drawRightString(541, 32, texte_page)

        self._underlying.restoreState()


@admin_required
def exporter_rapports_pdf(request):
    """
    Génère un rapport administratif PDF professionnel et institutionnel :
    En-tête SantéSN, métadonnées, 4 KPI, graphiques vectoriels ReportLab, tableaux consolidés,
    et pagination officielle "Page X sur Y".
    """
    donnees = _donnees_rapports()
    reponse = HttpResponse(content_type="application/pdf")
    reponse["Content-Disposition"] = 'attachment; filename="rapports_santesn.pdf"'

    document = SimpleDocTemplate(
        reponse,
        pagesize=A4,
        title="Rapports SantéSN",
        leftMargin=54,
        rightMargin=54,
        topMargin=58,
        bottomMargin=58,
    )
    styles = getSampleStyleSheet()

    style_titre = ParagraphStyle(
        "DocTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0b2027"),
        alignment=0,
    )
    style_sous_titre = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#0e7c86"),
        spaceAfter=4,
    )
    style_meta = ParagraphStyle(
        "DocMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor("#5f6f7d"),
    )
    style_h2 = ParagraphStyle(
        "SectionH2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#0b2027"),
        spaceBefore=8,
        spaceAfter=3,
    )

    maintenant = timezone.now().strftime("%d/%m/%Y à %H:%M")
    auteur = request.user.get_full_name() or request.user.email

    elements = [
        Paragraph("RAPPORT D'ACTIVITÉ & SYNTHÈSE ADMINISTRATIVE", style_titre),
        Paragraph("SantéSN — Système National de Tiers-Payant et de Gestion Médicale", style_sous_titre),
        Paragraph(f"Période : 12 derniers mois consolidés · Généré le {maintenant} · Par : <b>{auteur}</b>", style_meta),
        Spacer(1, 10),
    ]

    # Bloc KPI Clés (4 encadrés)
    kpi_data = [
        [
            Paragraph(f"<font size=7 color='#5f6f7d'>CONSULTATIONS</font><br/><font size=13 color='#095059'><b>{donnees['total_consultations']}</b></font><br/><font size=6.5 color='#64748b'>Actes réalisés</font>", styles["Normal"]),
            Paragraph(f"<font size=7 color='#5f6f7d'>ORDONNANCES</font><br/><font size=13 color='#095059'><b>{donnees['total_ordonnances']}</b></font><br/><font size=6.5 color='#64748b'>Prescriptions</font>", styles["Normal"]),
            Paragraph(f"<font size=7 color='#5f6f7d'>DÉLIVRANCES</font><br/><font size=13 color='#095059'><b>{donnees['total_delivrances']}</b></font><br/><font size=6.5 color='#64748b'>En pharmacie</font>", styles["Normal"]),
            Paragraph(f"<font size=7 color='#5f6f7d'>PRESTATAIRES</font><br/><font size=13 color='#095059'><b>{donnees['total_prestataires_partenaires']}</b></font><br/><font size=6.5 color='#64748b'>Partenaires</font>", styles["Normal"]),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[121, 121, 121, 121])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff5f5")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 12))

    # Graphique 1 : Consultations
    elements.append(Paragraph("1. Évolution des consultations médicales", style_h2))
    graphe_cons = _creer_graphique_consultations_pdf(
        donnees["consultations_par_mois"]["labels"],
        donnees["consultations_par_mois"]["totaux"],
    )
    elements.append(graphe_cons)
    elements.append(Spacer(1, 10))

    # Graphique 2 : Prises en charge
    elements.append(Paragraph("2. Répartition des demandes de prise en charge", style_h2))
    graphe_pec = _creer_graphique_pec_pdf(donnees["prises_en_charge_par_statut"])
    elements.append(graphe_pec)
    elements.append(Spacer(1, 10))

    # Tableaux analytiques détaillés
    def ajouter_tableau(titre, entetes, lignes, largeurs=[242, 242]):
        elements.append(Paragraph(titre, style_h2))
        tableau = Table([entetes] + lignes, hAlign="LEFT", colWidths=largeurs)
        tableau.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7c86")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(tableau)
        elements.append(Spacer(1, 10))

    ajouter_tableau(
        "3. Utilisateurs par rôle",
        ["Rôle", "Nombre de comptes"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["utilisateurs_par_role"]],
    )
    ajouter_tableau(
        "4. Assurés par type de bénéficiaire",
        ["Type de bénéficiaire", "Total enregistrés"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["patients_par_type"]],
    )
    ajouter_tableau(
        "5. Rendez-vous par statut",
        ["Statut", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["rendez_vous_par_statut"]],
    )
    ajouter_tableau(
        "6. Prises en charge par statut",
        ["Statut", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["prises_en_charge_par_statut"]],
    )
    ajouter_tableau(
        "7. Consultations par mois (Données chiffrées)",
        ["Mois", "Consultations"],
        [
            [label, str(total)]
            for label, total in zip(
                donnees["consultations_par_mois"]["labels"], donnees["consultations_par_mois"]["totaux"]
            )
        ],
    )

    document.build(elements, canvasmaker=NumberedCanvas)
    return reponse
