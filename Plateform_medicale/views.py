import csv
import datetime
import json
import unicodedata
from decimal import Decimal, InvalidOperation
import urllib.error
import urllib.parse
import urllib.request
from functools import wraps

import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.sessions.models import Session
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Case, Count, IntegerField, Q, Sum, Value, When
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
from django.http import Http404, HttpResponse, JsonResponse
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from .forms import (
    AyantDroitForm,
    ConsultationForm,
    EnvoyerNotificationForm,
    LoginForm,
    MedecinForm,
    MedecinProfilForm,
    MonCompteForm,
    OrdonnanceForm,
    PaiementReglementForm,
    PatientCreationForm,
    PatientForm,
    PharmacienAffectationForm,
    PlanCouvertureForm,
    PrestataireForm,
    PriseEnChargeForm,
    ProfilAssureForm,
    RendezVousAssureForm,
    RendezVousForm,
    ServiceMedicalForm,
    SetupWizardForm,
    UtilisateurCreationForm,
    UtilisateurModificationForm,
    generer_mot_de_passe,
    lier_fiche_medecin,
    lier_fiche_pharmacien,
)
from .models import (
    Consultation,
    Delivrance,
    Medecin,
    Notification,
    Ordonnance,
    Paiement,
    Patient,
    Pharmacien,
    PlanCouverture,
    Prestataire,
    PriseEnCharge,
    RendezVous,
    ServiceMedical,
    TentativeConnexion,
    User,
    distance_km,
    valider_telephone,
)


# ---------------------------------------------------------------------------
# Pagination des listes admin
# ---------------------------------------------------------------------------

TAILLE_PAGE_LISTE = 20

# Recherche de repli du pharmacien (scanner_ordonnance) : longueur minimale
# pour ne pas enumerer les patients, et plafond d'affichage au comptoir.
RECHERCHE_ORDONNANCE_MIN = 3
RECHERCHE_ORDONNANCE_MAX = 20


def _paginer(request, queryset):
    """Pagine un queryset pour une liste admin (parametre GET 'page', taille
    fixe). get_page() plutot que page() : replie silencieusement sur la
    derniere/premiere page si le numero demande est hors limites, plutot que
    de lever une exception pour un lien de pagination perime ou trafique."""
    paginateur = Paginator(queryset, TAILLE_PAGE_LISTE)
    return paginateur.get_page(request.GET.get("page"))


def _trier(request, queryset, champs_autorises, defaut):
    """Trie un queryset de liste admin depuis le parametre GET 'tri' (ex.
    'nom' ou '-nom'), restreint a `champs_autorises` (sans le signe -) pour
    ne jamais passer un champ arbitraire a order_by(). Retombe sur `defaut`
    (nom de champ ou tuple/liste de noms) si absent ou hors liste."""
    tri = request.GET.get("tri", "")
    if tri.lstrip("-") in champs_autorises:
        return queryset.order_by(tri)
    if isinstance(defaut, (list, tuple)):
        return queryset.order_by(*defaut)
    return queryset.order_by(defaut)


# ---------------------------------------------------------------------------
# Permissions par rôle
# ---------------------------------------------------------------------------

def role_required(*roles):
    """
    Restreint une vue aux rôles indiqués.

    Exemple :
        @role_required(User.Role.ADMIN, User.Role.MEDECIN)
        def ma_vue(request): ...
    """

    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped(request, *args, **kwargs):
            if request.user.role not in roles:
                raise PermissionDenied
            return view_func(request, *args, **kwargs)

        return _wrapped

    return decorator


def admin_required(view_func):
    """Restreint une vue au rôle ADMIN uniquement."""
    return role_required(User.Role.ADMIN)(view_func)


def compteurs_files_attente(request=None):
    """Compteurs des files d'attente administrateur.

    Source UNIQUE : les pastilles du menu lateral (via le context processor
    user_role) et le bandeau "A traiter" du dashboard affichent les memes
    nombres. Sans ce cache, la vue et le context processor lancaient chacun
    les memes deux requetes au meme rendu -- et surtout, deux noms coexistaient
    pour un meme chiffre, donc deux endroits a corriger le jour ou la regle
    metier change.

    Le cache est pose sur la requete : le context processor s'execute au rendu
    du gabarit, donc apres la vue, et reutilise ce qu'elle a deja calcule.
    """
    if request is not None and hasattr(request, "_compteurs_files_attente"):
        return request._compteurs_files_attente

    compteurs = {
        "prises_en_charge_attente": PriseEnCharge.objects.filter(statut="en_attente").count(),
        "paiements_non_regles": Paiement.objects.filter(statut=Paiement.Statut.NON_REGLE).count(),
    }
    if request is not None:
        request._compteurs_files_attente = compteurs
    return compteurs


def user_role(request):
    """Context processor : role, notifications non lues, et compteurs de file
    d'attente pour les pastilles du menu lateral administrateur.

    Les deux compteurs admin portent sur des champs indexes (db_index sur
    PriseEnCharge.statut et Paiement.statut) et ne sont calcules que pour le
    role ADMIN : les autres roles n'ont pas ces ecrans. Un visiteur anonyme ne
    declenche aucune requete (landing, connexion).
    """
    user = getattr(request, 'user', None)
    if user is None or not user.is_authenticated:
        return {'current_role': None, 'current_role_label': None, 'notifications_non_lues': 0}

    contexte = {
        'current_role': user.role,
        'current_role_label': user.get_role_display(),
        'notifications_non_lues': user.notifications.filter(lue=False).count(),
    }
    if user.role == User.Role.ADMIN:
        compteurs = compteurs_files_attente(request)
        contexte['nb_prises_en_charge_attente'] = compteurs["prises_en_charge_attente"]
        contexte['nb_paiements_non_regles'] = compteurs["paiements_non_regles"]
    return contexte


# ---------------------------------------------------------------------------
# Authentification et tableaux de bord par rôle
# ---------------------------------------------------------------------------

def _admin_exists():
    return User.objects.filter(role=User.Role.ADMIN).exists()


def login_view(request):
    """Connexion par email et mot de passe. Le role est detecte en base."""
    if not _admin_exists():
        return redirect('setup_wizard')

    # 'next' est pose par @login_required quand une page protegee redirige
    # ici (session expiree, ou lien direct sans etre connecte). Valide avant
    # de s'en servir comme cible (empeche un lien ?next=https://... trafique
    # de rediriger hors du site) ; sa seule presence sert aussi a expliquer
    # explicitement la redirection sur l'ecran de connexion plutot que de
    # rester muet (le formulaire n'a pas d'action explicite, la query string
    # de la page suit donc telle quelle jusqu'au POST).
    next_url = request.GET.get('next', '')
    next_valide = bool(next_url) and url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    )
    destination = next_url if next_valide else 'post_login_redirect'

    if request.user.is_authenticated:
        return redirect(destination)

    form = LoginForm(request=request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.user)
        return redirect(destination)

    return render(request, 'login.html', {'form': form, 'session_expiree': next_valide})


@require_POST
def logout_view(request):
    logout(request)
    messages.success(request, 'Vous avez été déconnecté.')
    return redirect('login')


@login_required
def post_login_redirect(request):
    """Redirection automatique vers le dashboard correspondant au role."""
    role = request.user.role
    if role == User.Role.ADMIN:
        return redirect('dashboard')
    if role == User.Role.ASSURE:
        return redirect('dashboard_assure')
    if role == User.Role.MEDECIN:
        return redirect('dashboard_medecin')
    if role == User.Role.PHARMACIEN:
        return redirect('dashboard_pharmacien')

    logout(request)
    messages.error(request, "Rôle inconnu. Contactez l'administration.")
    return redirect('login')


def setup_wizard(request):
    """
    Assistant de premiere installation.

    Accessible uniquement si aucun administrateur n'existe. Une fois le premier
    administrateur cree, l'assistant redirige toujours vers la connexion.
    """
    if _admin_exists():
        return redirect('login')

    form = SetupWizardForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        login(request, user)
        messages.success(
            request,
            'Bienvenue ! Votre compte Super Administrateur a été créé.',
        )
        return redirect('post_login_redirect')

    return render(request, 'setup_wizard.html', {'form': form})


# Sections de la page Parametres : (slug, libelle, icone, role requis).
# Une seule table pilote a la fois le menu de gauche ET le controle d'acces,
# pour qu'ils ne puissent pas diverger.
SECTIONS_PARAMETRES = [
    ("general", "Général", "user-circle", None),
    ("apparence", "Apparence", "eye", None),
    ("securite", "Sécurité", "lock", None),
    ("donnees", "Données", "download", "ADMIN"),
    ("avance", "Avancé", "filter", None),
]


@login_required
def parametres(request, section="general"):
    """Page Parametres, decoupee en sections qui ont chacune leur URL.

    Un clic dans le menu de gauche ouvre reellement la page correspondante
    (et non un simple defilement) : l'adresse est partageable, le bouton
    Retour du navigateur fonctionne, et chaque ecran reste court.

    Regle de contenu inchangee : la page n'affiche QUE des reglages adosses a
    du code reel. La section "general" montre la configuration de la
    plateforme en LECTURE SEULE (langue, fuseau, format de date) : ce sont de
    vraies valeurs, lues dans settings.py, qui repondent a une question
    legitime -- les presenter comme modifiables serait un mensonge, les
    cacher priverait l'administrateur d'une information utile.
    """
    autorisees = [s for s in SECTIONS_PARAMETRES
                  if s[3] is None or request.user.role == s[3]]
    slugs = {s[0] for s in autorisees}
    if section not in slugs:
        raise Http404("Section de paramètres inconnue.")

    contexte = {
        "sections": [
            {"slug": slug, "libelle": libelle, "icone": icone}
            for slug, libelle, icone, _ in autorisees
        ],
        "section": section,
        "section_libelle": next(s[1] for s in autorisees if s[0] == section),
    }

    if section == "general":
        contexte.update({
            "langue_plateforme": "Français",
            "fuseau_horaire": settings.TIME_ZONE,
        })
    elif section == "securite":
        contexte["duree_session_heures"] = settings.SESSION_COOKIE_AGE // 3600
        if request.user.role == User.Role.ADMIN:
            comptes, role_choisi, recherche = _comptes_bloques(request)
            contexte.update({
                "comptes_bloques": comptes,
                "role_choisi": role_choisi,
                "recherche_bloques": recherche,
                "roles_disponibles": User.Role.choices,
                "minutes_blocage": int(TentativeConnexion.DUREE_BLOCAGE.total_seconds() // 60),
                "max_tentatives": TentativeConnexion.MAX_TENTATIVES,
            })

    return render(request, "parametres.html", contexte)


def _comptes_bloques(request):
    """Comptes reellement bloques, filtres par role et par recherche.

    Renvoie (liste, role_choisi, recherche). Chaque element est un dict pret
    a afficher : on ne passe au gabarit que ce qui est necessaire a l'ecran
    (identite, role, echecs, temps restant) -- rien d'authentification.
    """
    role_choisi = request.GET.get("role", "")
    recherche = request.GET.get("q", "").strip()
    maintenant = timezone.now()

    comptes = []
    for utilisateur, ligne in TentativeConnexion.comptes_bloques():
        if role_choisi and utilisateur.role != role_choisi:
            continue
        if recherche:
            cible = f"{utilisateur.first_name} {utilisateur.last_name} {utilisateur.email}".lower()
            if recherche.lower() not in cible:
                continue
        restant = ligne.secondes_restantes(maintenant)
        comptes.append({
            "utilisateur": utilisateur,
            "tentatives": ligne.tentatives,
            "minutes_restantes": max(1, -(-restant // 60)),  # arrondi au superieur
        })
    comptes.sort(key=lambda c: c["minutes_restantes"], reverse=True)
    return comptes, role_choisi, recherche


@admin_required
@require_POST
def debloquer_compte(request, pk):
    """Deblocage manuel par un administrateur.

    Supprime la ligne de comptage, ce qui est exactement ce que fait une
    connexion reussie : le compte peut se reconnecter immediatement. Le
    deblocage automatique par expiration continue de fonctionner en parallele
    -- il ne depend que de dernier_echec.

    Deblocage INDIVIDUEL uniquement : pas de "tout debloquer". Un blocage
    massif est souvent le signe d'une attaque en cours ; tout relacher d'un
    clic annulerait la protection au pire moment.
    """
    utilisateur = get_object_or_404(User, pk=pk)
    supprimees = TentativeConnexion.objects.filter(email=utilisateur.email.lower()).delete()[0]
    if supprimees:
        messages.success(
            request,
            f"Le compte de {utilisateur} peut de nouveau se connecter.",
            extra_tags="succes-critique",
        )
    else:
        # Expiration survenue entre l'affichage et le clic : ce n'est pas une
        # erreur, le resultat voulu est deja atteint.
        messages.info(request, f"Le compte de {utilisateur} n'était plus bloqué.")
    return redirect(f"{reverse('parametres_section', args=['securite'])}#comptes-bloques")


@login_required
@require_POST
def deconnecter_partout(request):
    """Ferme toutes les sessions de l'utilisateur, y compris la courante.

    Django n'offre pas de primitive pour cela sans changer le mot de passe :
    on parcourt les sessions NON EXPIREES et on supprime celles dont
    _auth_user_id correspond. Suppose le backend de sessions en base (celui par
    defaut ; ce projet n'en change pas). Le filtre sur expire_date evite de
    decoder des sessions deja mortes.

    Le message est pose APRES logout() : logout() vide la session courante, un
    message ajoute avant serait perdu.
    """
    identifiant = str(request.user.pk)
    fermees = 0
    for session in Session.objects.filter(expire_date__gte=timezone.now()):
        if session.get_decoded().get("_auth_user_id") == identifiant:
            session.delete()
            fermees += 1

    logout(request)
    messages.success(
        request,
        f"{fermees} session(s) fermée(s) sur l'ensemble de vos appareils. "
        "Reconnectez-vous pour continuer.",
    )
    return redirect("login")


@login_required
def mon_compte(request):
    """Modification par l'utilisateur de ses propres informations.

    Ne permet pas de changer son role (regle metier : le role est stocke en
    base, jamais choisi par l'utilisateur). Changer l'email exige le mot de
    passe actuel, cf. MonCompteForm.
    """
    form = MonCompteForm(request.POST or None, instance=request.user)
    if request.method == "POST" and form.is_valid():
        ancien_email = request.user.email
        form.save()
        if form.cleaned_data["email"].lower() != ancien_email.lower():
            messages.success(
                request,
                "Informations enregistrées. Votre adresse de connexion est "
                f"désormais {form.cleaned_data['email']}.",
            )
        else:
            messages.success(request, "Informations enregistrées.")
        return redirect("mon_compte")

    return render(request, "mon_compte.html", {"form": form})


@login_required
def changer_mot_de_passe(request):
    """
    Changement du mot de passe par l'utilisateur connecte (tous roles).

    Distinct de la reinitialisation par l'admin (Gestion des utilisateurs) :
    ici, l'utilisateur doit connaitre son mot de passe actuel.
    """
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, 'Mot de passe modifié avec succès.')
            return redirect('post_login_redirect')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'changer_mot_de_passe.html', {'form': form})


# ---------------------------------------------------------------------------
# Vitrine publique
# ---------------------------------------------------------------------------

def landing(request):
    """Page d'accueil publique de SantéSN (vitrine)."""
    return render(request, "landing.html")


def robots_txt(request):
    """Une seule page publique (landing) : tout le reste (espaces authentifies)
    n'a pas vocation a etre indexe."""
    contenu = (
        "User-agent: *\n"
        "Allow: /$\n"
        "Disallow: /\n\n"
        f"Sitemap: {request.build_absolute_uri('/sitemap.xml')}\n"
    )
    return HttpResponse(contenu, content_type="text/plain")


def sitemap_xml(request):
    contenu = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f"  <url><loc>{request.build_absolute_uri('/')}</loc>"
        "<changefreq>monthly</changefreq><priority>1.0</priority></url>\n"
        "</urlset>\n"
    )
    return HttpResponse(contenu, content_type="application/xml")


# ---------------------------------------------------------------------------
# Dashboard administrateur
# ---------------------------------------------------------------------------

@admin_required
def dashboard(request):
    # Les demandes en attente remontent en tete : triee par date seule, la
    # liste noyait l'urgence sous des dossiers deja clos.
    dernieres_prises_en_charge = (
        PriseEnCharge.objects.select_related("patient")
        .annotate(
            priorite=Case(
                When(statut="en_attente", then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
        .order_by("priorite", "-date_demande")[:5]
    )

    # Bandeau financier : meme agregat que liste_paiements (Sum filtre par
    # statut), pour un signal de sante financiere absent jusqu'ici du
    # dashboard alors que c'est la donnee la plus parlante pour une
    # compagnie d'assurance/IPM qui evaluerait la plateforme.
    totaux_paiements = Paiement.objects.aggregate(
        total_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.REGLE)),
        total_non_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.NON_REGLE)),
    )
    montant_regle = totaux_paiements["total_regle"] or 0
    montant_non_regle = totaux_paiements["total_non_regle"] or 0
    montant_total_paiements = montant_regle + montant_non_regle
    taux_reglement = round((montant_regle / montant_total_paiements) * 100) if montant_total_paiements else None

    # Activite du jour, plateforme entiere (pas un seul medecin) : pouls de
    # l'activite absent jusqu'ici, seuls des totaux globaux sans notion de
    # temps etaient affiches.
    maintenant = timezone.now()
    debut_jour = maintenant.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_jour = debut_jour + datetime.timedelta(days=1)
    total_rendez_vous_aujourd_hui = (
        RendezVous.objects.filter(date_heure__gte=debut_jour, date_heure__lt=fin_jour)
        .exclude(statut=RendezVous.Statut.ANNULE)
        .count()
    )
    total_consultations_aujourd_hui = Consultation.objects.filter(
        date_consultation__gte=debut_jour, date_consultation__lt=fin_jour
    ).count()

    # Tendance des paiements regles sur les 7 derniers jours, pour le delta
    # affiche sous le montant regle du hero (direction A+, retour utilisateur
    # 2026-08-05 : la version initiale de "poste de pilotage" n'affichait
    # qu'un instantane, jamais de notion de mouvement). Reutilise
    # Paiement.date_reglement, deja renseigne a chaque reglement -- aucune
    # migration necessaire.
    il_y_a_7_jours = maintenant - datetime.timedelta(days=7)
    montant_regle_7j = Paiement.objects.filter(
        statut=Paiement.Statut.REGLE, date_reglement__gte=il_y_a_7_jours
    ).aggregate(total=Sum("montant_part_patient"))["total"] or 0

    # Borne haute explicite, contrairement aux deux agregats voisins :
    # Consultation.date_consultation est saisie par le medecin (et
    # ConsultationForm ne refuse pas une date future), alors que
    # Paiement.date_reglement et Ordonnance.date_creation sont posees par le
    # code. Sans elle, une consultation datee par erreur dans le futur
    # gonflerait un delta libelle "7 derniers jours".
    consultations_7j = Consultation.objects.filter(
        date_consultation__gte=il_y_a_7_jours, date_consultation__lte=maintenant
    ).count()
    ordonnances_7j = Ordonnance.objects.filter(date_creation__gte=il_y_a_7_jours).count()

    # Anciennete de la file d'attente : le compte seul
    # (total_prises_en_charge_attente) ne dit pas si c'est urgent. Meme champ
    # (filtre sur statut, qui est indexe ; date_demande ne l'est pas) que le
    # tri de liste_prises_en_charge, aucune migration necessaire.
    plus_ancienne_attente = (
        PriseEnCharge.objects.filter(statut="en_attente").order_by("date_demande").first()
    )
    jours_attente_max = (
        (maintenant - plus_ancienne_attente.date_demande).days if plus_ancienne_attente else None
    )

    # Derniers comptes crees, hors assures : les assures (y compris les
    # ayants droit, qui n'ont jamais de User) sont deja couverts par
    # liste_patients (triable par N, cf. la Huitieme passe dans
    # FONCTIONNEMENT.txt) -- pas de raison de les dupliquer ici. Les
    # prestataires n'ont pas de compte utilisateur, liste separee.
    derniers_comptes = (
        User.objects.exclude(role=User.Role.ASSURE).order_by("-date_joined")[:5]
    )

    tendance_paiements = _montants_regles_par_jour()

    # Un seul aller-retour pour les trois comptages de Patient plutot que trois
    # count() sur la meme table. La repartition principaux / ayants droit est
    # le coeur du sujet du projet et n'etait affichee nulle part.
    repartition_patients = Patient.objects.aggregate(
        total=Count("id"),
        principaux=Count("id", filter=Q(type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL)),
        sans_plan=Count(
            "id",
            filter=Q(
                type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL,
                plan_couverture__isnull=True,
            ),
        ),
    )
    patients_principaux = repartition_patients["principaux"]
    ayants_droit = repartition_patients["total"] - patients_principaux

    # Les quatre files d'attente du bandeau "A traiter". Chacune mene a une
    # liste filtree reellement existante (cf. liste_rendez_vous et
    # liste_ordonnances, ajoutees pour cette refonte).
    # Meme source que les pastilles du menu lateral : compteurs_files_attente()
    # met son resultat en cache sur la requete, le context processor le relit.
    compteurs = compteurs_files_attente(request)
    total_prises_en_charge_attente = compteurs["prises_en_charge_attente"]
    paiements_non_regles_nb = compteurs["paiements_non_regles"]
    rdv_a_confirmer = RendezVous.objects.filter(statut=RendezVous.Statut.DEMANDE).count()
    ordonnances_non_delivrees = Ordonnance.objects.filter(delivrance__isnull=True).count()

    # Libelles au pluriel : .values() ne rend que la valeur brute de l'enum
    # ("HOPITAL"), et les pluriels francais concernes sont irreguliers.
    pluriels_prestataire = {
        Prestataire.Type.HOPITAL: "hôpitaux",
        Prestataire.Type.CLINIQUE: "cliniques",
        Prestataire.Type.PHARMACIE: "pharmacies",
        Prestataire.Type.CABINET: "cabinets",
    }
    # Alerte du tableau de bord : uniquement s'il y a matiere. Un compteur a
    # zero n'apprend rien et occupe une place visible pour rien.
    nb_comptes_bloques = len(TentativeConnexion.comptes_bloques())

    prestataires_par_type = [
        {
            "libelle": pluriels_prestataire.get(ligne["type_prestataire"], "autres"),
            "total": ligne["total"],
        }
        for ligne in Prestataire.objects.values("type_prestataire")
        .annotate(total=Count("id"))
        .order_by("-total")
    ]

    contexte = {
        "total_patients": Patient.objects.count(),
        "total_medecins": Medecin.objects.count(),
        "total_pharmaciens": Pharmacien.objects.count(),
        "total_prestataires": Prestataire.objects.filter(partenaire=True).count(),
        "total_prises_en_charge_attente": total_prises_en_charge_attente,
        "jours_attente_max": jours_attente_max,
        "total_consultations": Consultation.objects.count(),
        "total_ordonnances": Ordonnance.objects.count(),
        "montant_regle": montant_regle,
        "montant_regle_7j": montant_regle_7j,
        "montant_non_regle": montant_non_regle,
        "taux_reglement": taux_reglement,
        "tendance_paiements": tendance_paiements,
        "consultations_7j": consultations_7j,
        "ordonnances_7j": ordonnances_7j,
        "total_comptes_actifs": User.objects.filter(is_active=True).count(),
        "total_comptes_inactifs": User.objects.filter(is_active=False).count(),
        "total_rendez_vous_aujourd_hui": total_rendez_vous_aujourd_hui,
        "total_consultations_aujourd_hui": total_consultations_aujourd_hui,
        "dernieres_prises_en_charge": dernieres_prises_en_charge,
        "derniers_comptes": derniers_comptes,
        "patients_principaux": patients_principaux,
        "ayants_droit": ayants_droit,
        "assures_sans_plan": repartition_patients["sans_plan"],
        "rdv_a_confirmer": rdv_a_confirmer,
        "ordonnances_non_delivrees": ordonnances_non_delivrees,
        "total_delivrances": Delivrance.objects.count(),
        "paiements_non_regles_nb": paiements_non_regles_nb,
        "montant_total_facture": montant_total_paiements,
        "medecins_sans_prestataire": Medecin.objects.filter(prestataire__isnull=True).count(),
        "pharmaciens_sans_prestataire": Pharmacien.objects.filter(prestataire__isnull=True).count(),
        "prestataires_sans_coordonnees": Prestataire.objects.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        ).count(),
        "prestataires_par_type": prestataires_par_type,
        "nb_comptes_bloques": nb_comptes_bloques,
        "file_totale": (
            total_prises_en_charge_attente
            + rdv_a_confirmer
            + ordonnances_non_delivrees
            + paiements_non_regles_nb
        ),
    }
    return render(request, "dashboard.html", contexte)


MOIS_ABREGES = ["Jan", "Fev", "Mar", "Avr", "Mai", "Jun", "Jul", "Aou", "Sep", "Oct", "Nov", "Dec"]


def _consultations_par_mois(nombre_mois=6):
    """Nombre de consultations par mois, sur les `nombre_mois` derniers mois (mois courant inclus)."""
    annee, mois = timezone.now().year, timezone.now().month
    mois_reference = []
    for _ in range(nombre_mois):
        mois_reference.append((annee, mois))
        mois -= 1
        if mois == 0:
            mois, annee = 12, annee - 1
    mois_reference.reverse()

    comptages = (
        Consultation.objects.annotate(mois=TruncMonth("date_consultation"))
        .values("mois")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {(c["mois"].year, c["mois"].month): c["total"] for c in comptages if c["mois"]}

    return {
        "labels": [f"{MOIS_ABREGES[m - 1]} {a}" for a, m in mois_reference],
        "totaux": [totaux_par_cle.get(cle, 0) for cle in mois_reference],
    }


def _consultations_par_jour(nombre_jours=30, queryset=None):
    """Nombre de consultations par jour, sur les `nombre_jours` derniers jours (jour courant inclus).

    `queryset` permet de restreindre le comptage (ex. aux consultations d'un
    seul medecin) ; par defaut, porte sur toutes les consultations.
    """
    aujourd_hui = timezone.now().date()
    jours_reference = [aujourd_hui - datetime.timedelta(days=delta) for delta in range(nombre_jours - 1, -1, -1)]

    base = Consultation.objects.all() if queryset is None else queryset
    comptages = (
        base.annotate(jour=TruncDate("date_consultation"))
        .values("jour")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {c["jour"]: c["total"] for c in comptages if c["jour"]}

    return {
        "labels": [jour.strftime("%d/%m") for jour in jours_reference],
        "totaux": [totaux_par_cle.get(jour, 0) for jour in jours_reference],
    }


def _montants_regles_par_jour(nombre_jours=30):
    """Montant total regle (Paiement.montant_part_patient) par jour, sur les
    `nombre_jours` derniers jours (jour courant inclus) -- meme forme que
    _consultations_par_jour, sur Paiement.date_reglement plutot que
    Consultation.date_consultation."""
    aujourd_hui = timezone.now().date()
    jours_reference = [aujourd_hui - datetime.timedelta(days=delta) for delta in range(nombre_jours - 1, -1, -1)]

    montants = (
        Paiement.objects.filter(statut=Paiement.Statut.REGLE, date_reglement__date__gte=jours_reference[0])
        .annotate(jour=TruncDate("date_reglement"))
        .values("jour")
        .annotate(total=Sum("montant_part_patient"))
    )
    totaux_par_cle = {m["jour"]: m["total"] for m in montants if m["jour"]}

    return {
        "labels": [jour.strftime("%d/%m") for jour in jours_reference],
        "totaux": [float(totaux_par_cle.get(jour, 0)) for jour in jours_reference],
    }


def _consultations_par_annee(nombre_annees=5):
    """Nombre de consultations par annee, sur les `nombre_annees` dernieres annees (annee courante incluse)."""
    annee_courante = timezone.now().year
    annees_reference = list(range(annee_courante - nombre_annees + 1, annee_courante + 1))

    comptages = (
        Consultation.objects.annotate(annee=TruncYear("date_consultation"))
        .values("annee")
        .annotate(total=Count("id"))
    )
    totaux_par_cle = {c["annee"].year: c["total"] for c in comptages if c["annee"]}

    return {
        "labels": [str(annee) for annee in annees_reference],
        "totaux": [totaux_par_cle.get(annee, 0) for annee in annees_reference],
    }


def _donnees_rapports():
    """Comptages et agregats de synthese de l'activite de la plateforme (Phase 13)."""
    return {
        "utilisateurs_par_role": [
            {"label": label, "total": User.objects.filter(role=value).count()}
            for value, label in User.Role.choices
        ],
        "patients_par_type": [
            {"label": label, "total": Patient.objects.filter(type_beneficiaire=value).count()}
            for value, label in Patient.TypeBeneficiaire.choices
        ],
        "rendez_vous_par_statut": [
            {"label": label, "total": RendezVous.objects.filter(statut=value).count()}
            for value, label in RendezVous.Statut.choices
        ],
        "prises_en_charge_par_statut": [
            {"label": label, "total": PriseEnCharge.objects.filter(statut=value).count()}
            for value, label in PriseEnCharge.STATUT_CHOICES
        ],
        "total_consultations": Consultation.objects.count(),
        "total_ordonnances": Ordonnance.objects.count(),
        "total_delivrances": Delivrance.objects.count(),
        "total_prestataires_partenaires": Prestataire.objects.filter(partenaire=True).count(),
        "consultations_par_jour": _consultations_par_jour(),
        "consultations_par_mois": _consultations_par_mois(),
        "consultations_par_annee": _consultations_par_annee(),
    }


@admin_required
def rapports(request):
    """Synthese de l'activite de la plateforme : comptages et graphiques (Phase 13)."""
    return render(request, "rapports.html", _donnees_rapports())


@admin_required
def exporter_rapports_excel(request):
    donnees = _donnees_rapports()
    classeur = openpyxl.Workbook()
    classeur.remove(classeur.active)

    def ajouter_feuille(nom, entetes, lignes):
        feuille = classeur.create_sheet(nom)
        feuille.append(entetes)
        for ligne in lignes:
            feuille.append(ligne)
        for index, nom_colonne in enumerate(entetes, start=1):
            feuille.column_dimensions[get_column_letter(index)].width = max(len(str(nom_colonne)), 18)

    ajouter_feuille("Chiffres cles", ["Indicateur", "Total"], [
        ["Consultations", donnees["total_consultations"]],
        ["Ordonnances", donnees["total_ordonnances"]],
        ["Delivrances", donnees["total_delivrances"]],
        ["Prestataires partenaires", donnees["total_prestataires_partenaires"]],
    ])
    ajouter_feuille(
        "Utilisateurs par role", ["Role", "Total"],
        [[ligne["label"], ligne["total"]] for ligne in donnees["utilisateurs_par_role"]],
    )
    ajouter_feuille(
        "Assures par type", ["Type", "Total"],
        [[ligne["label"], ligne["total"]] for ligne in donnees["patients_par_type"]],
    )
    ajouter_feuille(
        "Rendez-vous par statut", ["Statut", "Total"],
        [[ligne["label"], ligne["total"]] for ligne in donnees["rendez_vous_par_statut"]],
    )
    ajouter_feuille(
        "Prises en charge par statut", ["Statut", "Total"],
        [[ligne["label"], ligne["total"]] for ligne in donnees["prises_en_charge_par_statut"]],
    )
    ajouter_feuille(
        "Consultations par mois", ["Mois", "Total"],
        list(zip(donnees["consultations_par_mois"]["labels"], donnees["consultations_par_mois"]["totaux"])),
    )

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="rapports_santesn.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def exporter_rapports_pdf(request):
    donnees = _donnees_rapports()
    reponse = HttpResponse(content_type="application/pdf")
    reponse["Content-Disposition"] = 'attachment; filename="rapports_santesn.pdf"'

    document = SimpleDocTemplate(reponse, pagesize=A4, title="Rapports SanteSN")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Rapports SanteSN", styles["Title"]), Spacer(1, 12)]

    def ajouter_tableau(titre, entetes, lignes):
        elements.append(Paragraph(titre, styles["Heading2"]))
        tableau = Table([entetes] + lignes, hAlign="LEFT")
        tableau.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12885a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(tableau)
        elements.append(Spacer(1, 16))

    ajouter_tableau("Chiffres cles", ["Indicateur", "Total"], [
        ["Consultations", str(donnees["total_consultations"])],
        ["Ordonnances", str(donnees["total_ordonnances"])],
        ["Delivrances", str(donnees["total_delivrances"])],
        ["Prestataires partenaires", str(donnees["total_prestataires_partenaires"])],
    ])
    ajouter_tableau(
        "Utilisateurs par role", ["Role", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["utilisateurs_par_role"]],
    )
    ajouter_tableau(
        "Assures par type", ["Type", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["patients_par_type"]],
    )
    ajouter_tableau(
        "Rendez-vous par statut", ["Statut", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["rendez_vous_par_statut"]],
    )
    ajouter_tableau(
        "Prises en charge par statut", ["Statut", "Total"],
        [[ligne["label"], str(ligne["total"])] for ligne in donnees["prises_en_charge_par_statut"]],
    )
    ajouter_tableau(
        "Consultations par mois", ["Mois", "Total"],
        [
            [label, str(total)]
            for label, total in zip(
                donnees["consultations_par_mois"]["labels"], donnees["consultations_par_mois"]["totaux"]
            )
        ],
    )

    document.build(elements)
    return reponse


@admin_required
def liste_patients(request):
    patients = Patient.objects.select_related("assure_principal", "plan_couverture").all()

    type_beneficiaire = request.GET.get("type", "")
    if type_beneficiaire:
        patients = patients.filter(type_beneficiaire=type_beneficiaire)

    patients = _trier(
        request, patients,
        ["id", "nom", "type_beneficiaire", "assure_principal__nom", "numero_carte", "plan_couverture__nom"],
        ["nom", "prenom"],
    )

    contexte = {
        "patients": _paginer(request, patients),
        "types_beneficiaire": Patient.TypeBeneficiaire.choices,
        "type_selectionne": type_beneficiaire,
    }
    return render(request, "liste_patients.html", contexte)


@admin_required
def ajouter_patient(request):
    if request.method == "POST":
        form = PatientCreationForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            if patient.type_beneficiaire == Patient.TypeBeneficiaire.PRINCIPAL:
                mot_de_passe = generer_mot_de_passe()
                utilisateur = User.objects.create_user(
                    email=form.cleaned_data['email'],
                    password=mot_de_passe,
                    role=User.Role.ASSURE,
                    first_name=patient.prenom,
                    last_name=patient.nom,
                    phone_number=patient.telephone,
                )
                patient.user = utilisateur
                patient.save()
                return render(
                    request,
                    "mot_de_passe_genere.html",
                    {"utilisateur": utilisateur, "mot_de_passe": mot_de_passe, "action": "creation"},
                )
            patient.save()
            messages.success(request, "Assuré ajouté.")
            return redirect("liste_patients")
    else:
        form = PatientCreationForm()
    return render(request, "ajouter_patient.html", {"form": form})


@admin_required
def liste_medecins(request):
    medecins = _trier(request, Medecin.objects.all(), ["nom", "specialite", "email"], ["nom", "prenom"])
    return render(request, "liste_medecins.html", {"medecins": _paginer(request, medecins)})


@admin_required
def ajouter_medecin(request):
    if request.method == "POST":
        form = MedecinForm(request.POST)
        if form.is_valid():
            medecin = form.save(commit=False)
            mot_de_passe = generer_mot_de_passe()
            utilisateur = User.objects.create_user(
                email=medecin.email,
                password=mot_de_passe,
                role=User.Role.MEDECIN,
                first_name=medecin.prenom,
                last_name=medecin.nom,
                phone_number=medecin.telephone,
            )
            medecin.user = utilisateur
            medecin.save()
            return render(
                request,
                "mot_de_passe_genere.html",
                {"utilisateur": utilisateur, "mot_de_passe": mot_de_passe, "action": "creation"},
            )
    else:
        form = MedecinForm()
    return render(request, "ajouter_medecin.html", {"form": form})


@admin_required
def liste_services(request):
    services = _trier(request, ServiceMedical.objects.all(), ["nom", "prix"], "nom")
    return render(request, "liste_services.html", {"services": _paginer(request, services)})


@admin_required
def ajouter_service(request):
    if request.method == "POST":
        form = ServiceMedicalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Service ajouté.")
            return redirect("liste_services")
    else:
        form = ServiceMedicalForm()
    return render(request, "ajouter_service.html", {"form": form})


@admin_required
def liste_prises_en_charge(request):
    prises_en_charge = PriseEnCharge.objects.select_related("patient")

    recherche = request.GET.get("q", "").strip()
    if recherche:
        prises_en_charge = prises_en_charge.filter(
            Q(patient__nom__icontains=recherche) | Q(patient__prenom__icontains=recherche)
        )

    statut = request.GET.get("statut", "")
    if statut:
        prises_en_charge = prises_en_charge.filter(statut=statut)

    prises_en_charge = _trier(
        request, prises_en_charge, ["patient__nom", "date_demande", "statut"], "-date_demande",
    )
    return render(
        request,
        "liste_prises_en_charge.html",
        {
            "prises_en_charge": _paginer(request, prises_en_charge),
            "recherche": recherche,
            "statut_choisi": statut,
            "statuts": PriseEnCharge.STATUT_CHOICES,
        },
    )


@admin_required
def liste_rendez_vous(request):
    """Liste administrateur des rendez-vous, en lecture seule.

    L'administrateur suit le flux sans y intervenir : confirmer ou annuler un
    rendez-vous reste l'affaire du medecin (changer_statut_rendez_vous) et de
    l'assure (annuler_rendez_vous_assure). Aucune action d'ecriture n'est donc
    exposee ici.
    """
    rendez_vous = RendezVous.objects.select_related("patient", "medecin", "prestataire")

    recherche = request.GET.get("q", "").strip()
    if recherche:
        rendez_vous = rendez_vous.filter(
            Q(patient__nom__icontains=recherche)
            | Q(patient__prenom__icontains=recherche)
            | Q(medecin__nom__icontains=recherche)
            | Q(medecin__prenom__icontains=recherche)
        )

    statut = request.GET.get("statut", "")
    if statut:
        rendez_vous = rendez_vous.filter(statut=statut)

    rendez_vous = _trier(
        request,
        rendez_vous,
        ["date_heure", "patient__nom", "medecin__nom", "statut"],
        "-date_heure",
    )
    return render(
        request,
        "liste_rendez_vous.html",
        {
            "rendez_vous": _paginer(request, rendez_vous),
            "recherche": recherche,
            "statut_choisi": statut,
            "statuts": RendezVous.Statut.choices,
        },
    )


@admin_required
def liste_ordonnances(request):
    """Liste administrateur des ordonnances, en lecture seule.

    Le filtre "delivrance" repond a un angle mort : une ordonnance emise mais
    jamais retiree en pharmacie n'apparaissait sur aucun ecran. Valider une
    delivrance reste l'affaire du pharmacien (valider_delivrance). Le QR n'est
    pas affiche ici : il n'a de sens qu'au comptoir.
    """
    ordonnances = Ordonnance.objects.select_related(
        "consultation__patient", "consultation__medecin", "delivrance__pharmacien"
    )

    recherche = request.GET.get("q", "").strip()
    if recherche:
        ordonnances = ordonnances.filter(
            Q(consultation__patient__nom__icontains=recherche)
            | Q(consultation__patient__prenom__icontains=recherche)
            | Q(code_qr__icontains=recherche)
        )

    delivrance = request.GET.get("delivrance", "")
    if delivrance == "non":
        ordonnances = ordonnances.filter(delivrance__isnull=True)
    elif delivrance == "oui":
        ordonnances = ordonnances.filter(delivrance__isnull=False)

    ordonnances = _trier(
        request,
        ordonnances,
        ["date_creation", "consultation__patient__nom", "code_qr"],
        "-date_creation",
    )
    return render(
        request,
        "liste_ordonnances.html",
        {
            "ordonnances": _paginer(request, ordonnances),
            "recherche": recherche,
            "delivrance_choisie": delivrance,
        },
    )


@admin_required
def ajouter_prise_en_charge(request):
    """A la creation, le statut est toujours 'en_attente' : le champ n'est pas propose."""
    if request.method == "POST":
        form = PriseEnChargeForm(request.POST)
        form.fields.pop("statut")
        if form.is_valid():
            prise_en_charge = form.save(commit=False)
            prise_en_charge.statut = "en_attente"
            prise_en_charge.save()
            messages.success(request, "Prise en charge ajoutée.")
            return redirect("liste_prises_en_charge")
    else:
        form = PriseEnChargeForm()
        form.fields.pop("statut")
    return render(request, "ajouter_prise_en_charge.html", {"form": form})


@admin_required
def modifier_prise_en_charge(request, pk):
    prise_en_charge = get_object_or_404(PriseEnCharge, pk=pk)
    if request.method == "POST":
        form = PriseEnChargeForm(request.POST, instance=prise_en_charge)
        if form.is_valid():
            form.save()
            messages.success(request, "Prise en charge modifiée.")
            return redirect("liste_prises_en_charge")
    else:
        form = PriseEnChargeForm(instance=prise_en_charge)
    return render(
        request,
        "modifier_prise_en_charge.html",
        {"form": form, "prise_en_charge": prise_en_charge},
    )


@admin_required
def supprimer_prise_en_charge(request, pk):
    prise_en_charge = get_object_or_404(PriseEnCharge, pk=pk)
    if request.method == "POST":
        prise_en_charge.delete()
        messages.success(request, "Prise en charge supprimée.")
        return redirect("liste_prises_en_charge")
    return render(request, "confirmer_suppression.html", {"objet": prise_en_charge, "type": "Prise en charge"})


def _filtrer_paiements(request):
    """Filtres partages entre la liste et l'export CSV des paiements."""
    paiements = Paiement.objects.select_related(
        "consultation", "consultation__patient", "consultation__service"
    )

    statut = request.GET.get("statut", "")
    if statut:
        paiements = paiements.filter(statut=statut)

    recherche = request.GET.get("q", "").strip()
    if recherche:
        paiements = paiements.filter(
            Q(consultation__patient__nom__icontains=recherche)
            | Q(consultation__patient__prenom__icontains=recherche)
        )

    paiements = _trier(
        request, paiements,
        ["consultation__patient__nom", "montant_total", "montant_part_assurance", "montant_part_patient", "statut"],
        "-consultation__date_consultation",
    )

    return paiements, {"statut": statut, "recherche": recherche}


@admin_required
def liste_paiements(request):
    paiements, filtres = _filtrer_paiements(request)

    totaux = Paiement.objects.aggregate(
        total_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.REGLE)),
        total_non_regle=Sum("montant_part_patient", filter=Q(statut=Paiement.Statut.NON_REGLE)),
    )

    contexte = {
        "paiements": _paginer(request, paiements),
        "statut_choisi": filtres["statut"],
        "statuts": Paiement.Statut.choices,
        "recherche": filtres["recherche"],
        "total_regle": totaux["total_regle"] or 0,
        "total_non_regle": totaux["total_non_regle"] or 0,
    }
    return render(request, "liste_paiements.html", contexte)


def _cellule_csv(valeur):
    """Neutralise l'injection de formule CSV : un tableur (Excel, LibreOffice)
    execute une cellule commencant par =, +, -, @ ou une tabulation/retour
    chariot comme une formule. Nom/prenom/email etc. viennent de champs
    remplis par un administrateur (pas d'inscription publique), mais rien
    n'empeche une saisie malveillante ou accidentelle -- le prefixe
    (apostrophe) neutralise la formule sans changer la valeur affichee."""
    texte = "" if valeur is None else str(valeur)
    if texte and texte[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + texte
    return texte


@admin_required
def exporter_paiements_csv(request):
    paiements, _ = _filtrer_paiements(request)

    reponse = HttpResponse(content_type="text/csv")
    reponse["Content-Disposition"] = 'attachment; filename="paiements_santesn.csv"'
    reponse.write("﻿")  # BOM : Excel (FR) detecte l'UTF-8 sans le confondre avec l'encodage local.
    ecrivain = csv.writer(reponse, delimiter=";")
    ecrivain.writerow([
        "Reference", "Patient", "Date de consultation", "Montant total",
        "Part assurance", "Part patient", "Statut", "Mode de reglement",
        "Date de reglement",
    ])
    for paiement in paiements:
        ecrivain.writerow([_cellule_csv(v) for v in [
            paiement.pk,
            str(paiement.consultation.patient),
            paiement.consultation.date_consultation.strftime("%d/%m/%Y %H:%M"),
            paiement.montant_total,
            paiement.montant_part_assurance,
            paiement.montant_part_patient,
            paiement.get_statut_display(),
            paiement.get_mode_reglement_display() if paiement.mode_reglement else "",
            paiement.date_reglement.strftime("%d/%m/%Y %H:%M") if paiement.date_reglement else "",
        ]])
    return reponse


@admin_required
def marquer_paiement_regle(request, pk):
    paiement = get_object_or_404(Paiement, pk=pk)
    if paiement.statut == Paiement.Statut.REGLE:
        messages.info(request, "Ce paiement est déjà réglé.")
        return redirect("liste_paiements")

    if request.method == "POST":
        form = PaiementReglementForm(request.POST, instance=paiement)
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.statut = Paiement.Statut.REGLE
            paiement.date_reglement = timezone.now()
            paiement.save()
            messages.success(request, "Paiement marqué comme réglé.", extra_tags="succes-critique")
            return redirect("liste_paiements")
    else:
        form = PaiementReglementForm(instance=paiement)
    return render(request, "marquer_paiement_regle.html", {"form": form, "paiement": paiement})


@admin_required
def liste_prestataires(request):
    prestataires = Prestataire.objects.all()

    localisation = request.GET.get("localisation", "")
    if localisation == "sans":
        prestataires = prestataires.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        )
    elif localisation == "avec":
        prestataires = prestataires.filter(
            latitude__isnull=False, longitude__isnull=False
        )

    prestataires = _trier(
        request, prestataires, ["id", "nom", "type_prestataire", "ville", "partenaire"], "nom",
    )

    # Reseau de partenaires geolocalises, pour la carte (deplacee ici
    # depuis le dashboard admin) : seuls les champs necessaires au rendu.
    prestataires_carte = list(
        Prestataire.objects.filter(partenaire=True, latitude__isnull=False, longitude__isnull=False)
        .values("nom", "type_prestataire", "ville", "latitude", "longitude")
    )

    contexte = {
        "prestataires": _paginer(request, prestataires),
        "localisation_choisie": localisation,
        "prestataires_carte": prestataires_carte,
    }
    return render(request, "liste_prestataires.html", contexte)


@admin_required
def recherche_lieu_prestataire(request):
    """
    Relais serveur vers Nominatim (recherche OpenStreetMap) pour le bouton
    "Rechercher sur la carte" des formulaires prestataire. Un appel direct
    navigateur -> Nominatim est sujet a des echecs intermittents (CORS
    incoherent derriere leur cache, User-Agent de fetch() non identifiant) ;
    en passant par le serveur, la requete est same-origin cote navigateur et
    peut porter un User-Agent conforme a la politique d'usage de Nominatim.
    """
    requete = request.GET.get("q", "").strip()
    if not requete:
        return JsonResponse({"trouve": False})

    parametres = urllib.parse.urlencode({
        "format": "json",
        "limit": 1,
        "countrycodes": "sn",
        "q": f"{requete}, Senegal",
    })
    url = f"https://nominatim.openstreetmap.org/search?{parametres}"
    requete_http = urllib.request.Request(
        url, headers={"User-Agent": "SanteSN-PlateformeMedicale/1.0"}
    )

    try:
        with urllib.request.urlopen(requete_http, timeout=5) as reponse:
            resultats = json.loads(reponse.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, ValueError):
        return JsonResponse({"trouve": False, "erreur": "service_indisponible"})

    if not resultats:
        return JsonResponse({"trouve": False})

    lieu = resultats[0]
    return JsonResponse({
        "trouve": True,
        "lat": lieu.get("lat"),
        "lon": lieu.get("lon"),
        "nom": lieu.get("display_name", requete),
    })


@admin_required
def ajouter_prestataire(request):
    if request.method == "POST":
        form = PrestataireForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Prestataire ajouté.")
            return redirect("liste_prestataires")
    else:
        form = PrestataireForm()
    return render(request, "ajouter_prestataire.html", {"form": form})


@admin_required
def modifier_prestataire(request, pk):
    prestataire = get_object_or_404(Prestataire, pk=pk)
    if request.method == "POST":
        form = PrestataireForm(request.POST, instance=prestataire)
        if form.is_valid():
            form.save()
            messages.success(request, "Prestataire modifié.")
            return redirect("liste_prestataires")
    else:
        form = PrestataireForm(instance=prestataire)
    return render(request, "modifier_prestataire.html", {"form": form, "prestataire": prestataire})


@admin_required
def supprimer_prestataire(request, pk):
    prestataire = get_object_or_404(Prestataire, pk=pk)
    if request.method == "POST":
        prestataire.delete()
        messages.success(request, "Prestataire supprimé.")
        return redirect("liste_prestataires")
    return render(request, "confirmer_suppression.html", {"objet": prestataire, "type": "Prestataire"})


@admin_required
def liste_plans_couverture(request):
    plans = _trier(request, PlanCouverture.objects.all(), ["nom", "taux_couverture", "plafond_annuel"], "nom")
    return render(request, "liste_plans_couverture.html", {"plans": _paginer(request, plans)})


@admin_required
def ajouter_plan_couverture(request):
    if request.method == "POST":
        form = PlanCouvertureForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Plan de couverture ajouté.")
            return redirect("liste_plans_couverture")
    else:
        form = PlanCouvertureForm()
    return render(request, "ajouter_plan_couverture.html", {"form": form})


@admin_required
def modifier_plan_couverture(request, pk):
    plan = get_object_or_404(PlanCouverture, pk=pk)
    if request.method == "POST":
        form = PlanCouvertureForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "Plan de couverture modifié.")
            return redirect("liste_plans_couverture")
    else:
        form = PlanCouvertureForm(instance=plan)
    return render(request, "modifier_plan_couverture.html", {"form": form, "plan": plan})


@admin_required
def supprimer_plan_couverture(request, pk):
    plan = get_object_or_404(PlanCouverture, pk=pk)
    if request.method == "POST":
        plan.delete()
        messages.success(request, "Plan de couverture supprimé.")
        return redirect("liste_plans_couverture")
    avertissement = _avertissement_cascade({"assuré(s)/ayant(s) droit rattachés (plan retiré, pas supprimés)": plan.beneficiaires.count()})
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": plan, "type": "Plan de couverture", "avertissement": avertissement},
    )


# MODIFIER VUES
@admin_required
def modifier_patient(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == "POST":
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, "Assuré modifié.")
            return redirect("liste_patients")
    else:
        form = PatientForm(instance=patient)
    return render(request, "modifier_patient.html", {"form": form, "patient": patient})


@admin_required
def modifier_medecin(request, pk):
    medecin = get_object_or_404(Medecin, pk=pk)
    if request.method == "POST":
        form = MedecinForm(request.POST, instance=medecin)
        if form.is_valid():
            form.save()
            messages.success(request, "Médecin modifié.")
            return redirect("liste_medecins")
    else:
        form = MedecinForm(instance=medecin)
    return render(request, "modifier_medecin.html", {"form": form, "medecin": medecin})


@admin_required
def modifier_service(request, pk):
    service = get_object_or_404(ServiceMedical, pk=pk)
    if request.method == "POST":
        form = ServiceMedicalForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, "Service modifié.")
            return redirect("liste_services")
    else:
        form = ServiceMedicalForm(instance=service)
    return render(request, "modifier_service.html", {"form": form, "service": service})


def _avertissement_cascade(compteurs):
    """Construit un message d'avertissement a partir d'un dict {libelle: total}."""
    parties = [f"{total} {libelle}" for libelle, total in compteurs.items() if total]
    if not parties:
        return None
    return "Seront aussi supprimés : " + ", ".join(parties) + "."


# SUPPRIMER VUES
@admin_required
def supprimer_patient(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == "POST":
        # Le User (assure principal uniquement, jamais un ayant droit) doit
        # etre desactive : sinon la fiche disparait mais le compte de
        # connexion reste actif, et mon_profil_assure se recree tout seul
        # une fiche Patient a la prochaine connexion.
        if patient.user:
            patient.user.is_active = False
            patient.user.save(update_fields=["is_active"])
        patient.delete()
        messages.success(request, "Assuré supprimé.")
        return redirect("liste_patients")
    avertissement = _avertissement_cascade({
        "ayant(s) droit": patient.ayants_droit.count(),
        "consultation(s)": patient.consultation_set.count(),
        "prise(s) en charge": patient.priseencharge_set.count(),
        "rendez-vous": patient.rendez_vous.count(),
    })
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": patient, "type": "Patient", "avertissement": avertissement},
    )


@admin_required
def supprimer_medecin(request, pk):
    medecin = get_object_or_404(Medecin, pk=pk)
    if request.method == "POST":
        # Desactive le User lie : sinon la fiche Medecin disparait mais le
        # compte de connexion reste actif (voir supprimer_patient, meme
        # raisonnement).
        if medecin.user:
            medecin.user.is_active = False
            medecin.user.save(update_fields=["is_active"])
        medecin.delete()
        messages.success(request, "Médecin supprimé.")
        return redirect("liste_medecins")
    avertissement = _avertissement_cascade({
        "consultation(s)": medecin.consultation_set.count(),
        "rendez-vous": medecin.rendez_vous.count(),
        "paiement(s)": Paiement.objects.filter(consultation__medecin=medecin).count(),
        "ordonnance(s)": Ordonnance.objects.filter(consultation__medecin=medecin).count(),
    })
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": medecin, "type": "Medecin", "avertissement": avertissement},
    )


@admin_required
def liste_pharmaciens(request):
    pharmaciens = Pharmacien.objects.select_related("user", "prestataire")
    pharmaciens = _trier(
        request, pharmaciens, ["user__last_name", "user__email", "prestataire__nom"], "id",
    )
    return render(request, "liste_pharmaciens.html", {"pharmaciens": _paginer(request, pharmaciens)})


@admin_required
def modifier_pharmacien(request, pk):
    pharmacien = get_object_or_404(Pharmacien, pk=pk)
    if request.method == "POST":
        form = PharmacienAffectationForm(request.POST, instance=pharmacien)
        if form.is_valid():
            form.save()
            messages.success(request, "Pharmacien modifié.")
            return redirect("liste_pharmaciens")
    else:
        form = PharmacienAffectationForm(instance=pharmacien)
    return render(request, "modifier_pharmacien.html", {"form": form, "pharmacien": pharmacien})


@admin_required
def supprimer_service(request, pk):
    service = get_object_or_404(ServiceMedical, pk=pk)
    if request.method == "POST":
        service.delete()
        messages.success(request, "Service supprimé.")
        return redirect("liste_services")
    return render(request, "confirmer_suppression.html", {"objet": service, "type": "Service"})


# ---------------------------------------------------------------------------
# Gestion des utilisateurs (Administrateur)
# ---------------------------------------------------------------------------

def _filtrer_utilisateurs(request):
    """Filtres partages entre la liste et l'export Excel des utilisateurs."""
    utilisateurs = User.objects.all()

    role = request.GET.get("role", "")
    statut = request.GET.get("statut", "")
    recherche = request.GET.get("q", "").strip()

    if role:
        utilisateurs = utilisateurs.filter(role=role)
    if statut == "actif":
        utilisateurs = utilisateurs.filter(is_active=True)
    elif statut == "inactif":
        utilisateurs = utilisateurs.filter(is_active=False)
    if recherche:
        utilisateurs = utilisateurs.filter(
            Q(email__icontains=recherche)
            | Q(first_name__icontains=recherche)
            | Q(last_name__icontains=recherche)
        )

    utilisateurs = _trier(request, utilisateurs, ["last_name", "email", "role", "is_active"], ["last_name", "first_name"])

    return utilisateurs, {"role": role, "statut": statut, "recherche": recherche}


@admin_required
def liste_utilisateurs(request):
    utilisateurs, filtres = _filtrer_utilisateurs(request)
    contexte = {
        "utilisateurs": _paginer(request, utilisateurs),
        "roles": User.Role.choices,
        "role_selectionne": filtres["role"],
        "statut_selectionne": filtres["statut"],
        "recherche": filtres["recherche"],
    }
    return render(request, "liste_utilisateurs.html", contexte)


@admin_required
def exporter_utilisateurs_excel(request):
    utilisateurs, _ = _filtrer_utilisateurs(request)

    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Utilisateurs"
    entetes = ["Email", "Prenom", "Nom", "Telephone", "Role", "Statut", "Date de creation"]
    feuille.append(entetes)

    for utilisateur in utilisateurs:
        feuille.append([
            utilisateur.email,
            utilisateur.first_name,
            utilisateur.last_name,
            utilisateur.phone_number,
            utilisateur.get_role_display(),
            "Actif" if utilisateur.is_active else "Inactif",
            utilisateur.date_joined.strftime("%d/%m/%Y %H:%M"),
        ])

    for index, nom_colonne in enumerate(entetes, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(len(nom_colonne), 18)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="utilisateurs_santesn.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def exporter_utilisateurs_csv(request):
    utilisateurs, _ = _filtrer_utilisateurs(request)

    reponse = HttpResponse(content_type="text/csv")
    reponse["Content-Disposition"] = 'attachment; filename="utilisateurs_santesn.csv"'
    reponse.write("﻿")  # BOM : Excel (FR) detecte l'UTF-8 sans le confondre avec l'encodage local.
    ecrivain = csv.writer(reponse, delimiter=";")
    ecrivain.writerow(["Email", "Prenom", "Nom", "Telephone", "Role", "Statut", "Date de creation"])
    for utilisateur in utilisateurs:
        ecrivain.writerow([_cellule_csv(v) for v in [
            utilisateur.email,
            utilisateur.first_name,
            utilisateur.last_name,
            utilisateur.phone_number,
            utilisateur.get_role_display(),
            "Actif" if utilisateur.is_active else "Inactif",
            utilisateur.date_joined.strftime("%d/%m/%Y %H:%M"),
        ]])
    return reponse


COLONNES_IMPORT_UTILISATEURS = [
    "Email", "Prenom", "Nom", "Telephone", "Role",
    "Date de naissance", "Specialite", "Prestataire", "Plan de couverture",
]


def _normaliser_texte_import(valeur):
    """Normalise une valeur de cellule pour une comparaison insensible aux accents/majuscules."""
    texte = "" if valeur is None else str(valeur).strip()
    texte = unicodedata.normalize("NFKD", texte).encode("ascii", "ignore").decode("ascii")
    return texte.upper()


_ROLES_PAR_LIBELLE_IMPORT = {}
for _valeur_role, _label_role in User.Role.choices:
    _ROLES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_valeur_role)] = _valeur_role
    _ROLES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_label_role)] = _valeur_role


def _analyser_ligne_import_utilisateurs(numero_ligne, valeurs):
    """
    Valide une ligne du fichier d'import (voir COLONNES_IMPORT_UTILISATEURS).

    Retourne (donnees, None) si la ligne est valide, ou (None, message_erreur)
    sinon. Ne touche jamais la base : l'import est valide en integralite
    avant toute creation (regle "tout ou rien").
    """
    valeurs = (tuple(valeurs) + (None,) * len(COLONNES_IMPORT_UTILISATEURS))[:len(COLONNES_IMPORT_UTILISATEURS)]
    email, prenom, nom, telephone, role_brut, date_naissance_brute, specialite, prestataire_brut, plan_brut = valeurs

    email = (email or "").strip()
    prenom = (prenom or "").strip()
    nom = (nom or "").strip()
    telephone = (telephone or "").strip() if telephone else ""

    if not email or not prenom or not nom:
        return None, f"Ligne {numero_ligne} : email, prenom et nom sont obligatoires."

    role = _ROLES_PAR_LIBELLE_IMPORT.get(_normaliser_texte_import(role_brut))
    if not role:
        return None, (
            f"Ligne {numero_ligne} : role '{role_brut}' inconnu "
            "(attendu : Administrateur, Assure, Medecin ou Pharmacien)."
        )

    if telephone:
        try:
            valider_telephone(telephone)
        except ValidationError:
            return None, f"Ligne {numero_ligne} : numero de telephone invalide."

    prestataire = None
    prestataire_nom = (prestataire_brut or "").strip() if prestataire_brut else ""
    if prestataire_nom:
        prestataire = Prestataire.objects.filter(nom__iexact=prestataire_nom).first()
        if not prestataire:
            return None, f"Ligne {numero_ligne} : prestataire '{prestataire_nom}' introuvable."

    plan_couverture = None
    plan_nom = (plan_brut or "").strip() if plan_brut else ""
    if plan_nom:
        plan_couverture = PlanCouverture.objects.filter(nom__iexact=plan_nom).first()
        if not plan_couverture:
            return None, f"Ligne {numero_ligne} : plan de couverture '{plan_nom}' introuvable."

    donnees = {
        "email": email,
        "prenom": prenom,
        "nom": nom,
        "telephone": telephone,
        "role": role,
        "prestataire": prestataire,
        "plan_couverture": plan_couverture,
    }

    if role == User.Role.MEDECIN:
        if not telephone:
            return None, f"Ligne {numero_ligne} : le telephone est obligatoire pour un medecin."
        specialite = (specialite or "").strip()
        if not specialite:
            return None, f"Ligne {numero_ligne} : la specialite est obligatoire pour un medecin."
        donnees["specialite"] = specialite
    elif role == User.Role.ASSURE:
        if not date_naissance_brute:
            return None, f"Ligne {numero_ligne} : la date de naissance est obligatoire pour un assure."
        if isinstance(date_naissance_brute, datetime.datetime):
            donnees["date_naissance"] = date_naissance_brute.date()
        elif isinstance(date_naissance_brute, datetime.date):
            donnees["date_naissance"] = date_naissance_brute
        else:
            try:
                donnees["date_naissance"] = datetime.datetime.strptime(
                    str(date_naissance_brute).strip(), "%d/%m/%Y"
                ).date()
            except ValueError:
                return None, (
                    f"Ligne {numero_ligne} : date de naissance invalide (format attendu JJ/MM/AAAA)."
                )

    return donnees, None


def _creer_comptes_import_utilisateurs(lignes_validees):
    """Cree en une transaction tous les comptes (et fiches metier) valides par l'import."""
    resultats = []
    with transaction.atomic():
        for donnees in lignes_validees:
            mot_de_passe = generer_mot_de_passe()
            utilisateur = User.objects.create_user(
                email=donnees["email"],
                password=mot_de_passe,
                role=donnees["role"],
                first_name=donnees["prenom"],
                last_name=donnees["nom"],
                phone_number=donnees["telephone"],
            )
            if donnees["role"] == User.Role.MEDECIN:
                Medecin.objects.create(
                    user=utilisateur,
                    nom=donnees["nom"],
                    prenom=donnees["prenom"],
                    specialite=donnees["specialite"],
                    telephone=donnees["telephone"],
                    email=donnees["email"],
                    prestataire=donnees["prestataire"],
                )
            elif donnees["role"] == User.Role.PHARMACIEN:
                Pharmacien.objects.create(user=utilisateur, prestataire=donnees["prestataire"])
            elif donnees["role"] == User.Role.ASSURE:
                Patient.objects.create(
                    user=utilisateur,
                    nom=donnees["nom"],
                    prenom=donnees["prenom"],
                    date_naissance=donnees["date_naissance"],
                    telephone=donnees["telephone"],
                    type_beneficiaire=Patient.TypeBeneficiaire.PRINCIPAL,
                    plan_couverture=donnees["plan_couverture"],
                )
            resultats.append({
                "email": utilisateur.email,
                "nom_complet": f"{utilisateur.first_name} {utilisateur.last_name}",
                "role": utilisateur.get_role_display(),
                "mot_de_passe": mot_de_passe,
            })
    return resultats


@admin_required
def importer_utilisateurs_excel(request):
    """
    Creation en masse de comptes (Assure principal / Medecin / Pharmacien /
    Administrateur) depuis un fichier Excel : voir COLONNES_IMPORT_UTILISATEURS.

    Regle "tout ou rien" : la moindre ligne invalide bloque tout l'import
    (aucun compte cree), pour eviter un import partiel difficile a auditer.
    """
    erreurs = []
    if request.method == "POST":
        fichier = request.FILES.get("fichier")
        if not fichier:
            erreurs.append("Choisissez un fichier Excel (.xlsx) a importer.")
        else:
            try:
                classeur = openpyxl.load_workbook(fichier, data_only=True)
            except Exception:
                erreurs.append(
                    "Fichier illisible : verifiez qu'il s'agit bien d'un fichier Excel (.xlsx) valide."
                )
            else:
                feuille = classeur.active
                entetes = next(feuille.iter_rows(min_row=1, max_row=1, values_only=True), ())
                entetes_normalisees = [_normaliser_texte_import(entete) for entete in entetes]
                entetes_attendues = [_normaliser_texte_import(colonne) for colonne in COLONNES_IMPORT_UTILISATEURS]

                if entetes_normalisees[:len(entetes_attendues)] != entetes_attendues:
                    erreurs.append(
                        "En-tetes de colonnes invalides : utilisez le modele telechargeable ci-dessous."
                    )
                else:
                    lignes_brutes = [
                        (numero, valeurs)
                        for numero, valeurs in enumerate(
                            feuille.iter_rows(min_row=2, values_only=True), start=2
                        )
                        if valeurs and not all(valeur in (None, "") for valeur in valeurs)
                    ]
                    if not lignes_brutes:
                        erreurs.append("Le fichier ne contient aucune ligne a importer.")
                    else:
                        donnees_valides = []
                        emails_vus = set()
                        for numero_ligne, valeurs in lignes_brutes:
                            donnees, erreur = _analyser_ligne_import_utilisateurs(numero_ligne, valeurs)
                            if erreur:
                                erreurs.append(erreur)
                                continue
                            email_normalise = donnees["email"].lower()
                            if email_normalise in emails_vus:
                                erreurs.append(
                                    f"Ligne {numero_ligne} : email '{donnees['email']}' en double dans le fichier."
                                )
                                continue
                            if User.objects.filter(email__iexact=donnees["email"]).exists():
                                erreurs.append(
                                    f"Ligne {numero_ligne} : email '{donnees['email']}' "
                                    "deja utilise par un compte existant."
                                )
                                continue
                            emails_vus.add(email_normalise)
                            donnees_valides.append(donnees)

                        if not erreurs:
                            resultats = _creer_comptes_import_utilisateurs(donnees_valides)
                            return render(
                                request, "resultat_import_utilisateurs.html", {"resultats": resultats}
                            )

    return render(request, "importer_utilisateurs.html", {"erreurs": erreurs})


@admin_required
def telecharger_modele_import_utilisateurs(request):
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Import utilisateurs"
    feuille.append(COLONNES_IMPORT_UTILISATEURS)
    feuille.append([
        "email@domaine.com", "Prenom", "Nom", "770000000",
        "Assure / Medecin / Pharmacien / Administrateur",
        "JJ/MM/AAAA (uniquement pour un Assure)",
        "Ex: Medecine generale (uniquement pour un Medecin)",
        "Nom exact d'un prestataire existant (optionnel)",
        "Nom exact d'un plan de couverture existant (optionnel)",
    ])

    for index, nom_colonne in enumerate(COLONNES_IMPORT_UTILISATEURS, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(len(nom_colonne), 20)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="modele_import_utilisateurs.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def ajouter_utilisateur(request):
    if request.method == "POST":
        form = UtilisateurCreationForm(request.POST)
        if form.is_valid():
            utilisateur = form.save()
            return render(
                request,
                "mot_de_passe_genere.html",
                {
                    "utilisateur": utilisateur,
                    "mot_de_passe": form.mot_de_passe_genere,
                    "action": "creation",
                },
            )
    else:
        form = UtilisateurCreationForm()
    return render(request, "ajouter_utilisateur.html", {"form": form})


@admin_required
def modifier_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        form = UtilisateurModificationForm(request.POST, instance=utilisateur)
        if form.is_valid():
            nouveau_role = form.cleaned_data["role"]
            if utilisateur.pk == request.user.pk and nouveau_role != request.user.role:
                form.add_error("role", "Vous ne pouvez pas modifier votre propre rôle.")
            else:
                utilisateur_modifie = form.save()
                lier_fiche_medecin(utilisateur_modifie)
                lier_fiche_pharmacien(utilisateur_modifie)
                messages.success(request, "Utilisateur modifié avec succès.")
                return redirect("liste_utilisateurs")
    else:
        form = UtilisateurModificationForm(instance=utilisateur)
    return render(
        request,
        "modifier_utilisateur.html",
        {"form": form, "utilisateur": utilisateur},
    )


@admin_required
@require_POST
def activer_desactiver_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if utilisateur.pk == request.user.pk:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
        return redirect("liste_utilisateurs")

    # Un seul administrateur actif a la fois (comme dans un veritable SaaS) :
    # reactiver un ancien compte ADMIN ne doit pas en creer un second en
    # silence. Les creations/promotions ADMIN sont deja bloquees cote
    # formulaire (UtilisateurCreationForm/UtilisateurModificationForm) ; ce
    # garde-fou couvre le seul autre chemin qui pourrait faire exister un
    # deuxieme administrateur actif.
    if (
        utilisateur.role == User.Role.ADMIN
        and not utilisateur.is_active
        and User.objects.filter(role=User.Role.ADMIN, is_active=True).exclude(pk=utilisateur.pk).exists()
    ):
        messages.error(
            request,
            "Un seul administrateur peut être actif à la fois. "
            "Désactivez d'abord l'administrateur actuel.",
        )
        return redirect("liste_utilisateurs")

    utilisateur.is_active = not utilisateur.is_active
    utilisateur.save(update_fields=["is_active"])
    if utilisateur.is_active:
        messages.success(request, f"Compte de {utilisateur} activé.")
    else:
        messages.success(request, f"Compte de {utilisateur} désactivé.")
    return redirect("liste_utilisateurs")


@admin_required
def reinitialiser_mot_de_passe(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        nouveau_mot_de_passe = generer_mot_de_passe()
        utilisateur.set_password(nouveau_mot_de_passe)
        utilisateur.save(update_fields=["password"])
        return render(
            request,
            "mot_de_passe_genere.html",
            {
                "utilisateur": utilisateur,
                "mot_de_passe": nouveau_mot_de_passe,
                "action": "reinitialisation",
            },
        )
    return render(request, "reinitialiser_mot_de_passe.html", {"utilisateur": utilisateur})


@admin_required
def supprimer_utilisateur(request, pk):
    utilisateur = get_object_or_404(User, pk=pk)
    if utilisateur.pk == request.user.pk:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
        return redirect("liste_utilisateurs")

    if request.method == "POST":
        utilisateur.delete()
        messages.success(request, "Utilisateur supprimé.")
        return redirect("liste_utilisateurs")
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": utilisateur, "type": "Utilisateur"},
    )


# ---------------------------------------------------------------------------
# Espace Medecin
# ---------------------------------------------------------------------------

def _medecin_courant(request):
    return getattr(request.user, "medecin", None)


def _patients_du_medecin(medecin):
    return Patient.objects.filter(
        Q(rendez_vous__medecin=medecin) | Q(consultation__medecin=medecin)
    ).distinct().order_by("nom", "prenom")


@role_required(User.Role.MEDECIN)
def dashboard_medecin(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    maintenant = timezone.now()
    rendez_vous_a_venir = RendezVous.objects.filter(
        medecin=medecin, date_heure__gte=maintenant
    ).exclude(statut=RendezVous.Statut.ANNULE)

    contexte = {
        "total_patients": _patients_du_medecin(medecin).count(),
        "total_rendez_vous_a_venir": rendez_vous_a_venir.count(),
        "total_consultations": Consultation.objects.filter(medecin=medecin).count(),
        "total_ordonnances": Ordonnance.objects.filter(consultation__medecin=medecin).count(),
        "prochains_rendez_vous": rendez_vous_a_venir.select_related("patient").order_by("date_heure")[:5],
        "medecin": medecin,
        "tendance_consultations": _consultations_par_jour(queryset=Consultation.objects.filter(medecin=medecin)),
    }
    return render(request, "dashboard_medecin.html", contexte)


@role_required(User.Role.MEDECIN)
def agenda_medecin(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    # order_by explicite : voir mes_rendez_vous_assure.
    rendez_vous = RendezVous.objects.filter(medecin=medecin).select_related(
        "patient", "prestataire"
    ).order_by("-date_heure")
    return render(request, "agenda_medecin.html",
                  {"rendez_vous": _paginer(request, rendez_vous)})


@role_required(User.Role.MEDECIN)
def ajouter_rendez_vous(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    if request.method == "POST":
        form = RendezVousForm(request.POST)
        if form.is_valid():
            rendez_vous = form.save(commit=False)
            rendez_vous.medecin = medecin
            rendez_vous.save()
            messages.success(request, "Rendez-vous créé.")
            return redirect("agenda_medecin")
    else:
        form = RendezVousForm()
    return render(request, "ajouter_rendez_vous.html", {"form": form})


@role_required(User.Role.MEDECIN)
@require_POST
def changer_statut_rendez_vous(request, pk):
    medecin = _medecin_courant(request)
    rendez_vous = get_object_or_404(RendezVous, pk=pk, medecin=medecin)
    nouveau_statut = request.POST.get("statut")
    if nouveau_statut in RendezVous.Statut.values:
        rendez_vous.statut = nouveau_statut
        rendez_vous.save(update_fields=["statut"])
        messages.success(request, "Statut du rendez-vous mis à jour.")
    return redirect("agenda_medecin")


@role_required(User.Role.MEDECIN)
def mes_patients(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")
    return render(request, "mes_patients.html",
                  {"patients": _paginer(request, _patients_du_medecin(medecin))})


@role_required(User.Role.MEDECIN)
def rechercher_patients_medecin(request):
    """
    Recherche live pour la barre de recherche rapide du medecin (numero de
    carte, nom, prenom, identifiant numerique). Renvoie du JSON, jamais de
    donnee medicale : seulement de quoi identifier le bon patient avant
    d'ouvrir sa fiche (voir fiche_patient_medecin).
    """
    medecin = _medecin_courant(request)
    if medecin is None:
        return JsonResponse({"resultats": []})

    requete = request.GET.get("q", "").strip()
    if len(requete) < 2:
        return JsonResponse({"resultats": []})

    filtre = (
        Q(numero_carte__icontains=requete)
        | Q(nom__icontains=requete)
        | Q(prenom__icontains=requete)
    )
    if requete.isdecimal():
        filtre |= Q(pk=requete)

    patients = list(
        Patient.objects.filter(filtre)
        .annotate(
            priorite=Case(
                When(numero_carte__iexact=requete, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
        .order_by("priorite", "nom", "prenom")[:8]
    )

    patients_lies = set(
        _patients_du_medecin(medecin)
        .filter(pk__in=[patient.pk for patient in patients])
        .values_list("pk", flat=True)
    )

    resultats = [
        {
            "id": patient.pk,
            "nom": patient.nom,
            "prenom": patient.prenom,
            "numero_carte": patient.numero_carte,
            "type_beneficiaire": patient.get_type_beneficiaire_display(),
            "date_naissance": patient.date_naissance.isoformat(),
            "deja_vu": patient.pk in patients_lies,
        }
        for patient in patients
    ]
    return JsonResponse({"resultats": resultats})


@role_required(User.Role.MEDECIN)
def fiche_patient_medecin(request, pk):
    """
    Fiche d'un patient, ouverte depuis la recherche rapide. L'historique
    n'expose que les consultations du medecin connecte avec ce patient
    (jamais celles d'un autre medecin - voir la spec, section "Decisions
    actees").
    """
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    patient = get_object_or_404(
        Patient.objects.select_related(
            "assure_principal", "assure_principal__plan_couverture", "plan_couverture"
        ),
        pk=pk,
    )
    historique = (
        Consultation.objects.filter(medecin=medecin, patient=patient)
        .select_related("service", "prise_en_charge")
        .prefetch_related("ordonnance_set")
        .order_by("-date_consultation")
    )
    if patient.type_beneficiaire == Patient.TypeBeneficiaire.PRINCIPAL:
        ayants_droit = patient.ayants_droit.all()
    else:
        ayants_droit = Patient.objects.none()

    prochains_rendez_vous = (
        RendezVous.objects.filter(medecin=medecin, patient=patient, date_heure__gte=timezone.now())
        .exclude(statut=RendezVous.Statut.ANNULE)
        .order_by("date_heure")
    )

    contexte = {
        "patient": patient,
        "historique": historique,
        "ayants_droit": ayants_droit,
        "prochains_rendez_vous": prochains_rendez_vous,
        "deja_vu": _patients_du_medecin(medecin).filter(pk=patient.pk).exists(),
    }
    return render(request, "fiche_patient_medecin.html", contexte)


@role_required(User.Role.MEDECIN)
def historique_consultations(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    consultations = Consultation.objects.filter(medecin=medecin).select_related(
        "patient", "service", "prise_en_charge"
    ).prefetch_related("ordonnance_set").order_by("-date_consultation")

    patient_id = request.GET.get("patient", "")
    if patient_id.isdigit():
        consultations = consultations.filter(patient_id=patient_id)

    date_filtre = request.GET.get("date", "")
    if date_filtre:
        try:
            date_valide = datetime.date.fromisoformat(date_filtre)
        except ValueError:
            date_filtre = ""
        else:
            consultations = consultations.filter(date_consultation__date=date_valide)

    contexte = {
        "consultations": _paginer(request, consultations),
        "patients_du_medecin": _patients_du_medecin(medecin),
        "patient_selectionne": patient_id,
        "date_selectionnee": date_filtre,
    }
    return render(request, "historique_consultations.html", contexte)


@role_required(User.Role.MEDECIN)
def ajouter_consultation_medecin(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    if request.method == "POST":
        form = ConsultationForm(request.POST)
        if form.is_valid():
            consultation = form.save(commit=False)
            consultation.medecin = medecin
            consultation.save()
            Paiement.calculer_pour(consultation).save()
            messages.success(request, "Consultation enregistrée.")
            return redirect("ajouter_ordonnance_medecin", consultation_pk=consultation.pk)
    else:
        patient_id = request.GET.get("patient", "")
        initial = {}
        if patient_id.isdecimal() and Patient.objects.filter(pk=patient_id).exists():
            initial["patient"] = patient_id
        form = ConsultationForm(initial=initial)
    return render(request, "ajouter_consultation_medecin.html", {"form": form})


@role_required(User.Role.MEDECIN)
def ajouter_ordonnance_medecin(request, consultation_pk):
    medecin = _medecin_courant(request)
    consultation = get_object_or_404(Consultation, pk=consultation_pk, medecin=medecin)

    if request.method == "POST":
        form = OrdonnanceForm(request.POST)
        if form.is_valid():
            ordonnance = form.save(commit=False)
            ordonnance.consultation = consultation
            ordonnance.save()
            messages.success(request, "Ordonnance créée.", extra_tags="succes-critique")
            return redirect("voir_ordonnance_medecin", pk=ordonnance.pk)
    else:
        form = OrdonnanceForm()
    return render(
        request,
        "ajouter_ordonnance_medecin.html",
        {"form": form, "consultation": consultation},
    )


@role_required(User.Role.MEDECIN)
def voir_ordonnance_medecin(request, pk):
    medecin = _medecin_courant(request)
    ordonnance = get_object_or_404(Ordonnance, pk=pk, consultation__medecin=medecin)
    return render(
        request,
        "voir_ordonnance.html",
        {"ordonnance": ordonnance, "retour_url": "historique_consultations"},
    )


@role_required(User.Role.MEDECIN)
def modifier_profil_medecin(request):
    medecin = _medecin_courant(request)
    if medecin is None:
        return render(request, "medecin_fiche_manquante.html")

    if request.method == "POST":
        form = MedecinProfilForm(request.POST, instance=medecin)
        if form.is_valid():
            form.save()
            messages.success(request, "Profil mis à jour.")
            return redirect("modifier_profil_medecin")
    else:
        form = MedecinProfilForm(instance=medecin)
    return render(request, "modifier_profil_medecin.html", {"form": form, "medecin": medecin})


# ---------------------------------------------------------------------------
# Espace Pharmacien
# ---------------------------------------------------------------------------

def _pharmacien_courant(request):
    return getattr(request.user, "pharmacien", None)


@role_required(User.Role.PHARMACIEN)
def dashboard_pharmacien(request):
    pharmacien = _pharmacien_courant(request)
    if pharmacien is None:
        return render(request, "pharmacien_fiche_manquante.html")

    delivrances = Delivrance.objects.filter(pharmacien=pharmacien)
    aujourd_hui = timezone.localdate()
    debut_semaine = aujourd_hui - datetime.timedelta(days=aujourd_hui.weekday())
    contexte = {
        "total_delivrances": delivrances.count(),
        "delivrances_aujourd_hui": delivrances.filter(date_delivrance__date=aujourd_hui).count(),
        "delivrances_semaine": delivrances.filter(date_delivrance__date__gte=debut_semaine).count(),
        "patients_servis": delivrances.values(
            "ordonnance__consultation__patient"
        ).distinct().count(),
        "dernieres_delivrances": delivrances.select_related(
            "ordonnance__consultation__patient"
        ).order_by("-date_delivrance")[:5],
        "pharmacien": pharmacien,
    }
    return render(request, "dashboard_pharmacien.html", contexte)


@role_required(User.Role.PHARMACIEN)
def scanner_ordonnance(request):
    """Comptoir du pharmacien : verification d'une ordonnance avant delivrance.

    DEUX chemins, volontairement dissymetriques :

    1. Le code (scanne ou saisi) fait une correspondance EXACTE. Un code
       identifie une ordonnance et une seule : on peut donc l'ouvrir
       directement. C'est le chemin normal, inchange.

    2. La recherche manuelle (nom du patient ou fragment de code) est le
       repli quand le QR est illisible, l'impression pale ou le code mal
       recopie. Elle ne selectionne JAMAIS d'ordonnance, meme s'il n'y a
       qu'un seul resultat : elle affiche une liste et le pharmacien
       designe explicitement la bonne. Delivrer le mauvais traitement
       parce qu'un logiciel a "devine" est un risque qu'on n'accepte pas.

    Le bouton de selection d'un resultat renvoie simplement le code exact
    dans le chemin 1 : une seule logique d'ouverture, donc une seule
    surface a securiser.
    """
    pharmacien = _pharmacien_courant(request)
    if pharmacien is None:
        return render(request, "pharmacien_fiche_manquante.html")

    ordonnance = None
    resultats = None
    recherche = ""
    trop_de_resultats = False

    if request.method == "POST":
        code = request.POST.get("code_qr", "").strip().upper()
        recherche = request.POST.get("recherche", "").strip()

        if code:
            try:
                ordonnance = Ordonnance.objects.select_related(
                    "consultation__patient", "consultation__medecin", "delivrance"
                ).get(code_qr=code)
            except Ordonnance.DoesNotExist:
                messages.error(request, "Aucune ordonnance ne correspond à ce code.")

        elif recherche:
            # Longueur minimale : une recherche d'un caractere listerait une
            # bonne partie des patients de la plateforme. Ce sont des donnees
            # medicales, on ne les enumere pas.
            if len(recherche) < RECHERCHE_ORDONNANCE_MIN:
                messages.error(
                    request,
                    f"Saisissez au moins {RECHERCHE_ORDONNANCE_MIN} caractères "
                    "pour lancer une recherche.",
                )
            else:
                trouvees = list(
                    Ordonnance.objects.select_related(
                        "consultation__patient", "consultation__medecin", "delivrance"
                    )
                    .filter(
                        Q(consultation__patient__nom__icontains=recherche)
                        | Q(consultation__patient__prenom__icontains=recherche)
                        | Q(code_qr__icontains=recherche)
                    )
                    .order_by("-date_creation")[: RECHERCHE_ORDONNANCE_MAX + 1]
                )
                # On demande un element de plus que la limite : sa presence
                # signale qu'il y en avait davantage, sans second COUNT.
                trop_de_resultats = len(trouvees) > RECHERCHE_ORDONNANCE_MAX
                resultats = trouvees[:RECHERCHE_ORDONNANCE_MAX]
                if not resultats:
                    messages.error(
                        request,
                        "Aucune ordonnance ne correspond à cette recherche.",
                    )
                elif trop_de_resultats:
                    messages.warning(
                        request,
                        f"Plus de {RECHERCHE_ORDONNANCE_MAX} ordonnances correspondent. "
                        "Précisez le nom du patient.",
                    )

    return render(
        request,
        "scanner_ordonnance.html",
        {
            "ordonnance": ordonnance,
            "resultats": resultats,
            "recherche": recherche,
            "trop_de_resultats": trop_de_resultats,
        },
    )


@role_required(User.Role.PHARMACIEN)
@require_POST
def valider_delivrance(request, pk):
    pharmacien = _pharmacien_courant(request)
    if pharmacien is None:
        return render(request, "pharmacien_fiche_manquante.html")
    code_qr = request.POST.get("code_qr", "").strip().upper()
    ordonnance = get_object_or_404(Ordonnance, pk=pk, code_qr=code_qr)
    if hasattr(ordonnance, "delivrance"):
        messages.error(request, "Cette ordonnance a déjà été délivrée.")
    else:
        Delivrance.objects.create(ordonnance=ordonnance, pharmacien=pharmacien)
        messages.success(request, "Délivrance validée.", extra_tags="succes-critique")
    return redirect("historique_delivrances")


@role_required(User.Role.PHARMACIEN)
def historique_delivrances(request):
    pharmacien = _pharmacien_courant(request)
    if pharmacien is None:
        return render(request, "pharmacien_fiche_manquante.html")

    delivrances = Delivrance.objects.filter(pharmacien=pharmacien).select_related(
        "ordonnance__consultation__patient", "ordonnance__consultation__medecin"
    ).order_by("-date_delivrance")
    return render(request, "historique_delivrances.html",
                  {"delivrances": _paginer(request, delivrances)})


# ---------------------------------------------------------------------------
# Espace Assure
# ---------------------------------------------------------------------------

def _patient_principal(request):
    return getattr(request.user, "patient", None)


def _beneficiaires(patient):
    return Patient.objects.filter(
        Q(pk=patient.pk) | Q(assure_principal=patient)
    ).order_by("nom", "prenom")


@role_required(User.Role.ASSURE)
def prestataires_proches(request):
    prestataires_partenaires = Prestataire.objects.filter(partenaire=True)
    avec_coordonnees = prestataires_partenaires.filter(
        latitude__isnull=False, longitude__isnull=False
    )
    sans_coordonnees = list(
        prestataires_partenaires.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        ).order_by("ville", "nom")
    )
    for prestataire in sans_coordonnees:
        morceaux = [valeur for valeur in (prestataire.nom, prestataire.ville) if valeur]
        prestataire.lien_itineraire = (
            "https://www.google.com/maps/dir/?api=1&destination="
            + urllib.parse.quote(", ".join(morceaux) + ", Senegal")
        )

    lat_param = request.GET.get("lat")
    lng_param = request.GET.get("lng")
    lat_utilisateur = lng_utilisateur = None
    if lat_param and lng_param:
        try:
            lat_utilisateur = float(lat_param)
            lng_utilisateur = float(lng_param)
        except ValueError:
            lat_utilisateur = lng_utilisateur = None
    localisation_active = lat_utilisateur is not None and lng_utilisateur is not None

    if localisation_active:
        prestataires_tries = sorted(
            (
                (prestataire, round(distance_km(
                    lat_utilisateur, lng_utilisateur,
                    float(prestataire.latitude), float(prestataire.longitude),
                ), 1))
                for prestataire in avec_coordonnees
            ),
            key=lambda item: item[1],
        )
    else:
        prestataires_tries = [
            (prestataire, None)
            for prestataire in avec_coordonnees.order_by("ville", "nom")
        ]

    prestataires_geojson = [
        {
            "pk": prestataire.pk,
            "nom": prestataire.nom,
            "type": prestataire.get_type_prestataire_display(),
            "type_code": prestataire.type_prestataire,
            "ville": prestataire.ville,
            "telephone": prestataire.telephone,
            "latitude": float(prestataire.latitude),
            "longitude": float(prestataire.longitude),
            "medecin_count": prestataire.medecins.count(),
        }
        for prestataire in avec_coordonnees
    ]

    return render(request, "prestataires_proches.html", {
        "prestataires_tries": prestataires_tries,
        "prestataires_sans_coordonnees": sans_coordonnees,
        "prestataires_geojson": prestataires_geojson,
        "localisation_active": localisation_active,
    })


@role_required(User.Role.ASSURE)
def dashboard_assure(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")

    beneficiaires = _beneficiaires(patient)
    maintenant = timezone.now()
    rendez_vous_a_venir = RendezVous.objects.filter(
        patient__in=beneficiaires, date_heure__gte=maintenant
    ).exclude(statut=RendezVous.Statut.ANNULE)

    contexte = {
        "patient": patient,
        "total_ayants_droit": beneficiaires.exclude(pk=patient.pk).count(),
        "total_rendez_vous_a_venir": rendez_vous_a_venir.count(),
        "total_ordonnances": Ordonnance.objects.filter(consultation__patient__in=beneficiaires).count(),
        "prochains_rendez_vous": rendez_vous_a_venir.select_related(
            "medecin", "prestataire", "patient"
        ).order_by("date_heure")[:5],
    }
    return render(request, "dashboard_assure.html", contexte)


@role_required(User.Role.ASSURE)
def mon_profil_assure(request):
    patient = _patient_principal(request)

    if request.method == "POST":
        form = ProfilAssureForm(request.POST, instance=patient)
        if form.is_valid():
            profil = form.save(commit=False)
            profil.user = request.user
            profil.type_beneficiaire = Patient.TypeBeneficiaire.PRINCIPAL
            profil.save()
            messages.success(request, "Profil enregistré.")
            return redirect("dashboard_assure")
    else:
        initial = {}
        if patient is None:
            initial = {"nom": request.user.last_name, "prenom": request.user.first_name}
        form = ProfilAssureForm(instance=patient, initial=initial)

    return render(request, "mon_profil_assure.html", {"form": form, "patient": patient})


@role_required(User.Role.ASSURE)
def liste_ayants_droit(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    return render(
        request,
        "liste_ayants_droit.html",
        {"ayants_droit": patient.ayants_droit.all().order_by("nom", "prenom")},
    )


@role_required(User.Role.ASSURE)
def ajouter_ayant_droit(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")

    if request.method == "POST":
        form = AyantDroitForm(request.POST)
        if form.is_valid():
            ayant_droit = form.save(commit=False)
            ayant_droit.type_beneficiaire = Patient.TypeBeneficiaire.AYANT_DROIT
            ayant_droit.assure_principal = patient
            ayant_droit.save()
            messages.success(request, "Ayant droit ajouté.")
            return redirect("liste_ayants_droit")
    else:
        form = AyantDroitForm()
    return render(request, "ajouter_ayant_droit.html", {"form": form})


@role_required(User.Role.ASSURE)
def modifier_ayant_droit(request, pk):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    ayant_droit = get_object_or_404(Patient, pk=pk, assure_principal=patient)

    if request.method == "POST":
        form = AyantDroitForm(request.POST, instance=ayant_droit)
        if form.is_valid():
            form.save()
            messages.success(request, "Ayant droit modifié.")
            return redirect("liste_ayants_droit")
    else:
        form = AyantDroitForm(instance=ayant_droit)
    return render(request, "modifier_ayant_droit.html", {"form": form, "ayant_droit": ayant_droit})


@role_required(User.Role.ASSURE)
def supprimer_ayant_droit(request, pk):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    ayant_droit = get_object_or_404(Patient, pk=pk, assure_principal=patient)

    if request.method == "POST":
        ayant_droit.delete()
        messages.success(request, "Ayant droit supprimé.")
        return redirect("liste_ayants_droit")
    return render(
        request,
        "confirmer_suppression.html",
        {"objet": ayant_droit, "type": "Ayant droit"},
    )


@role_required(User.Role.ASSURE)
def mes_rendez_vous_assure(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    beneficiaires = _beneficiaires(patient)
    # order_by explicite : paginer un queryset non ordonne rend l'ordre des
    # pages instable (et Django emet UnorderedObjectListWarning).
    rendez_vous = RendezVous.objects.filter(patient__in=beneficiaires).select_related(
        "patient", "medecin", "prestataire"
    ).order_by("-date_heure")
    return render(request, "mes_rendez_vous.html",
                  {"rendez_vous": _paginer(request, rendez_vous)})


@role_required(User.Role.ASSURE)
def ajouter_rendez_vous_assure(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    beneficiaires = _beneficiaires(patient)

    if request.method == "POST":
        form = RendezVousAssureForm(request.POST, beneficiaires=beneficiaires)
        if form.is_valid():
            form.save()
            messages.success(request, "Demande de rendez-vous envoyée.")
            return redirect("mes_rendez_vous_assure")
    else:
        initial = {}
        prestataire_id = request.GET.get("prestataire")
        if prestataire_id and prestataire_id.isdigit():
            if Prestataire.objects.filter(pk=prestataire_id, partenaire=True).exists():
                initial["prestataire"] = prestataire_id
        form = RendezVousAssureForm(beneficiaires=beneficiaires, initial=initial)
    return render(request, "ajouter_rendez_vous_assure.html", {"form": form})


@role_required(User.Role.ASSURE)
@require_POST
def annuler_rendez_vous_assure(request, pk):
    patient = _patient_principal(request)
    beneficiaires = _beneficiaires(patient) if patient else Patient.objects.none()
    rendez_vous = get_object_or_404(RendezVous, pk=pk, patient__in=beneficiaires)

    if rendez_vous.statut in (RendezVous.Statut.DEMANDE, RendezVous.Statut.CONFIRME):
        rendez_vous.statut = RendezVous.Statut.ANNULE
        rendez_vous.save(update_fields=["statut"])
        messages.success(request, "Rendez-vous annulé.")
    else:
        messages.error(request, "Ce rendez-vous ne peut plus être annulé.")
    return redirect("mes_rendez_vous_assure")


@role_required(User.Role.ASSURE)
def mes_ordonnances_assure(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    beneficiaires = _beneficiaires(patient)
    ordonnances = Ordonnance.objects.filter(
        consultation__patient__in=beneficiaires
    ).select_related("consultation__patient", "consultation__medecin").order_by("-date_creation")
    return render(request, "mes_ordonnances.html",
                  {"ordonnances": _paginer(request, ordonnances)})


@role_required(User.Role.ASSURE)
def voir_ordonnance_assure(request, pk):
    patient = _patient_principal(request)
    beneficiaires = _beneficiaires(patient) if patient else Patient.objects.none()
    ordonnance = get_object_or_404(Ordonnance, pk=pk, consultation__patient__in=beneficiaires)
    return render(
        request,
        "voir_ordonnance.html",
        {"ordonnance": ordonnance, "retour_url": "mes_ordonnances_assure"},
    )


@role_required(User.Role.ASSURE)
def mon_historique_assure(request):
    patient = _patient_principal(request)
    if patient is None:
        return redirect("mon_profil_assure")
    beneficiaires = _beneficiaires(patient)
    consultations = Consultation.objects.filter(patient__in=beneficiaires).select_related(
        "patient", "medecin", "service", "paiement"
    ).order_by("-date_consultation")
    return render(request, "mon_historique.html",
                  {"consultations": _paginer(request, consultations)})


# ---------------------------------------------------------------------------
# Notifications
# ---------------------------------------------------------------------------

@admin_required
def envoyer_notification(request):
    if request.method == "POST":
        form = EnvoyerNotificationForm(request.POST)
        if form.is_valid():
            message = form.cleaned_data["message"]
            destinataire = form.cleaned_data["destinataire"]
            role = form.cleaned_data["role"]

            if destinataire:
                destinataires = [destinataire]
            else:
                destinataires = list(User.objects.filter(role=role, is_active=True))

            Notification.objects.bulk_create(
                [Notification(destinataire=u, message=message) for u in destinataires]
            )
            messages.success(request, f"Notification envoyée à {len(destinataires)} utilisateur(s).")
            return redirect("liste_notifications_envoyees")
    else:
        form = EnvoyerNotificationForm()
    return render(request, "envoyer_notification.html", {"form": form})


@admin_required
def liste_notifications_envoyees(request):
    notifications = Notification.objects.select_related("destinataire").all()

    lue = request.GET.get("lue", "")
    if lue == "oui":
        notifications = notifications.filter(lue=True)
    elif lue == "non":
        notifications = notifications.filter(lue=False)

    recherche = request.GET.get("q", "").strip()
    if recherche:
        notifications = notifications.filter(
            Q(message__icontains=recherche)
            | Q(destinataire__email__icontains=recherche)
            | Q(destinataire__first_name__icontains=recherche)
            | Q(destinataire__last_name__icontains=recherche)
        )

    # Le plafond des 200 plus recentes s'applique une fois les filtres
    # appliques (sinon une recherche pourrait manquer une notification plus
    # ancienne que les 200 dernieres notifications tous destinataires confondus).
    notifications = notifications[:200]

    contexte = {
        "notifications": _paginer(request, notifications),
        "lue_choisie": lue,
        "recherche": recherche,
    }
    return render(request, "liste_notifications_envoyees.html", contexte)


@login_required
def mes_notifications(request):
    notifications = request.user.notifications.all()

    lue = request.GET.get("lue", "")
    if lue == "oui":
        notifications = notifications.filter(lue=True)
    elif lue == "non":
        notifications = notifications.filter(lue=False)

    contexte = {
        "notifications": _paginer(request, notifications),
        "lue_choisie": lue,
    }
    return render(request, "mes_notifications.html", contexte)


@login_required
@require_POST
def marquer_notification_lue(request, pk):
    notification = get_object_or_404(Notification, pk=pk, destinataire=request.user)
    notification.lue = True
    notification.save(update_fields=["lue"])
    return redirect("mes_notifications")


# ---------------------------------------------------------------------------
# Import des reglements de paiement
# ---------------------------------------------------------------------------

# Un Paiement ne peut PAS etre cree par import : il est en relation 1-1 avec
# une Consultation et tous ses montants sont derives (Paiement.calculer_pour --
# prix du service, taux du plan si la prise en charge est validee). Importer
# des montants reviendrait a ecraser cette regle metier.
# Ce que l'import fait, et c'est le besoin reel : enregistrer EN MASSE le
# reglement de paiements existants (rapprochement de fin de mois). C'est le
# pendant en masse de marquer_paiement_regle.
COLONNES_IMPORT_REGLEMENTS = [
    "Reference", "Patient", "Date de consultation", "Part patient",
    "Mode de reglement", "Date de reglement",
]

_MODES_PAR_LIBELLE_IMPORT = {}
for _valeur_mode, _label_mode in Paiement.ModeReglement.choices:
    _MODES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_valeur_mode)] = _valeur_mode
    _MODES_PAR_LIBELLE_IMPORT[_normaliser_texte_import(_label_mode)] = _valeur_mode


def _lire_date_import(valeur):
    """Accepte une date Excel native ou un texte JJ/MM/AAAA. Renvoie None si
    la valeur est inexploitable -- l'appelant produit alors l'erreur de ligne."""
    if valeur in (None, ""):
        return None
    if isinstance(valeur, datetime.datetime):
        return valeur.date()
    if isinstance(valeur, datetime.date):
        return valeur
    texte = str(valeur).strip()
    for format_date in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(texte, format_date).date()
        except ValueError:
            continue
    return None


def _analyser_ligne_import_reglement(numero, valeurs, references_vues):
    """Valide une ligne du fichier de reglements.

    Renvoie (donnees, erreurs). `donnees` vaut None des qu'une erreur est
    trouvee : on ne construit jamais un reglement a moitie valide.
    """
    erreurs = []
    reference, _patient, _date_consultation, part_patient, mode, date_reglement = (
        list(valeurs) + [None] * 6
    )[:6]

    # --- Reference : seule colonne qui identifie le paiement ---
    texte_reference = "" if reference is None else str(reference).strip()
    if not texte_reference:
        erreurs.append(f"Ligne {numero} : la référence est obligatoire.")
        return None, erreurs
    try:
        identifiant = int(float(texte_reference))
    except (TypeError, ValueError):
        erreurs.append(f"Ligne {numero} : référence « {texte_reference} » invalide.")
        return None, erreurs

    if identifiant in references_vues:
        erreurs.append(
            f"Ligne {numero} : la référence {identifiant} apparaît déjà "
            f"ligne {references_vues[identifiant]}."
        )
        return None, erreurs

    paiement = Paiement.objects.filter(pk=identifiant).select_related(
        "consultation__patient"
    ).first()
    if paiement is None:
        erreurs.append(f"Ligne {numero} : aucun paiement ne porte la référence {identifiant}.")
        return None, erreurs
    if paiement.statut == Paiement.Statut.REGLE:
        erreurs.append(
            f"Ligne {numero} : le paiement {identifiant} "
            f"({paiement.consultation.patient}) est déjà réglé."
        )
        return None, erreurs

    # --- Part patient : colonne de CONTROLE, facultative mais verifiee ---
    if part_patient not in (None, ""):
        texte_montant = str(part_patient).strip().replace(" ", "").replace(",", ".")
        try:
            montant = Decimal(texte_montant)
        except (InvalidOperation, ValueError):
            erreurs.append(f"Ligne {numero} : montant « {part_patient} » invalide.")
        else:
            if montant != paiement.montant_part_patient:
                erreurs.append(
                    f"Ligne {numero} : la part patient indiquée ({montant}) ne correspond "
                    f"pas au paiement {identifiant} ({paiement.montant_part_patient})."
                )

    # --- Mode de reglement : obligatoire, comme dans la saisie manuelle ---
    cle_mode = _normaliser_texte_import(mode)
    if not cle_mode:
        erreurs.append(f"Ligne {numero} : le mode de règlement est obligatoire.")
        mode_valide = None
    else:
        mode_valide = _MODES_PAR_LIBELLE_IMPORT.get(cle_mode)
        if mode_valide is None:
            attendus = ", ".join(label for _, label in Paiement.ModeReglement.choices)
            erreurs.append(
                f"Ligne {numero} : mode de règlement « {mode} » inconnu. Attendu : {attendus}."
            )

    # --- Date de reglement : obligatoire, jamais dans le futur ---
    date_valide = _lire_date_import(date_reglement)
    if date_reglement in (None, ""):
        erreurs.append(f"Ligne {numero} : la date de règlement est obligatoire.")
    elif date_valide is None:
        erreurs.append(
            f"Ligne {numero} : date de règlement « {date_reglement} » invalide (JJ/MM/AAAA)."
        )
    elif date_valide > timezone.localdate():
        erreurs.append(f"Ligne {numero} : la date de règlement ne peut pas être dans le futur.")

    if erreurs:
        return None, erreurs

    references_vues[identifiant] = numero
    return {"paiement": paiement, "mode": mode_valide, "date": date_valide}, []


@admin_required
def telecharger_modele_import_reglements(request):
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "Import reglements"
    feuille.append(COLONNES_IMPORT_REGLEMENTS)
    feuille.append([
        "Reference du paiement (colonne de l'export CSV)",
        "Nom du patient (indicatif, non utilise)",
        "Date de la consultation (indicatif, non utilise)",
        "Part patient attendue (controle, optionnel)",
        " / ".join(label for _, label in Paiement.ModeReglement.choices),
        "JJ/MM/AAAA",
    ])
    for index, nom_colonne in enumerate(COLONNES_IMPORT_REGLEMENTS, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(len(nom_colonne), 22)

    reponse = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    reponse["Content-Disposition"] = 'attachment; filename="modele_import_reglements.xlsx"'
    classeur.save(reponse)
    return reponse


@admin_required
def importer_reglements_excel(request):
    """Enregistrement en masse du reglement de paiements existants.

    Meme regle "tout ou rien" que l'import des comptes : une seule ligne
    invalide annule l'ensemble. Sur des ecritures financieres, un import
    partiel laisserait une caisse impossible a rapprocher.
    """
    erreurs = []
    if request.method == "POST":
        fichier = request.FILES.get("fichier")
        if not fichier:
            erreurs.append("Choisissez un fichier Excel (.xlsx) à importer.")
        else:
            try:
                classeur = openpyxl.load_workbook(fichier, data_only=True)
            except Exception:
                erreurs.append(
                    "Fichier illisible : vérifiez qu'il s'agit bien d'un fichier Excel (.xlsx) valide."
                )
            else:
                feuille = classeur.active
                entetes = next(feuille.iter_rows(min_row=1, max_row=1, values_only=True), ())
                entetes_normalisees = [_normaliser_texte_import(e) for e in entetes]
                attendues = [_normaliser_texte_import(c) for c in COLONNES_IMPORT_REGLEMENTS]

                if entetes_normalisees[: len(attendues)] != attendues:
                    erreurs.append(
                        "En-têtes de colonnes invalides : utilisez le modèle téléchargeable ci-dessous."
                    )
                else:
                    lignes = [
                        (numero, valeurs)
                        for numero, valeurs in enumerate(
                            feuille.iter_rows(min_row=2, values_only=True), start=2
                        )
                        if valeurs and not all(v in (None, "") for v in valeurs)
                    ]
                    if not lignes:
                        erreurs.append("Le fichier ne contient aucune ligne à importer.")
                    else:
                        references_vues = {}
                        a_regler = []
                        for numero, valeurs in lignes:
                            donnees, erreurs_ligne = _analyser_ligne_import_reglement(
                                numero, valeurs, references_vues
                            )
                            erreurs.extend(erreurs_ligne)
                            if donnees is not None:
                                a_regler.append(donnees)

                        if not erreurs:
                            with transaction.atomic():
                                for donnees in a_regler:
                                    paiement = donnees["paiement"]
                                    paiement.statut = Paiement.Statut.REGLE
                                    paiement.mode_reglement = donnees["mode"]
                                    # On conserve la date du releve, pas celle
                                    # de l'import : c'est elle qui fait foi.
                                    paiement.date_reglement = timezone.make_aware(
                                        datetime.datetime.combine(
                                            donnees["date"], datetime.time()
                                        )
                                    )
                                    paiement.save(update_fields=[
                                        "statut", "mode_reglement", "date_reglement",
                                    ])
                            messages.success(
                                request,
                                f"{len(a_regler)} règlement(s) enregistré(s).",
                                extra_tags="succes-critique",
                            )
                            return redirect("liste_paiements")

    return render(request, "importer_reglements.html", {"erreurs": erreurs})
